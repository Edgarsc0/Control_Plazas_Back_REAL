import datetime
import json
import re
import logging
from pathlib import Path
from zoneinfo import ZoneInfo

from django.conf import settings
from django.core import exceptions
from django.db.models import (
    Aggregate,
    Avg,
    Case,
    CharField,
    Count,
    DateField,
    DateTimeField,
    ExpressionWrapper,
    F,
    IntegerField,
    Q,
    Sum,
    Value,
    When,
)
from django.db import transaction
from django.db.models.functions import Cast, Coalesce, Concat, ExtractHour, Substr, Trim, TruncDate
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    AlineacionOrganizacionalHistorico,
    AnuenciaAnexo,
    AnuenciaAnexo3Version,
    AnuenciaAnexoCambio,
    AnuenciaJustificacionCatalogo,
    CatAcciones,
    CatAccionesMotivos,
    CatCorreccionPosicion,
    COLUMNAS_CORRECCION_VALIDAS,
    CatNivelJerarquicoPlaza,
    CatPtoFunc,
    COLUMNAS_QUINCENAL_VALIDAS,
    COLUMNAS_SOLICITUD_VACANTE,
    CuadroVacancia,
    DatosPersonales,
    DESCRIPCION_NJ_CHOICES,
    EmpleadosCompletosSig,
    FiltroGuardado,
    MovPos,
    MovPosLatest,
    NIVEL_3_PREFIJO_TITULAR_ADUANA,
    NIVEL_JERARQUICO_LABELS,
    NIVEL_JERARQUICO_LABELS_NIVEL_3_TITULAR_ADUANA,
    NIVELES_JERARQUICOS,
    NivelJerarquicoPrioridadConfig,
    OrganigramaAnam,
    Plantilla1800Plazas,
    REPORTADA_CHOICES,
    RcCatCodPresupuestal,
    SuscripcionNotificacionPosicion,
    TblColumnasPlantillaQuincenal,
)
from .celda_override import (
    TABLA_EMPLEADOS,
    TABLA_MOV_POS,
    borrar_contenido_celda,
    borrar_override_fecha_anuencia,
    borrar_override_quincenal,
    get_fecha_anuencia_baseline_map,
    get_fecha_anuencia_baseline_texto_map,
    get_fecha_anuencia_overrides_map,
    get_fecha_anuencia_overrides_texto_map,
    get_quincenal_overrides_map,
    notificar_cambio_celda,
    obtener_estadisticas_overrides,
    obtener_historial_overrides,
    registrar_override_fecha_anuencia,
    registrar_y_aplicar_override_datos_personales,
    registrar_y_aplicar_override_empleado,
    registrar_y_aplicar_override_quincenal,
    serializar_override,
    TABLA_QUINCENAL,
)
from .nivel_jerarquico_sync import aplicar_prioridad_nivel_jerarquico
from .serializers import (
    AnuenciaAnexo3VersionDetailSerializer,
    AnuenciaAnexo3VersionListSerializer,
    AnuenciaAnexoCambioSerializer,
    AnuenciaAnexoDetailSerializer,
    AnuenciaAnexoListSerializer,
    AnuenciaJustificacionCatalogoSerializer,
    CatAccionesMotivosSerializer,
    CatAccionesSerializer,
    CatNivelJerarquicoPlazaSerializer,
    CatPtoFuncSerializer,
    OrganigramaAnamSerializer,
    RcCatCodPresupuestalSerializer,
)

logger = logging.getLogger(__name__)


class GroupConcat(Aggregate):
    """Agregado GROUP_CONCAT(DISTINCT ...) para MySQL (no incluido en Django).

    Separador 0x1f (unit separator) para no chocar con comas en los valores.
    """

    function = "GROUP_CONCAT"
    template = "%(function)s(DISTINCT %(expressions)s SEPARATOR 0x1f)"
    output_field = CharField()


# Sentinel que el frontend manda en vez de "" para seleccionar "(Vacío)" en un
# filtro de columna: una cadena vacía real en la URL se descarta antes de
# llegar aquí (buildQuery del front, y el chequeo `not val` de abajo).
EMPTY_VALUE_TOKEN = "__EMPTY__"

# Parámetros de control que NO son columnas filtrables (paginación, orden, etc.).
FILTER_SKIP_PARAMS = frozenset(
    {
        "distinct_field",
        "distinct_search",
        "page",
        "page_size",
        "search",
        "is_latest",
        "no_pagination",
        "sort_by",
        "sort_order",
        "oficio",
        "nivel",
        "advanced_filters",
    }
)


def apply_text_search(queryset, query, fields):
    """Filtro de búsqueda libre: OR de ``icontains`` sobre ``fields``."""
    query = (query or "").strip()
    if not query:
        return queryset
    q = Q()
    for field in fields:
        q |= Q(**{f"{field}__icontains": query})
    return queryset.filter(q)


def annotate_fecha_anuencia(queryset, source_field="fecha_vacancia", overrides=None, baseline=None):
    """Anota ``fecha_anuencia`` = ``source_field`` + 30 días, calculada al
    vuelo en cada consulta (nunca persistida en BD, nunca toca el SP
    ``sp_obtener_todas_vacancias``): es una simple suma de días sobre un
    valor que ese SP ya calcula con una lógica mucho más compleja (recorre
    historial de movimientos, detecta insubsistencias, etc.) — duplicar esa
    lógica solo para sumarle 30 días sería reinventar algo que ya existe, y
    persistirla en una columna aparte obligaría a mantener dos columnas
    sincronizadas para siempre.

    ``source_field`` es un varchar con fechas ISO ('YYYY-MM-DD') o vacío/NULL
    (nunca se llena para posiciones ocupadas) — el regex en el ``When``
    descarta cualquier valor que no sea una fecha real (vacío, NULL, el
    sentinel histórico '1900-01-01' se deja pasar tal cual, igual que ya
    ocurre con ``fecha_vacancia``) devolviendo NULL en vez de reventar el
    CAST. Al ser una expresión SQL (no un post-proceso en Python), participa
    igual que un campo real en filtros/orden/`distinct_field` — mismo
    tratamiento que ya recibe `fecha_vacancia` (campo real del modelo).

    Prioridad de 3 niveles (de mayor a menor):
      1. ``overrides``: dict opcional ``{no_pos_actual: date}`` (ver
         ``celda_override.get_fecha_anuencia_overrides_map``) — edición
         manual del usuario, gana siempre.
      2. ``baseline``: dict opcional ``{no_pos_actual: date}`` (ver
         ``celda_override.get_fecha_anuencia_baseline_map``) — valor cargado
         del Excel de plantilla (columna AL) cuando trae una fecha real (la
         mayoría de posiciones no traen nada ahí y caen al nivel 3).
      3. ``computed_default``: fecha_vacancia + 30 días.
    Se resuelve con un ``Case`` adicional por nivel (una rama ``When`` por
    entrada activa, comparando contra el campo real e indexado
    ``no_pos_actual``) en vez de un ``Subquery`` correlacionado contra
    ``CeldaOverride``/``TblColumnasPlantillaQuincenal`` — más simple y más
    barato ya que en la práctica son pocas entradas activas a la vez.
    """
    from datetime import timedelta

    computed_default = Case(
        When(
            **{f"{source_field}__regex": r"^[0-9]{4}-[0-9]{2}-[0-9]{2}$"},
            then=ExpressionWrapper(
                Cast(source_field, output_field=DateField()) + timedelta(days=30),
                output_field=DateField(),
            ),
        ),
        default=Value(None),
        output_field=DateField(),
    )

    layered_default = computed_default
    if baseline:
        whens_baseline = [
            When(no_pos_actual=pos, then=Value(fecha, output_field=DateField()))
            for pos, fecha in baseline.items()
        ]
        layered_default = Case(*whens_baseline, default=computed_default, output_field=DateField())

    if not overrides:
        return queryset.annotate(fecha_anuencia=layered_default)

    whens = [
        When(no_pos_actual=pos, then=Value(fecha, output_field=DateField()))
        for pos, fecha in overrides.items()
    ]
    return queryset.annotate(
        fecha_anuencia=Case(*whens, default=layered_default, output_field=DateField())
    )


def corregir_fecha_anuencia_row(row, pos, posiciones_ocupadas, overrides=None, overrides_texto=None, baseline_texto=None):
    """Aplica a ``row["fecha_anuencia"]`` la MISMA regla que ya aplica
    ``fecha_vacancia`` en estas vistas (vacío si la posición está ocupada —
    el SP nunca limpia `FECHA VACANCIA` al volver a ocuparse, así que esta
    corrección en Python es la única barrera contra una fecha de vacancia/
    anuencia obsoleta) y la formatea como 'YYYY-MM-DD' (la anotación declara
    ``output_field=DateField()``, pero el driver de MySQL no siempre lo
    entrega como objeto `date`/`datetime` — para ~7.6% de las filas lo
    devuelve ya como string 'YYYY-MM-DD HH:MM:SS.ffffff'; como los strings no
    tienen `.strftime`, antes se colaban tal cual. El frontend calcula los
    días restantes parseando por '-' un valor que asume 'YYYY-MM-DD' — con
    el sufijo de hora, ese parseo fallaba silenciosamente, y como el
    semáforo de color Y el tooltip de días restantes dependen del mismo
    cálculo, ambos desaparecían juntos exactamente en esas filas. Cortar en
    el primer espacio o 'T' normaliza cualquier variante que llegue.

    También agrega ``row["fecha_anuencia_override"]`` (bool): si la posición
    tiene un override manual activo (ver ``overrides``, el mismo dict que ya
    se le pasó a ``annotate_fecha_anuencia``) — el frontend lo usa para
    resaltar en azul las fechas editadas manualmente. `False` si la posición
    está ocupada (el override deja de aplicar junto con la fecha misma).

    ``overrides_texto``/``baseline_texto`` (``{pos: str}``, ver
    ``celda_override.get_fecha_anuencia_overrides_texto_map``/
    ``get_fecha_anuencia_baseline_texto_map``): "Fecha de Anuencia" no
    siempre es una fecha real — el Excel de origen trae, para algunas
    posiciones, una de 4 categorías de texto fijo (p.ej. "Nueva Creación")
    que no puede representarse en la anotación SQL ``DateField`` de
    ``fecha_anuencia`` (por eso ``overrides``/el baseline SQL las omiten).
    Aquí, a nivel Python, tienen la prioridad final sobre el valor ya
    resuelto arriba (fecha real calculada u "").
    """
    if pos in posiciones_ocupadas:
        row["fecha_anuencia"] = ""
        row["fecha_anuencia_override"] = False
        return
    fa = row.get("fecha_anuencia")
    if hasattr(fa, "strftime"):
        row["fecha_anuencia"] = fa.strftime("%Y-%m-%d")
    elif fa:
        row["fecha_anuencia"] = str(fa).split(" ")[0].split("T")[0]
    else:
        row["fecha_anuencia"] = ""
    row["fecha_anuencia_override"] = bool(overrides) and pos in overrides

    if overrides_texto and pos in overrides_texto:
        row["fecha_anuencia"] = overrides_texto[pos]
        row["fecha_anuencia_override"] = True
    elif baseline_texto and pos in baseline_texto:
        row["fecha_anuencia"] = baseline_texto[pos]


def _annotate_full_name(queryset, alias="full_name", fields=("nombre", "ap_pat", "ap_mat")):
    """Anota ``alias`` = concatenación de ``fields`` unidos por espacio y recortados.

    El front unificó nombre + ap. paterno + ap. materno en una sola columna
    "Nombre Completo" (ver ``buildFullName`` en ``MovimientosPersonalTab.jsx``),
    pero el modelo no tiene ese campo combinado en BD. Este helper reproduce
    esa concatenación en el ORM para que filtros/sugerencias sobre esa columna
    unificada busquen en los tres campos reales a la vez.
    """
    if alias in queryset.query.annotations:
        return queryset
    parts = []
    for i, field in enumerate(fields):
        if i:
            parts.append(Value(" "))
        parts.append(Coalesce(Trim(field), Value("")))
    return queryset.annotate(**{alias: Trim(Concat(*parts, output_field=CharField()))})


def apply_dynamic_column_filters(
    queryset,
    request,
    model,
    skip_params=FILTER_SKIP_PARAMS,
    extra_valid_fields=None,
    full_name_column=None,
    full_name_fields=("nombre", "ap_pat", "ap_mat"),
    computed_field_resolver=None,
):
    """Aplica los filtros dinámicos de columna del frontend.

    Soporta ``?campo=val``, selección múltiple (``val1,val2`` -> ``__in``),
    sufijos explícitos (``__istartswith``/``__iexact``/...) y negación
    (``exclude__campo=val``). Las columnas de texto se filtran sobre
    ``Trim(campo)`` (anotando ``trimmed_<campo>``), que es lo que aprovechan los
    índices funcionales. Lógica única compartida por MovPosDetalleView y
    MovimientosPersonalListView (antes estaba duplicada).

    ``extra_valid_fields``: nombres de columnas calculadas (no campos reales
    del modelo, ej. ``fecha_anuencia``) que de todas formas deben aceptarse
    para filtrar — el caller ya se aseguró de anotarlas en el queryset antes
    de llamar esta función (ver ``annotate_fecha_anuencia``).

    ``full_name_column``: si se da, ese nombre de columna (no es un campo real
    del modelo) se trata como texto sobre la concatenación de
    ``full_name_fields`` (ver `_annotate_full_name`) en vez de exigir que sea
    un campo válido del modelo.

    ``computed_field_resolver(base_field, suffix, val_list, is_exclude) -> Q | None``:
    para columnas que sí pasaron ``extra_valid_fields`` pero cuyo valor no vive
    tal cual en la anotación del queryset (ej. ``fecha_anuencia``: es DateField
    a nivel SQL, pero puede mostrar 4 categorías de texto fijo resueltas sólo
    en Python — filtrar `fecha_anuencia__in=[...]` directo contra la anotación
    tiraba 500 en cuanto el valor no era una fecha real). Si el resolver
    devuelve un ``Q``, se aplica en vez de la lógica genérica de abajo (que
    asume que el campo es filtrable tal cual).
    """
    valid_fields = {f.name for f in model._meta.get_fields()} | set(extra_valid_fields or ())
    text_fields = {
        f.name
        for f in model._meta.get_fields()
        if f.get_internal_type() in ("CharField", "TextField")
    }
    skip = set(skip_params)

    for param, val in request.query_params.items():
        if param in skip:
            continue
        is_exclude = param.startswith("exclude__")
        actual_param = param[9:] if is_exclude else param
        base_field = actual_param.split("__")[0]
        is_full_name = full_name_column is not None and base_field == full_name_column
        if not is_full_name and (base_field not in valid_fields or not val):
            continue
        if is_full_name and not val:
            continue

        if computed_field_resolver is not None:
            suffix_for_resolver = actual_param.split("__", 1)[1] if "__" in actual_param else None
            val_list_for_resolver = [
                "" if v.strip() == EMPTY_VALUE_TOKEN else v.strip()
                for v in val.split(",")
                if v.strip()
            ]
            resolved_q = computed_field_resolver(base_field, suffix_for_resolver, val_list_for_resolver, is_exclude)
            if resolved_q is not None:
                queryset = queryset.exclude(resolved_q) if is_exclude else queryset.filter(resolved_q)
                continue

        is_text = True if is_full_name else base_field in text_fields
        if is_full_name:
            target_field = "full_name"
            queryset = _annotate_full_name(queryset, target_field, full_name_fields)
        else:
            target_field = f"trimmed_{base_field}" if is_text else base_field
            if is_text and target_field not in queryset.query.annotations:
                queryset = queryset.annotate(**{target_field: Trim(base_field)})

        if "__" in actual_param:
            suffix = actual_param.split("__", 1)[1]
            actual_param_target = f"{target_field}__{suffix}"
        else:
            suffix = None
            actual_param_target = target_field

        val_list = [
            "" if v.strip() == EMPTY_VALUE_TOKEN else v.strip()
            for v in val.split(",")
            if v.strip()
        ]
        apply = queryset.exclude if is_exclude else queryset.filter

        if suffix == "in" or (not suffix and len(val_list) > 1):
            if "" in val_list:
                if is_text:
                    # "" en trimmed_<campo> no matchea filas con NULL real en
                    # MySQL (Trim(NULL) es NULL); hay que cubrir ambos casos.
                    q = Q(**{f"{target_field}__in": val_list}) | Q(
                        **{f"{target_field}__isnull": True}
                    )
                else:
                    # target_field no es texto (ej. DateField como
                    # fecha_ocupacion): "" no es un valor válido para __in —
                    # Django intenta parsearlo como fecha y tira
                    # ValidationError (500 en el request), que el front
                    # traga en su .catch, así que "Aplicar Filtro" con sólo
                    # "(Vacío)" marcado parecía no hacer nada (bug real,
                    # mismo patrón que ya se cubrió para fecha_anuencia vía
                    # fecha_anuencia_column_resolver). El token vacío en un
                    # campo tipado sólo puede significar NULL.
                    non_empty = [v for v in val_list if v != ""]
                    q = Q(**{f"{target_field}__isnull": True})
                    if non_empty:
                        q |= Q(**{f"{target_field}__in": non_empty})
                queryset = queryset.exclude(q) if is_exclude else queryset.filter(q)
            else:
                queryset = apply(**{f"{target_field}__in": val_list})
        elif suffix:
            value = val_list[0] if len(val_list) == 1 else val_list
            queryset = apply(**{actual_param_target: value})
        elif is_text:
            queryset = apply(**{f"{target_field}__icontains": val_list[0]})
        else:
            queryset = apply(**{target_field: val_list[0]})

    return queryset


def apply_advanced_filters(
    queryset,
    request,
    model,
    computed_resolver=None,
    extra_valid_fields=None,
    full_name_column=None,
    full_name_fields=("nombre", "ap_pat", "ap_mat"),
):
    """Aplica las condiciones del modal "Filtros Avanzados" (``?advanced_filters=``).

    JSON array de nodos: condición ``{ column, condition, compareType, compareColumn, value, logic }``
    o grupo ``{ type: "group", logic, children: [condición, ...] }`` (paréntesis
    explícito — un solo nivel de anidamiento, un grupo no contiene otros
    grupos). ``logic`` en el item i combina (AND/OR) con el Q acumulado de los
    items 0..i-1 de esa misma lista (top-level o `children` de un grupo); un
    nodo sin ``type`` (filtros guardados viejos) se trata como condición.
    ``computed_resolver(column, condition, value) -> Q | None`` permite que el
    caller resuelva columnas calculadas que no son campos reales del modelo
    (p. ej. "ocupacion"/"total_movimientos" en MovPos). Lógica única compartida
    por MovPosDetalleView y MovimientosPersonalListView (antes solo vivía,
    inline, en MovPosDetalleView).

    ``extra_valid_fields``: nombres de columnas calculadas ya anotadas en el
    queryset (ej. ``fecha_anuencia``, ver ``annotate_fecha_anuencia``) que se
    tratan como campo real para condiciones simples (no de fecha con
    before/after — esas siguen resolviéndose vía ``computed_resolver`` si
    hace falta ese detalle).

    ``full_name_column``: si se da (p. ej. ``"nombre"`` en Movimientos), las
    condiciones de texto sobre esa columna buscan sobre la concatenación de
    ``full_name_fields`` en vez de sólo el campo real (ver `_annotate_full_name`).
    """
    advanced_filters_raw = request.query_params.get("advanced_filters", "").strip()
    if not advanced_filters_raw:
        return queryset

    try:
        advanced_conditions = json.loads(advanced_filters_raw)
    except (ValueError, TypeError):
        return queryset

    if not isinstance(advanced_conditions, list):
        return queryset
    advanced_conditions = advanced_conditions[:20]  # sanity cap

    valid_fields = {f.name for f in model._meta.get_fields()} | set(extra_valid_fields or ())
    text_fields = {
        f.name
        for f in model._meta.get_fields()
        if f.get_internal_type() in ("CharField", "TextField")
    }
    # Columnas de fecha "reales" del modelo (DateField/DateTimeField). Se
    # comparan por fecha truncada, no por texto, para que "igual"/"no igual"
    # funcionen sobre un DateTimeField sin que la hora embebida tumbe el match
    # (BUG-F04).
    date_fields = {
        f.name
        for f in model._meta.get_fields()
        if f.get_internal_type() in ("DateField", "DateTimeField")
    }
    # Campos numéricos "reales" del modelo — la mayoría de las tablas
    # importadas (MovPos, etc.) guardan hasta los montos como CharField, así
    # que las condiciones >, <, >=, <= solo se ofrecen (front) y aplican
    # (aquí) donde el dato de verdad está tipado como número; el resto sigue
    # comparándose como texto.
    numeric_fields = {
        f.name
        for f in model._meta.get_fields()
        if f.get_internal_type() in (
            "IntegerField", "SmallIntegerField", "BigIntegerField",
            "PositiveIntegerField", "PositiveSmallIntegerField", "PositiveBigIntegerField",
            "FloatField", "DecimalField",
        )
    }
    datetime_fields = {
        f.name
        for f in model._meta.get_fields()
        if f.get_internal_type() == "DateTimeField"
    }
    # CharField que en realidad guarda fecha+hora como texto plano
    # (`YYYY-MM-DD-HH.MM.SS.ffffff`, p. ej. `MovPosBase.fh_ult_actz`), a
    # diferencia del resto de columnas de fecha del mismo modelo que son
    # `YYYY-MM-DD` puro. Se compara por los primeros 10 caracteres, no por el
    # texto completo (BUG-F03).
    TEXT_DATETIME_FIELDS = {"fh_ult_actz"}

    date_lookup_by_condition = {
        "before": "lt",
        "after": "gt",
        "before_or_equal": "lte",
        "after_or_equal": "gte",
    }
    number_lookup_by_condition = {
        "greater_than": "gt",
        "less_than": "lt",
        "greater_or_equal": "gte",
        "less_or_equal": "lte",
    }
    # Unifica fecha+número para compareType=='campo': ahí el lookup Django
    # (lt/gt/lte/gte) es el mismo sin importar el tipo de campo, así que da
    # igual si el front mandó una u otra familia de claves de condición.
    ORDERING_LOOKUP_BY_CONDITION = {**date_lookup_by_condition, **number_lookup_by_condition}
    text_lookup_by_condition = {
        "contains": ("icontains", False),
        "not_contains": ("icontains", True),
        "starts_with": ("istartswith", False),
        "not_starts_with": ("istartswith", True),
        "ends_with": ("iendswith", False),
        "not_ends_with": ("iendswith", True),
        "equals": ("iexact", False),
        "not_equals": ("iexact", True),
    }

    def resolve_target_field(field_name):
        nonlocal queryset
        if field_name in text_fields:
            target = f"trimmed_{field_name}"
            if target not in queryset.query.annotations:
                queryset = queryset.annotate(**{target: Trim(field_name)})
            return target
        return field_name

    def resolve_date_field(field_name):
        """Campo anotado apto para comparar contra un valor `YYYY-MM-DD`."""
        nonlocal queryset
        if field_name in TEXT_DATETIME_FIELDS:
            target = f"date10_{field_name}"
            if target not in queryset.query.annotations:
                queryset = queryset.annotate(**{target: Substr(field_name, 1, 10)})
            return target
        if field_name in datetime_fields:
            target = f"trunc_{field_name}"
            if target not in queryset.query.annotations:
                # `TruncDate(field_name)` nativo compila a
                # `CONVERT_TZ(field, 'UTC', 'America/Mexico_City')` (USE_TZ=True):
                # el MySQL de este entorno tiene `mysql.time_zone_name` vacía (sin
                # `mysql_tzinfo_to_sql` cargado), así que CONVERT_TZ con nombre de
                # zona devuelve NULL para TODAS las filas — truncaba a NULL
                # siempre, dejando `equals`/`not_equals` en 0/0. Se calcula el
                # offset con `zoneinfo` (Python puro, no depende de las tablas de
                # MySQL), se resta como intervalo, y se trunca forzando
                # `tzinfo=UTC` para que Django nunca emita CONVERT_TZ.
                offset = datetime.datetime.now(ZoneInfo(settings.TIME_ZONE)).utcoffset()
                shifted = ExpressionWrapper(F(field_name) + offset, output_field=DateTimeField())
                queryset = queryset.annotate(**{target: TruncDate(shifted, tzinfo=datetime.timezone.utc)})
            return target
        return field_name

    def build_condition_q(cond):
        if not isinstance(cond, dict):
            return None
        column = cond.get("column")

        if column not in valid_fields:
            if computed_resolver is None:
                return None
            if cond.get("compareType", "valor") == "campo":
                return None  # comparing a computed column to another field isn't supported
            value = cond.get("value", "")
            if value is None or str(value).strip() == "":
                return None
            return computed_resolver(
                column, cond.get("condition", "contains"), str(value).strip()
            )

        condition = cond.get("condition", "contains")
        compare_type = cond.get("compareType", "valor")
        if full_name_column is not None and column == full_name_column:
            nonlocal queryset
            target_field = "full_name"
            queryset = _annotate_full_name(queryset, target_field, full_name_fields)
            is_text = True
        else:
            target_field = resolve_target_field(column)
            is_text = column in text_fields

        if condition in ("empty", "not_empty"):
            # "Vacío"/"No vacío" ignoran compareType/value/compareColumn (el
            # front ya los oculta). `target_field` para columnas de texto ya
            # viene envuelto en Trim (ver resolve_target_field), así que NULL
            # y " " caen ambos en el mismo chequeo == "".
            if is_text:
                q = Q(**{f"{target_field}__isnull": True}) | Q(**{target_field: ""})
            else:
                q = Q(**{f"{target_field}__isnull": True})
            return q if condition == "empty" else ~q

        if compare_type == "campo":
            compare_column = cond.get("compareColumn")
            if compare_column not in valid_fields:
                return None
            target_compare_field = resolve_target_field(compare_column)
            f_expr = F(target_compare_field)

            if condition == "equals":
                return Q(**{target_field: f_expr})
            if condition == "not_equals":
                return ~Q(**{target_field: f_expr})
            lookup = ORDERING_LOOKUP_BY_CONDITION.get(condition)
            if lookup:
                return Q(**{f"{target_field}__{lookup}": f_expr})
            return None

        # compare_type == 'valor'
        value = cond.get("value", "")
        if value is None or str(value).strip() == "":
            return None
        value = str(value).strip()

        is_date_like = column in date_fields or column in TEXT_DATETIME_FIELDS
        if is_date_like:
            date_target = resolve_date_field(column)
            lookup = date_lookup_by_condition.get(condition)
            if lookup:
                return Q(**{f"{date_target}__{lookup}": value})
            if condition == "equals":
                return Q(**{date_target: value})
            if condition == "not_equals":
                # Una fecha NULL nunca es igual a `value`: debe contar como "no
                # es igual a", no desaparecer por "NOT (NULL = x)" = NULL en SQL
                # (mismo patrón que BUG-F06, aquí invertido: NULL sí debe pasar).
                return Q(**{f"{date_target}__isnull": True}) | ~Q(**{date_target: value})
            return None

        if column in numeric_fields:
            # El front ofrece ADEMÁS las condiciones de texto (Contiene,
            # Empieza con...) sobre columnas numéricas (ver
            # ADV_NUMBER_TEXT_CONDITIONS) — sin esto, elegir una de ésas caía
            # al `numeric_value = float(value)` de abajo, nunca encontraba un
            # lookup para "contains" y terminaba en `return None`, que
            # silenciosamente NO filtraba nada (bug real: "Días Ocupada
            # contiene 66" devolvía el set completo sin filtrar). MySQL sí
            # resuelve `__icontains`/`__istartswith`/etc. sobre una columna
            # entera (cast implícito a texto), así que basta con reusar el
            # mismo lookup de texto directo sobre `target_field` (que para un
            # campo numérico es la columna cruda, sin Trim). 'equals'/
            # 'not_equals' se EXCLUYEN a propósito: son ambiguos (también
            # existen en `number_lookup_by_condition`/ADV_NUMBER_CONDITIONS,
            # ver ADV_NUMBER_TEXT_CONDITIONS en advancedFilters.js) y deben
            # resolverse como comparación numérica más abajo (`float(value)
            # == numeric_value`), no como texto — si no, "066" no matchearía
            # 66 (iexact compara strings) aunque numéricamente sí sean iguales.
            if condition in text_lookup_by_condition and condition not in ("equals", "not_equals"):
                lookup, negate = text_lookup_by_condition[condition]
                q = Q(**{f"{target_field}__{lookup}": value})
                if not negate:
                    return q
                return Q(**{f"{target_field}__isnull": True}) | ~q
            try:
                numeric_value = float(value)
            except ValueError:
                return None
            lookup = number_lookup_by_condition.get(condition)
            if lookup:
                return Q(**{f"{target_field}__{lookup}": numeric_value})
            if condition == "equals":
                return Q(**{target_field: numeric_value})
            if condition == "not_equals":
                return ~Q(**{target_field: numeric_value})
            return None

        if is_text and condition in text_lookup_by_condition:
            lookup, negate = text_lookup_by_condition[condition]
            q = Q(**{f"{target_field}__{lookup}": value})
            if not negate:
                return q
            # "NOT (NULL LIKE ...)" es NULL en SQL, no TRUE: sin este OR las
            # filas con el campo vacío desaparecían tanto de "Contiene" como
            # de "No contiene" (BUG-F06). Los filtros `__in` (checkbox) ya
            # son NULL-aware vía EMPTY_VALUE_TOKEN; esto iguala el
            # comportamiento en Filtros Avanzados.
            return Q(**{f"{target_field}__isnull": True}) | ~q

        if condition == "equals":
            return Q(**{target_field: value})
        if condition == "not_equals":
            return ~Q(**{target_field: value})

        return None

    def fold_q(nodes):
        """Mismo fold AND/OR secuencial (i combina con el Q acumulado de 0..i-1),
        reutilizado para el top-level y para los `children` de un grupo."""
        combined = None
        for node in nodes:
            q = build_node_q(node)
            if q is None:
                continue
            if combined is None:
                combined = q
            elif (node.get("logic") or "AND").upper() == "OR":
                combined = combined | q
            else:
                combined = combined & q
        return combined

    def build_node_q(node):
        """Condición u grupo (paréntesis explícito, ver `AdvancedFiltersModal`).
        Un grupo resuelve su propio fold sobre `children` ANTES de combinarse
        con el resto de la lista donde vive — un solo nivel de anidamiento
        (un grupo no contiene otros grupos)."""
        if not isinstance(node, dict):
            return None
        if node.get("type") == "group":
            return fold_q(node.get("children", [])[:20])
        return build_condition_q(node)

    combined_q = fold_q(advanced_conditions)

    if combined_q is not None:
        queryset = queryset.filter(combined_q)

    return queryset


MOV_POS_COLUMN_LABELS = {
    "no_pos_actual": "No. Posición",
    "total_movimientos": "Histórico",
    "ocupacion": "Ocupación",
    "fecha_vacancia": "Fecha de Vacancia",
    "fecha_anuencia": "Fecha de Anuencia",
    "estado_psn": "Estado (A/I)",
    "f_efva": "Fecha Efectiva",
    "cd_motivo": "Cod. Motivo",
    "motivo": "Motivo",
    "cd_un": "Cod. UN",
    "unidad_de_negocio": "Unidad Negocio",
    "unidad_adva": "Unidad Adva",
    "cd_departamento": "Cod. Depto",
    "cd_puesto": "Cod. Puesto",
    "puesto_ptal": "Puesto Ptal",
    "estado_ptal": "Estado Ptal",
    "fecha_est": "Fecha Est",
    "maximo": "Máximo",
    "depnd_drt": "Depnd Drt",
    "depnd_indrt": "Depnd Indrt",
    "ubicacion": "Ubicación",
    "nvl_direc": "Nvl Direc",
    "plan_sal": "Plan Sal",
    "grado": "Grado",
    "esc": "Esc",
    "partida_ptal": "Partida Ptal",
    "gp_pago": "Gp Pago",
    "prog_beneficios": "Prog Beneficios",
    "fecha_captura": "Fecha Captura",
    "fh_ult_actz": "F/H Últ Actz",
    "por": "Por",
    "hr_estd_semn": "Hr Estd/Semn",
    "descr": "Descr",
    "gp_trabajo": "Gp Trabajo",
    "org_code": "Org Code",
    "grupo_cd_sal": "Grupo Cd Sal",
    "formal_desc": "Formal Desc",
    "pto_compt": "Pto Compt",
    "posn_clv": "Posn Clv",
    "presupuesto": "Presupuesto",
    "nombre_puesto": "Nombre Puesto",
    "categoria_vacancia": "Categoría Vacancia",
    "tuvo_insubsistencia": "Tuvo Insubsistencia",
}

MOV_POS_MONO_COLUMNS = {
    "no_pos_actual", "cd_un", "cd_departamento", "cd_puesto",
    "maximo", "grado", "esc", "partida_ptal",
}

# Columnas sintéticas del export de "Vacantes": desglosan el mismo registro
# decisivo que MovPosVacanciaDetalleView resuelve para el modal "Detalle de
# Vacancia", pero resuelto por lote para todas las filas del export.
VACANCIA_DETALLE_COLUMN_LABELS = {
    "vac_empleado_relacionado": "Empleado Relacionado",
    "vac_num_empleado": "Número de Empleado",
    "vac_accion": "Acción",
    "vac_motivo": "Motivo de la Vacancia",
    "vac_posicion_destino": "Posición Destino",
    "vac_fecha_efectiva_mov": "Fecha Efectiva del Movimiento",
    "vac_fecha_captura_mov": "Fecha de Captura del Movimiento",
    "vac_insub_persona": "Insubsistencia - Persona",
    "vac_insub_num_empleado": "Insubsistencia - Número de Empleado",
    "vac_insub_motivo": "Insubsistencia - Motivo",
    "vac_insub_fecha_efectiva": "Insubsistencia - Fecha Efectiva",
    "vac_insub_fecha_captura": "Insubsistencia - Fecha de Captura",
}

# Claves insertadas justo después de "categoria_vacancia" (detalle de la baja
# o el traslado que originó la vacancia).
VACANCIA_DETALLE_CATEGORIA_KEYS = [
    "vac_empleado_relacionado", "vac_num_empleado", "vac_accion", "vac_motivo",
    "vac_posicion_destino", "vac_fecha_efectiva_mov", "vac_fecha_captura_mov",
]

# Claves insertadas justo después de "tuvo_insubsistencia".
VACANCIA_DETALLE_INSUBSISTENCIA_KEYS = [
    "vac_insub_persona", "vac_insub_num_empleado", "vac_insub_motivo",
    "vac_insub_fecha_efectiva", "vac_insub_fecha_captura",
]


def _enrich_rows_with_vacancia_detalle(resultados):
    """Rellena en cada fila de MOV_POS (dicts de queryset.values()) las
    columnas sintéticas de VACANCIA_DETALLE_COLUMN_LABELS, resolviendo por
    lote (una sola query IN) los registros decisivos en
    cp_tbl_mov_completo_29_05_26 en vez de repetir la lógica N+1 de
    MovPosVacanciaDetalleView por cada fila."""
    from .models import CpTblMovCompleto290526

    decisivo_ids, insub_ids = set(), set()
    for r in resultados:
        if r.get("id_registro_desicivo"):
            decisivo_ids.add(r["id_registro_desicivo"])
        if (r.get("tuvo_insubsistencia") or "").strip().upper() == "S" and r.get("id_insubsistencia_detectada"):
            insub_ids.add(r["id_insubsistencia_detectada"])

    registros_by_id = {}
    all_ids = decisivo_ids | insub_ids
    if all_ids:
        registros_by_id = {
            reg.id: reg for reg in CpTblMovCompleto290526.objects.filter(id__in=all_ids)
        }

    def _nombre_completo(reg):
        return " ".join(p for p in [reg.nombre, reg.ap_pat, reg.ap_mat] if p).strip()

    for r in resultados:
        for key in VACANCIA_DETALLE_COLUMN_LABELS:
            r[key] = ""

        categoria = (r.get("categoria_vacancia") or "").strip().upper()
        id_decisivo = r.get("id_registro_desicivo")
        if categoria in ("A", "B") and id_decisivo:
            registro = registros_by_id.get(id_decisivo)
            if registro:
                r["vac_empleado_relacionado"] = _nombre_completo(registro)
                r["vac_num_empleado"] = registro.num_empleado
                r["vac_accion"] = registro.accion_nombre or registro.accion
                r["vac_motivo"] = registro.motivo_nombre or registro.motivo
                r["vac_fecha_efectiva_mov"] = registro.fecha_efectiva
                r["vac_fecha_captura_mov"] = registro.fecha_captura
                if categoria == "B":
                    r["vac_posicion_destino"] = registro.posicion

        if (r.get("tuvo_insubsistencia") or "").strip().upper() == "S":
            id_insub = r.get("id_insubsistencia_detectada")
            registro_ins = registros_by_id.get(id_insub) if id_insub else None
            if registro_ins:
                r["vac_insub_persona"] = _nombre_completo(registro_ins)
                r["vac_insub_num_empleado"] = registro_ins.num_empleado
                r["vac_insub_motivo"] = registro_ins.motivo_nombre or registro_ins.motivo
                r["vac_insub_fecha_efectiva"] = registro_ins.fecha_efectiva
                r["vac_insub_fecha_captura"] = registro_ins.fecha_captura


# Descripción de cada categoría de vacancia, igual a CATEGORIA_VACANCIA_TOOLTIP
# en MovimientosTab.jsx (modal "Detalle de Vacancia").
VACANCIA_CATEGORIA_DESCRIPCIONES = {
    "A": "Posición vacante porque empleado que la ocupaba causó baja",
    "B": "Posición vacante porque empleado que la ocupaba cambió a otra posición, vacancia = fecha en que tomó esa nueva posición",
    "C": "Posición vacante porque jamás tuvo ocupante, vacancia = fecha creación posición",
}

# Texto explicativo (2 renglones) mostrado en las filas 7-8 del export de
# "Vacantes", arriba del header real (fila 9). Clave ausente = columna sin nota.
VACANCIA_EXPORT_COLUMN_NOTES = {
    "total_movimientos": ("Cantidad de movimientos", "que ha tenido la posición"),
    "fecha_vacancia": ("Fecha de vacancia calculada por", "el Sistema de Control de Plazas (SCP)"),
    "categoria_vacancia": ("Identificador de causa", "de vacancia asignado por el SCP"),
    "vac_empleado_relacionado": ("Empleado que causó baja o fue trasladado", "a otra posición según aplique el caso (categoría de vacancia)"),
    "vac_num_empleado": ("Número de empleado que causó baja o fue trasladado", "a otra posición según aplique el caso (categoría de vacancia)"),
    "vac_accion": ("Acción del movimiento", "que provocó la baja o traslado"),
    "vac_motivo": ("Motivo del movimiento que provocó", "la baja o traslado"),
    "vac_posicion_destino": ("Posición Destino", "a la que fue trasladado el empleado"),
    "vac_fecha_efectiva_mov": ("Fecha Efectiva del Movimiento", "reportado en las columnas de Acción y Motivo"),
    "vac_fecha_captura_mov": ("Fecha de Captura del Movimiento", "reportado en las columnas de Acción y Motivo"),
    "tuvo_insubsistencia": ("¿Tuvo Insubsistencia", "la posición?"),
    "vac_insub_persona": ("Persona que causó", "la insubsistencia"),
    "vac_insub_num_empleado": ("No. Empleado de la persona que causó", "la insubsistencia"),
    "vac_insub_motivo": ("Motivo", "de la insubsistencia"),
    "vac_insub_fecha_efectiva": ("Fecha efectiva", "de la insubsistencia"),
    "vac_insub_fecha_captura": ("Fecha de captura", "de la insubsistencia"),
}

# Ancho mínimo de columna para que el texto largo de las notas explicativas
# (filas 7-8) y la leyenda (filas 1-5) no se vea cortado.
VACANCIA_EXPORT_MIN_COL_WIDTH = {
    "total_movimientos": 22,
    "fecha_vacancia": 24,
    "categoria_vacancia": 24,
    "vac_empleado_relacionado": 46,
    "vac_num_empleado": 30,
    "vac_accion": 20,
    "vac_motivo": 30,
    "vac_posicion_destino": 26,
    "vac_fecha_efectiva_mov": 30,
    "vac_fecha_captura_mov": 30,
    "tuvo_insubsistencia": 20,
    "vac_insub_persona": 38,
    "vac_insub_num_empleado": 34,
    "vac_insub_motivo": 26,
    "vac_insub_fecha_efectiva": 28,
    "vac_insub_fecha_captura": 28,
}

# La leyenda de categorías (E1:H5) vive en columnas fijas, sin relación con
# `visible_keys`; se les da un ancho mínimo propio para que quepan sus textos.
VACANCIA_LEGEND_MIN_COL_WIDTH = {"F": 60, "G": 22, "H": 40}

# (keys en orden, título del grupo para la fila 6, color de fondo de las filas 7-8)
VACANCIA_EXPORT_GROUPS = [
    (["fecha_vacancia", "categoria_vacancia"], "INFORMACIÓN GENERAL DE LA VACANCIA", "FF2A6099"),
    (VACANCIA_DETALLE_CATEGORIA_KEYS, "DETALLE DE LA VACANCIA", "FFBF0041"),
    (["tuvo_insubsistencia"] + VACANCIA_DETALLE_INSUBSISTENCIA_KEYS, "DETALLE DE LA INSUBSISTENCIA", "FF800080"),
]


def _categoria_vacancia_stats(resultados):
    """Cuenta, por cada fila ya exportada (solo vacantes), cuántas son de cada
    categoría de vacancia y cuántas de esas tuvieron insubsistencia."""
    stats = {c: {"total": 0, "insub": 0} for c in ("A", "B", "C")}
    for r in resultados:
        categoria = (r.get("categoria_vacancia") or "").strip().upper()
        if categoria not in stats:
            continue
        stats[categoria]["total"] += 1
        if (r.get("tuvo_insubsistencia") or "").strip().upper() == "S":
            stats[categoria]["insub"] += 1
    return stats


def _write_vacancia_report_cover(ws, visible_keys, resultados, row_offset=0):
    """Escribe, solo para el export de solo-"Vacantes", la portada del reporte:
    - Filas 1-5, columnas E-H: leyenda de categorías de vacancia + cantidad de
      plazas y cantidad con insubsistencia por categoría, con fila de totales.
    - Fila 6: encabezados de grupo (merge) sobre "información general",
      "detalle de la vacancia" y "detalle de la insubsistencia".
    - Filas 7-8: nota explicativa de 2 renglones por columna, coloreada según
      el grupo al que pertenece.
    Las filas 6-8 se ubican dinámicamente según la posición real de cada
    columna en `visible_keys`, ya que el usuario puede ocultar columnas del
    resto de la tabla (no. de posición, estado, etc.) desde la UI.

    `row_offset` corre todas las filas hacia abajo (usado cuando el membretado
    institucional ya ocupa las primeras filas de la hoja).
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    font_label = Font(name="Segoe UI", size=10, bold=True, color="FFFFFFFF")
    font_value = Font(name="Segoe UI", size=10, bold=False, color="FF000000")
    font_bold_black = Font(name="Segoe UI", size=10, bold=True, color="FF000000")

    legend_header_fill = PatternFill(start_color="FF2A6099", end_color="FF2A6099", fill_type="solid")
    legend_cat_fill = PatternFill(start_color="FFB4C7DC", end_color="FFB4C7DC", fill_type="solid")
    legend_total_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    def _cell(row, col, value, fill=None, font=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = border
        cell.alignment = align_center
        cell.font = font or font_value
        if fill:
            cell.fill = fill
        return cell

    # ── Leyenda de categorías (E1:H4) + fila de totales (E5:H5) ─────────────
    o = row_offset
    stats = _categoria_vacancia_stats(resultados)
    _cell(o + 1, 5, "CATEGORIA DE LA VACANCIA", legend_header_fill, font_label)
    _cell(o + 1, 6, "Descripción", legend_header_fill, font_label)
    _cell(o + 1, 7, "Cantidad de Plazas", legend_header_fill, font_label)
    _cell(o + 1, 8, "Cantidad de Plazas que Tuvieron Insubsistencia", legend_header_fill, font_label)

    for offset, categoria in enumerate(("A", "B", "C")):
        row = o + 2 + offset
        _cell(row, 5, categoria, legend_cat_fill, font_bold_black)
        _cell(row, 6, VACANCIA_CATEGORIA_DESCRIPCIONES[categoria])
        _cell(row, 7, stats[categoria]["total"])
        _cell(row, 8, stats[categoria]["insub"])

    total_plazas = sum(s["total"] for s in stats.values())
    total_insub = sum(s["insub"] for s in stats.values())
    _cell(o + 5, 5, "TOTAL", legend_total_fill, font_bold_black)
    ws.merge_cells(start_row=o + 5, start_column=5, end_row=o + 5, end_column=6)
    _cell(o + 5, 6, None, legend_total_fill, font_bold_black)
    _cell(o + 5, 7, total_plazas, legend_total_fill, font_bold_black)
    _cell(o + 5, 8, total_insub, legend_total_fill, font_bold_black)

    # ── Fila 6 (grupos) y filas 7-8 (notas explicativas) ────────────────────
    col_index = {key: idx for idx, key in enumerate(visible_keys, start=1)}
    group_fill_by_key = {}
    for keys, group_title, color in VACANCIA_EXPORT_GROUPS:
        cols = sorted(col_index[k] for k in keys if k in col_index)
        if not cols:
            continue
        group_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for k in keys:
            group_fill_by_key[k] = group_fill
        start_col, end_col = cols[0], cols[-1]
        black_fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
        _cell(o + 6, start_col, group_title, black_fill, font_label)
        if end_col > start_col:
            ws.merge_cells(start_row=o + 6, start_column=start_col, end_row=o + 6, end_column=end_col)

    yellow_fill = PatternFill(start_color="FFFFFF38", end_color="FFFFFF38", fill_type="solid")
    for key, (line1, line2) in VACANCIA_EXPORT_COLUMN_NOTES.items():
        col = col_index.get(key)
        if not col:
            continue
        fill = group_fill_by_key.get(key)
        # Columnas sin grupo (fondo amarillo, ej. "Histórico") usan texto negro;
        # el blanco quedaba invisible sobre amarillo.
        note_font = font_label if fill else font_bold_black
        fill = fill or yellow_fill
        _cell(o + 7, col, line1, fill, note_font)
        _cell(o + 8, col, line2, fill, note_font)

    # ── Título + fecha/hora de descarga (A1:D2) ─────────────────────────────
    from django.utils import timezone

    meses_es = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
    ]
    ahora = timezone.localtime(timezone.now())
    fecha_str = f"{ahora.day:02d} de {meses_es[ahora.month - 1]} de {ahora.year}, {ahora.strftime('%I:%M %p')}"

    title_fill = PatternFill(start_color="FF2B4C7E", end_color="FF2B4C7E", fill_type="solid")

    def _merge_box(row, value):
        _cell(row, 1, value, title_fill, font_label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for col in range(2, 5):
            _cell(row, col, None, title_fill, font_label)

    _merge_box(o + 1, "Reporte generado por el Sistema de Control de Plazas (SCP)")
    _merge_box(o + 2, f"Fecha y hora de descarga: {fecha_str}")


# Antes recalculaba ROW_NUMBER() OVER sobre toda MOV_POS en cada request
# (~300ms, ver AUDITORIA_BUGS_BACK.md BE2). MOV_POS_LATEST la materializa la
# tarea de importación de ZAFIRO una sola vez por import.
LATEST_MOVPOS_RAW_SQL = "SELECT id FROM MOV_POS_LATEST"

OCUPADAS_RAW_SQL = """
    SELECT DISTINCT e.`Posición`
    FROM EMPLEADOS_COMPLETOS_SIG e
    WHERE (e.`Id Empleado` IS NOT NULL AND TRIM(e.`Id Empleado`) <> '')
       OR (e.`Nombres` IS NOT NULL AND TRIM(e.`Nombres`) <> '');
"""

from django.core.cache import cache
from django.db import connection

from eje_central_back.renderers import orjson_dumps, orjson_response


def get_posiciones_ocupadas_set():
    """``{no_pos_actual}`` de posiciones ocupadas (``OCUPADAS_RAW_SQL``),
    cacheado 10 min — repetido inline en varias vistas de MovPos; ver
    ``corregir_fecha_anuencia_row``/``annotate_fecha_anuencia`` para el porqué
    (una posición ocupada nunca debe mostrar `fecha_anuencia`/`fecha_vacancia`
    obsoleta, el SP no las limpia al re-ocuparse)."""
    cache_key_ocupadas = "mov_pos_ocupadas_set"
    posiciones_ocupadas = cache.get(cache_key_ocupadas)
    if posiciones_ocupadas is None:
        with connection.cursor() as cursor:
            cursor.execute(OCUPADAS_RAW_SQL)
            posiciones_ocupadas = set(row[0] for row in cursor.fetchall() if row[0])
        cache.set(cache_key_ocupadas, posiciones_ocupadas, 600)
    return posiciones_ocupadas


def resolve_fecha_anuencia_value(pos, fa, posiciones_ocupadas, overrides_texto, baseline_texto):
    """Replica, campo a campo, la prioridad que aplica
    ``corregir_fecha_anuencia_row`` a nivel de fila: ocupada -> vacío;
    override de texto -> ese texto; baseline de texto -> ese texto; fecha
    real anotada -> 'YYYY-MM-DD'; si no, vacío. Usado para que filtrar/listar
    valores distintos de `fecha_anuencia` coincida exactamente con lo que
    pinta la tabla (incl. las 4 categorías de texto fijo que sólo viven en
    CeldaOverride, nunca en la anotación SQL)."""
    if pos in posiciones_ocupadas:
        return ""
    if pos in overrides_texto:
        return overrides_texto[pos]
    if pos in baseline_texto:
        return baseline_texto[pos]
    if hasattr(fa, "strftime"):
        return fa.strftime("%Y-%m-%d")
    if fa:
        return str(fa).split(" ")[0].split("T")[0]
    return ""


def obtener_posiciones_activas():
    # Cache active position codes for 60 seconds to speed up parallel requests on page load
    cache_key = "active_position_codes"
    cached_val = cache.get(cache_key)
    if cached_val is not None:
        return cached_val

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT `Nº Pos Actual` FROM MOV_POS_LATEST WHERE `Estado Psn` = 'A'
        """)
        result = [row[0] for row in cursor.fetchall() if row[0]]

    cache.set(cache_key, result, 1200)
    return result


def _get_mapa_correccion(columna):
    """
    Retorna un dict {posicion: valor} de una columna de corrección de solo
    lectura (``cat_correccion_posicion`` — Código, Tipo de Aduana, DG de
    Aduana compactada, ver CatCorreccionPosicion). Se cachea 30 minutos por
    columna — el catálogo es estático y solo cambia cuando el admin ejecuta
    `cargar_correcciones_plantilla` de nuevo.
    """
    cache_key = f"cat_correccion_posicion_mapa_{columna}"
    mapa = cache.get(cache_key)
    if mapa is None:
        mapa = dict(
            CatCorreccionPosicion.objects.filter(columna=columna).values_list("posicion", "valor")
        )
        cache.set(cache_key, mapa, 1800)  # 30 minutos
    return mapa


def _get_mapa_codigos():
    """{posicion: codigo} — ver `_get_mapa_correccion`."""
    return _get_mapa_correccion("codigo")


def _get_mapa_codigos_en_anuencia():
    """
    ``{codigo: [nombre_archivo, ...]}`` — para cada Código Federal de Puesto,
    los nombres de TODOS los Anexo 2 guardados (no eliminados) que lo traen
    en la lista `filas` de alguna de sus hojas. Usada en Mov. Posiciones
    (``MovPosDetalleView``) para marcar que una plaza está en proceso de
    anuencia — y en el flujo de "Agregar a Anexo 2" para avisarlo.

    Una plaza puede legítimamente aparecer en más de un Anexo 2 a lo largo
    del tiempo (se solicita, se ocupa, vuelve a quedar vacante, se vuelve a
    solicitar) — por eso esto es una lista y no un valor único, y por eso ya
    NO bloquea agregarla de nuevo en otro Anexo 2 (sólo informa en cuáles ya
    está). Los eliminados (soft delete, ver `AnuenciaAnexo.eliminado`) se
    excluyen: "eliminar" un Anexo 2 es precisamente la forma de decir "esta
    solicitud ya no cuenta" y liberar sus plazas sin perder el registro en la
    base de datos.

    Cacheado 5 min porque escanea el JSON de ``hojas`` de todos los Anexo 2;
    se invalida también al crear/actualizar/eliminar uno (ver
    ``AnuenciaAnexoViewSet``) para que el cambio se refleje sin esperar el TTL.
    """
    mapa = cache.get("mapa_codigos_en_anuencia")
    if mapa is not None:
        return mapa
    mapa = {}
    # `.order_by()` vacío cancela el `Meta.ordering` (-actualizado_en) del
    # modelo — sin esto, MySQL hace un filesort cargando la columna `hojas`
    # completa (JSON grande) por cada fila para poder ordenar, y con varios
    # Anexo 2 de cientos de plazas se queda sin sort_buffer_size (error 1038,
    # mismo caso ya resuelto en `AnuenciaAnexoViewSet.get_queryset` con
    # `.defer("hojas")` — aquí no se puede diferir porque sí se necesita leer
    # `hojas`, así que se cancela el ORDER BY en su lugar).
    qs = AnuenciaAnexo.objects.filter(eliminado=False).order_by()
    for nombre, hojas in qs.values_list("nombre_archivo", "hojas"):
        for hoja in hojas or []:
            for fila in hoja.get("filas") or []:
                codigo = str(fila.get("codigo") or "").strip()
                if not codigo:
                    continue
                lista = mapa.setdefault(codigo, [])
                if nombre not in lista:
                    lista.append(nombre)
    cache.set("mapa_codigos_en_anuencia", mapa, 300)
    return mapa


def _get_mapa_fecha_alta_solicitada_anuencia():
    """
    ``{codigo: {"fecha_alta_solicitada", "anexo_id", "hoja_id", "fila_id",
    "nombre_archivo"}}`` — para cada Código Federal de Puesto que esté en
    algún Anexo 2 guardado (no eliminado), la ``fecha_alta_solicitada`` que
    se capturó ahí, más las claves necesarias para poder editarla de vuelta
    (ver ``MovPosFechaAltaSolicitadaOverrideView``).

    Usado en Mov. Posiciones para mostrar la columna "Fecha de alta
    solicitada" (ver ``enriquecer_mov_pos_rows``) sincronizada con el Anexo 2
    sin duplicar el dato en un campo del modelo — el Anexo 2 (``hojas``) es
    la única fuente de verdad.

    Si un código está en más de un Anexo 2 (legítimo — ver
    ``_get_mapa_codigos_en_anuencia``), se usa el más recientemente
    actualizado: es el que refleja mejor el estado actual de la solicitud.

    Cacheado 5 min; se invalida junto con "mapa_codigos_en_anuencia" (ver
    ``_invalidar_cache_anuencia``).
    """
    mapa = cache.get("mapa_fecha_alta_solicitada_anuencia")
    if mapa is not None:
        return mapa
    mapa = {}
    # `.order_by()` vacío cancela el `Meta.ordering` (-actualizado_en) del
    # modelo — igual que en `_get_mapa_codigos_en_anuencia`: un ORDER BY con
    # la columna `hojas` (JSON grande) en juego fuerza un filesort en MySQL
    # que se queda sin sort_buffer_size con varios Anexo 2 de cientos de
    # plazas (error 1038, confirmado al probar este mismo cambio). El más
    # reciente se decide en Python, no en SQL (ver abajo).
    qs = AnuenciaAnexo.objects.filter(eliminado=False).order_by()
    filas = list(qs.values_list("id", "nombre_archivo", "hojas", "actualizado_en"))
    filas.sort(key=lambda f: f[3], reverse=True)
    for anexo_id, nombre, hojas, _actualizado_en in filas:
        for hoja in hojas or []:
            hoja_id = hoja.get("_id")
            for fila in hoja.get("filas") or []:
                codigo = str(fila.get("codigo") or "").strip()
                if not codigo or codigo in mapa:
                    continue
                fila_id = fila.get("_id")
                if not fila_id:
                    continue
                mapa[codigo] = {
                    "fecha_alta_solicitada": fila.get("fecha_alta_solicitada") or "",
                    "anexo_id": anexo_id,
                    "hoja_id": hoja_id,
                    "fila_id": fila_id,
                    "nombre_archivo": nombre,
                }
    cache.set("mapa_fecha_alta_solicitada_anuencia", mapa, 300)
    return mapa


def _invalidar_cache_anuencia():
    """Invalida los mapas cacheados que Mov. Posiciones deriva de los Anexo 2
    guardados (ver `_get_mapa_codigos_en_anuencia` /
    `_get_mapa_fecha_alta_solicitada_anuencia`) — llamar tras cualquier
    guardado, eliminación o reactivación de un Anexo 2."""
    cache.delete("mapa_codigos_en_anuencia")
    cache.delete("mapa_fecha_alta_solicitada_anuencia")


def _completar_aduana_row(r, mapa_tipo_aduana=None, mapa_dg_aduana=None):
    """
    Completa `tipo_de_aduana`/`dg_o_aduana_compactada` con el valor del
    catálogo de corrección SOLO si vienen vacíos de EMPLEADOS_COMPLETOS_SIG
    (~33% de las posiciones — confirmado que ZAFIRO no los trae completos,
    pero el Excel de referencia sí) — nunca se sobreescribe un valor real.

    Los mapas se pueden precargar una sola vez fuera del loop de filas que
    llama a esta función (ver `_enriquecer_empleados_completos_rows`): sin
    esto, cada fila vuelve a golpear el caché de Redis por su cuenta, lo que
    en listas de cientos de filas se nota (medido: ~5s extra en una búsqueda
    de ~430 filas antes de este cambio).
    """
    pos = r.get("posicion", "")
    if mapa_tipo_aduana is None:
        mapa_tipo_aduana = _get_mapa_correccion("tipo_de_aduana")
    if mapa_dg_aduana is None:
        mapa_dg_aduana = _get_mapa_correccion("dg_o_aduana_compactada")
    if not str(r.get("tipo_de_aduana") or "").strip():
        r["tipo_de_aduana"] = mapa_tipo_aduana.get(pos, "")
    if not str(r.get("dg_o_aduana_compactada") or "").strip():
        r["dg_o_aduana_compactada"] = mapa_dg_aduana.get(pos, "")


def _get_mapa_quincenal():
    """
    Retorna un dict anidado {posicion: {columna: valor}} con las 10 columnas
    quincenal editables (AL–AV, excepto fecha_anuencia_detalle — ver
    ``_get_fecha_anuencia_bulk_map``), mergeando el baseline cargado del Excel
    (``TblColumnasPlantillaQuincenal``) con las ediciones manuales activas
    (``celda_override.get_quincenal_overrides_map``, tabla PLANTILLA_QUINCENAL)
    — el override del usuario tiene prioridad sobre el valor del Excel.

    Se cachea 30 minutos — se invalida desde ``ColumnasQuincenalView`` tras
    cada edición y desde ``cargar_columnas_quincenal`` tras cada recarga del
    baseline.
    """
    cache_key = "tbl_columnas_quincenal_mapa"
    mapa = cache.get(cache_key)
    if mapa is None:
        mapa = {}
        for row in TblColumnasPlantillaQuincenal.objects.exclude(
            columna="fecha_anuencia_detalle"
        ).values("posicion", "columna", "valor"):
            mapa.setdefault(row["posicion"], {})[row["columna"]] = row["valor"] or ""

        for pos, overrides_pos in get_quincenal_overrides_map().items():
            mapa.setdefault(pos, {}).update(overrides_pos)

        cache.set(cache_key, mapa, 1800)  # 30 minutos
    return mapa


def _get_posiciones_ocupadas_set():
    """
    ``set`` de posiciones actualmente ocupadas — mismo cache/criterio
    (``mov_pos_ocupadas_set``/``OCUPADAS_RAW_SQL``) que ya usan
    ``_get_fecha_anuencia_bulk_map``/``_get_fecha_vacancia_bulk_map``. Se
    reutiliza para blanquear ``COLUMNAS_SOLICITUD_VACANTE`` (ver
    ``EmpleadosCompletosActivosDetalleView``/``_aplicar_mapeos_detalle_excel``)
    una vez que la plaza ya se ocupó.
    """
    posiciones_ocupadas = cache.get("mov_pos_ocupadas_set")
    if posiciones_ocupadas is None:
        with connection.cursor() as cursor:
            cursor.execute(OCUPADAS_RAW_SQL)
            posiciones_ocupadas = set(row[0] for row in cursor.fetchall() if row[0])
        cache.set("mov_pos_ocupadas_set", posiciones_ocupadas, 600)
    return posiciones_ocupadas


def _get_fecha_anuencia_bulk_map(posiciones):
    """
    Calcula ``fecha_anuencia`` para un conjunto de posiciones con la MISMA
    prioridad y fuente de datos que ``MovPosDetalleView`` (override manual >
    valor cargado del Excel > fecha_vacancia + 30 días) y el MISMO criterio
    de "ocupada" (``mov_pos_ocupadas_set``/``OCUPADAS_RAW_SQL``) — así
    Plantilla Detalle y Mov. Posiciones siempre muestran la misma fecha para
    la misma posición. Devuelve ``{posicion: 'YYYY-MM-DD'|""}``.
    """
    if not posiciones:
        return {}

    sub_ids = cache.get("latest_movpos_sub_ids")
    if sub_ids is None:
        with connection.cursor() as cursor:
            cursor.execute(LATEST_MOVPOS_RAW_SQL)
            sub_ids = [row[0] for row in cursor.fetchall() if row[0]]
        cache.set("latest_movpos_sub_ids", sub_ids, 600)

    posiciones_ocupadas = cache.get("mov_pos_ocupadas_set")
    if posiciones_ocupadas is None:
        with connection.cursor() as cursor:
            cursor.execute(OCUPADAS_RAW_SQL)
            posiciones_ocupadas = set(row[0] for row in cursor.fetchall() if row[0])
        cache.set("mov_pos_ocupadas_set", posiciones_ocupadas, 600)

    qs = MovPos.objects.filter(id__in=sub_ids, no_pos_actual__in=list(posiciones))
    qs = annotate_fecha_anuencia(
        qs,
        overrides=get_fecha_anuencia_overrides_map(),
        baseline=get_fecha_anuencia_baseline_map(),
    )

    # "Fecha de Anuencia" no siempre es una fecha real (ver
    # corregir_fecha_anuencia_row) — estas 2 tienen prioridad final sobre lo
    # que resuelve la anotación SQL (que solo puede representar fechas).
    overrides_texto = get_fecha_anuencia_overrides_texto_map()
    baseline_texto = get_fecha_anuencia_baseline_texto_map()

    mapa = {}
    for pos, fecha in qs.values_list("no_pos_actual", "fecha_anuencia"):
        if pos in posiciones_ocupadas:
            mapa[pos] = ""
            continue
        if pos in overrides_texto:
            mapa[pos] = overrides_texto[pos]
        elif pos in baseline_texto:
            mapa[pos] = baseline_texto[pos]
        elif hasattr(fecha, "strftime"):
            mapa[pos] = fecha.strftime("%Y-%m-%d")
        elif fecha:
            mapa[pos] = str(fecha).split(" ")[0].split("T")[0]
        else:
            mapa[pos] = ""
    return mapa


def _get_fecha_anuencia_override_map(posiciones):
    """
    ``{posicion: bool}`` — True si la posición tiene un override manual
    activo de fecha_anuencia (fecha real o categoría de texto), para
    resaltar en azul la celda en Plantilla Detalle — mismo criterio que
    ``row["fecha_anuencia_override"]`` en ``corregir_fecha_anuencia_row``
    (usado por Mov. Posiciones), aquí en versión bulk.
    """
    if not posiciones:
        return {}
    overrides_texto = get_fecha_anuencia_overrides_texto_map()
    return {pos: (pos in overrides_texto) for pos in posiciones}


def _get_fecha_vacancia_bulk_map(posiciones):
    """
    ``fecha_vacancia`` real de MOV_POS (calculada por el SP de ZAFIRO
    ``sp_obtener_todas_vacancias``, ver ``MovPosDetalleView``) para un
    conjunto de posiciones — SIN override ni baseline del Excel, a
    diferencia de ``fecha_anuencia``: el Excel tiene errores conocidos en su
    columna "Fecha que se genera la vacante", así que Plantilla Detalle debe
    mostrar SIEMPRE este valor calculado, idéntico al que ya muestra
    Mov. Posiciones en su columna "Fecha de Vacancia" (mismo criterio de
    "ocupada" — se pone en blanco igual que allá). Devuelve
    ``{posicion: 'YYYY-MM-DD'|""}``.
    """
    if not posiciones:
        return {}

    sub_ids = cache.get("latest_movpos_sub_ids")
    if sub_ids is None:
        with connection.cursor() as cursor:
            cursor.execute(LATEST_MOVPOS_RAW_SQL)
            sub_ids = [row[0] for row in cursor.fetchall() if row[0]]
        cache.set("latest_movpos_sub_ids", sub_ids, 600)

    posiciones_ocupadas = cache.get("mov_pos_ocupadas_set")
    if posiciones_ocupadas is None:
        with connection.cursor() as cursor:
            cursor.execute(OCUPADAS_RAW_SQL)
            posiciones_ocupadas = set(row[0] for row in cursor.fetchall() if row[0])
        cache.set("mov_pos_ocupadas_set", posiciones_ocupadas, 600)

    qs = MovPos.objects.filter(id__in=sub_ids, no_pos_actual__in=list(posiciones))

    mapa = {}
    for pos, fecha in qs.values_list("no_pos_actual", "fecha_vacancia"):
        if pos in posiciones_ocupadas or not fecha:
            mapa[pos] = ""
        else:
            mapa[pos] = str(fecha).split(" ")[0].split("T")[0]
    return mapa


def _get_mov_pos_id_bulk_map(posiciones):
    """
    ``{posicion: mov_pos_id}`` — id (PK) de la fila MÁS RECIENTE de MOV_POS
    para cada posición (misma resolución "latest" que
    ``_get_fecha_vacancia_bulk_map``). Plantilla Detalle lo necesita para
    abrir el modal de Detalle de Vacancia (``VacanciaDetalleModal``, mismo
    componente que ya usa Mov. Posiciones) al hacer clic en "Fecha que se
    genera la vacante" — ``MovPosVacanciaDetalleView`` exige el id de ESE
    renglón específico de MOV_POS (categoria_vacancia, id_registro_desicivo,
    tuvo_insubsistencia son campos por-renglón, no derivables solo de la
    posición).
    """
    if not posiciones:
        return {}

    sub_ids = cache.get("latest_movpos_sub_ids")
    if sub_ids is None:
        with connection.cursor() as cursor:
            cursor.execute(LATEST_MOVPOS_RAW_SQL)
            sub_ids = [row[0] for row in cursor.fetchall() if row[0]]
        cache.set("latest_movpos_sub_ids", sub_ids, 600)

    qs = MovPos.objects.filter(id__in=sub_ids, no_pos_actual__in=list(posiciones))
    return dict(qs.values_list("no_pos_actual", "id"))


def _get_datos_mov_pos_bulk_map(posiciones):
    """
    ``{posicion: {cd_puesto, puesto_ptal, esc, partida_ptal, unidad_de_negocio}}``
    de la fila MÁS RECIENTE de MOV_POS para cada posición — fuente para el
    autollenado del Anexo 2 (``AnuenciaLookupView``) y para la Unidad de
    Negocio del Anexo 3 (``AnuenciaAnexo3View``), en vez de Plantilla Detalle
    (``EmpleadosCompletosSig``). ``cd_puesto``/``puesto_ptal`` son las llaves
    para cruzar contra ``CAT_PTO_FUNC``/``rc_cat_cod_presupuestal`` (ver
    ``_get_mapa_puesto_funcional``/``_get_mapa_cod_presupuestal``).
    """
    mapa_ids = _get_mov_pos_id_bulk_map(posiciones)
    if not mapa_ids:
        return {}
    filas = {
        f["id"]: f
        for f in MovPos.objects.filter(id__in=mapa_ids.values()).values(
            "id", "cd_puesto", "puesto_ptal", "esc", "partida_ptal", "unidad_de_negocio"
        )
    }
    return {pos: filas[mov_id] for pos, mov_id in mapa_ids.items() if mov_id in filas}


def _get_mapa_puesto_funcional():
    """``{cd_puesto: nombre_puesto_funcional}`` desde CAT_PTO_FUNC, catálogo
    completo cacheado 30 min (551 filas — trivial de tener entero en memoria).

    Antes filtraba por los códigos presentes en cada request
    (``CatPtoFunc.objects.filter(cd_pto_funcional__in=codigos)``), lo que
    repetía esa consulta sin cachear en cada apertura del dropdown de
    "Puesto" en Mov. Posiciones (~1.3s extra medidos sólo en esta consulta,
    sobre las ~13k filas de `is_latest=true`) — cachear el catálogo entero
    es más simple y más rápido que cachear por combinación de códigos.
    Invalidado en ``CatPtoFuncViewSet`` al crear/editar/borrar.
    """
    mapa = cache.get("mapa_puesto_funcional_completo")
    if mapa is None:
        mapa = dict(
            CatPtoFunc.objects.values_list("cd_pto_funcional", "nombre_puesto_funcional")
        )
        cache.set("mapa_puesto_funcional_completo", mapa, 1800)
    return mapa


def _get_mapa_cod_presupuestal():
    """``{(puesto_ptal, escala): {nivel, smn}}`` desde rc_cat_cod_presupuestal
    (catálogo "Códigos Presupuestales" en Catálogos > Estructura
    Organizacional) — llave compuesta código+escala, igual que en BD.

    Catálogo completo cacheado 30 min (192 filas), mismo motivo que
    ``_get_mapa_puesto_funcional``. Invalidado en
    ``RcCatCodPresupuestalViewSet`` al crear/editar/borrar.
    """
    mapa = cache.get("mapa_cod_presupuestal_completo")
    if mapa is None:
        filas = RcCatCodPresupuestal.objects.values(
            "codigo_presupuestal", "escala", "nivel", "smn"
        )
        mapa = {(f["codigo_presupuestal"], f["escala"]): f for f in filas}
        cache.set("mapa_cod_presupuestal_completo", mapa, 1800)
    return mapa


def _escala_entera(valor):
    try:
        return int(str(valor or "").strip())
    except ValueError:
        return None


def enriquecer_mov_pos_rows(resultados):
    """
    Agrega a cada fila de MOV_POS (ya con sus campos crudos vía `.values()`)
    las mismas columnas "resueltas" que ya autollena el Anexo 2 — Puesto,
    Nivel Salarial, Salario Mensual Neto y Tipo de contratación (ver
    `AnuenciaLookupView`/`_get_datos_mov_pos_bulk_map`) — y si la posición ya
    está en algún Anexo 2 guardado, en cuáles. Requiere que cada fila ya
    traiga `codigo` (bolt-on de `_get_mapa_codigos`, ver los tres bucles de
    `MovPosDetalleView`). Muta `resultados` in-place; no hace nada si viene
    vacío.

    `anuencia_anexo_nombre` es el texto para mostrar en la celda (nombres
    unidos con ", " — una plaza puede estar en más de un Anexo 2, ver
    `_get_mapa_codigos_en_anuencia`); `anuencia_anexo_nombres` es la lista
    cruda, la que realmente hay que usar para filtrar/agrupar (comparar
    contra el texto unido rompería en cuanto una plaza esté en 2+ anexos —
    ver `_match_lista_anuencia`).
    """
    if not resultados:
        return

    mapa_puesto_funcional = _get_mapa_puesto_funcional()
    mapa_presupuestal = _get_mapa_cod_presupuestal()
    mapa_en_anuencia = _get_mapa_codigos_en_anuencia()
    mapa_fecha_alta = _get_mapa_fecha_alta_solicitada_anuencia()

    for r in resultados:
        r["denominacion_puesto"] = mapa_puesto_funcional.get(r.get("cd_puesto"), "")
        datos_presupuestal = mapa_presupuestal.get(
            (r.get("puesto_ptal"), _escala_entera(r.get("esc"))), {}
        )
        r["nivel_salarial"] = datos_presupuestal.get("nivel") or ""
        smn = datos_presupuestal.get("smn")
        r["salario_mensual_neto"] = str(smn) if smn is not None else ""
        r["tipo_contratacion"] = _TIPO_CONTRATACION_ANEXO2.get(
            str(r.get("partida_ptal") or "").strip(), ""
        )
        codigo = str(r.get("codigo") or "").strip()
        nombres = mapa_en_anuencia.get(codigo, [])
        r["anuencia_anexo_nombres"] = nombres
        r["anuencia_anexo_nombre"] = ", ".join(nombres)
        # "Fecha de alta solicitada": vive únicamente dentro de `hojas` del
        # Anexo 2 (no hay campo propio en MOV_POS) — sólo se puede editar
        # desde aquí si el código está en alguno (ver
        # MovPosFechaAltaSolicitadaOverrideView).
        info_alta = mapa_fecha_alta.get(codigo)
        r["fecha_alta_solicitada"] = (info_alta or {}).get("fecha_alta_solicitada") or ""
        r["fecha_alta_solicitada_editable"] = info_alta is not None


# Columnas de MOV_POS que NO son campos reales del modelo, calculadas en
# Python después de la consulta — filtrar/ordenar/mostrar el dropdown de
# estas columnas no puede resolverse con SQL directo (a diferencia de
# "ocupacion"/"fecha_anuencia", que si tiran a una anotación o un mapa
# barato); necesitan materializar la fila completa vía `enriquecer_mov_pos_rows`
# (ver `_calcular_columnas_derivadas_mov_pos` y su uso en `MovPosDetalleView`
# para filtros de columna, filtros avanzados y el dropdown de cada una).
CAMPOS_DERIVADOS_MOV_POS = {
    "codigo",
    "denominacion_puesto",
    "nivel_salarial",
    "salario_mensual_neto",
    "tipo_contratacion",
    "anuencia_anexo_nombre",
    "fecha_alta_solicitada",
}


def _calcular_columnas_derivadas_mov_pos(queryset):
    """
    ``{no_pos_actual: {codigo, denominacion_puesto, nivel_salarial,
    salario_mensual_neto, tipo_contratacion, anuencia_anexo_nombre}}`` para
    TODAS las filas de `queryset` (sin paginar) — reusa exactamente
    `enriquecer_mov_pos_rows` (la misma que rellena esas columnas para
    mostrarlas en pantalla) para que un filtro sobre una de ellas nunca
    pueda "mentir" respecto a lo que se ve en la tabla.
    """
    filas = list(queryset.values("no_pos_actual", "cd_puesto", "puesto_ptal", "esc", "partida_ptal"))
    if not filas:
        return {}
    mapa_codigos = _get_mapa_codigos()
    for f in filas:
        f["codigo"] = mapa_codigos.get(f["no_pos_actual"], "")
    enriquecer_mov_pos_rows(filas)
    return {f["no_pos_actual"]: f for f in filas}


def populate_movpos_occupant_details(resultados, posiciones_ocupadas):
    if not resultados:
        return
    pos_list = [r.get("no_pos_actual") for r in resultados if r.get("no_pos_actual")]
    occupants = {}
    if pos_list:
        with connection.cursor() as cursor:
            format_strings = ','.join(['%s'] * len(pos_list))
            query = f"""
                SELECT `Posición`, `Id Empleado`, `Nombres`
                FROM EMPLEADOS_COMPLETOS_SIG
                WHERE `Posición` IN ({format_strings})
            """
            cursor.execute(query, pos_list)
            for row in cursor.fetchall():
                pos_code = row[0]
                id_emp = row[1]
                name = row[2]
                id_emp_str = str(id_emp).strip() if id_emp is not None else ""
                name_str = str(name).strip() if name is not None else ""
                if id_emp_str or name_str:
                    occupants[pos_code] = {
                        "id_empleado": id_emp_str,
                        "nombres": name_str
                    }

    for r in resultados:
        pos = r.get("no_pos_actual")
        is_occupied = pos in posiciones_ocupadas
        occ = occupants.get(pos)
        if is_occupied and occ:
            r["ocupante_id"] = occ["id_empleado"]
            r["ocupante_nombre"] = occ["nombres"]
        else:
            r["ocupante_id"] = ""
            r["ocupante_nombre"] = ""


# Create your views here.
import io

import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .excel_letterhead import add_excel_letterhead, LETTERHEAD_ROWS


class ExportExcelView(APIView):
    """
    Vista genérica para exportar datos JSON a un archivo Excel (.xlsx) real con estilos institucionales.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        filename = request.query_params.get("filename", "Export.xlsx")
        # Congelar columna A además del encabezado es el default histórico de
        # este endpoint (varios consumidores ya cuentan con ello); se deja
        # opt-out vía query param para quien no quiera ninguna columna sticky
        # (ver tab "Historial de movimientos" de EmployeesModal.jsx) sin tocar
        # el comportamiento por defecto del resto de llamadores.
        sticky_column = request.query_params.get("sticky_column", "1").strip().lower() not in ("0", "false", "no")

        if not data or not isinstance(data, list):
            return Response(
                {"error": "Se requiere una lista de objetos para exportar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Convertir JSON a DataFrame de Pandas
            df = pd.DataFrame(data)

            # Crear el archivo Excel en memoria
            output = io.BytesIO()

            # Forzar conversión de todas las columnas a tipos básicos para evitar errores de serialización
            for col in df.columns:
                if df[col].dtype == "object":
                    df[col] = df[col].fillna("")

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Plantilla", startrow=LETTERHEAD_ROWS)

                workbook = writer.book
                worksheet = writer.sheets["Plantilla"]
                add_excel_letterhead(worksheet, len(df.columns))
                header_row_num = LETTERHEAD_ROWS + 1

                # --- ESTILOS ROBUSTOS ---
                # Usamos códigos ARGB completos (FF + Hex) para máxima compatibilidad
                header_fill = PatternFill(
                    start_color="FF621F32", end_color="FF621F32", fill_type="solid"
                )
                zebra_fill = PatternFill(
                    start_color="FFF9FAFB", end_color="FFF9FAFB", fill_type="solid"
                )
                header_font = Font(color="FFFFFFFF", bold=True, size=11, name="Calibri")
                data_font = Font(size=10, name="Calibri")

                side = Side(style="thin", color="FFD1D5DB")
                thin_border = Border(left=side, right=side, top=side, bottom=side)

                align_center = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                align_left = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                # --- PROCESAR ENCABEZADOS ---
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=header_row_num, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
                    cell.alignment = align_center

                    # Cálculo de ancho ultra-seguro
                    try:
                        # Obtenemos el máximo largo de los datos en esta columna
                        # Filtramos nulos y convertimos a string antes de medir
                        lengths = df[column_title].astype(str).map(len)
                        max_val_len = lengths.max() if not lengths.empty else 0

                        # Manejo de NaN o valores no numéricos en el cálculo
                        if pd.isna(max_val_len):
                            max_val_len = 0

                        header_len = len(str(column_title))
                        final_width = max(float(max_val_len), float(header_len)) + 3

                        worksheet.column_dimensions[
                            get_column_letter(col_num)
                        ].width = min(final_width, 60)
                    except:
                        worksheet.column_dimensions[
                            get_column_letter(col_num)
                        ].width = 20

                # --- PROCESAR DATOS ---
                # Limitamos el procesamiento de estilos si el dataset es masivo para evitar timeouts
                max_styled_rows = 5000
                rows_to_process = min(len(df), max_styled_rows)

                for row_num in range(header_row_num + 1, header_row_num + 1 + rows_to_process):
                    is_zebra = row_num % 2 == 0
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = align_left
                        cell.font = data_font
                        if is_zebra:
                            cell.fill = zebra_fill

                # Congelar paneles: fila de encabezado siempre, columna A solo
                # si sticky_column no viene desactivado (ver arriba).
                freeze_col = "B" if sticky_column else "A"
                worksheet.freeze_panes = f"{freeze_col}{header_row_num + 1}"

            output.seek(0)
            file_data = output.read()

            if not file_data:
                raise ValueError("El archivo generado está vacío.")

            # Preparar la respuesta HTTP
            response = HttpResponse(
                file_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            return response

        except Exception:
            logger.exception("Fallo crítico al generar Excel")
            return Response(
                {"error": "Fallo crítico al generar Excel"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class PlantillaVacantesPorNivelView(APIView):
    permission_classes = [IsAuthenticated]

    # Devuelve el resumen de las posiciones ocupadas y vacantes por nivel
    def get(self, request):
        cache_key = "plantilla_vacantes_por_nivel"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        active_position_codes = obtener_posiciones_activas()

        resultados = list(
            EmpleadosCompletosSig.objects.filter(posicion__in=active_position_codes)
            .values("nivel")
            .annotate(
                Activo=Count("estado_nomina", filter=Q(estado_nomina="Activo")),
                Vacante=Count("estado_nomina", filter=Q(estado_nomina="Vacante")),
                Suspendido=Count("estado_nomina", filter=Q(estado_nomina="Suspendido")),
                Permiso_Retribuido=Count(
                    "estado_nomina", filter=Q(estado_nomina="Permiso Retribuido")
                ),
                Permiso=Count("estado_nomina", filter=Q(estado_nomina="Permiso")),
            )
            .order_by("nivel")
        )
        cache.set(cache_key, resultados, 1200)
        return Response(resultados, status=status.HTTP_200_OK)


class PlantillaVacantesPorNivelResumenView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def obtener_resumen_dinamico():
        cache_key = "plantilla_vacantes_por_nivel_resumen"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return cached_data

        active_position_codes = obtener_posiciones_activas()

        base_qs = EmpleadosCompletosSig.objects.filter(
            posicion__in=active_position_codes
        ).exclude(Q(estado_nomina__isnull=True) | Q(estado_nomina="Estado Nomina"))

        estados_unicos = base_qs.values_list("estado_nomina", flat=True).distinct()

        # total_niveles: Conteo de niveles distintos
        # total_registros: Conteo total de filas válidas
        agregaciones = {
            "total_niveles": Count("nivel", distinct=True),
            "total_registros": Count("*"),
        }

        # 3. Iteramos sobre los estados para crear el equivalente al SUM(CASE...)
        for estado in estados_unicos:
            # Usamos el nombre del estado con la primera letra en mayúscula y sin espacios
            # para ser consistentes con la otra vista y evitar colisiones con campos del modelo (que son minúsculas)
            llave = estado.replace(" ", "_")

            agregaciones[llave] = Sum(
                Case(
                    When(estado_nomina=estado, then=1),
                    default=0,
                    output_field=IntegerField(),
                )
            )

        # 4. Ejecutamos la consulta pasándole el diccionario desempaquetado (**agregaciones)
        resultado = base_qs.aggregate(**agregaciones)
        cache.set(cache_key, resultado, 1200)

        return resultado

    def get(self, request, *args, **kwargs):
        try:
            datos = self.obtener_resumen_dinamico()
            return Response(datos, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EmpleadosCompletosEstatusNominaResumenView(APIView):
    view_permission = "authentication.view_plantilla_detalle"

    def get(self, request, *args, **kwargs):
        try:
            # Se agrega SIEMPRE sobre el mismo dataset cacheado que alimenta
            # la tabla de Detalle (`_obtener_detalle_activos_cacheado`), en
            # vez de correr su propia query con su propio caché de 20 min:
            # antes, como sólo el caché de Detalle se invalida al editar
            # (`_invalidar_cache_detalle_plantilla`), esta tarjeta podía
            # quedar hasta 20 minutos desfasada frente al conteo real de la
            # tabla (p. ej. justo después de una sincronización de nómina).
            # Al derivarse del mismo dataset ya cacheado, tarjeta y tabla
            # quedan sincronizadas por construcción, sin depender de acordarse
            # de invalidar dos cachés en paralelo.
            filas = _obtener_detalle_activos_cacheado()

            resumen = {
                "total_registros": len(filas),
                "Activo": 0,
                "Vacante": 0,
                "Suspendido": 0,
                "Licencia": 0,
                "Licencia_Medica": 0,
            }

            for fila in filas:
                estado = fila.get("estado_nomina")
                val_estat = fila.get("val_estat")

                # Normalizar estados según el mapeo solicitado. `val_estat`
                # ("Ocupada"/"Vacante") es un respaldo para cuando
                # `estado_nomina` llega vacío/código no reconocido pese a que
                # la plaza sí tiene ocupante real (acciones de personal muy
                # recientes donde nómina aún no asigna el código A/S/L/P) —
                # mismo criterio que mapEstadoNomina en el frontend
                # (PlantillaDetalleTab.jsx), para que esta tarjeta de resumen
                # no contradiga el conteo real de la tabla.
                vacante_fallback = "Activo" if val_estat == "Ocupada" else "Vacante"
                if not estado or estado.strip() == "":
                    label = vacante_fallback
                else:
                    estado_upper = estado.strip().upper()
                    if estado_upper == "A":
                        label = "Activo"
                    elif estado_upper == "S":
                        label = "Suspendido"
                    elif estado_upper == "L":
                        label = "Licencia"
                    elif estado_upper == "P":
                        label = "Licencia_Medica"
                    else:
                        label = vacante_fallback

                resumen[label] = resumen.get(label, 0) + 1

            return Response(resumen, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def _enriquecer_empleados_completos_rows(resultados):
    """Inyecta código federal + columnas quincenal + fecha_anuencia + fecha_vacancia
    + mov_pos_id + aduana en cada fila de ``EmpleadosCompletosSig.objects.values()``.

    Compartido por las 3 rutas de `EmpleadosCompletosActivosDetalleView` (todo,
    oficio/nivel, search) — antes estaba duplicado en las primeras dos.
    """
    mapa_codigos = _get_mapa_codigos()
    mapa_quincenal = _get_mapa_quincenal()
    _posiciones = [r.get("posicion", "") for r in resultados]
    mapa_fa = _get_fecha_anuencia_bulk_map(_posiciones)
    mapa_fa_override = _get_fecha_anuencia_override_map(_posiciones)
    mapa_fv = _get_fecha_vacancia_bulk_map(_posiciones)
    mapa_mov_id = _get_mov_pos_id_bulk_map(_posiciones)
    posiciones_ocupadas = _get_posiciones_ocupadas_set()
    # Precargados una sola vez (no por fila) — ver docstring de
    # `_completar_aduana_row` sobre el costo de no hacerlo.
    mapa_tipo_aduana = _get_mapa_correccion("tipo_de_aduana")
    mapa_dg_aduana = _get_mapa_correccion("dg_o_aduana_compactada")
    _COLS_QUINCENAL = sorted(COLUMNAS_QUINCENAL_VALIDAS - {"fecha_anuencia_detalle"})
    for r in resultados:
        pos = r.get("posicion", "")
        r["codigo"] = mapa_codigos.get(pos, "")
        custom = mapa_quincenal.get(pos, {})
        for col in _COLS_QUINCENAL:
            r[col] = custom.get(col, "")
        # Datos de un candidato "solicitado" solo tienen sentido mientras la
        # plaza siga vacante — si ya se ocupó, se blanquean sin importar lo
        # que haya quedado guardado.
        if pos in posiciones_ocupadas:
            for col in COLUMNAS_SOLICITUD_VACANTE:
                r[col] = ""
        r["fecha_anuencia_detalle"] = mapa_fa.get(pos, "")
        # Resalta en azul la celda si tiene un override manual activo — mismo
        # criterio que Mov. Posiciones.
        r["fecha_anuencia_detalle_override"] = mapa_fa_override.get(pos, False)
        # "Fecha que se genera la vacante": SIEMPRE la calculada (fecha_vacancia
        # de MOV_POS), nunca la del Excel — ver docstring de
        # _get_fecha_vacancia_bulk_map.
        r["fecha_genera_vacante"] = mapa_fv.get(pos, "")
        # id de MOV_POS para abrir el modal de Detalle de Vacancia
        # (VacanciaDetalleModal, mismo que Mov. Posiciones).
        r["mov_pos_id"] = mapa_mov_id.get(pos)
        _completar_aduana_row(r, mapa_tipo_aduana, mapa_dg_aduana)
    return resultados


# Campos sobre los que busca `search` en EmpleadosCompletosActivosDetalleView —
# mismo criterio que MovimientosPersonalListView (apply_text_search con una
# lista acotada, no el blob completo de ~80 columnas): cubre lo que el front
# promete ("nombre, RFC, CURP, unidad administrativa, etc.", ver TableroRH.jsx).
CAMPOS_BUSQUEDA_EMPLEADOS_DETALLE = [
    "nombres", "rfc", "curp", "posicion", "numempleado", "id_empleado",
    "unidad_administrativa", "nombre_puesto_funcional", "departamento",
    "dependencia_directa", "ua", "ua2",
]


class DefaultApiPagination(PageNumberPagination):
    """Paginador genérico para endpoints de dataset completo (EMPLEADOS_COMPLETOS_SIG,
    BAJAS_SIG, ...) consumidos hoy sin paginar por eje_central_front y, opcionalmente
    (?pagination=true), por clientes externos como rendicionCuentasBack."""
    page_size = 500
    page_size_query_param = "page_size"
    max_page_size = 10000


def _paginated_or_full_response(request, data):
    """Devuelve `data` paginado (envelope count/next/previous/results) si viene
    ?pagination=true en el query string; si no, exactamente igual que hoy (lista
    plana) — eje_central_front nunca manda este flag, así que su contrato no cambia."""
    if request.query_params.get("pagination", "false").strip().lower() == "true":
        paginator = DefaultApiPagination()
        page = paginator.paginate_queryset(data, request)
        return paginator.get_paginated_response(page)
    return Response(data, status=status.HTTP_200_OK)


def _obtener_detalle_activos_cacheado():
    """
    Filas enriquecidas de EmpleadosCompletosSig acotadas a posiciones activas
    — mismo cache_key/TTL que la ruta sin filtros de
    EmpleadosCompletosActivosDetalleView. Única fuente de verdad compartida
    con EmpleadosCompletosEstatusNominaResumenView para que la tarjeta de
    resumen nunca contradiga el conteo real de la tabla de Detalle (ver
    docstring de esa vista).
    """
    cache_key = "empleados_completos_activos_detalle"
    cached_data = cache.get(cache_key)
    if cached_data is not None:
        return cached_data

    active_position_codes = obtener_posiciones_activas()
    queryset = EmpleadosCompletosSig.objects.filter(posicion__in=active_position_codes)
    resultados = _enriquecer_empleados_completos_rows(list(queryset.values()))
    cache.set(cache_key, resultados, 1200)
    return resultados


class EmpleadosCompletosActivosDetalleView(APIView):
    # Dataset base compartido por 3 tabs (Detalle/Estatus/Mov. Posiciones cruzan
    # contra `detalle`) — cualquiera de los 3 permisos basta, no solo Detalle.
    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
    )

    def get(self, request, *args, **kwargs):
        oficio = request.query_params.get("oficio")
        nivel = request.query_params.get("nivel")
        search = (request.query_params.get("search") or "").strip()

        if search:
            # Búsqueda libre sobre el universo de posiciones activas (mismo
            # scope que la ruta "todo" de abajo), sin caché: a diferencia de
            # oficio/nivel (catálogo acotado) o "todo" (una sola clave), el
            # espacio de términos de búsqueda es ilimitado — mismo criterio
            # que MovimientosPersonalListView, que tampoco cachea `search`.
            # La usa TableroRH para no tener que traer el dataset completo
            # (~11 mil filas) al cliente en cada carga solo para filtrarlo ahí.
            try:
                active_position_codes = obtener_posiciones_activas()
                queryset = EmpleadosCompletosSig.objects.filter(
                    posicion__in=active_position_codes
                )
                queryset = apply_text_search(queryset, search, CAMPOS_BUSQUEDA_EMPLEADOS_DETALLE)
                resultados = _enriquecer_empleados_completos_rows(list(queryset.values()))
                return _paginated_or_full_response(request, resultados)
            except Exception:
                logger.exception("Error inesperado en {}".format(request.path))
                return Response(
                    {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        if oficio or nivel:
            cache_key = f"empleados_completos_activos_detalle_{oficio}_{nivel}"
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return _paginated_or_full_response(request, cached_data)

            try:
                # Obtener posiciones de Plantilla1800Plazas que cumplan los filtros
                posiciones_qs = Plantilla1800Plazas.objects.all()
                if oficio:
                    if oficio == "(vacío)":
                        posiciones_qs = posiciones_qs.filter(
                            Q(of_de_solicitud__isnull=True) | Q(of_de_solicitud="")
                        )
                    else:
                        posiciones_qs = posiciones_qs.filter(of_de_solicitud=oficio)
                if nivel:
                    posiciones_qs = posiciones_qs.filter(nivel=nivel)

                posiciones_list = list(posiciones_qs.values_list("posición", flat=True))

                # Filtrar EmpleadosCompletosSig
                queryset = EmpleadosCompletosSig.objects.filter(
                    posicion__in=posiciones_list
                )
                resultados = _enriquecer_empleados_completos_rows(list(queryset.values()))

                cache.set(cache_key, resultados, 300)
                return _paginated_or_full_response(request, resultados)
            except Exception:
                logger.exception("Error inesperado en {}".format(request.path))
                return Response(
                    {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        try:
            resultados = _obtener_detalle_activos_cacheado()
            return _paginated_or_full_response(request, resultados)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


FOTOS_EMPLEADOS_CONTENT_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
}


def _invalidar_cache_detalle_plantilla():
    """
    Invalida las variantes cacheadas de EmpleadosCompletosActivosDetalleView
    (cache_key "empleados_completos_activos_detalle" sin filtros +
    "empleados_completos_activos_detalle_{oficio}_{nivel}" con filtros, más
    "active_employees_filtered" de otro consumidor) tras cualquier edición
    que afecte esa respuesta: overrides de EMPLEADOS_COMPLETOS_SIG, columnas
    quincenal o fecha_anuencia. Función compartida — antes duplicada
    (imperfectamente, sin "active_employees_filtered") entre
    EmpleadosCompletosCeldaOverrideView y ColumnasQuincenalView.
    """
    cache.delete_many(["empleados_completos_activos_detalle", "active_employees_filtered"])
    try:
        import redis as redis_lib

        r = redis_lib.Redis.from_url(settings.CELERY_BROKER_URL)
        for key in r.scan_iter("*empleados_completos_activos_detalle_*"):
            r.delete(key)
    except Exception:
        logger.exception("Error al invalidar cache filtrada tras edición de Plantilla Detalle")


class ColumnasQuincenalView(APIView):
    """
    Edición manual de las columnas quincenal (AL–AV del Excel, excepto
    fecha_anuencia_detalle — ver MovPosFechaAnuenciaOverrideView, unificada
    con el sistema de Mov. Posiciones) desde el tab Plantilla Detalle.

    Registra el cambio en CeldaOverride (tabla=PLANTILLA_QUINCENAL, ver
    celda_override.registrar_y_aplicar_override_quincenal) — el MISMO
    mecanismo de auditoría (valor_original/valor_nuevo/usuario/activo +
    historial, ver EmpleadosCompletosCeldaHistorialView) que ya existe para
    EMPLEADOS_COMPLETOS_SIG. El baseline cargado del Excel
    (TblColumnasPlantillaQuincenal, ver management command
    `cargar_columnas_quincenal`) nunca se modifica aquí: solo se lee como
    "valor original" de la primera edición y como fallback en la lectura
    (ver `_get_mapa_quincenal`) — así una recarga futura del Excel actualiza
    el valor de referencia sin pisar lo que el usuario ya editó.

    GET  ?posicion=X                 → {columna: valor_efectivo, ...} (override > baseline).
    PATCH {posicion, columna, valor} → registra el override.
    DELETE {posicion, columna}       → revierte al valor baseline del Excel (o vacío).

    Solo admite columnas declaradas en COLUMNAS_QUINCENAL_EDITABLES.
    'reportada' solo acepta '' | 'Si' | 'No'.
    """

    edit_permission = "authentication.edit_plantilla_detalle"
    view_permission = "authentication.view_plantilla_detalle"

    @staticmethod
    def _invalidar_cache():
        cache.delete("tbl_columnas_quincenal_mapa")
        _invalidar_cache_detalle_plantilla()

    def get(self, request):
        posicion = request.query_params.get("posicion", "").strip()
        if not posicion:
            return Response({"detail": "posicion es requerida."}, status=status.HTTP_400_BAD_REQUEST)
        baseline = dict(
            TblColumnasPlantillaQuincenal.objects.filter(posicion=posicion)
            .exclude(columna="fecha_anuencia_detalle")
            .values_list("columna", "valor")
        )
        overrides = get_quincenal_overrides_map().get(posicion, {})
        resultado = {**baseline, **overrides}
        return Response(resultado)

    def patch(self, request):
        posicion = request.data.get("posicion", "").strip()
        columna = request.data.get("columna", "").strip()
        valor = request.data.get("valor")  # puede ser None → se convierte a ""
        if not posicion or not columna:
            return Response({"detail": "posicion y columna son requeridos."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            override = registrar_y_aplicar_override_quincenal(posicion, columna, valor, request.user)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if override is None:
            # Valor idéntico al vigente — no se registró como cambio (mismo
            # criterio 8.10 QA que EMPLEADOS_COMPLETOS_SIG/fecha_anuencia).
            return Response({
                "posicion": posicion,
                "columna": columna,
                "valor": "" if valor is None else str(valor).strip(),
                "sin_cambios": True,
            })

        self._invalidar_cache()
        return Response({
            "posicion": posicion,
            "columna": columna,
            "valor": override.valor_nuevo,
            "valor_original": override.valor_original,
            "usuario": request.user.username,
            "fecha_modificacion": override.fecha_modificacion,
            "created": True,
        })

    def delete(self, request):
        posicion = request.data.get("posicion", "").strip()
        columna = request.data.get("columna", "").strip()
        if not posicion or not columna:
            return Response({"detail": "posicion y columna son requeridos."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            deleted = borrar_override_quincenal(posicion, columna)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        self._invalidar_cache()
        return Response({"posicion": posicion, "columna": columna, "deleted": deleted})


class EmpleadoFotoView(APIView):
    """
    Sirve la fotografía de un empleado, cargada bajo demanda desde el tab
    Detalle (botón "Ver") — no se precarga para toda la tabla a propósito,
    para no pagar miles de lecturas de disco en cada carga de la tabla (ver
    discusión de diseño: prioriza que ESTE endpoint sea rápido, ya que si se
    pide es porque el usuario ya está esperando verla).

    Resuelve en este orden (ver `excel_fotos.resolver_foto_empleado`):
      1. Archivo ``<numempleado>.<ext>`` directo en disco — convención
         estándar, cubre toda foto nueva sin ningún paso extra ni tabla de
         por medio (nunca queda desactualizado). Se prueban varias
         VARIANTES del número (ver ``_variantes_numempleado``): verificado
         contra datos reales que el archivo casi nunca está guardado con el
         mismo padding de ceros que trae `numempleado` en BD (de una
         muestra de 2000, 0 coincidían exacto — el 83% solo coincidía
         quitando los ceros a la izquierda), igual que ya contemplaba
         `excel_fotos_vba_guide.md` para el export a Excel.
      2. RFC en vivo: SICRE nombra la foto por RFC mientras el empleado no
         tiene numempleado asignado, y la renombra sola en cuanto sí lo
         tiene — se recalcula contra el RFC actual en BD en cada llamada
         (sin caché) para no quedar desfasado cuando SICRE haga ese rename.
      3. Alias en `EmpleadoFotoAlias` (excepciones históricas irregulares
         que ni 1 ni 2 cubren) — se llena con el management command
         `cargar_fotos_empleados`, red de seguridad, no la vía principal.
      4. 404 si ninguna resuelve.
    """

    # Un solo endpoint compartido por todos los tabs que muestran fotografía
    # (Detalle, Estatus Nómina, Mov. Posiciones, Movimientos, Bajas,
    # Distribución Geográfica) — cada uno tiene su propio permiso "ver
    # fotografía" independiente del permiso de ver el tab en sí, así que
    # aquí basta con tener CUALQUIERA de ellos (OR, ya soportado por
    # HasModulePermission cuando view_permission es una tupla).
    view_permission = (
        "authentication.view_plantilla_detalle_foto",
        "authentication.view_plantilla_estatus_nomina_foto",
        "authentication.view_plantilla_mov_posiciones_foto",
        "authentication.view_plantilla_movimientos_foto",
        "authentication.view_plantilla_bajas_foto",
        "authentication.view_plantilla_geografia_foto",
    )

    def get(self, request, numempleado):
        from django.http import FileResponse, HttpResponseNotFound
        from .excel_fotos import resolver_foto_empleado

        numempleado = str(numempleado).strip()
        if not numempleado:
            return HttpResponseNotFound()

        ruta = resolver_foto_empleado(numempleado)
        if ruta is None:
            return HttpResponseNotFound()

        # ETag barato (mtime+tamaño, sin leer/hashear el archivo completo)
        # para que el navegador no vuelva a pedir la misma foto dos veces.
        stat = ruta.stat()
        etag = f'"{stat.st_mtime_ns:x}-{stat.st_size:x}"'
        if request.headers.get("If-None-Match") == etag:
            return HttpResponse(status=304)

        ext_lower = ruta.suffix.lower().lstrip(".")
        response = FileResponse(
            open(ruta, "rb"),
            content_type=FOTOS_EMPLEADOS_CONTENT_TYPES.get(ext_lower, "application/octet-stream"),
        )
        response["ETag"] = etag
        response["Cache-Control"] = "private, max-age=86400"
        return response


class DatosPersonalesEmpleadoView(APIView):
    """
    Datos personales (tabla DATOS_PERSONALES, importada por Celery desde el
    CSV de ZAFIRO) de un solo empleado — consumida por el tab "Datos
    personales" del expediente (EmployeeRecordModal), cargados bajo demanda
    igual que la fotografía. Comparte permisos con el expediente en general:
    basta con poder ver cualquiera de los tabs que lo abren.
    """

    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
        "authentication.view_plantilla_bajas",
        "authentication.view_plantilla_geografia",
    )

    def get(self, request, no_empleado):
        no_empleado = str(no_empleado).strip()
        if not no_empleado:
            return Response(
                {"detail": "no_empleado es requerido."}, status=status.HTTP_400_BAD_REQUEST
            )

        registro = DatosPersonales.objects.filter(no_empleado=no_empleado).first()
        if registro is None:
            return Response(
                {"detail": "No se encontraron datos personales para este empleado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        data = {
            "no_empleado": registro.no_empleado,
            "hr_id_persona": registro.hr_id_persona,
            "position_nbr": registro.position_nbr,
            "nombre_completo": registro.nombre_completo,
            "rfc": registro.rfc,
            "curp": registro.curp,
            "puesto": registro.puesto,
            "puesto_estructural": registro.puesto_estructural,
            "puesto_funcional": registro.puesto_funcional,
            "deptid": registro.deptid,
            "unidad_administrativa": registro.unidad_administrativa,
            "humanos_status": registro.humanos_status,
            "estatus_nomina": registro.estatus_nomina,
            "escolaridad_tipo": registro.escolaridad_tipo,
            "escolaridad_nivrl": registro.escolaridad_nivrl,
            "escolaridad_area": registro.escolaridad_area,
            "carrera": registro.carrera,
            "centro_escolar": registro.centro_escolar,
            "phone": registro.phone,
            "phone1": registro.phone1,
            "extension": registro.extension,
            "conmutador": registro.conmutador,
            "email_addr": registro.email_addr,
            "email_addr2": registro.email_addr2,
            "calle": registro.calle,
            "hr_numero_exterior": registro.hr_numero_exterior,
            "hr_numero_interior": registro.hr_numero_interior,
            "colonia": registro.colonia,
            "postal": registro.postal,
            "hr_municipio": registro.hr_municipio,
            "estado": registro.estado,
        }
        return Response(data, status=status.HTTP_200_OK)


# Campos de DATOS_PERSONALES ofrecidos por la opción "Incluir datos
# personales" del export a Excel de Plantilla Detalle (ver
# DatosPersonalesBulkView y _aplicar_mapeos_detalle_excel). Se excluye
# no_empleado (ya es la clave de cruce, viene como "numempleado" en
# Plantilla Detalle) y se etiquetan con "(Datos Personales)" los campos que
# también existen con otro significado/valor en la plantilla, para no
# confundirlos en el Excel.
DATOS_PERSONALES_EXPORT_FIELDS = [
    ("hr_id_persona", "HR ID Persona"),
    ("position_nbr", "Position NBR (Datos Personales)"),
    ("nombre_completo", "Nombre Completo (Datos Personales)"),
    ("rfc", "RFC (Datos Personales)"),
    ("curp", "CURP (Datos Personales)"),
    ("puesto_estructural", "Puesto Estructural (Datos Personales)"),
    ("puesto_funcional", "Puesto Funcional (Datos Personales)"),
    ("puesto", "Puesto (Datos Personales)"),
    ("escolaridad_tipo", "Escolaridad Tipo"),
    ("escolaridad_nivrl", "Escolaridad Nivel"),
    ("escolaridad_area", "Escolaridad Área"),
    ("carrera", "Carrera"),
    ("centro_escolar", "Centro Escolar"),
    ("humanos_status", "Status RH"),
    ("estatus_nomina", "Estatus Nómina (Datos Personales)"),
    ("phone", "Teléfono"),
    ("phone1", "Teléfono 2"),
    ("extension", "Extensión"),
    ("email_addr", "Correo Electrónico"),
    ("email_addr2", "Correo Electrónico 2"),
    ("calle", "Calle"),
    ("hr_numero_exterior", "Número Exterior"),
    ("hr_numero_interior", "Número Interior"),
    ("colonia", "Colonia"),
    ("postal", "Código Postal"),
    ("hr_municipio", "Municipio"),
    ("estado", "Estado (Domicilio)"),
    ("deptid", "Dept ID (Datos Personales)"),
    ("unidad_administrativa", "Unidad Administrativa (Datos Personales)"),
]


def _get_datos_personales_bulk_map(no_empleados):
    """no_empleado -> dict de DATOS_PERSONALES_EXPORT_FIELDS, para varios
    empleados en una sola consulta (cruce por numempleado en el export)."""
    limpios = sorted({str(n).strip() for n in no_empleados if str(n or "").strip()})
    if not limpios:
        return {}
    campos = [f for f, _ in DATOS_PERSONALES_EXPORT_FIELDS]
    registros = DatosPersonales.objects.filter(no_empleado__in=limpios).values(
        "no_empleado", *campos
    )
    return {r["no_empleado"]: r for r in registros}


class DatosPersonalesBulkView(APIView):
    """
    Datos personales de VARIOS empleados en una sola consulta — usado por la
    opción "Incluir datos personales" del export a Excel de Plantilla
    Detalle (100% client-side vía ExcelJS en PlantillaDetalleTab.jsx): el
    front manda los `numempleado` visibles y arma las columnas extra
    localmente. Mismos permisos que DatosPersonalesEmpleadoView (consulta
    individual del expediente).
    """

    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
        "authentication.view_plantilla_bajas",
        "authentication.view_plantilla_geografia",
    )

    def post(self, request):
        no_empleados = request.data.get("no_empleados")
        if not isinstance(no_empleados, list) or not no_empleados:
            return Response(
                {"detail": "no_empleados (lista) es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        mapa = _get_datos_personales_bulk_map(no_empleados)
        return Response({"results": mapa}, status=status.HTTP_200_OK)


class DatosPersonalesCeldaOverrideView(APIView):
    """
    Edición manual de una celda de DATOS_PERSONALES (Escolaridad/Contacto/
    Domicilio) desde el tab "Datos personales" del expediente. Registra el
    cambio en CeldaOverride (tabla=DATOS_PERSONALES, clave no_empleado) y lo
    aplica de inmediato sobre la fila viva (ver plantilla.celda_override).
    Se reaplica tras cada importar_zafiro, ya que la tabla se trunca y
    recarga completa (blue-green swap).
    """

    edit_permission = "authentication.edit_datos_personales"

    def post(self, request):
        no_empleado = request.data.get("no_empleado")
        columna = request.data.get("columna")
        valor_nuevo = request.data.get("valor_nuevo")
        if not no_empleado or not columna:
            return Response(
                {"detail": "no_empleado y columna son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            override = registrar_y_aplicar_override_datos_personales(
                no_empleado, columna, valor_nuevo, request.user
            )
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if override is None:
            return Response({
                "no_empleado": no_empleado,
                "columna": columna,
                "valor_nuevo": valor_nuevo,
                "sin_cambios": True,
            })

        return Response({
            "no_empleado": no_empleado,
            "columna": columna,
            "valor_nuevo": override.valor_nuevo,
            "valor_original": override.valor_original,
            "usuario": request.user.username,
            "fecha_modificacion": override.fecha_modificacion,
        })


def _mapear_estado_nomina_excel(val):
    """Réplica exacta de ``mapEstadoNomina`` (frontend, PlantillaDetalleTab.jsx)
    para que el export server-side muestre las mismas etiquetas que la tabla."""
    if not val or not str(val).strip():
        return "Vacante"
    codigo = str(val).strip().upper()
    return {"A": "Activo", "S": "Suspendido", "L": "Licencia", "P": "Licencia Médica"}.get(codigo, "Vacante")


# Réplicas exactas de mapPartida/mapTipoContratacion/displayRango (frontend,
# PlantillaDetalleTab.jsx) para que el export server-side (con fotos) muestre
# las MISMAS etiquetas que la tabla en pantalla, en vez de los códigos crudos.
_TIPO_CONTRATACION_LABELS_EXCEL = {"SAT_CFZA": "Confianza", "SAT_BSE": "Base"}


def _mapear_partida_excel(codigo, posicion):
    codigo = str(codigo or "").strip()
    if codigo == "11301":
        return "Permanente"
    if codigo == "11401":
        return "PASEM"
    if codigo == "12201":
        return "Eventual N.C." if str(posicion or "").strip().startswith("2026") else "Eventual"
    return codigo


def _mapear_tipo_contratacion_excel(val):
    codigo = str(val or "").strip()
    return _TIPO_CONTRATACION_LABELS_EXCEL.get(codigo, codigo)


def _completar_rango_excel(rango, tipo_personal):
    val = str(rango or "").strip()
    if val:
        return val
    return "Civil" if str(tipo_personal or "").strip().lower() == "civil" else ""


def _aplicar_mapeos_detalle_excel(rows):
    """Aplica in-place los mapeos de Partida/Tipo de Contratación/Rango/Aduana
    y la inyección de Código/columnas quincenal/Fecha de Anuencia/Fecha de
    Vacancia a filas de EMPLEADOS_COMPLETOS_SIG antes de generar el Excel —
    para que la descarga (con o sin fotos) siempre coincida con lo que se ve
    en pantalla (ver mismas funciones en PlantillaDetalleTab.jsx y la
    inyección equivalente en EmpleadosCompletosActivosDetalleView). Código,
    quincenal, Fecha de Anuencia y Fecha de Vacancia NO son columnas reales
    de EMPLEADOS_COMPLETOS_SIG — antes de esto, incluirlas en un export
    producía celdas vacías."""
    posiciones = [r.get("posicion", "") for r in rows]
    mapa_codigos = _get_mapa_codigos()
    mapa_quincenal = _get_mapa_quincenal()
    mapa_fa = _get_fecha_anuencia_bulk_map(posiciones)
    mapa_fv = _get_fecha_vacancia_bulk_map(posiciones)
    posiciones_ocupadas = _get_posiciones_ocupadas_set()
    cols_quincenal = sorted(COLUMNAS_QUINCENAL_VALIDAS - {"fecha_anuencia_detalle"})

    for r in rows:
        pos = r.get("posicion", "")
        if "partida" in r:
            r["partida"] = _mapear_partida_excel(r.get("partida"), pos)
        if "tipo_de_contratacion" in r:
            r["tipo_de_contratacion"] = _mapear_tipo_contratacion_excel(r.get("tipo_de_contratacion"))
        if "rango" in r:
            r["rango"] = _completar_rango_excel(r.get("rango"), r.get("tipo_de_personal_sedena_semar"))
        if "tipo_de_aduana" in r or "dg_o_aduana_compactada" in r:
            _completar_aduana_row(r)

        r["codigo"] = mapa_codigos.get(pos, "")
        custom = mapa_quincenal.get(pos, {})
        for col in cols_quincenal:
            r[col] = custom.get(col, "")
        # Mismo criterio que EmpleadosCompletosActivosDetalleView: el export
        # tampoco debe mostrar un candidato "solicitado" para una plaza que ya
        # se ocupó.
        if pos in posiciones_ocupadas:
            for col in COLUMNAS_SOLICITUD_VACANTE:
                r[col] = ""
        r["fecha_anuencia_detalle"] = mapa_fa.get(pos, "")
        r["fecha_genera_vacante"] = mapa_fv.get(pos, "")
        # Alias bajo la clave "fecha_vacancia" — es la que usa el export de
        # ExportarEmpleadosPorPosicionConFotosView (columna del EmployeesModal
        # compartido, ver ALL_AVAILABLE_COLUMNS en EmployeesModal.jsx); sin
        # esto la columna salía vacía porque solo se poblaba "fecha_genera_vacante".
        r["fecha_vacancia"] = mapa_fv.get(pos, "")
    return rows


_EXCEL_CON_FOTOS_CONTENT_TYPE_XLSM = "application/vnd.ms-excel.sheet.macroEnabled.12"
_EXCEL_CON_FOTOS_CONTENT_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _excel_con_fotos_response(buffer, incluir_fotos, nombre_base):
    from .excel_fotos import VBA_HABILITADO

    es_xlsm = incluir_fotos and VBA_HABILITADO
    extension = "xlsm" if es_xlsm else "xlsx"
    content_type = _EXCEL_CON_FOTOS_CONTENT_TYPE_XLSM if es_xlsm else _EXCEL_CON_FOTOS_CONTENT_TYPE_XLSX
    sufijo = "_ConFotos" if incluir_fotos else ""
    response = HttpResponse(buffer.read(), content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{nombre_base}{sufijo}.{extension}"'
    return response


class ExportarPlantillaDetalleConFotosView(APIView):
    """
    Genera el Excel de Plantilla Detalle, opcionalmente con fotografías de
    empleados embebidas en la celda (ver excel_fotos_vba_guide.md). Camino
    NUEVO y opt-in: el export normal (sin fotos) sigue siendo 100%
    client-side (ExcelJS) en PlantillaDetalleTab.jsx — este endpoint solo se
    llama cuando el usuario, con permiso view_plantilla_detalle_foto, elige
    incluir fotos en el modal de confirmación.

    PlantillaDetalleTab filtra/ordena 100% client-side (no hay equivalente
    server-side de sus filtros), así que el frontend manda las claves de
    negocio (`posicion`) de las filas ya filtradas, en su orden actual, más
    la lista de columnas visibles ({key,label}) tal como las tiene el usuario.
    """

    view_permission = "authentication.view_plantilla_detalle"

    def post(self, request):
        from .excel_fotos import generar_workbook_excel_con_fotos

        posiciones = request.data.get("posiciones") or []
        columnas = request.data.get("columnas") or []
        incluir_fotos = bool(request.data.get("incluir_fotos")) and request.user.has_perm(
            "authentication.view_plantilla_detalle_foto"
        )
        incluir_datos_personales = bool(request.data.get("incluir_datos_personales"))

        if not posiciones or not columnas:
            return Response({"error": "Faltan 'posiciones' o 'columnas'."}, status=status.HTTP_400_BAD_REQUEST)

        rows_by_posicion = {
            str(r["posicion"]): r
            for r in EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values()
        }
        rows = _aplicar_mapeos_detalle_excel(
            [rows_by_posicion[str(p)] for p in posiciones if str(p) in rows_by_posicion]
        )

        # Tope de seguridad: la plantilla activa completa son ~13,300 filas
        # hoy (medido) — con margen. Sin este tope, "sin querer" no hay
        # riesgo real aquí (a diferencia de Movimientos, un log histórico sin
        # límite natural), pero se deja por consistencia con las otras vistas.
        if incluir_fotos and len(rows) > 15000:
            return Response(
                {"error": f"El conjunto tiene {len(rows)} filas — el máximo para incluir fotografías es 15,000."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # "Incluir datos personales": cruce por numempleado con DATOS_PERSONALES,
        # agregando sus columnas (prefijo "dp_" para no pisar columnas homónimas
        # de Plantilla Detalle, ver DATOS_PERSONALES_EXPORT_FIELDS) al final.
        if incluir_datos_personales:
            mapa_dp = _get_datos_personales_bulk_map([r.get("numempleado") for r in rows])
            for r in rows:
                registro_dp = mapa_dp.get(str(r.get("numempleado") or "").strip(), {})
                for campo, _label in DATOS_PERSONALES_EXPORT_FIELDS:
                    r[f"dp_{campo}"] = registro_dp.get(campo, "")
            columnas = list(columnas) + [
                {"key": f"dp_{campo}", "label": label} for campo, label in DATOS_PERSONALES_EXPORT_FIELDS
            ]

        buffer = generar_workbook_excel_con_fotos(
            columnas=columnas, rows=rows, incluir_fotos=incluir_fotos,
            sheet_name="Plantilla_Empleados", numero_empleado_key="numempleado",
            mono_keys=["posicion", "id_empleado", "rfc", "curp", "nivel", "codigo_presupuestal",
                       "ua", "cent", "dir", "subd", "jd", "depto", "numeral"],
            estado_nomina_key="estado_nomina", mapear_estado_nomina=_mapear_estado_nomina_excel,
        )
        return _excel_con_fotos_response(buffer, incluir_fotos, "Plantilla_Empleados")


class ExportarPlantillaHistoricaConFotosView(APIView):
    """Export "con fotos" de la reconstrucción histórica (botón "Consultar
    plantillas pasadas" en Plantilla Detalle) — hermano de
    ExportarPlantillaDetalleConFotosView, pero SIN re-consultar
    EMPLEADOS_COMPLETOS_SIG por posición: esa vista re-deriva el ocupante
    ACTUAL de cada posición, lo que en modo histórico mostraría a la persona
    equivocada (la que ocupa la plaza HOY, no la de la fecha consultada).

    El front ya tiene las filas correctas (`historicoFilas`, resueltas por
    PlantillaHistoricaView) y ya las mapea a valores de pantalla (Partida,
    Tipo de Contratación, Rango, Estado Nómina — mismo mapeo que el export
    sin fotos, ver handleExportExcel en PlantillaDetalleTab.jsx) antes de
    mandarlas aquí: esta vista sólo arma el Excel (membretado + fotos vía
    `numempleado`, que sí corresponde a la persona histórica correcta) sin
    tocar la DB salvo para el cruce opcional de Datos Personales.
    """

    view_permission = "authentication.view_plantilla_historico"

    def post(self, request):
        from .excel_fotos import generar_workbook_excel_con_fotos

        fecha = (request.data.get("fecha") or "").strip()
        rows = request.data.get("rows") or []
        columnas = request.data.get("columnas") or []
        incluir_fotos = bool(request.data.get("incluir_fotos")) and request.user.has_perm(
            "authentication.view_plantilla_detalle_foto"
        )
        incluir_datos_personales = bool(request.data.get("incluir_datos_personales"))

        if not fecha or not rows or not columnas:
            return Response({"error": "Faltan 'fecha', 'rows' o 'columnas'."}, status=status.HTTP_400_BAD_REQUEST)

        # Mismo tope que el export en vivo (ver ExportarPlantillaDetalleConFotosView).
        if incluir_fotos and len(rows) > 15000:
            return Response(
                {"error": f"El conjunto tiene {len(rows)} filas — el máximo para incluir fotografías es 15,000."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if incluir_datos_personales:
            mapa_dp = _get_datos_personales_bulk_map([r.get("numempleado") for r in rows])
            for r in rows:
                registro_dp = mapa_dp.get(str(r.get("numempleado") or "").strip(), {})
                for campo, _label in DATOS_PERSONALES_EXPORT_FIELDS:
                    r[f"dp_{campo}"] = registro_dp.get(campo, "")
            columnas = list(columnas) + [
                {"key": f"dp_{campo}", "label": label} for campo, label in DATOS_PERSONALES_EXPORT_FIELDS
            ]

        try:
            fecha_legend = datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            fecha_legend = fecha

        buffer = generar_workbook_excel_con_fotos(
            columnas=columnas, rows=rows, incluir_fotos=incluir_fotos,
            sheet_name="Plantilla_Empleados", numero_empleado_key="numempleado",
            mono_keys=["posicion", "id_empleado", "rfc", "curp", "nivel", "codigo_presupuestal",
                       "ua", "cent", "dir", "subd", "jd", "depto", "numeral"],
            extra_legend=f"Plantilla histórica al {fecha_legend}",
        )
        return _excel_con_fotos_response(buffer, incluir_fotos, f"Plantilla de Empleados ({fecha})")


class ExportarMovimientosPersonalConFotosView(APIView):
    """
    Genera el Excel de Movimientos (personal, tab "Movimientos" —
    MovimientosPersonalTab.jsx), opcionalmente con fotografías embebidas.

    A diferencia de Detalle, MovimientosPersonalListView YA filtra/ordena
    server-side vía query params (apply_dynamic_column_filters,
    apply_advanced_filters, search, sort_by) — este endpoint reusa esa MISMA
    lógica llamándola internamente con no_pagination forzado, en vez de
    reimplementar los filtros (evita que ambas implementaciones diverjan).
    """

    view_permission = "authentication.view_plantilla_movimientos"

    def get(self, request):
        from .excel_fotos import generar_workbook_excel_con_fotos

        try:
            columnas = json.loads(request.query_params.get("columnas", "") or "[]")
        except (ValueError, TypeError):
            columnas = []
        incluir_fotos = (
            request.query_params.get("incluir_fotos", "false").strip().lower() == "true"
            and request.user.has_perm("authentication.view_plantilla_movimientos_foto")
        )

        if not columnas:
            return Response({"error": "Falta 'columnas'."}, status=status.HTTP_400_BAD_REQUEST)

        request._request.GET = request._request.GET.copy()
        request._request.GET["no_pagination"] = "true"
        inner_response = MovimientosPersonalListView().get(request)
        rows = list(inner_response.data) if inner_response.status_code == 200 else []

        # Tope de seguridad: a diferencia de Detalle (censo actual, acotado
        # por el tamaño de la plantilla), esta tabla es un LOG histórico sin
        # límite natural — medido en ~152,160 filas hoy, y solo crece. Sin
        # filtro, incluir fotos aquí sería una operación desproporcionada
        # (confirmado: una prueba sin filtros nunca terminó en varios
        # minutos). Se exige que el usuario acote antes de incluir fotos.
        if incluir_fotos and len(rows) > 5000:
            return Response(
                {"error": f"El conjunto filtrado tiene {len(rows)} filas — el máximo para incluir fotografías es 5,000. Aplica un filtro más específico."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for row in rows:
            row["nombre"] = " ".join(filter(None, [row.get("nombre"), row.get("ap_pat"), row.get("ap_mat")]))

        buffer = generar_workbook_excel_con_fotos(
            columnas=columnas, rows=rows, incluir_fotos=incluir_fotos,
            sheet_name="Movimientos_Personal", numero_empleado_key="num_empleado",
            mono_keys=["posicion", "num_empleado", "rfc", "curp", "nv_jerarquico", "un", "id_persona"],
        )
        return _excel_con_fotos_response(buffer, incluir_fotos, "Movimientos_Personal")


class ExportarBajasConFotosView(APIView):
    """
    Genera el Excel de Empleados Bajas, opcionalmente con fotografías
    embebidas. BajasTab filtra 100% client-side sobre `bajasData` (sin
    refetch), así que el frontend manda la lista de `id` (PK de BajasSig)
    de las filas visibles tras su filtro/orden actual.
    """

    view_permission = "authentication.view_plantilla_bajas"

    def post(self, request):
        from .excel_fotos import generar_workbook_excel_con_fotos

        ids = request.data.get("ids") or []
        columnas = request.data.get("columnas") or []
        incluir_fotos = bool(request.data.get("incluir_fotos")) and request.user.has_perm(
            "authentication.view_plantilla_bajas_foto"
        )

        if not ids or not columnas:
            return Response({"error": "Faltan 'ids' o 'columnas'."}, status=status.HTTP_400_BAD_REQUEST)

        rows_by_id = {r["id"]: r for r in BajasSig.objects.filter(id__in=ids).values()}
        rows = [rows_by_id[i] for i in ids if i in rows_by_id]

        # Tope de seguridad — histórico de bajas medido en ~5,710 filas hoy;
        # margen razonable para su crecimiento (ver mismo criterio en Detalle
        # y Movimientos).
        if incluir_fotos and len(rows) > 8000:
            return Response(
                {"error": f"El conjunto tiene {len(rows)} filas — el máximo para incluir fotografías es 8,000."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        buffer = generar_workbook_excel_con_fotos(
            columnas=columnas, rows=rows, incluir_fotos=incluir_fotos,
            sheet_name="Empleados_Bajas", numero_empleado_key="no_empleado",
            mono_keys=["posicion", "no_empleado", "partida", "grado", "escala", "nivel"],
        )
        return _excel_con_fotos_response(buffer, incluir_fotos, "Empleados_Bajas")


# Whitelist de permisos "ver fotografía" aceptados por
# ExportarEmpleadosPorPosicionConFotosView — el frontend manda cuál aplica
# (el mismo que ya usa para decidir `canViewPhoto` en EmployeesModal), pero
# se valida contra esta lista fija antes de usarlo en has_perm(...) para que
# el cliente no pueda pedir la verificación de un permiso arbitrario.
PERMISOS_FOTO_VALIDOS = {
    "view_plantilla_detalle_foto",
    "view_plantilla_estatus_nomina_foto",
    "view_plantilla_mov_posiciones_foto",
    "view_plantilla_movimientos_foto",
    "view_plantilla_bajas_foto",
    "view_plantilla_geografia_foto",
}


class ExportarEmpleadosPorPosicionConFotosView(APIView):
    """
    Genera un Excel, opcionalmente con fotografías, para las filas
    actualmente mostradas en el componente compartido `EmployeesModal.jsx`
    (listado de empleados por nivel/estatus o por UA) — usado tanto por
    Estatus Nómina (drill-down de un donut) como por Cuadros de Vacancia
    (drill-down "Ocupación por Nivel Jerárquico"/"por Nivel Tabular").

    Proceso NUEVO y aislado — NO toca los exports complejos ya existentes
    de esos dos tabs (ExportarEstatusExcelView/generar_excel_estatus_task,
    generateCuadroVacanciaExcel), que siguen intactos.

    Ambas fuentes de EmployeesModal comparten la misma clave de negocio
    (`posicion`) contra EMPLEADOS_COMPLETOS_SIG (ver EmpleadosPorNivelYEstatusView
    y mapVacanteRowToEmployeeRow), así que se reutiliza el mismo query y el
    mismo generador de Fase 1 (ExportarPlantillaDetalleConFotosView) — las
    filas vacantes simplemente no tienen `numempleado` resoluble y no llevan
    foto, sin necesitar lógica extra de "ocupada vs vacante".
    """

    view_permission = (
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
    )

    def post(self, request):
        from .excel_fotos import generar_workbook_excel_con_fotos

        posiciones = request.data.get("posiciones") or []
        columnas = request.data.get("columnas") or []
        permiso_foto = request.data.get("permiso_foto") or ""

        incluir_fotos = (
            bool(request.data.get("incluir_fotos"))
            and permiso_foto in PERMISOS_FOTO_VALIDOS
            and request.user.has_perm(f"authentication.{permiso_foto}")
        )

        if not posiciones or not columnas:
            return Response({"error": "Faltan 'posiciones' o 'columnas'."}, status=status.HTTP_400_BAD_REQUEST)

        rows_by_posicion = {
            str(r["posicion"]): r
            for r in EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values()
        }
        rows = _aplicar_mapeos_detalle_excel(
            [rows_by_posicion[str(p)] for p in posiciones if str(p) in rows_by_posicion]
        )

        if incluir_fotos and len(rows) > 15000:
            return Response(
                {"error": f"El conjunto tiene {len(rows)} filas — el máximo para incluir fotografías es 15,000."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        buffer = generar_workbook_excel_con_fotos(
            columnas=columnas, rows=rows, incluir_fotos=incluir_fotos,
            sheet_name="Listado_Empleados", numero_empleado_key="numempleado",
            mono_keys=["posicion", "id_empleado", "rfc", "curp", "nivel", "codigo_presupuestal",
                       "ua", "cent", "dir", "subd", "jd", "depto", "numeral"],
            estado_nomina_key="estado_nomina", mapear_estado_nomina=_mapear_estado_nomina_excel,
        )
        return _excel_con_fotos_response(buffer, incluir_fotos, "Listado_Empleados")


class EmpleadosCompletosCeldaOverrideView(APIView):
    """
    Edición manual de una celda de EMPLEADOS_COMPLETOS_SIG desde el tab
    Detalle. Registra el cambio en CeldaOverride y lo aplica de inmediato
    sobre la fila viva (ver plantilla.celda_override). Se reaplica solo tras
    cada importar_zafiro, ya que la tabla se trunca y recarga cada 30 min.
    """

    edit_permission = "authentication.edit_plantilla_detalle"

    def post(self, request):
        posicion = request.data.get("posicion")
        columna = request.data.get("columna")
        valor_nuevo = request.data.get("valor_nuevo")
        if not posicion or not columna:
            return Response(
                {"detail": "posicion y columna son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            override = registrar_y_aplicar_override_empleado(
                posicion, columna, valor_nuevo, request.user
            )
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if override is None:
            # 8.10 QA: valor idéntico al actual — no se registró como cambio
            # (ver registrar_y_aplicar_override_empleado), así que tampoco se
            # invalida caché ni se notifica en tiempo real a otros usuarios.
            return Response({
                "posicion": posicion,
                "columna": columna,
                "valor_nuevo": valor_nuevo,
                "sin_cambios": True,
            })

        self._invalidar_cache_detalle()
        notificar_cambio_celda(
            posicion, columna, override.valor_nuevo, request.user, override.fecha_modificacion
        )

        return Response({
            "posicion": posicion,
            "columna": columna,
            "valor_nuevo": override.valor_nuevo,
            "valor_original": override.valor_original,
            "usuario": request.user.username,
            "fecha_modificacion": override.fecha_modificacion,
        })

    def delete(self, request):
        posicion = request.data.get("posicion")
        columna = request.data.get("columna")
        if not posicion or not columna:
            return Response(
                {"detail": "posicion y columna son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            borrar_contenido_celda(posicion, columna)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        self._invalidar_cache_detalle()
        from django.utils import timezone

        notificar_cambio_celda(posicion, columna, None, request.user, timezone.now())

        return Response({"posicion": posicion, "columna": columna, "valor_nuevo": None})

    @staticmethod
    def _invalidar_cache_detalle():
        cache.delete_many(["empleados_completos_activos_detalle", "active_employees_filtered"])
        # Variante cacheada por filtros oficio/nivel (ver EmpleadosCompletosActivosDetalleView.get,
        # cache_key = f"empleados_completos_activos_detalle_{oficio}_{nivel}") — misma barrida por
        # patrón que hace importar_zafiro tras cada import.
        try:
            import redis as redis_lib

            r = redis_lib.Redis.from_url(settings.CELERY_BROKER_URL)
            for key in r.scan_iter("*empleados_completos_activos_detalle_*"):
                r.delete(key)
        except Exception:
            logger.exception("Error al invalidar cache filtrada tras override de celda")


class _CeldaHistorialBaseView(APIView):
    """
    Base del modal "Historial de Cambios": historial de ediciones manuales
    (CeldaOverride) de una o más tablas. Solo lectura — no reaplica ni
    modifica nada (ver plantilla.celda_override.obtener_historial_overrides).

    Las subclases fijan `tabla` (valor de la columna `tabla` por el que se
    filtra, o una LISTA de varios — ver EmpleadosCompletosCeldaHistorialView,
    que unifica en un solo historial las 3 tablas editables desde Plantilla
    Detalle), `clave_key` (campo de `clave_negocio` que identifica la fila) y
    su propio `view_permission`.
    """

    tabla = None
    clave_key = "posicion"

    def get(self, request):
        search = (request.query_params.get("search") or "").strip() or None
        columna = (request.query_params.get("columna") or "").strip() or None
        posicion = (request.query_params.get("posicion") or "").strip() or None

        activo_param = request.query_params.get("activo")
        activo = None
        if activo_param in ("true", "1"):
            activo = True
        elif activo_param in ("false", "0"):
            activo = False

        try:
            limit = min(max(int(request.query_params.get("limit", 100)), 1), 500)
        except (TypeError, ValueError):
            limit = 100
        try:
            offset = max(int(request.query_params.get("offset", 0)), 0)
        except (TypeError, ValueError):
            offset = 0

        resultados, total = obtener_historial_overrides(
            tabla=self.tabla,
            clave_key=self.clave_key,
            search=search,
            columna=columna,
            posicion=posicion,
            activo=activo,
            limit=limit,
            offset=offset,
        )

        return Response(
            {
                "count": total,
                "limit": limit,
                "offset": offset,
                "resultados": [
                    serializar_override(o, self.clave_key) for o in resultados
                ],
                "estadisticas": obtener_estadisticas_overrides(self.tabla),
            },
            status=status.HTTP_200_OK,
        )


class EmpleadosCompletosCeldaHistorialView(_CeldaHistorialBaseView):
    """
    Historial de ediciones manuales sobre las 3 tablas editables desde el
    tab Plantilla Detalle: EMPLEADOS_COMPLETOS_SIG (columnas normales),
    PLANTILLA_QUINCENAL (columnas AL–AV) y MOV_POS (Fecha de Anuencia,
    unificada con el override de Mov. Posiciones) — se muestran juntas en un
    solo historial porque las 3 se editan desde la misma pantalla.
    """

    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_mov_posiciones",
    )
    tabla = [TABLA_EMPLEADOS, TABLA_QUINCENAL, TABLA_MOV_POS]
    clave_key = "posicion"


class MovPosCeldaHistorialView(_CeldaHistorialBaseView):
    """
    Historial de ediciones manuales sobre MOV_POS (tab Mov. Posiciones): hoy
    solo `fecha_anuencia`, ver MovPosFechaAnuenciaOverrideView. La clave de
    negocio es `no_pos_actual`, no `posicion`.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"
    tabla = TABLA_MOV_POS
    clave_key = "no_pos_actual"


class EmpleadosPorNivelYEstatusView(APIView):
    """
    Retorna empleados filtrados por nivel y estado de nómina.
    Query params: nivel, estado_nomina
    Ejemplo: /api/empleados/?nivel=C1&estado_nomina=Activo
    """

    view_permission = "authentication.view_plantilla_estatus_nomina"

    def get(self, request):
        nivel = request.query_params.get("nivel")
        estado_nomina = request.query_params.get("estado_nomina")

        if not nivel or not estado_nomina:
            return Response(
                {"error": "Los parámetros 'nivel' y 'estado_nomina' son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Map full names back to letters for EmpleadosCompletosSig DB query
        estatus_map_reverse = {
            "Activo": "A",
            "Suspendido": "S",
            "Licencia": "L",
            "Licencia Médica": "P",
            "Vacante": "V",
        }
        db_estado_nomina = estatus_map_reverse.get(estado_nomina, estado_nomina)

        # Categorías del drill-down "Ocupadas vs Vacantes por familia de nivel"
        # (DesgloseJerarquicoCharts.jsx, gráfica 3): mismas reglas de negocio de
        # partida presupuestal que el resto del cuadro de vacancia, pero exactas
        # (no la heurística por prefijo de Posición que usan las gráficas).
        # Nueva Creación es subconjunto de Eventuales por partida, así que
        # Eventuales la excluye para no duplicar registros entre categorías.
        # Igual criterio para las 3 divisiones de Ocupadas, solo que con
        # estado_nomina != ' ' en vez de = ' '.
        CATEGORIAS_VACANCIA = {
            "Ocupadas",
            "Ocupadas Eventuales",
            "Ocupadas Permanentes",
            "Ocupadas Eventuales Nueva Creación",
            "Vacantes Eventuales",
            "Vacantes Permanentes",
            "Vacantes Eventuales Nueva Creación",
        }

        try:
            # 1. Obtener posiciones actualmente activas
            active_position_codes = obtener_posiciones_activas()

            # 2. Obtener los registros de EMPLEADOS_COMPLETOS_SIG correspondientes al nivel y estatus
            base_qs = EmpleadosCompletosSig.objects.filter(
                posicion__in=active_position_codes
            )

            if nivel == "SIN NIVEL":
                base_qs = base_qs.filter(Q(nivel__isnull=True) | Q(nivel__exact=""))
            else:
                base_qs = base_qs.filter(nivel=nivel)

            if estado_nomina in CATEGORIAS_VACANCIA:
                base_qs = base_qs.annotate(
                    partida_trim=Trim("partida"), posicion_trim=Trim("posicion")
                )
                if estado_nomina == "Ocupadas":
                    queryset = base_qs.exclude(estado_nomina=" ")
                elif estado_nomina == "Ocupadas Permanentes":
                    queryset = base_qs.exclude(estado_nomina=" ").filter(
                        partida_trim="11301", posicion_trim__startswith="103"
                    )
                elif estado_nomina == "Ocupadas Eventuales Nueva Creación":
                    queryset = base_qs.exclude(estado_nomina=" ").filter(
                        partida_trim="12201",
                        posicion_trim__startswith="2026",
                    )
                elif estado_nomina == "Ocupadas Eventuales":
                    queryset = (
                        base_qs.exclude(estado_nomina=" ")
                        .filter(partida_trim="12201")
                        .exclude(posicion_trim__startswith="2026")
                    )
                elif estado_nomina == "Vacantes Permanentes":
                    queryset = base_qs.filter(
                        estado_nomina=" ", partida_trim="11301"
                    )
                elif estado_nomina == "Vacantes Eventuales Nueva Creación":
                    queryset = base_qs.filter(
                        estado_nomina=" ",
                        partida_trim="12201",
                        posicion_trim__startswith="2026",
                    )
                else:  # "Vacantes Eventuales"
                    queryset = base_qs.filter(
                        estado_nomina=" ", partida_trim="12201"
                    ).exclude(posicion_trim__startswith="2026")
            elif estado_nomina == "Vacante":
                # La UI agrupa bajo "Vacante" todo lo que no sea A, S, L, P
                queryset = base_qs.exclude(
                    estado_nomina__in=["A", "a", "S", "s", "L", "l", "P", "p"]
                )
            else:
                queryset = base_qs.filter(estado_nomina__iexact=db_estado_nomina)

            return Response(
                {
                    "total": queryset.count(),
                    "resultados": list(queryset.values()),
                },
                status=status.HTTP_200_OK,
            )
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class OcupacionPorOficiosResumenView(APIView):
    """
    Devuelve un resumen dinámico de ocupación por 'Of. De Solicitud' con desglose por Nivel.

    Usa Django ORM para construir dinámicamente el equivalente a la query SQL con ROLLUP.
    Las columnas se generan automáticamente según los valores únicos de Nivel encontrados.

    Retorna:
    {
        "filas": [
            {
                "Of. De Solicitud": "Oficina A",
                "A212": 5,
                "D312": 3,
                "(vacío)": 1,
                "Total Resultado": 11
            },
            ...
        ],
        "columnas": ["Of. De Solicitud", "A212", "D312", "(vacío)", "Total Resultado"],
        "total_general": 1857
    }
    """

    # Alimenta las 3 tabs de Ocupación (Sankey/Tabla/Estadísticas) con el mismo dato.
    view_permission = (
        "authentication.view_ocupacion_sankey",
        "authentication.view_ocupacion_tabla",
        "authentication.view_ocupacion_estadisticas",
    )

    @staticmethod
    def obtener_resumen_dinamico():
        """
        Construye dinámicamente el resumen usando Django ORM de forma optimizada en O(1) queries.
        Equivalente a la query SQL con GROUP BY ... WITH ROLLUP
        """

        # 1. Obtener los niveles únicos (DISTINCT)
        niveles_unicos = (
            Plantilla1800Plazas.objects.exclude(Q(nivel__isnull=True) | Q(nivel=""))
            .values_list("nivel", flat=True)
            .distinct()
            .order_by("nivel")
        )
        niveles = list(niveles_unicos)

        # 2. Hacer una consulta agrupada para obtener los conteos totales de oficina y nivel
        conteo_agrupado = (
            Plantilla1800Plazas.objects.values("of_de_solicitud", "nivel")
            .annotate(cantidad=Count("id"))
            .order_by("of_de_solicitud", "nivel")
        )

        # 3. Procesar en memoria en Python los totales
        data_dict = {}
        for item in conteo_agrupado:
            oficina = item["of_de_solicitud"] or "(vacío)"
            nivel = item["nivel"]
            if not nivel:
                nivel = "(vacío)"
            cantidad = item["cantidad"]

            if oficina not in data_dict:
                data_dict[oficina] = {}
            data_dict[oficina][nivel] = data_dict[oficina].get(nivel, 0) + cantidad

        # 4. Obtener los conteos de ocupados agrupados
        conteo_ocupado = (
            Plantilla1800Plazas.objects.exclude(
                Q(rfc__isnull=True)
                | Q(rfc="")
                | Q(curp__isnull=True)
                | Q(curp="")
                | Q(num_empleado__isnull=True)
                | Q(num_empleado="")
                | Q(nombres__isnull=True)
                | Q(nombres="")
            )
            .values("of_de_solicitud", "nivel")
            .annotate(cantidad=Count("id"))
            .order_by("of_de_solicitud", "nivel")
        )

        ocupadas_dict = {}
        for item in conteo_ocupado:
            oficina = item["of_de_solicitud"] or "(vacío)"
            nivel = item["nivel"]
            if not nivel:
                nivel = "(vacío)"
            cantidad = item["cantidad"]

            if oficina not in ocupadas_dict:
                ocupadas_dict[oficina] = {}
            ocupadas_dict[oficina][nivel] = (
                ocupadas_dict[oficina].get(nivel, 0) + cantidad
            )

        # 5. Construir las filas
        filas = []
        totales_generales = {nivel: 0 for nivel in niveles}
        totales_generales["(vacío)"] = 0

        totales_ocupados_generales = {nivel: 0 for nivel in niveles}
        totales_ocupados_generales["(vacío)"] = 0

        total_gral = 0
        total_ocupadas_gral = 0

        # Ordenar oficinas
        oficinas_ordenadas = sorted(
            list(data_dict.keys()), key=lambda x: (x == "(vacío)", x)
        )

        for oficina in oficinas_ordenadas:
            conteos_nivel = data_dict[oficina]
            total_oficina = sum(conteos_nivel.values())

            ocupadas_oficina = ocupadas_dict.get(oficina, {})
            total_ocupadas_oficina = sum(ocupadas_oficina.values())

            fila = {"Of. De Solicitud": oficina}

            for nivel in niveles:
                count_nivel = conteos_nivel.get(nivel, 0)
                fila[nivel] = count_nivel
                totales_generales[nivel] += count_nivel

                count_ocupadas = ocupadas_oficina.get(nivel, 0)
                fila[f"ocupadas_{nivel}"] = count_ocupadas
                totales_ocupados_generales[nivel] += count_ocupadas

            # Nivel vacío en esta oficina
            count_vacio = conteos_nivel.get("(vacío)", 0)
            fila["(vacío)"] = count_vacio
            totales_generales["(vacío)"] += count_vacio

            count_ocupadas_vacio = ocupadas_oficina.get("(vacío)", 0)
            fila["ocupadas_(vacío)"] = count_ocupadas_vacio
            totales_ocupados_generales["(vacío)"] += count_ocupadas_vacio

            fila["Total Resultado"] = total_oficina
            fila["ocupadas_Total Resultado"] = total_ocupadas_oficina

            total_gral += total_oficina
            total_ocupadas_gral += total_ocupadas_oficina

            filas.append(fila)

        # 6. Agregar fila de totales (equivalente a ROLLUP)
        fila_total = {"Of. De Solicitud": "Total Resultado"}
        for nivel in niveles:
            fila_total[nivel] = totales_generales[nivel]
            fila_total[f"ocupadas_{nivel}"] = totales_ocupados_generales[nivel]

        fila_total["(vacío)"] = totales_generales["(vacío)"]
        fila_total["ocupadas_(vacío)"] = totales_ocupados_generales["(vacío)"]

        fila_total["Total Resultado"] = total_gral
        fila_total["ocupadas_Total Resultado"] = total_ocupadas_gral

        filas.append(fila_total)

        # 7. Definir columnas en el orden correcto
        columnas = ["Of. De Solicitud"] + niveles + ["(vacío)", "Total Resultado"]

        # 8. Conteo de posiciones ocupadas que inician con 2026
        ocupadas_2026 = (
            Plantilla1800Plazas.objects.filter(posición__startswith="2026")
            .exclude(rfc__isnull=True)
            .exclude(rfc__exact="")
            .exclude(curp__isnull=True)
            .exclude(curp__exact="")
            .exclude(num_empleado__isnull=True)
            .exclude(num_empleado__exact="")
            .exclude(nombres__isnull=True)
            .exclude(nombres__exact="")
            .count()
        )

        # 9. Conteo de empleados ocupados en EmpleadosCompletosSig que inician con 2026
        ocupadas_sig = EmpleadosCompletosSig.objects.filter(
            val_estat="Ocupada", posicion__startswith="2026"
        ).count()

        return {
            "filas": filas,
            "columnas": columnas,
            "total_general": Plantilla1800Plazas.objects.count(),
            "ocupadas_2026": ocupadas_2026,
            "ocupadas_sig": ocupadas_sig,
        }

    def get(self, request, *args, **kwargs):
        """
        GET /api/plantilla/ocupacion_por_oficios_resumen/

        Retorna el resumen dinámico de ocupación por oficios con desglose por nivel,
        construido usando Django ORM.
        """
        try:
            datos = self.obtener_resumen_dinamico()
            return Response(datos, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error al generar el resumen de ocupación por oficios")
            return Response(
                {"error": "Error al generar el resumen de ocupación por oficios"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RegistrosPorOficio1800PlazasView(APIView):
    """
    Devuelve los registros detallados del modelo Plantilla1800Plazas filtrados por 'Of. De Solicitud' y opcionalmente por 'Nivel'.
    """

    # Mismo dataset que OcupacionPorOficiosResumenView, filtrado por oficio/nivel.
    view_permission = (
        "authentication.view_ocupacion_sankey",
        "authentication.view_ocupacion_tabla",
        "authentication.view_ocupacion_estadisticas",
    )

    def get(self, request):
        no_oficio = request.query_params.get("oficio")
        nivel = request.query_params.get("nivel")
        resumen = request.query_params.get("resumen") == "true"

        if not no_oficio and not nivel:
            return Response(
                {
                    "error": "No se especificó ningun filtro. Se requiere al menos 'oficio' o 'nivel' como parámetro de consulta."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = Plantilla1800Plazas.objects.all()

        # Filtro dinámico por oficio
        if no_oficio:
            if no_oficio == "(vacío)":
                queryset = queryset.filter(
                    Q(of_de_solicitud__isnull=True) | Q(of_de_solicitud="")
                )
            else:
                queryset = queryset.filter(of_de_solicitud=no_oficio)

        # Filtro dinámico por nivel
        if nivel:
            queryset = queryset.filter(nivel=nivel)

        if resumen:
            # Agrupamos por nivel y oficio, y contamos
            resultados = list(
                queryset.values("nivel", "of_de_solicitud")
                .annotate(total=Count("*"))
                .order_by("of_de_solicitud", "nivel")
            )
            total_registros = sum(r["total"] for r in resultados)
            total_count = len(resultados)
        else:
            # Traemos todos los campos (registros completos)
            resultados = list(queryset.order_by("nivel").values())
            total_registros = len(resultados)
            total_count = total_registros

        if not resultados:
            return Response(
                {
                    "mensaje": "No se encontraron registros con los filtros proporcionados."
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        # Respuesta consolidada
        data = {
            "total_registros": total_registros,
            "resultados": resultados,
            "tipo_vista": "resumen" if resumen else "detallado",
        }
        if not resumen:
            data["total"] = total_count
        else:
            data["total_grupos"] = total_count

        if no_oficio:
            data["oficio"] = no_oficio
        if nivel:
            data["nivel"] = nivel

        return Response(data, status=status.HTTP_200_OK)


class Plantilla1800PlazasListView(APIView):
    """
    Vista para listar y actualizar registros de la plantilla de 1800 plazas.
    """

    view_permission = "authentication.edit_ocupacion_plazas"

    CACHE_KEY = "plantilla_1800_list_json"

    def get(self, request):
        # El dataset cambia sólo en el sync de ZAFIRO o en PATCH. Servimos los
        # bytes JSON cacheados (orjson) para evitar re-consultar y re-serializar
        # ~12k filas en cada request.
        payload = cache.get(self.CACHE_KEY)
        if payload is None:
            resultados = list(
                Plantilla1800Plazas.objects.all().order_by("id").values()
            )
            payload = orjson_dumps(resultados)
            cache.set(self.CACHE_KEY, payload, 3600)
        return orjson_response(payload)

    def patch(self, request):
        """
        Actualización parcial de registros.
        Se espera un objeto con el ID y los campos a cambiar, o una lista de ellos.
        """
        data = request.data
        if not isinstance(data, list):
            data = [data]

        errores = []

        # Campos reales del modelo (no persistir atributos que no son columnas).
        valid_fields = {f.name for f in Plantilla1800Plazas._meta.get_fields()}

        # Normaliza los ids a int (la PK es entera) y reporta los inválidos.
        ids = []
        for item in data:
            rid = item.get("id")
            if rid in (None, ""):
                errores.append({"error": "ID no proporcionado", "item": item})
                continue
            try:
                ids.append(int(rid))
            except (TypeError, ValueError):
                errores.append(
                    {"error": f"Registro con ID {rid} no existe", "id": rid}
                )

        # 1 sola query para traer todos los registros (antes: 1 get por item).
        registros = Plantilla1800Plazas.objects.in_bulk(ids)

        a_actualizar = []
        campos = set()
        for item in data:
            rid = item.get("id")
            if rid in (None, ""):
                continue
            try:
                key = int(rid)
            except (TypeError, ValueError):
                continue
            registro = registros.get(key)
            if registro is None:
                errores.append(
                    {"error": f"Registro con ID {rid} no existe", "id": rid}
                )
                continue
            for field, value in item.items():
                if field != "id" and field in valid_fields:
                    setattr(registro, field, value)
                    campos.add(field)
            a_actualizar.append(registro)

        # 1 sola query (en lotes) para todas las actualizaciones
        # (antes: 1 save por item).
        if a_actualizar and campos:
            Plantilla1800Plazas.objects.bulk_update(
                a_actualizar, list(campos), batch_size=500
            )
        actualizados = len(a_actualizar)

        cache.delete(self.CACHE_KEY)
        return Response(
            {
                "mensaje": f"{actualizados} registros actualizados correctamente.",
                "errores": errores,
            },
            status=status.HTTP_200_OK if not errores else status.HTTP_207_MULTI_STATUS,
        )


class EmpleadosEstatusPorNivelUaView(APIView):
    """
    Vista para resumir el estatus de la nómina por nivel y por unidad administrativa (UA)
    de los empleados correspondientes a las posiciones activas.
    """

    view_permission = "authentication.view_plantilla_estatus_nomina"

    def get(self, request, *args, **kwargs):
        cache_key = "empleados_estatus_por_nivel_ua"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        try:
            # 1. Obtener posiciones actualmente activas
            active_position_codes = obtener_posiciones_activas()

            # 2. Obtener todos los registros de EMPLEADOS_COMPLETOS_SIG en esas posiciones
            active_employees = EmpleadosCompletosSig.objects.filter(
                posicion__in=active_position_codes
            )

            # 3. Agrupación por Unidad Administrativa, Nivel y Estado de Nómina
            # — una sola query sobre active_employees; por_nivel se deriva
            # sumando sobre las UA en vez de repetir el mismo scan con un
            # segundo GROUP BY (nivel, estado_nomina) equivalente.
            ua_data = active_employees.values(
                "unidad_administrativa", "nivel", "estado_nomina"
            ).annotate(count=Count("id"))

            por_nivel = {}
            por_ua = {}
            for item in ua_data:
                ua_name = item["unidad_administrativa"] or "SIN UA"
                nv = item["nivel"] or "SIN NIVEL"
                est = item["estado_nomina"] or "SIN ESTATUS"
                count = item["count"]

                por_nivel.setdefault(nv, {})
                por_nivel[nv][est] = por_nivel[nv].get(est, 0) + count

                por_ua.setdefault(ua_name, {}).setdefault(nv, {})
                por_ua[ua_name][nv][est] = count

            res_data = {"por_nivel": por_nivel, "por_ua": por_ua}
            cache.set(cache_key, res_data, 1200)
            return Response(res_data, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EmpleadosDistribucionGeograficaView(APIView):
    """
    Retorna la distribución geográfica agrupada por coordenadas para los empleados activos.
    """

    view_permission = "authentication.view_plantilla_geografia"

    def get(self, request, *args, **kwargs):
        cache_key = "empleados_distribucion_geografica"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        try:
            active_position_codes = obtener_posiciones_activas()

            # Agregacion en SQL: agrupa ~13k empleados por coordenada en la DB
            # (antes se iteraban todos en Python). GROUP_CONCAT(DISTINCT) trae los
            # valores unicos por grupo; el separador 0x1f evita choques con comas.
            with connection.cursor() as cursor:
                cursor.execute("SET SESSION group_concat_max_len = 1000000")

            grupos = (
                EmpleadosCompletosSig.objects.filter(posicion__in=active_position_codes)
                .exclude(latitud__isnull=True)
                .exclude(latitud="")
                .exclude(longitud__isnull=True)
                .exclude(longitud="")
                .annotate(lat=Trim("latitud"), lng=Trim("longitud"))
                .values("lat", "lng")
                .annotate(
                    n=Count("id"),
                    descripciones=GroupConcat("descripcion_ubicacion"),
                    aduanas=GroupConcat("aduana"),
                    tipos=GroupConcat("tipo"),
                    uas=GroupConcat("unidad_administrativa"),
                )
            )

            def _split(s):
                return [x for x in (s.split("\x1f") if s else []) if x]

            resultados = []
            for g in grupos:
                lat, lng = g["lat"], g["lng"]
                if not lat or not lng:
                    continue
                try:
                    float(lat)
                    float(lng)
                except (ValueError, TypeError):
                    continue

                descripciones = _split(g["descripciones"])
                aduanas = _split(g["aduanas"])
                tipos = _split(g["tipos"])
                uas = _split(g["uas"])

                is_aduana = any(a.strip().upper().startswith("ADUANA") for a in aduanas)
                aduana_principal = next(
                    (a for a in aduanas if a.strip().upper().startswith("ADUANA")),
                    aduanas[0] if aduanas else "",
                )
                tipo_principal = tipos[0] if tipos else ""
                desc_principal = descripciones[0] if descripciones else ""

                nombre_principal = (
                    aduana_principal if (is_aduana and aduana_principal) else desc_principal
                )
                if not nombre_principal and descripciones:
                    nombre_principal = descripciones[0]

                resultados.append(
                    {
                        "latitud": float(lat),
                        "longitud": float(lng),
                        "nombre": nombre_principal or "Ubicación sin nombre",
                        "is_aduana": is_aduana,
                        "tipo": tipo_principal,
                        "count": g["n"],
                        "descripciones": descripciones,
                        "aduanas": aduanas,
                        "tipos": tipos,
                        "uas": uas,
                    }
                )

            cache.set(cache_key, resultados, 1200)
            return Response(resultados, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def get_mov_pos_stats():
    cache_key = "mov_pos_card_stats"
    stats = cache.get(cache_key)
    if stats is None:
        from django.db import connection

        query = """
            SELECT
                (SELECT COUNT(*) FROM MOV_POS) as total_movimientos,
                COUNT(*) as todas_posiciones,
                SUM(CASE WHEN `Estado Psn` = 'A' THEN 1 ELSE 0 END) as posiciones_activas,
                SUM(CASE WHEN `Estado Psn` = 'I' THEN 1 ELSE 0 END) as posiciones_inactivas
            FROM MOV_POS_LATEST;
        """
        try:
            with connection.cursor() as cursor:
                cursor.execute(query)
                row = cursor.fetchone()
                if row:
                    stats = {
                        "total_movimientos": int(row[0]) if row[0] is not None else 0,
                        "todas_posiciones": int(row[1]) if row[1] is not None else 0,
                        "posiciones_activas": int(row[2]) if row[2] is not None else 0,
                        "posiciones_inactivas": int(row[3])
                        if row[3] is not None
                        else 0,
                    }
                else:
                    stats = {
                        "total_movimientos": 0,
                        "todas_posiciones": 0,
                        "posiciones_activas": 0,
                        "posiciones_inactivas": 0,
                    }
        except Exception:
            stats = {
                "total_movimientos": 0,
                "todas_posiciones": 0,
                "posiciones_activas": 0,
                "posiciones_inactivas": 0,
            }
        cache.set(cache_key, stats, 600)  # Cache for 10 minutes
    return stats


class MovPosPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 10000

    def get_paginated_response(self, data):
        stats = get_mov_pos_stats()
        return Response(
            {
                "next": self.get_next_link(),
                "previous": self.get_previous_link(),
                "count": self.page.paginator.count,
                "results": data,
                "stats": stats,
            }
        )


class MovPosDetalleView(APIView):
    view_permission = "authentication.view_plantilla_mov_posiciones"
    pagination_class = MovPosPagination

    def get(self, request, *args, **kwargs):
        from django.db import connection
        from django.db.models import Count, Q
        from django.db.models.functions import Trim

        from .models import MovPos, Plantilla1800Plazas

        # FECHA_OCUPACION/ID_REGISTRO_DES_FECHA_OCUPACION se recalculan una
        # sola vez por import, en InvalidarCacheZafiroView (justo cuando
        # Celery avisa que importar_zafiro terminó) — no aquí. Antes este
        # GET disparaba `CALL sp_actualizar_fecha_ocupacion_mov_pos()` la
        # primera vez que alguien abría este tab tras cada import (probe +
        # CALL lazy), sin ningún lock: si dos requests llegaban a la vez,
        # ambas disparaban el SP en paralelo, y su UPDATE (con window
        # functions sobre cp_tbl_mov_completo_29_05_26) tardaba varios
        # minutos sosteniendo locks de fila en MOV_POS — chocando con
        # `_reaplicar_prioridad_nivel_jerarquico` y los pasos de vacancia del
        # pipeline de ZAFIRO (error 1205 Lock wait timeout, ver ZafiroBitacora).

        queryset = MovPos.objects.all()
        # `fecha_anuencia` = fecha_vacancia + 30 días por default, salvo que el
        # usuario haya editado manualmente esa fecha para la posición (ver
        # celda_override.get_fecha_anuencia_overrides_map) — se anota desde el
        # inicio para que participe en filtros/orden/distinct_field igual que
        # un campo real. Se guarda en una variable (no solo inline) para
        # reusarla más abajo en corregir_fecha_anuencia_row, que necesita el
        # mismo mapa para marcar `fecha_anuencia_override` por fila.
        fecha_anuencia_overrides = get_fecha_anuencia_overrides_map()
        fecha_anuencia_baseline = get_fecha_anuencia_baseline_map()
        fecha_anuencia_overrides_texto = get_fecha_anuencia_overrides_texto_map()
        fecha_anuencia_baseline_texto = get_fecha_anuencia_baseline_texto_map()
        queryset = annotate_fecha_anuencia(
            queryset, overrides=fecha_anuencia_overrides, baseline=fecha_anuencia_baseline
        )

        oficio = request.query_params.get("oficio")
        nivel = request.query_params.get("nivel")

        if oficio or nivel:
            posiciones_qs = Plantilla1800Plazas.objects.all()
            if oficio:
                if oficio == "(vacío)":
                    posiciones_qs = posiciones_qs.filter(
                        Q(of_de_solicitud__isnull=True) | Q(of_de_solicitud="")
                    )
                else:
                    posiciones_qs = posiciones_qs.filter(of_de_solicitud=oficio)
            if nivel:
                posiciones_qs = posiciones_qs.filter(nivel=nivel)

            # Subquery en lugar de materializar miles de posiciones en una lista
            # Python + un IN gigante (evita armar/transferir la lista y deja que
            # MySQL resuelva el filtro).
            queryset = queryset.filter(
                no_pos_actual__in=posiciones_qs.values("posición")
            )

        # is_latest filter (defaults to True unless explicitly requested as 'false')
        is_latest = request.query_params.get("is_latest", "true").lower() != "false"
        if is_latest:
            cache_key_latest = "latest_movpos_sub_ids"
            sub_ids = cache.get(cache_key_latest)
            if sub_ids is None:
                with connection.cursor() as cursor:
                    cursor.execute(LATEST_MOVPOS_RAW_SQL)
                    sub_ids = [row[0] for row in cursor.fetchall() if row[0]]
                cache.set(cache_key_latest, sub_ids, 600)  # Cache for 10 minutes
            queryset = queryset.filter(id__in=sub_ids)

        # Search query
        queryset = apply_text_search(
            queryset,
            request.query_params.get("search", ""),
            [
                "no_pos_actual",
                "motivo",
                "unidad_de_negocio",
                "unidad_adva",
                "puesto_ptal",
                "descr",
                "nombre_puesto",
            ],
        )

        # Dynamic Column Filters
        # "fecha_anuencia" no es un campo real del modelo (es la anotación de
        # annotate_fecha_anuencia) — se suma aquí para que filtros/orden/
        # distinct_field más abajo la traten igual que un campo real.
        valid_fields = [f.name for f in MovPos._meta.get_fields()] + ["fecha_anuencia"]
        text_fields = [
            f.name
            for f in MovPos._meta.get_fields()
            if f.get_internal_type() in ["CharField", "TextField"]
        ]

        _cache_columnas_derivadas = {}

        def _match_columna_derivada(valor, suffix, val_list):
            valor_str = str(valor if valor is not None else "")
            if suffix in (None, "in"):
                return valor_str in set(val_list)
            needle = str(val_list[0] if val_list else "").lower()
            s = valor_str.lower()
            if suffix == "icontains":
                return needle in s
            if suffix == "istartswith":
                return s.startswith(needle)
            if suffix == "iendswith":
                return s.endswith(needle)
            if suffix == "iexact":
                return s == needle
            return False

        def mov_pos_column_filter_resolver(base_field, suffix, val_list, is_exclude):
            # Sólo intercepta la selección múltiple del dropdown (__in / valor
            # único, que es como llega desde "Aplicar Filtro" en el front) —
            # ver docstring de apply_dynamic_column_filters. Sin esto,
            # `fecha_anuencia__in=[...]` se filtraba tal cual contra la
            # anotación DateField y tiraba 500 en cuanto la selección incluía
            # una categoría de texto (p. ej. "Nueva Creación") o el token de
            # vacío (bug reportado: aplicar el filtro no actualizaba la tabla).
            if base_field == "fecha_anuencia":
                posiciones_ocupadas = get_posiciones_ocupadas_set()
                wanted = set(val_list)
                match_pos = [
                    pos for pos, fa in queryset.values_list("no_pos_actual", "fecha_anuencia")
                    if resolve_fecha_anuencia_value(
                        pos, fa, posiciones_ocupadas,
                        fecha_anuencia_overrides_texto, fecha_anuencia_baseline_texto,
                    ) in wanted
                ]
                return Q(no_pos_actual__in=match_pos) if match_pos else Q(pk__in=[])

            # "codigo"/"denominacion_puesto"/"nivel_salarial"/
            # "salario_mensual_neto"/"tipo_contratacion"/"anuencia_anexo_nombre"
            # no son campos reales de MOV_POS — sin este resolver, la rama
            # genérica de apply_dynamic_column_filters los saltaba en
            # silencio (`base_field not in valid_fields`) y "Aplicar Filtro"
            # no hacía nada en esas columnas (el dropdown sí mostraba
            # opciones porque `distinct_field` ya las resolvía aparte, pero
            # aplicar el filtro quedaba mudo). Memoizado: si se filtra más de
            # una de estas columnas en la misma request, sólo se calcula una
            # vez (el valor de cada columna es propio de la posición, no
            # depende de qué tan angosto esté `queryset` en ese momento).
            if base_field not in CAMPOS_DERIVADOS_MOV_POS:
                return None
            if not _cache_columnas_derivadas:
                _cache_columnas_derivadas.update(_calcular_columnas_derivadas_mov_pos(queryset))
            # "anuencia_anexo_nombre" es multi-valor (una plaza puede estar en
            # varios Anexo 2 a la vez, ver `_get_mapa_codigos_en_anuencia`):
            # comparar contra el texto unido (`anuencia_anexo_nombre`) haría
            # que marcar la casilla "Anexo A" nunca matcheara una plaza que
            # está en "Anexo A, Anexo B" (el valor exacto de esa celda es la
            # combinación completa, no "Anexo A" sola) — hay que comparar
            # contra cada nombre de `anuencia_anexo_nombres` por separado.
            if base_field == "anuencia_anexo_nombre":
                match_pos = [
                    pos for pos, datos in _cache_columnas_derivadas.items()
                    if any(
                        _match_columna_derivada(n, suffix, val_list)
                        for n in (datos.get("anuencia_anexo_nombres") or [""])
                    )
                ]
            else:
                match_pos = [
                    pos for pos, datos in _cache_columnas_derivadas.items()
                    if _match_columna_derivada(datos.get(base_field), suffix, val_list)
                ]
            return Q(no_pos_actual__in=match_pos) if match_pos else Q(pk__in=[])

        queryset = apply_dynamic_column_filters(
            queryset, request, MovPos,
            extra_valid_fields={"fecha_anuencia", *CAMPOS_DERIVADOS_MOV_POS},
            computed_field_resolver=mov_pos_column_filter_resolver,
        )

        # "ocupacion" is a computed column (not a real model field), so the
        # generic Dynamic Column Filters loop above silently skips it.
        # Apply it explicitly here, before pagination/sorting/distinct.
        # Covers both the dropdown (__in) and the free-text column filter
        # (__icontains/__istartswith/__iendswith/__iexact, incl. exclude__).
        ocupacion_param_key = None
        for k in request.query_params.keys():
            base = k[9:] if k.startswith("exclude__") else k
            if base == "ocupacion" or base.startswith("ocupacion__"):
                ocupacion_param_key = k
                break

        if ocupacion_param_key:
            cache_key_ocupadas = "mov_pos_ocupadas_set"
            posiciones_ocupadas = cache.get(cache_key_ocupadas)
            if posiciones_ocupadas is None:
                with connection.cursor() as cursor:
                    cursor.execute(OCUPADAS_RAW_SQL)
                    posiciones_ocupadas = set(
                        [row[0] for row in cursor.fetchall() if row[0]]
                    )
                cache.set(cache_key_ocupadas, posiciones_ocupadas, 600)

            ocupacion_raw = request.query_params.get(ocupacion_param_key, "")
            is_exclude = ocupacion_param_key.startswith("exclude__")
            suffix = (
                ocupacion_param_key.split("__", 1)[1]
                if "__"
                in (ocupacion_param_key[9:] if is_exclude else ocupacion_param_key)
                else "in"
            )

            if suffix == "in":
                selected_vals = set(
                    v.strip() for v in ocupacion_raw.split(",") if v.strip()
                )
            else:
                # Free-text condition: evaluate against the two possible values.
                needle = ocupacion_raw.strip().lower()
                candidates = ["Ocupada", "Vacante"]
                if suffix in ("icontains",):
                    selected_vals = {c for c in candidates if needle in c.lower()}
                elif suffix in ("istartswith",):
                    selected_vals = {
                        c for c in candidates if c.lower().startswith(needle)
                    }
                elif suffix in ("iendswith",):
                    selected_vals = {
                        c for c in candidates if c.lower().endswith(needle)
                    }
                elif suffix in ("iexact",):
                    selected_vals = {c for c in candidates if c.lower() == needle}
                else:
                    selected_vals = set(candidates)

            if is_exclude:
                selected_vals = {"Ocupada", "Vacante"} - selected_vals

            want_ocupada = "Ocupada" in selected_vals
            want_vacante = "Vacante" in selected_vals

            if want_ocupada and not want_vacante:
                queryset = queryset.filter(no_pos_actual__in=list(posiciones_ocupadas))
            elif want_vacante and not want_ocupada:
                queryset = queryset.exclude(no_pos_actual__in=list(posiciones_ocupadas))
            elif not want_ocupada and not want_vacante:
                queryset = queryset.none()

        # "total_movimientos" is also a computed column (count of historical
        # rows per posicion), so it needs the same explicit handling.
        total_mov_raw = request.query_params.get(
            "total_movimientos__in"
        ) or request.query_params.get("total_movimientos")
        if total_mov_raw:
            selected_counts = set()
            for v in total_mov_raw.split(","):
                v = v.strip()
                if v:
                    try:
                        selected_counts.add(int(v))
                    except ValueError:
                        pass
            if selected_counts:
                pos_list = list(
                    queryset.values_list("no_pos_actual", flat=True).distinct()
                )
                full_counts = dict(
                    MovPos.objects.filter(no_pos_actual__in=pos_list)
                    .values("no_pos_actual")
                    .annotate(c=Count("id"))
                    .values_list("no_pos_actual", "c")
                )
                match_pos = [p for p, c in full_counts.items() if c in selected_counts]
                queryset = queryset.filter(no_pos_actual__in=match_pos)
            else:
                queryset = queryset.none()

        # Advanced filters (built from the "Filtros Avanzados" modal). Se aplica
        # ANTES de las ramas distinct_field de abajo: si no, "Vista actual" en el
        # dropdown de columna ignoraría las condiciones del modal (mismo bug que
        # en MovimientosPersonalListView). "ocupacion" y "total_movimientos" son
        # columnas calculadas (no campos reales del modelo), por eso
        # apply_advanced_filters() necesita este resolver; el resto es genérico
        # (ver apply_advanced_filters).
        _cache_columnas_derivadas_avanzados = {}

        def mov_pos_computed_resolver(column, condition, value):
            def get_posiciones_ocupadas():
                cache_key_ocupadas = "mov_pos_ocupadas_set"
                posiciones_ocupadas = cache.get(cache_key_ocupadas)
                if posiciones_ocupadas is None:
                    with connection.cursor() as cursor:
                        cursor.execute(OCUPADAS_RAW_SQL)
                        posiciones_ocupadas = set(
                            [row[0] for row in cursor.fetchall() if row[0]]
                        )
                    cache.set(cache_key_ocupadas, posiciones_ocupadas, 600)
                return posiciones_ocupadas

            def text_condition_matches(haystack, condition, needle):
                s = str(haystack).lower()
                n = str(needle).lower()
                if condition == "contains":
                    return n in s
                if condition == "not_contains":
                    return n not in s
                if condition == "starts_with":
                    return s.startswith(n)
                if condition == "not_starts_with":
                    return not s.startswith(n)
                if condition == "ends_with":
                    return s.endswith(n)
                if condition == "not_ends_with":
                    return not s.endswith(n)
                if condition == "equals":
                    return s == n
                if condition == "not_equals":
                    return s != n
                return False

            if column == "ocupacion":
                posiciones_ocupadas = get_posiciones_ocupadas()
                candidates = ["Ocupada", "Vacante"]
                selected = {
                    c for c in candidates if text_condition_matches(c, condition, value)
                }

                want_ocupada = "Ocupada" in selected
                want_vacante = "Vacante" in selected
                if want_ocupada and want_vacante:
                    return Q(no_pos_actual__isnull=False) | Q(no_pos_actual__isnull=True)
                if want_ocupada:
                    return Q(no_pos_actual__in=list(posiciones_ocupadas))
                if want_vacante:
                    return ~Q(no_pos_actual__in=list(posiciones_ocupadas))
                return Q(pk__in=[])

            if column == "total_movimientos":
                pos_list = list(
                    queryset.values_list("no_pos_actual", flat=True).distinct()
                )
                if not pos_list:
                    return Q(pk__in=[])
                full_counts = dict(
                    MovPos.objects.filter(no_pos_actual__in=pos_list)
                    .values("no_pos_actual")
                    .annotate(c=Count("id"))
                    .values_list("no_pos_actual", "c")
                )
                match_pos = [
                    p for p, c in full_counts.items()
                    if text_condition_matches(c, condition, value)
                ]
                if not match_pos:
                    return Q(pk__in=[])
                return Q(no_pos_actual__in=match_pos)

            # Mismas 6 columnas calculadas que en el resolver de "Aplicar
            # Filtro" de arriba (ver `mov_pos_column_filter_resolver`) —
            # aquí también faltaban, así que "Filtros Avanzados" tampoco
            # filtraba nada al usarlas.
            if column in CAMPOS_DERIVADOS_MOV_POS:
                if "datos" not in _cache_columnas_derivadas_avanzados:
                    _cache_columnas_derivadas_avanzados["datos"] = _calcular_columnas_derivadas_mov_pos(queryset)
                datos_por_pos = _cache_columnas_derivadas_avanzados["datos"]
                # Mismo motivo que en el resolver de "Aplicar Filtro" de
                # arriba: "anuencia_anexo_nombre" es multi-valor, así que
                # "es igual a 'Anexo A'" tiene que probar contra cada nombre
                # de la lista, no contra el texto unido de la celda.
                if column == "anuencia_anexo_nombre":
                    match_pos = [
                        pos for pos, datos in datos_por_pos.items()
                        if any(
                            text_condition_matches(n, condition, value)
                            for n in (datos.get("anuencia_anexo_nombres") or [""])
                        )
                    ]
                else:
                    match_pos = [
                        pos for pos, datos in datos_por_pos.items()
                        if text_condition_matches(datos.get(column, ""), condition, value)
                    ]
                if not match_pos:
                    return Q(pk__in=[])
                return Q(no_pos_actual__in=match_pos)

            return None

        queryset = apply_advanced_filters(
            # OJO: a diferencia de apply_dynamic_column_filters, aquí
            # `extra_valid_fields` hace que la columna se trate como campo
            # REAL en la lógica genérica — `computed_resolver` sólo se llama
            # para columnas que NO estén en `valid_fields` (ver
            # `build_condition_q`: "if column not in valid_fields"). Por eso
            # las 6 columnas calculadas de MOV_POS (CAMPOS_DERIVADOS_MOV_POS)
            # NO van aquí, igual que "ocupacion"/"total_movimientos" tampoco
            # van — deben quedar AFUERA de `valid_fields` para que caigan al
            # resolver en vez de a la rama genérica, que las trataría como un
            # campo real inexistente y filtraría en silencio sin resultado.
            queryset, request, MovPos, computed_resolver=mov_pos_computed_resolver,
            extra_valid_fields={"fecha_anuencia"},
        )

        # If distinct_field requested, return distinct values directly
        distinct_field = request.query_params.get("distinct_field", "").strip()

        # Special handling for computed columns not present in the model
        if distinct_field == "ocupacion":
            cache_key_ocupadas = "mov_pos_ocupadas_set"
            posiciones_ocupadas = cache.get(cache_key_ocupadas)
            if posiciones_ocupadas is None:
                with connection.cursor() as cursor:
                    cursor.execute(OCUPADAS_RAW_SQL)
                    posiciones_ocupadas = set(
                        [row[0] for row in cursor.fetchall() if row[0]]
                    )
                cache.set(cache_key_ocupadas, posiciones_ocupadas, 600)
            all_pos = list(queryset.values_list("no_pos_actual", flat=True))
            ocupadas = sum(1 for p in all_pos if p in posiciones_ocupadas)
            vacantes = len(all_pos) - ocupadas
            results = []
            if ocupadas > 0:
                results.append({"value": "Ocupada", "count": ocupadas})
            if vacantes > 0:
                results.append({"value": "Vacante", "count": vacantes})
            return Response(results)

        if distinct_field == "total_movimientos":
            pos_list = list(queryset.values_list("no_pos_actual", flat=True).distinct())
            if pos_list:
                full_counts = dict(
                    MovPos.objects.filter(no_pos_actual__in=pos_list)
                    .values("no_pos_actual")
                    .annotate(c=Count("id"))
                    .values_list("no_pos_actual", "c")
                )
                count_dist = {}
                for c in full_counts.values():
                    count_dist[c] = count_dist.get(c, 0) + 1
                results = [
                    {"value": str(k), "count": v} for k, v in sorted(count_dist.items())
                ]
            else:
                results = []
            return Response(results)

        # "codigo"/"denominacion_puesto"/"nivel_salarial"/
        # "salario_mensual_neto"/"tipo_contratacion"/"anuencia_anexo_nombre"
        # se calculan post-query (ver CAMPOS_DERIVADOS_MOV_POS), no son campos
        # reales de MovPos -> no están en `valid_fields` y sin esta rama
        # caían al fallback genérico de abajo, devolviendo el listado
        # paginado completo en vez de valores distintos (rompía el dropdown
        # de filtro de esas columnas).
        if distinct_field in CAMPOS_DERIVADOS_MOV_POS:
            datos_por_pos = _calcular_columnas_derivadas_mov_pos(queryset)
            counts = {}
            if distinct_field == "anuencia_anexo_nombre":
                # Multi-valor: cada nombre de Anexo 2 es su propia opción del
                # dropdown (no la combinación completa) — una plaza en 2
                # anexos suma 1 al conteo de CADA UNO, no a una entrada
                # "Anexo A, Anexo B" aparte. Ver `_get_mapa_codigos_en_anuencia`.
                for datos in datos_por_pos.values():
                    for n in (datos.get("anuencia_anexo_nombres") or [""]):
                        counts[n] = counts.get(n, 0) + 1
            else:
                for datos in datos_por_pos.values():
                    val = datos.get(distinct_field, "")
                    counts[val] = counts.get(val, 0) + 1

            distinct_search = request.query_params.get("distinct_search", "").strip()
            if distinct_search:
                needle = distinct_search.lower()
                counts = {k: v for k, v in counts.items() if needle in str(k).lower()}

            results = [
                {"value": k, "count": v} for k, v in sorted(counts.items())
            ]
            return Response(results)

        # "fecha_anuencia" es anotación DateField (no campo real de MovPos):
        # `MovPos._meta.get_field("fecha_anuencia")` en la rama genérica de
        # abajo lanza `FieldDoesNotExist` en cuanto hay `distinct_search`
        # (bug reportado: buscar cualquier texto en este dropdown tiraba 500).
        # Además, esa rama genérica solo conoce el valor SQL (fecha real o
        # vacío) — las 4 categorías de texto fijo (p.ej. "Nueva Creación")
        # viven en `CeldaOverride` y sólo se resuelven en Python (ver
        # `corregir_fecha_anuencia_row`), así que nunca aparecían como
        # opción/resultado de búsqueda aquí. Replica esa misma resolución por
        # fila para que el dropdown liste y busque exactamente lo que se ve
        # en la tabla.
        if distinct_field == "fecha_anuencia":
            posiciones_ocupadas = get_posiciones_ocupadas_set()

            counts = {}
            for pos, fa in queryset.values_list("no_pos_actual", "fecha_anuencia"):
                val = resolve_fecha_anuencia_value(
                    pos, fa, posiciones_ocupadas,
                    fecha_anuencia_overrides_texto, fecha_anuencia_baseline_texto,
                )
                counts[val] = counts.get(val, 0) + 1

            distinct_search = request.query_params.get("distinct_search", "").strip()
            if distinct_search:
                needle = distinct_search.lower()
                counts = {k: v for k, v in counts.items() if needle in k.lower()}

            results = [
                {"value": k, "count": v} for k, v in sorted(counts.items())
            ]
            return Response(results)

        if distinct_field in valid_fields:
            is_text = distinct_field in text_fields
            target_distinct_field = (
                f"trimmed_{distinct_field}" if is_text else distinct_field
            )
            if is_text and target_distinct_field not in queryset.query.annotations:
                queryset = queryset.annotate(
                    **{target_distinct_field: Trim(distinct_field)}
                )

            # Apply search filter on the distinct field if present
            distinct_search = request.query_params.get("distinct_search", "").strip()
            if distinct_search:
                if is_text:
                    queryset = queryset.filter(
                        **{f"{target_distinct_field}__icontains": distinct_search}
                    )
                else:
                    # DateTimeField guarda hora exacta: un match exacto contra
                    # solo "YYYY-MM-DD" (lo unico que manda el front) nunca
                    # coincide salvo que la hora sea 00:00:00. `__date` no
                    # sirve: convierte a TIME_ZONE local (America/Mexico_City,
                    # UTC-6) antes de truncar, y estos timestamps casi siempre
                    # se guardaron como medianoche UTC -> se corren un dia
                    # hacia atras (mismo bug de -1 dia documentado en todo el
                    # front para columnas de fecha). Se trunca en UTC, sin
                    # conversion de zona horaria, para comparar la fecha tal
                    # cual esta guardada.
                    field_internal_type = MovPos._meta.get_field(
                        distinct_field
                    ).get_internal_type()
                    if field_internal_type == "DateTimeField":
                        utc_date_field = f"{target_distinct_field}__utc_date"
                        if utc_date_field not in queryset.query.annotations:
                            queryset = queryset.annotate(
                                **{utc_date_field: TruncDate(
                                    distinct_field, tzinfo=datetime.timezone.utc
                                )}
                            )
                        search_lookup = utc_date_field
                    else:
                        search_lookup = target_distinct_field
                    try:
                        queryset = queryset.filter(
                            **{search_lookup: distinct_search}
                        )
                    except (ValueError, exceptions.ValidationError):
                        queryset = queryset.none()

            distinct_qs = (
                queryset.values(target_distinct_field)
                .annotate(count=Count("*"))
                .order_by(target_distinct_field)
            )

            results = []
            for item in distinct_qs:
                val = item[target_distinct_field]
                # `fecha_anuencia` (anotación DateField) llega como
                # date/datetime, no como el string plano que ya son los demás
                # campos (varchar) — se homologa al mismo formato que se
                # muestra en la tabla ('YYYY-MM-DD').
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                results.append(
                    {"value": val if val is not None else "", "count": item["count"]}
                )
            return Response(results)

        # Sorting
        sort_by_param = request.query_params.get("sort_by", "").strip()
        sort_order = request.query_params.get("sort_order", "desc").strip().lower()
        sort_fields = [f.strip() for f in sort_by_param.split(",")] if sort_by_param else []
        # Columnas calculadas en Python DESPUÉS de la consulta (ver
        # CAMPOS_DERIVADOS_MOV_POS/enriquecer_mov_pos_rows — Puesto, Nivel
        # Salarial, Salario Mensual Neto, Tipo de Contratación, En Anuencia,
        # Fecha de alta solicitada): no son campos ni anotaciones reales de
        # MovPos, así que nunca entran a `valid_fields` y un ORDER BY de SQL
        # sobre ellas no tiene efecto — ANTES esto se ignoraba en silencio
        # (la petición se mandaba pero la tabla no se reordenaba). Se
        # detectan aquí para ordenarlas aparte, en Python, sobre el
        # resultado ya enriquecido (ver más abajo).
        campos_derivados_sort = [f for f in sort_fields if f in CAMPOS_DERIVADOS_MOV_POS]
        if sort_fields:
            order_by_args = []
            for field in sort_fields:
                if field in valid_fields:
                    is_text = field in text_fields
                    target_sort_field = f"trimmed_{field}" if is_text else field
                    if is_text and target_sort_field not in queryset.query.annotations:
                        queryset = queryset.annotate(**{target_sort_field: Trim(field)})
                    if sort_order == "desc":
                        order_by_args.append(f"-{target_sort_field}")
                    else:
                        order_by_args.append(target_sort_field)
            if order_by_args:
                queryset = queryset.order_by(*order_by_args)
            else:
                queryset = queryset.order_by(
                    "-f_efva", "-fecha_captura", "no_pos_actual"
                )
        else:
            # Default ordering requested by the user:
            # SELECT * FROM MOV_POS ORDERY BY fecha efectiva DESC, FECHA CAPTURA DESC, y ordenar tambien por posicion
            queryset = queryset.order_by("-f_efva", "-fecha_captura", "no_pos_actual")

        def _enriquecer_resultados(resultados):
            """Agrega a cada fila `total_movimientos`, `estatus_ocupacion`/
            `ocupacion`, corrige `fecha_vacancia`/`fecha_anuencia` para las
            ocupadas y resuelve las columnas derivadas (ver
            `enriquecer_mov_pos_rows`) — mismo trabajo que antes se repetía
            casi idéntico en las 3 ramas de respuesta de abajo."""
            if not resultados:
                return resultados
            counts = dict(
                MovPos.objects.values_list("no_pos_actual").annotate(c=Count("id"))
            )
            posiciones_ocupadas = cache.get("mov_pos_ocupadas_set")
            if posiciones_ocupadas is None:
                with connection.cursor() as cursor:
                    cursor.execute(OCUPADAS_RAW_SQL)
                    posiciones_ocupadas = set(
                        [row[0] for row in cursor.fetchall() if row[0]]
                    )
                cache.set("mov_pos_ocupadas_set", posiciones_ocupadas, 600)

            populate_movpos_occupant_details(resultados, posiciones_ocupadas)
            mapa_codigos = _get_mapa_codigos()
            for r in resultados:
                pos = r.get("no_pos_actual")
                r["total_movimientos"] = counts.get(pos, 1)
                r["estatus_ocupacion"] = (
                    "Ocupada" if pos in posiciones_ocupadas else "Vacante"
                )
                r["ocupacion"] = r["estatus_ocupacion"]
                r["fecha_vacancia"] = (
                    "" if pos in posiciones_ocupadas else r.get("fecha_vacancia", "")
                )
                r["codigo"] = mapa_codigos.get(pos, "")
                corregir_fecha_anuencia_row(
                    r, pos, posiciones_ocupadas, fecha_anuencia_overrides,
                    fecha_anuencia_overrides_texto, fecha_anuencia_baseline_texto,
                )
            enriquecer_mov_pos_rows(resultados)
            return resultados

        # Excel download or full list without pagination (bypass pagination if is_latest is true)
        no_pagination = (
            request.query_params.get("no_pagination", "false").strip().lower() == "true"
            or is_latest
        )
        # `pagination=true`: usado sólo por clientes externos (rendicionCuentasBack vía
        # eje_central_client.py) para forzar el paginador real de abajo aun cuando
        # is_latest quedó en su default true. eje_central_front nunca manda este flag,
        # así que su respuesta (envelope completo, sin slicing real) no cambia.
        if request.query_params.get("pagination", "false").strip().lower() == "true":
            no_pagination = False

        if campos_derivados_sort:
            # No se puede ordenar en SQL por una columna calculada en Python
            # (ver arriba) — se materializa TODO lo filtrado, se enriquece y
            # se ordena en memoria, y recién después se pagina/recorta según
            # lo que haya pedido el cliente. `sort()` es estable: se ordena
            # de la columna menos prioritaria a la más prioritaria para que
            # un `sort_by` con varias columnas se resuelva en el orden
            # correcto (la primera manda, las demás desempatan).
            resultados = _enriquecer_resultados(list(queryset.values()))
            reverse = sort_order == "desc"

            def _clave_ordenable(campo):
                # Valor "puro" (sin marca de vacío): a un vacío se le da un
                # placeholder de un tipo comparable con el resto (0.0 para
                # numéricas, "" para texto) — no importa dónde caiga en este
                # paso, el segundo `sort()` de abajo lo manda al final sin
                # importar la dirección pedida.
                es_numerico = campo == "salario_mensual_neto"

                def _clave(r):
                    val = r.get(campo)
                    if val in (None, ""):
                        return 0.0 if es_numerico else ""
                    if es_numerico:
                        try:
                            return float(val)
                        except (TypeError, ValueError):
                            return 0.0
                    return str(val).lower()

                return _clave

            for campo in reversed(campos_derivados_sort):
                resultados.sort(key=_clave_ordenable(campo), reverse=reverse)

            # Los vacíos SIEMPRE al final, sin importar asc/desc (si se
            # horneara la marca de vacío dentro de la clave de arriba,
            # `reverse=True` los mandaría al frente en descendente — no es lo
            # que espera el usuario al ordenar una columna). `sort()` es
            # estable, así que esto sólo reubica los vacíos sin desordenar lo
            # que el paso anterior ya dejó bien puesto.
            campo_principal = campos_derivados_sort[0]
            resultados.sort(key=lambda r: r.get(campo_principal) in (None, ""))

            if no_pagination:
                is_excel_mode = (
                    request.query_params.get("no_pagination", "false").strip().lower()
                    == "true"
                )
                if not is_excel_mode:
                    stats = get_mov_pos_stats()
                    return Response(
                        {
                            "next": None,
                            "previous": None,
                            "count": len(resultados),
                            "results": resultados,
                            "stats": stats,
                        }
                    )
                return Response(resultados)

            paginator = self.pagination_class()
            page = paginator.paginate_queryset(resultados, request, view=self)
            if page is not None:
                return paginator.get_paginated_response(page)
            return Response(resultados)

        if no_pagination:
            resultados = _enriquecer_resultados(list(queryset.values()))
            is_excel_mode = (
                request.query_params.get("no_pagination", "false").strip().lower()
                == "true"
            )
            if not is_excel_mode:
                stats = get_mov_pos_stats()
                return Response(
                    {
                        "next": None,
                        "previous": None,
                        "count": len(resultados),
                        "results": resultados,
                        "stats": stats,
                    }
                )
            return Response(resultados)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset.values(), request, view=self)
        if page is not None:
            resultados = _enriquecer_resultados(list(page))
            return paginator.get_paginated_response(resultados)

        resultados = _enriquecer_resultados(list(queryset.values()))
        return Response(resultados)


def _recalcular_fecha_anuencia_actual(no_pos_actual):
    """Recalcula el valor VIGENTE de `fecha_anuencia` para UNA posición
    (fecha_vacancia + 30 días, o vacío si está ocupada) reusando las mismas
    `annotate_fecha_anuencia`/`corregir_fecha_anuencia_row` que ya arman el
    listado — usado por `MovPosFechaAnuenciaOverrideView.delete()` para que
    el frontend pueda repintar la celda de inmediato al revertir a
    automático, sin esperar a que el usuario recargue la página (antes el
    DELETE devolvía `None` a secas, dejando la celda en blanco hasta el
    siguiente GET)."""
    row = {"fecha_anuencia": None}
    obj = (
        annotate_fecha_anuencia(
            MovPos.objects.filter(no_pos_actual=no_pos_actual),
            baseline=get_fecha_anuencia_baseline_map(),
        )
        .values("fecha_anuencia")
        .first()
    )
    if obj:
        row["fecha_anuencia"] = obj["fecha_anuencia"]

    posiciones_ocupadas = cache.get("mov_pos_ocupadas_set")
    if posiciones_ocupadas is None:
        with connection.cursor() as cursor:
            cursor.execute(OCUPADAS_RAW_SQL)
            posiciones_ocupadas = set(r[0] for r in cursor.fetchall() if r[0])
        cache.set("mov_pos_ocupadas_set", posiciones_ocupadas, 600)

    corregir_fecha_anuencia_row(
        row, no_pos_actual, posiciones_ocupadas,
        baseline_texto=get_fecha_anuencia_baseline_texto_map(),
    )
    return row["fecha_anuencia"]


class MovPosFechaAnuenciaOverrideView(APIView):
    """
    Edición manual de `fecha_anuencia` (sugerida por default como el valor
    cargado del Excel o, si no hay, fecha_vacancia + 30 días, ver
    `annotate_fecha_anuencia`) desde el tab Mov. Posiciones **y** desde
    Plantilla Detalle (columna "Fecha de Anuencia", misma clave de negocio
    `posicion` == `no_pos_actual` — ver PlantillaDetalleTab.jsx) — ambos tabs
    comparten el mismo override, así que siempre muestran la misma fecha.
    A diferencia de `EmpleadosCompletosCeldaOverrideView`: no hay tabla viva
    que actualizar ni caché que invalidar — el valor se calcula (override o
    default) en cada request. `delete()` además devuelve el valor automático
    recién recalculado (ver `_recalcular_fecha_anuencia_actual`) para que el
    frontend pinte la celda de inmediato al revertir, sin esperar a un
    GET/recarga nueva.
    """

    edit_permission = (
        "authentication.edit_plantilla_mov_posiciones",
        "authentication.edit_plantilla_detalle",
    )

    def post(self, request):
        no_pos_actual = request.data.get("no_pos_actual")
        valor_nuevo = request.data.get("valor_nuevo")
        if not no_pos_actual or not valor_nuevo:
            return Response(
                {"detail": "no_pos_actual y valor_nuevo son requeridos."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            override = registrar_override_fecha_anuencia(
                no_pos_actual, valor_nuevo, request.user
            )
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if override is None:
            return Response({
                "no_pos_actual": no_pos_actual,
                "fecha_anuencia": valor_nuevo,
                "sin_cambios": True,
            })

        return Response({
            "no_pos_actual": no_pos_actual,
            "fecha_anuencia": override.valor_nuevo,
            "valor_original": override.valor_original,
            "usuario": request.user.username,
            "fecha_modificacion": override.fecha_modificacion,
        })

    def delete(self, request):
        no_pos_actual = request.data.get("no_pos_actual")
        if not no_pos_actual:
            return Response(
                {"detail": "no_pos_actual es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        borrar_override_fecha_anuencia(no_pos_actual)
        fecha_anuencia = _recalcular_fecha_anuencia_actual(no_pos_actual)
        return Response({"no_pos_actual": no_pos_actual, "fecha_anuencia": fecha_anuencia})


class MovPosFechaAltaSolicitadaOverrideView(APIView):
    """
    Edición de "Fecha de alta solicitada" desde Mov. Posiciones — a
    diferencia de `MovPosFechaAnuenciaOverrideView` (que guarda en una tabla
    de overrides propia), este campo NO tiene un dueño propio: vive dentro
    de `hojas` de un AnuenciaAnexo (ver `_get_mapa_fecha_alta_solicitada_anuencia`),
    así que editarlo aquí escribe directamente sobre esa fila del Anexo 2 y
    dispara el mismo aviso SSE ("anuencia_anexo_updates") que usa
    AnuenciaTab.jsx — por eso una pestaña con ese Anexo 2 abierto ve el
    cambio sin recargar, igual que si se hubiera editado ahí mismo.

    Sólo se puede editar si el Código Federal de Puesto ya está en algún
    Anexo 2 guardado (no hay dónde más guardarlo); si no, 400.
    """

    edit_permission = (
        "authentication.edit_plantilla_mov_posiciones",
        "authentication.edit_plantilla_detalle",
    )

    def _aplicar(self, request, codigo, valor_nuevo):
        from django.shortcuts import get_object_or_404

        if not codigo:
            return Response({"detail": "codigo es requerido."}, status=status.HTTP_400_BAD_REQUEST)

        info = _get_mapa_fecha_alta_solicitada_anuencia().get(codigo)
        if not info:
            return Response(
                {"detail": "Esta posición no está en ningún Anexo 2 guardado; agrégala primero para poder capturar su fecha de alta solicitada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        anexo = get_object_or_404(AnuenciaAnexo, pk=info["anexo_id"], eliminado=False)
        hojas = anexo.hojas or []
        fila_encontrada = None
        nombre_hoja = None
        for hoja in hojas:
            if hoja.get("_id") != info["hoja_id"]:
                continue
            nombre_hoja = hoja.get("nombre") or "(sin nombre)"
            for fila in hoja.get("filas") or []:
                if fila.get("_id") == info["fila_id"]:
                    fila_encontrada = fila
                    break
            break

        if fila_encontrada is None:
            return Response(
                {"detail": "No se encontró la plaza dentro de ese Anexo 2 (puede haber cambiado); recarga e intenta de nuevo."},
                status=status.HTTP_409_CONFLICT,
            )

        valor_anterior = fila_encontrada.get("fecha_alta_solicitada") or ""
        valor_nuevo = valor_nuevo or ""
        if valor_anterior == valor_nuevo:
            return Response({"codigo": codigo, "fecha_alta_solicitada": valor_nuevo, "sin_cambios": True})

        fila_encontrada["fecha_alta_solicitada"] = valor_nuevo
        anexo.hojas = hojas
        anexo.actualizado_por = request.user
        anexo.save(update_fields=["hojas", "actualizado_por", "actualizado_en"])

        _invalidar_cache_anuencia()
        _notificar_actualizacion_anuencia_anexo(anexo, request.user)
        resumen = _resumen_fila_anuencia(fila_encontrada)
        etiqueta = "Editó" if valor_nuevo else "Borró"
        detalle = (
            f'{etiqueta} Fecha de alta solicitada de {resumen} en "{nombre_hoja}": '
            f'"{valor_anterior or "—"}" → "{valor_nuevo or "—"}" (desde Mov. Posiciones)'
        )
        _registrar_cambio_anuencia_anexo(anexo, request.user, [detalle])

        return Response({"codigo": codigo, "fecha_alta_solicitada": valor_nuevo, "anexo_id": anexo.id})

    def post(self, request):
        codigo = str(request.data.get("codigo") or "").strip()
        valor_nuevo = request.data.get("valor_nuevo")
        return self._aplicar(request, codigo, valor_nuevo)

    def delete(self, request):
        codigo = str(request.data.get("codigo") or "").strip()
        return self._aplicar(request, codigo, "")


class MovPosHistoriaView(APIView):
    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        posicion = request.query_params.get("posicion")
        if not posicion:
            return Response(
                {"error": "Parámetro 'posicion' es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Obtener todos los registros para la posición, ordenados del más reciente al más antiguo
            queryset = MovPos.objects.filter(no_pos_actual=posicion).order_by("-id")

            resultados = list(queryset.values())

            return Response(resultados, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MovPosVacanciaDetalleView(APIView):
    """
    Devuelve el detalle dinámico (por categoría de vacancia A/B/C) del
    registro decisivo que originó la fecha de vacancia de una posición.

    - Categoría A (baja): el registro decisivo vive en
      cp_tbl_mov_completo_29_05_26 y describe al empleado que causó la baja,
      el motivo, fecha efectiva y fecha de captura de esa baja.
    - Categoría B (cambio de posición): mismo origen de datos; la posición
      origen es la posición sobre la que se consulta y la posición destino
      es `posicion` del registro decisivo.
    - Categoría C (nunca ocupada): el registro decisivo vive en el propio
      MOV_POS y es el primer movimiento histórico de esa plaza.

    Parámetro: ?id=<id de MOV_POS> (el id del renglón de MOV_POS sobre el
    que se muestra la fecha de vacancia, NO el idRegistroDesicivo).
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from .notificaciones_posicion import construir_detalle_vacancia

        mov_id = request.query_params.get("id")
        if not mov_id:
            return Response(
                {"error": "Parámetro 'id' es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            mov_row = MovPos.objects.get(id=mov_id)
        except (MovPos.DoesNotExist, ValueError):
            return Response(
                {"error": "Registro de MOV_POS no encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(construir_detalle_vacancia(mov_row))


class MovPosOcupacionDetalleView(APIView):
    """
    Devuelve el detalle del registro decisivo (cp_tbl_mov_completo_29_05_26)
    que originó la fecha de ocupación vigente de una posición: el movimiento
    que abrió la racha ininterrumpida actual del ocupante en esa posición
    (ver sp_actualizar_fecha_ocupacion_mov_pos), no necesariamente la
    primera vez que el empleado apareció en ella.

    Parámetro: ?id=<id de MOV_POS> (el id del renglón de MOV_POS sobre el
    que se muestra la fecha de ocupación, NO el
    id_registro_des_fecha_ocupacion).
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from .notificaciones_posicion import construir_detalle_ocupacion

        mov_id = request.query_params.get("id")
        if not mov_id:
            return Response(
                {"error": "Parámetro 'id' es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            mov_row = MovPos.objects.get(id=mov_id)
        except (MovPos.DoesNotExist, ValueError):
            return Response(
                {"error": "Registro de MOV_POS no encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(construir_detalle_ocupacion(mov_row))


class SuscripcionesPosicionView(APIView):
    """
    Menú contextual "Notificarme cuando la posición quede vacante/se ocupe"
    de la columna Posición (PlantillaDetalleTab/MovimientosTab). Cualquier
    usuario autenticado puede suscribirse — no requiere permiso de módulo
    (sin `view_permission`, `HasModulePermission` sólo exige login).

    GET  -> suscripciones del usuario autenticado que le importan a la
            "campanita" de notificaciones (NotificacionesPosicionBell):
            las PENDIENTES (`activa=True`) y las YA LLEGADAS
            (`notificado_en` no nulo), con su `detalle_enviado` (snapshot
            de lo que decía el correo). Se excluyen las simplemente
            canceladas (`activa=False` y nunca notificadas) — esas ya no
            le sirven a nadie.
    POST -> crea una suscripción {posicion, tipo}. Snapshotea el estado
            actual (ocupada/vacante) como `estado_conocido_al_suscribir` —
            la notificación se dispara cuando el estado real deje de
            coincidir con este snapshot (ver `notificaciones_posicion.py`).
            Idempotente: MySQL no aplica el UniqueConstraint condicional del
            modelo (ver W036 en la migración), así que la duplicidad se
            evita aquí explícitamente devolviendo la existente en vez de
            crear una segunda.
    """

    CAMPOS_RESPUESTA = (
        "id", "posicion", "tipo", "activa", "estado_conocido_al_suscribir",
        "creado_en", "notificado_en", "detalle_enviado",
    )

    def get(self, request):
        subs = (
            SuscripcionNotificacionPosicion.objects
            .filter(usuario=request.user)
            .filter(Q(activa=True) | Q(notificado_en__isnull=False))
            .order_by("-creado_en")
            .values(*self.CAMPOS_RESPUESTA)
        )
        return Response(list(subs))

    def post(self, request):
        posicion = (request.data.get("posicion") or "").strip()
        tipo = (request.data.get("tipo") or "").strip().upper()
        if not posicion or tipo not in dict(SuscripcionNotificacionPosicion.TIPO_CHOICES):
            return Response(
                {"error": "Se requiere 'posicion' y 'tipo' ('VACANTE' u 'OCUPACION')."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existente = SuscripcionNotificacionPosicion.objects.filter(
            usuario=request.user, posicion=posicion, tipo=tipo, activa=True
        ).first()
        if existente:
            return Response(
                {
                    **{k: getattr(existente, k) for k in self.CAMPOS_RESPUESTA},
                    "ya_existia": True,
                },
                status=status.HTTP_200_OK,
            )

        posiciones_ocupadas = get_posiciones_ocupadas_set()
        estado_actual = "O" if posicion in posiciones_ocupadas else "V"

        sub = SuscripcionNotificacionPosicion.objects.create(
            usuario=request.user,
            posicion=posicion,
            tipo=tipo,
            estado_conocido_al_suscribir=estado_actual,
        )
        return Response(
            {
                **{k: getattr(sub, k) for k in self.CAMPOS_RESPUESTA},
                "ya_existia": False,
            },
            status=status.HTTP_201_CREATED,
        )


class SuscripcionPosicionDetalleView(APIView):
    """DELETE -> si la suscripción sigue pendiente (`activa=True`), la
    cancela (soft delete: `activa=False`, nunca se notificará). Si ya se
    notificó (o ya estaba cancelada), no hay nada que "cancelar" — se borra
    el renglón por completo (equivale a "descartar" la tarjeta en la
    campanita de notificaciones). Un usuario no puede tocar la suscripción
    de otro (filtra por `usuario=request.user`, un 404 en vez de 403 para
    no filtrar si el id pertenece a alguien más)."""

    def delete(self, request, pk):
        try:
            sub = SuscripcionNotificacionPosicion.objects.get(pk=pk, usuario=request.user)
        except SuscripcionNotificacionPosicion.DoesNotExist:
            return Response(
                {"error": "Suscripción no encontrada."}, status=status.HTTP_404_NOT_FOUND
            )
        if sub.activa:
            sub.activa = False
            sub.save(update_fields=["activa"])
        else:
            sub.delete()
        return Response({"status": "ok"})


class FiltrosGuardadosView(APIView):
    """
    Combinaciones de condiciones de `AdvancedFiltersModal` guardadas por el
    usuario autenticado, para reaplicarlas sin reconstruirlas a mano.
    `vista` identifica el tab de origen (cada uno tiene su propio set de
    columnas) — un filtro guardado en un tab no aplica a otro.

    GET  ?vista=<key> -> filtros del usuario en esa vista.
    POST {vista, nombre, condiciones} -> crea uno nuevo. 409 si ya existe
         un filtro con ese nombre para el usuario en esa vista.
    """

    CAMPOS_RESPUESTA = ("id", "vista", "nombre", "condiciones", "creado_en")

    def get(self, request):
        vista = request.query_params.get("vista")
        if not vista:
            return Response(
                {"error": "Falta parámetro 'vista'."}, status=status.HTTP_400_BAD_REQUEST
            )
        filtros = (
            FiltroGuardado.objects
            .filter(usuario=request.user, vista=vista)
            .values(*self.CAMPOS_RESPUESTA)
        )
        return Response(list(filtros))

    def post(self, request):
        vista = (request.data.get("vista") or "").strip()
        nombre = (request.data.get("nombre") or "").strip()
        condiciones = request.data.get("condiciones")

        if not vista or not nombre:
            return Response(
                {"error": "Se requiere 'vista' y 'nombre'."}, status=status.HTTP_400_BAD_REQUEST
            )
        if len(nombre) > 100:
            return Response(
                {"error": "El nombre no puede exceder 100 caracteres."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not isinstance(condiciones, list) or not condiciones:
            return Response(
                {"error": "'condiciones' debe ser una lista no vacía."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if FiltroGuardado.objects.filter(usuario=request.user, vista=vista, nombre=nombre).exists():
            return Response(
                {"error": "Ya existe un filtro con ese nombre en esta vista."},
                status=status.HTTP_409_CONFLICT,
            )

        filtro = FiltroGuardado.objects.create(
            usuario=request.user, vista=vista, nombre=nombre, condiciones=condiciones,
        )
        return Response(
            {k: getattr(filtro, k) for k in self.CAMPOS_RESPUESTA},
            status=status.HTTP_201_CREATED,
        )


class FiltroGuardadoDetalleView(APIView):
    """DELETE -> borra el filtro guardado. Un usuario no puede tocar el
    filtro de otro (filtra por `usuario=request.user`, 404 en vez de 403
    para no filtrar si el id pertenece a alguien más)."""

    def delete(self, request, pk):
        filtro = FiltroGuardado.objects.filter(pk=pk, usuario=request.user).first()
        if not filtro:
            return Response(
                {"error": "Filtro no encontrado."}, status=status.HTTP_404_NOT_FOUND
            )
        filtro.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Comprobar Alineación Organizacional (MOV_POS vs EMPLEADOS_COMPLETOS_SIG) ──
#
# 13 pares de columnas que deberían coincidir entre la plaza (MOV_POS, la
# fila más reciente y activa por posición vía MOV_POS_LATEST) y la persona/
# ocupante actual (EMPLEADOS_COMPLETOS_SIG, 1 fila por `Posición`, siempre
# presente aunque la plaza esté vacante). Ver ALINEACIÓN_PLAZA_PERSONA.
ALINEACION_CAMPOS = [
    {"key": "cd_un", "label": "Código Unidad de Negocio", "mov_field": "cd_un", "emp_field": "cd_un", "prefijo2": False},
    {"key": "cd_ua", "label": "Código Unidad Administrativa", "mov_field": "unidad_adva", "emp_field": "cd_ua", "prefijo2": False},
    {"key": "nivel_jerarquico", "label": "Nivel Jerárquico", "mov_field": "nvl_direc", "emp_field": "nj", "prefijo2": False},
    {"key": "id_departamento", "label": "Id Departamento", "mov_field": "cd_departamento", "emp_field": "id_departamento", "prefijo2": False},
    {"key": "cd_pto_funcional", "label": "Código Puesto Funcional", "mov_field": "cd_puesto", "emp_field": "cd_pto_funcional", "prefijo2": False},
    {"key": "nombre_pto_funcional", "label": "Nombre Puesto Funcional", "mov_field": "nombre_puesto", "emp_field": "nombre_puesto_funcional", "prefijo2": False},
    {"key": "codigo_presupuestal", "label": "Código Presupuestal", "mov_field": "puesto_ptal", "emp_field": "codigo_presupuestal", "prefijo2": False},
    {"key": "escala", "label": "Escala", "mov_field": "esc", "emp_field": "escala", "prefijo2": False},
    {"key": "grado", "label": "Grado", "mov_field": "grado", "emp_field": "nivel", "prefijo2": True},
    {"key": "partida", "label": "Partida", "mov_field": "partida_ptal", "emp_field": "partida", "prefijo2": False},
    {"key": "tipo_contratacion", "label": "Tipo de Contratación", "mov_field": "gp_trabajo", "emp_field": "tipo_de_contratacion", "prefijo2": False},
    {"key": "dependencia_directa", "label": "Dependencia Directa", "mov_field": "depnd_drt", "emp_field": "dependencia_directa", "prefijo2": False},
    {"key": "ubicacion", "label": "Ubicación", "mov_field": "ubicacion", "emp_field": "ubicacion", "prefijo2": False},
]

ALINEACION_DATASET_CACHE_KEY = "mov_pos_alineacion_dataset"
ALINEACION_STATS_CACHE_KEY = "mov_pos_alineacion_stats"
ALINEACION_CACHE_TTL = 600  # 10 minutos, igual que el resto de caches de Mov Pos


def _alineacion_normalizar(valor, prefijo2=False):
    """Trim + upper case antes de comparar (evita falsos "difiere" por
    espacios extra o mayúsculas/minúsculas; ambas tablas ya tienen ese
    problema conocido, de ahí los índices funcionales Trim(...))."""
    s = "" if valor is None else str(valor).strip().upper()
    return s[:2] if prefijo2 else s


# Campos de ALINEACION_CAMPOS que sí tienen un catálogo de descripción legible
# detrás del código (ver tooltip en frontend AlineacionOrganizacionalTab). Los
# demás (escala, grado, partida, etc.) son códigos sin catálogo asociado.
TOOLTIP_ALINEACION_CAMPOS = {"cd_un", "cd_ua", "nivel_jerarquico", "id_departamento"}


def _cargar_catalogos_alineacion():
    """Catálogos código→nombre para las tooltips de alineación. `ua_unidadadministrativa`
    y `ORGANIGRAMA_ANAM` no tienen modelo Django (mismo patrón raw SQL que
    OrganigramaDeptoView / DesgloseJerarquicoView)."""
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("SELECT codigo, nombre FROM ua_unidadadministrativa")
        ua_lookup = {(row[0] or "").strip().upper(): row[1] for row in cursor.fetchall()}
        cursor.execute("SELECT departamento, descripcion_larga FROM ORGANIGRAMA_ANAM WHERE isSIGInfo = 1")
        depto_lookup = {(row[0] or "").strip().upper(): row[1] for row in cursor.fetchall()}
    return {"ua": ua_lookup, "depto": depto_lookup}


def _nombre_nivel_jerarquico(codigo, nombre_puesto_funcional=None):
    s = (codigo or "").strip()
    if not s:
        return None
    try:
        nivel = int(s)
        # Nivel 3 es "Director" por default, salvo "Titular de Aduana" cuando
        # el puesto funcional es un Administrador de Aduana — mismo criterio
        # que `nivel_jerarquico_sync.aplicar_nivel_a_plazas`.
        if nivel == 3 and str(nombre_puesto_funcional or "").strip().upper().startswith(
            NIVEL_3_PREFIJO_TITULAR_ADUANA.upper()
        ):
            return NIVEL_JERARQUICO_LABELS_NIVEL_3_TITULAR_ADUANA["nombre_nj"]
        return NIVEL_JERARQUICO_LABELS[nivel]["nombre_nj"]
    except (ValueError, KeyError):
        return None


def _resolver_nombres_tooltip(key, mov_row, emp_row, catalogos):
    """(nombre_mov, nombre_emp) legibles para un campo con catálogo. `emp_row`
    ya viene resuelto a dict vacío si la plaza está vacante."""
    if key == "cd_un":
        return mov_row.get("unidad_de_negocio"), emp_row.get("unidad_de_negocio")
    if key == "cd_ua":
        mov_nombre = catalogos["ua"].get((mov_row.get("unidad_adva") or "").strip().upper())
        return mov_nombre, emp_row.get("unidad_administrativa")
    if key == "nivel_jerarquico":
        return (
            _nombre_nivel_jerarquico(mov_row.get("nvl_direc"), mov_row.get("nombre_puesto")),
            _nombre_nivel_jerarquico(emp_row.get("nj"), emp_row.get("nombre_puesto_funcional")),
        )
    if key == "id_departamento":
        mov_nombre = catalogos["depto"].get((mov_row.get("cd_departamento") or "").strip().upper())
        return mov_nombre, emp_row.get("departamento")
    return None, None


def _comparar_alineacion_plaza(mov_row, emp_row, catalogos=None):
    detalle = {}
    coincidentes = 0
    diferentes = []
    emp_row_dict = emp_row or {}
    for campo in ALINEACION_CAMPOS:
        key = campo["key"]
        mov_valor = mov_row.get(campo["mov_field"])
        emp_valor = emp_row.get(campo["emp_field"]) if emp_row else None
        mov_norm = _alineacion_normalizar(mov_valor, campo["prefijo2"])
        emp_norm = _alineacion_normalizar(emp_valor, campo["prefijo2"])
        coincide = mov_norm == emp_norm
        if coincide:
            coincidentes += 1
        else:
            diferentes.append(key)
        detalle[key] = {
            "label": campo["label"],
            "mov_pos_valor": mov_valor if mov_valor not in (None, "") else "(vacío)",
            "empleados_sig_valor": emp_valor if emp_valor not in (None, "") else "(vacío)",
            "coincide": coincide,
            "ambos_vacios": mov_norm == "" and emp_norm == "",
        }
        if catalogos and key in TOOLTIP_ALINEACION_CAMPOS:
            mov_nombre, emp_nombre = _resolver_nombres_tooltip(key, mov_row, emp_row_dict, catalogos)
            detalle[key]["mov_pos_nombre"] = mov_nombre or None
            detalle[key]["empleados_sig_nombre"] = emp_nombre or None
    total = len(ALINEACION_CAMPOS)
    return {
        "detalle": detalle,
        "campos_coincidentes": coincidentes,
        "campos_totales": total,
        "campos_diferentes": diferentes,
        "porcentaje_alineacion": round((coincidentes / total) * 100, 1) if total else 0.0,
        "es_alineada": coincidentes == total,
    }


def _construir_dataset_alineacion():
    """Cruce completo MOV_POS (activas/latest) x EMPLEADOS_COMPLETOS_SIG.
    Todo en memoria (~11.4k filas, mismo volumen que ya carga sin paginar
    la rama `is_latest=true` de MovPosDetalleView): 2 queries + comparación
    Python, cacheado 10 min."""
    cached = cache.get(ALINEACION_DATASET_CACHE_KEY)
    if cached is not None:
        return cached

    active_ids = list(
        MovPosLatest.objects.filter(estado_psn="A").values_list("mov_pos_id", flat=True)
    )

    mov_fields = [
        "id", "no_pos_actual", "nombre_puesto", "estado_psn",
        "unidad_de_negocio", "cd_un", "unidad_adva", "nvl_direc",
        "cd_departamento", "cd_puesto", "puesto_ptal", "esc", "grado",
        "partida_ptal", "gp_trabajo", "depnd_drt", "ubicacion",
    ]
    mov_rows = list(MovPos.objects.filter(id__in=active_ids).values(*mov_fields))

    posiciones = [r["no_pos_actual"] for r in mov_rows if r["no_pos_actual"]]
    emp_fields = [
        "posicion", "id_empleado", "nombres",
        "unidad_de_negocio", "cd_un", "cd_ua", "unidad_administrativa", "nj",
        "id_departamento", "departamento",
        "cd_pto_funcional", "nombre_puesto_funcional", "codigo_presupuestal",
        "escala", "nivel", "partida", "tipo_de_contratacion",
        "dependencia_directa", "ubicacion",
    ]
    emp_by_pos = {
        row["posicion"]: row
        for row in EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values(*emp_fields)
    }
    catalogos = _cargar_catalogos_alineacion()

    resultados = []
    for mov_row in mov_rows:
        pos = mov_row["no_pos_actual"]
        emp_row = emp_by_pos.get(pos)
        id_emp = (emp_row.get("id_empleado") or "").strip() if emp_row else ""
        nombres = (emp_row.get("nombres") or "").strip() if emp_row else ""
        es_vacante = not (id_emp or nombres)

        comparacion = _comparar_alineacion_plaza(mov_row, emp_row, catalogos)

        fila = dict(mov_row)
        fila["ocupante_id"] = id_emp
        fila["ocupante_nombre"] = nombres
        fila["ocupacion"] = "Vacante" if es_vacante else "Ocupada"
        fila["campos_coincidentes"] = comparacion["campos_coincidentes"]
        fila["campos_totales"] = comparacion["campos_totales"]
        fila["num_diferencias"] = comparacion["campos_totales"] - comparacion["campos_coincidentes"]
        fila["porcentaje_alineacion"] = comparacion["porcentaje_alineacion"]
        fila["estado_alineacion"] = "Alineada" if comparacion["es_alineada"] else "Con Diferencias"
        fila["campos_diferentes"] = comparacion["campos_diferentes"]
        fila["alineacion_detalle"] = comparacion["detalle"]
        for campo in ALINEACION_CAMPOS:
            det = comparacion["detalle"][campo["key"]]
            fila[f"match_{campo['key']}"] = "Coincide" if det["coincide"] else "Difiere"
            # Valores crudos aplanados (columnas separadas MOV_POS / EMPLEADOS_COMPLETOS_SIG
            # en el frontend) para que el filtro/orden genérico de columna funcione
            # también sobre ellos, igual que sobre match_<campo>.
            fila[f"mov_{campo['key']}"] = det["mov_pos_valor"]
            fila[f"emp_{campo['key']}"] = det["empleados_sig_valor"]
        resultados.append(fila)

    cache.set(ALINEACION_DATASET_CACHE_KEY, resultados, ALINEACION_CACHE_TTL)
    return resultados


def get_mov_pos_alineacion_stats(resultados=None):
    stats = cache.get(ALINEACION_STATS_CACHE_KEY)
    if stats is not None:
        return stats
    if resultados is None:
        resultados = _construir_dataset_alineacion()

    total = len(resultados)
    alineadas = sum(1 for r in resultados if r["estado_alineacion"] == "Alineada")
    vacantes = sum(1 for r in resultados if r["ocupacion"] == "Vacante")
    con_diferencias_vacantes = sum(
        1 for r in resultados if r["estado_alineacion"] != "Alineada" and r["ocupacion"] == "Vacante"
    )
    con_diferencias_ocupadas = sum(
        1 for r in resultados if r["estado_alineacion"] != "Alineada" and r["ocupacion"] != "Vacante"
    )

    por_campo = {}
    for campo in ALINEACION_CAMPOS:
        key = campo["key"]
        difieren = sum(1 for r in resultados if r[f"match_{key}"] == "Difiere")
        coinciden = total - difieren
        por_campo[key] = {
            "key": key,
            "label": campo["label"],
            "coinciden": coinciden,
            "difieren": difieren,
            "porcentaje_coincidencia": round((coinciden / total) * 100, 1) if total else 0.0,
        }

    stats = {
        "total_activas": total,
        "total_alineadas": alineadas,
        "total_con_diferencias": total - alineadas,
        "con_diferencias_ocupadas": con_diferencias_ocupadas,
        "con_diferencias_vacantes": con_diferencias_vacantes,
        "total_vacantes": vacantes,
        "total_ocupadas": total - vacantes,
        "porcentaje_alineacion_general": round((alineadas / total) * 100, 1) if total else 0.0,
        "por_campo": por_campo,
    }
    cache.set(ALINEACION_STATS_CACHE_KEY, stats, ALINEACION_CACHE_TTL)
    return stats


def _alineacion_valor_o_vacio(v):
    return "" if v is None else str(v)


def _alineacion_texto_coincide(hay, condicion, needle):
    s = _alineacion_valor_o_vacio(hay).strip().lower()
    n = (needle or "").strip().lower()
    if condicion == "istartswith":
        return s.startswith(n)
    if condicion == "iendswith":
        return s.endswith(n)
    if condicion == "iexact":
        return s == n
    return n in s  # icontains / default


def apply_dynamic_column_filters_dict(resultados, request):
    """Equivalente a `apply_dynamic_column_filters` pero sobre una lista de
    dicts en memoria en vez de un queryset (necesario porque las columnas de
    alineación -match_<campo>/estado_alineacion- no son campos reales de
    ningún modelo). Soporta el mismo contrato que ya consume el frontend:
    ``?campo=v1,v2`` (__in), sufijos __icontains/__istartswith/__iendswith/
    __iexact, negación ``exclude__campo=...`` y el sentinel de "(Vacío)"."""
    if not resultados:
        return resultados
    campos_validos = resultados[0].keys()

    for param, val in request.query_params.items():
        if param in FILTER_SKIP_PARAMS or not val:
            continue
        is_exclude = param.startswith("exclude__")
        actual_param = param[9:] if is_exclude else param
        field, _, suffix = actual_param.partition("__")
        if field not in campos_validos:
            continue

        val_list = [
            "" if v.strip() == EMPTY_VALUE_TOKEN else v.strip()
            for v in val.split(",") if v.strip()
        ]
        if not val_list:
            continue

        if suffix in ("icontains", "istartswith", "iendswith", "iexact"):
            needle = val_list[0]
            keep = lambda row: _alineacion_texto_coincide(row.get(field), suffix, needle)
        elif suffix == "in" or len(val_list) > 1:
            selected = set(val_list)
            keep = lambda row: _alineacion_valor_o_vacio(row.get(field)).strip() in selected
        else:
            needle = val_list[0]
            keep = lambda row: _alineacion_valor_o_vacio(row.get(field)).strip() == needle

        resultados = [r for r in resultados if not keep(r)] if is_exclude else [r for r in resultados if keep(r)]

    return resultados


def apply_text_search_dict(resultados, query, fields):
    query = (query or "").strip().lower()
    if not query:
        return resultados
    return [
        r for r in resultados
        if any(query in _alineacion_valor_o_vacio(r.get(f)).lower() for f in fields)
    ]


def apply_distinct_field_dict(resultados, field, search=""):
    search = (search or "").strip().lower()
    counts = {}
    for r in resultados:
        val = r.get(field)
        val = "" if val is None else val
        if search and search not in str(val).lower():
            continue
        counts[val] = counts.get(val, 0) + 1
    def _orden(kv):
        v = kv[0]
        return (0, v) if isinstance(v, (int, float)) else (1, str(v))

    return [
        {"value": v, "count": c}
        for v, c in sorted(counts.items(), key=_orden)
    ]


class MovPosAlineacionView(APIView):
    """
    Comprobar Alineación Organizacional: cruza cada plaza activa de MOV_POS
    (vía MOV_POS_LATEST) con su fila correspondiente en EMPLEADOS_COMPLETOS_SIG
    (join por no_pos_actual = Posición) y compara los 13 campos que deberían
    coincidir entre ambas tablas (ver ALINEACIÓN_PLAZA_PERSONA). Devuelve, por
    plaza, el detalle campo a campo (valor en cada tabla + si coincide) y
    agregados globales (% de alineación general y por columna).

    Filtros soportados (mismo contrato que el resto de tablas de Mov. Posiciones):
    ``search``, filtros de columna (``campo=v1,v2``, ``__icontains``, etc.,
    incl. ``exclude__``) sobre cualquier columna base o calculada
    (``estado_alineacion``, ``match_<campo>``, ``ocupacion``), ``distinct_field``/
    ``distinct_search``, ``sort_by``/``sort_order``, ``page``/``page_size``,
    ``no_pagination``. Además ``diff_fields`` (lista separada por comas de
    keys de ``ALINEACION_CAMPOS``, p.ej. ``id_departamento,dependencia_directa``):
    conserva solo las plazas que difieren en AL MENOS UNO de esos campos (OR),
    a diferencia de los filtros de columna normales que son AND entre columnas.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        try:
            dataset = _construir_dataset_alineacion()
            stats = get_mov_pos_alineacion_stats(dataset)

            resultados = apply_text_search_dict(
                dataset,
                request.query_params.get("search", ""),
                ["no_pos_actual", "nombre_puesto", "unidad_de_negocio", "depnd_drt", "ubicacion", "ocupante_nombre"],
            )

            diff_fields_param = request.query_params.get("diff_fields", "").strip()
            if diff_fields_param:
                campos_validos = {c["key"] for c in ALINEACION_CAMPOS}
                diff_keys = [k.strip() for k in diff_fields_param.split(",") if k.strip() in campos_validos]
                if diff_keys:
                    resultados = [
                        r for r in resultados
                        if any(r.get(f"match_{k}") == "Difiere" for k in diff_keys)
                    ]

            # Filtros de columna ANTES de la rama distinct_field: si no, "Vista
            # actual" en el dropdown de columna ignoraría el resto de filtros
            # de columna activos (mismo bug que en MovimientosPersonalListView
            # / MovPosDetalleView).
            resultados = apply_dynamic_column_filters_dict(resultados, request)

            distinct_field = request.query_params.get("distinct_field", "").strip()
            if distinct_field:
                return Response(
                    apply_distinct_field_dict(
                        resultados, distinct_field, request.query_params.get("distinct_search", "")
                    )
                )

            sort_by = request.query_params.get("sort_by", "").strip()
            sort_order = request.query_params.get("sort_order", "asc").strip().lower()
            if sort_by:
                sort_fields = [f.strip() for f in sort_by.split(",") if f.strip()]
                resultados = sorted(
                    resultados,
                    key=lambda row: tuple(_alineacion_valor_o_vacio(row.get(f)).lower() for f in sort_fields),
                    reverse=(sort_order == "desc"),
                )
            else:
                resultados = sorted(resultados, key=lambda r: _alineacion_valor_o_vacio(r.get("no_pos_actual")))

            count = len(resultados)

            no_pagination = request.query_params.get("no_pagination", "false").strip().lower() == "true"
            if no_pagination:
                return Response({"count": count, "next": None, "previous": None, "results": resultados, "stats": stats})

            try:
                page = max(1, int(request.query_params.get("page", 1)))
            except (TypeError, ValueError):
                page = 1
            try:
                page_size = max(1, min(int(request.query_params.get("page_size", 50)), 10000))
            except (TypeError, ValueError):
                page_size = 50

            start = (page - 1) * page_size
            return Response(
                {
                    "count": count,
                    "next": page * page_size < count,
                    "previous": page > 1,
                    "results": resultados[start:start + page_size],
                    "stats": stats,
                }
            )
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MovPosAlineacionHistoricoView(APIView):
    """Histórico diario del % de Alineación General, poblado por la tarea
    Celery `importar_zafiro` (ver `_actualizar_historico_alineacion_general`
    en plantilla/tasks.py: 1 fila por día, upsert). Alimenta la gráfica de
    tendencia en AlineacionOrganizacionalTab."""

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        try:
            from datetime import timedelta

            from django.utils import timezone

            qs = AlineacionOrganizacionalHistorico.objects.all().order_by("fecha")
            dias_param = request.query_params.get("dias", "").strip()
            if dias_param:
                try:
                    desde = timezone.localdate() - timedelta(days=int(dias_param))
                    qs = qs.filter(fecha__gte=desde)
                except ValueError:
                    pass

            resultados = [
                {
                    "fecha": r.fecha.isoformat(),
                    "porcentaje_alineacion_general": float(r.porcentaje_alineacion_general),
                    "total_activas": r.total_activas,
                    "total_alineadas": r.total_alineadas,
                    "total_con_diferencias": r.total_con_diferencias,
                }
                for r in qs
            ]
            return Response({"results": resultados})
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MovPosExportExcelView(APIView):
    """Genera y descarga directamente el Excel de Movimientos de Posiciones
    con los filtros activos en el frontend. Evita que el cliente descargue
    todos los datos en JSON y ejecute ExcelJS localmente."""

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from django.db import connection as db_connection
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        # ── 1. Columnas visibles ────────────────────────────────────────────
        visible_raw = request.query_params.get("visible_columns", "").strip()
        if visible_raw:
            visible_keys = [k.strip() for k in visible_raw.split(",") if k.strip()]
        else:
            visible_keys = list(MOV_POS_COLUMN_LABELS.keys())

        # ── 2. Construir queryset con los mismos filtros que MovPosDetalleView ──
        queryset = MovPos.objects.all()
        fecha_anuencia_overrides = get_fecha_anuencia_overrides_map()
        fecha_anuencia_baseline = get_fecha_anuencia_baseline_map()
        fecha_anuencia_overrides_texto = get_fecha_anuencia_overrides_texto_map()
        fecha_anuencia_baseline_texto = get_fecha_anuencia_baseline_texto_map()
        queryset = annotate_fecha_anuencia(
            queryset, overrides=fecha_anuencia_overrides, baseline=fecha_anuencia_baseline
        )

        oficio = request.query_params.get("oficio")
        nivel = request.query_params.get("nivel")
        if oficio or nivel:
            posiciones_qs = Plantilla1800Plazas.objects.all()
            if oficio:
                if oficio == "(vacío)":
                    posiciones_qs = posiciones_qs.filter(
                        Q(of_de_solicitud__isnull=True) | Q(of_de_solicitud="")
                    )
                else:
                    posiciones_qs = posiciones_qs.filter(of_de_solicitud=oficio)
            if nivel:
                posiciones_qs = posiciones_qs.filter(nivel=nivel)
            queryset = queryset.filter(
                no_pos_actual__in=posiciones_qs.values("posición")
            )

        is_latest = request.query_params.get("is_latest", "true").lower() != "false"
        if is_latest:
            cache_key_latest = "latest_movpos_sub_ids"
            sub_ids = cache.get(cache_key_latest)
            if sub_ids is None:
                with db_connection.cursor() as cursor:
                    cursor.execute(LATEST_MOVPOS_RAW_SQL)
                    sub_ids = [row[0] for row in cursor.fetchall() if row[0]]
                cache.set(cache_key_latest, sub_ids, 600)
            queryset = queryset.filter(id__in=sub_ids)

        queryset = apply_text_search(
            queryset,
            request.query_params.get("search", ""),
            ["no_pos_actual", "motivo", "unidad_de_negocio", "unidad_adva",
             "puesto_ptal", "descr", "nombre_puesto"],
        )

        def fecha_anuencia_column_resolver(base_field, suffix, val_list, is_exclude):
            # Mismo fix que en MovPosDetalleView.get — ver comentario ahí.
            if base_field != "fecha_anuencia":
                return None
            posiciones_ocupadas = get_posiciones_ocupadas_set()
            wanted = set(val_list)
            match_pos = [
                pos for pos, fa in queryset.values_list("no_pos_actual", "fecha_anuencia")
                if resolve_fecha_anuencia_value(
                    pos, fa, posiciones_ocupadas,
                    fecha_anuencia_overrides_texto, fecha_anuencia_baseline_texto,
                ) in wanted
            ]
            return Q(no_pos_actual__in=match_pos) if match_pos else Q(pk__in=[])

        queryset = apply_dynamic_column_filters(
            queryset, request, MovPos, extra_valid_fields={"fecha_anuencia"},
            computed_field_resolver=fecha_anuencia_column_resolver,
        )

        # ocupacion computed column
        ocupacion_param_key = None
        for k in request.query_params.keys():
            base = k[9:] if k.startswith("exclude__") else k
            if base == "ocupacion" or base.startswith("ocupacion__"):
                ocupacion_param_key = k
                break

        cache_key_ocupadas = "mov_pos_ocupadas_set"

        def _get_posiciones_ocupadas():
            pos_ocup = cache.get(cache_key_ocupadas)
            if pos_ocup is None:
                with db_connection.cursor() as cursor:
                    cursor.execute(OCUPADAS_RAW_SQL)
                    pos_ocup = set(row[0] for row in cursor.fetchall() if row[0])
                cache.set(cache_key_ocupadas, pos_ocup, 600)
            return pos_ocup

        is_vacantes_only = False
        if ocupacion_param_key:
            posiciones_ocupadas = _get_posiciones_ocupadas()
            ocupacion_raw = request.query_params.get(ocupacion_param_key, "")
            is_exclude = ocupacion_param_key.startswith("exclude__")
            suffix = (
                ocupacion_param_key.split("__", 1)[1]
                if "__" in (ocupacion_param_key[9:] if is_exclude else ocupacion_param_key)
                else "in"
            )
            if suffix == "in":
                selected_vals = set(v.strip() for v in ocupacion_raw.split(",") if v.strip())
            else:
                needle = ocupacion_raw.strip().lower()
                candidates = ["Ocupada", "Vacante"]
                if suffix in ("icontains",):
                    selected_vals = {c for c in candidates if needle in c.lower()}
                elif suffix in ("istartswith",):
                    selected_vals = {c for c in candidates if c.lower().startswith(needle)}
                elif suffix in ("iendswith",):
                    selected_vals = {c for c in candidates if c.lower().endswith(needle)}
                elif suffix in ("iexact",):
                    selected_vals = {c for c in candidates if c.lower() == needle}
                else:
                    selected_vals = set(candidates)
            if is_exclude:
                selected_vals = {"Ocupada", "Vacante"} - selected_vals
            want_ocupada = "Ocupada" in selected_vals
            want_vacante = "Vacante" in selected_vals
            if want_ocupada and not want_vacante:
                queryset = queryset.filter(no_pos_actual__in=list(posiciones_ocupadas))
            elif want_vacante and not want_ocupada:
                queryset = queryset.exclude(no_pos_actual__in=list(posiciones_ocupadas))
                is_vacantes_only = True
            elif not want_ocupada and not want_vacante:
                queryset = queryset.none()

        total_mov_raw = (
            request.query_params.get("total_movimientos__in")
            or request.query_params.get("total_movimientos")
        )
        if total_mov_raw:
            selected_counts = set()
            for v in total_mov_raw.split(","):
                v = v.strip()
                if v:
                    try:
                        selected_counts.add(int(v))
                    except ValueError:
                        pass
            if selected_counts:
                pos_list = list(queryset.values_list("no_pos_actual", flat=True).distinct())
                full_counts = dict(
                    MovPos.objects.filter(no_pos_actual__in=pos_list)
                    .values("no_pos_actual")
                    .annotate(c=Count("id"))
                    .values_list("no_pos_actual", "c")
                )
                match_pos = [p for p, c in full_counts.items() if c in selected_counts]
                queryset = queryset.filter(no_pos_actual__in=match_pos)
            else:
                queryset = queryset.none()

        # Advanced filters
        def _computed_resolver(column, condition, value):
            def _text_match(haystack, cond, needle):
                s, n = str(haystack).lower(), str(needle).lower()
                if cond == "contains": return n in s
                if cond == "not_contains": return n not in s
                if cond == "starts_with": return s.startswith(n)
                if cond == "not_starts_with": return not s.startswith(n)
                if cond == "ends_with": return s.endswith(n)
                if cond == "not_ends_with": return not s.endswith(n)
                if cond == "equals": return s == n
                if cond == "not_equals": return s != n
                return False

            if column == "ocupacion":
                pos_ocup = _get_posiciones_ocupadas()
                selected = {c for c in ["Ocupada", "Vacante"] if _text_match(c, condition, value)}
                want_o = "Ocupada" in selected
                want_v = "Vacante" in selected
                if want_o and want_v:
                    return Q(no_pos_actual__isnull=False) | Q(no_pos_actual__isnull=True)
                if want_o:
                    return Q(no_pos_actual__in=list(pos_ocup))
                if want_v:
                    return ~Q(no_pos_actual__in=list(pos_ocup))
                return Q(pk__in=[])
            if column == "total_movimientos":
                pos_list = list(queryset.values_list("no_pos_actual", flat=True).distinct())
                if not pos_list:
                    return Q(pk__in=[])
                full_counts = dict(
                    MovPos.objects.filter(no_pos_actual__in=pos_list)
                    .values("no_pos_actual")
                    .annotate(c=Count("id"))
                    .values_list("no_pos_actual", "c")
                )
                match_pos = [p for p, c in full_counts.items() if _text_match(c, condition, value)]
                return Q(no_pos_actual__in=match_pos) if match_pos else Q(pk__in=[])
            return None

        queryset = apply_advanced_filters(
            queryset, request, MovPos, computed_resolver=_computed_resolver,
            extra_valid_fields={"fecha_anuencia"},
        )

        # Sorting
        valid_fields = [f.name for f in MovPos._meta.get_fields()] + ["fecha_anuencia"]
        text_fields_set = {
            f.name for f in MovPos._meta.get_fields()
            if f.get_internal_type() in ("CharField", "TextField")
        }
        sort_by_param = request.query_params.get("sort_by", "").strip()
        sort_order = request.query_params.get("sort_order", "desc").strip().lower()
        if sort_by_param:
            sort_fields = [f.strip() for f in sort_by_param.split(",")]
            order_by_args = []
            for field in sort_fields:
                if field in valid_fields:
                    is_text = field in text_fields_set
                    target_sort_field = f"trimmed_{field}" if is_text else field
                    if is_text and target_sort_field not in queryset.query.annotations:
                        queryset = queryset.annotate(**{target_sort_field: Trim(field)})
                    order_by_args.append(f"-{target_sort_field}" if sort_order == "desc" else target_sort_field)
            if order_by_args:
                queryset = queryset.order_by(*order_by_args)
            else:
                queryset = queryset.order_by("-f_efva", "-fecha_captura", "no_pos_actual")
        else:
            queryset = queryset.order_by("-f_efva", "-fecha_captura", "no_pos_actual")

        # ── 3. Materializar datos ───────────────────────────────────────────
        resultados = list(queryset.values())
        counts = dict(MovPos.objects.values_list("no_pos_actual").annotate(c=Count("id")))
        posiciones_ocupadas = _get_posiciones_ocupadas()
        for r in resultados:
            pos = r.get("no_pos_actual")
            r["total_movimientos"] = counts.get(pos, 1)
            r["ocupacion"] = "Ocupada" if pos in posiciones_ocupadas else "Vacante"
            r["fecha_vacancia"] = "" if pos in posiciones_ocupadas else r.get("fecha_vacancia", "")
            corregir_fecha_anuencia_row(
                r, pos, posiciones_ocupadas, fecha_anuencia_overrides,
                fecha_anuencia_overrides_texto, fecha_anuencia_baseline_texto,
            )

        # Export de solo "Vacantes": desglosa el registro decisivo (baja o
        # traslado) y la insubsistencia de cada posición vacante en columnas
        # extra, insertadas junto a "Categoría Vacancia" y "Tuvo Insubsistencia"
        # aunque el usuario las haya ocultado en la UI de la tabla.
        if is_vacantes_only:
            _enrich_rows_with_vacancia_detalle(resultados)

            for anchor_key in ("fecha_vacancia", "categoria_vacancia", "tuvo_insubsistencia"):
                if anchor_key not in visible_keys:
                    visible_keys.append(anchor_key)

            def _insert_after(keys, anchor, new_keys):
                new_keys = [k for k in new_keys if k not in keys]
                if not new_keys:
                    return keys
                idx = keys.index(anchor) + 1
                return keys[:idx] + new_keys + keys[idx:]

            visible_keys = _insert_after(visible_keys, "categoria_vacancia", VACANCIA_DETALLE_CATEGORIA_KEYS)
            visible_keys = _insert_after(visible_keys, "tuvo_insubsistencia", VACANCIA_DETALLE_INSUBSISTENCIA_KEYS)

        # ── 4. Generar Excel ────────────────────────────────────────────────
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos_Posiciones"

            header_fill = PatternFill(start_color="FF2B4C7E", end_color="FF2B4C7E", fill_type="solid")
            zebra_fill = PatternFill(start_color="FFF4F7FA", end_color="FFF4F7FA", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFFFF")
            data_font = Font(name="Segoe UI", size=9)
            gold_side = Side(style="thin", color="FFBC955C")
            gold_border = Border(left=gold_side, right=gold_side, top=gold_side, bottom=gold_side)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")

            letterhead_off = add_excel_letterhead(ws, len(visible_keys))

            # Export de solo "Vacantes": portada con leyenda de categorías +
            # conteos y grupos de columnas explicados, encima del header real.
            if is_vacantes_only:
                _write_vacancia_report_cover(ws, visible_keys, resultados, row_offset=letterhead_off)
            header_row = letterhead_off + (9 if is_vacantes_only else 1)
            data_start_row = header_row + 1

            # Header row
            for col_idx, key in enumerate(visible_keys, start=1):
                label = MOV_POS_COLUMN_LABELS.get(key) or VACANCIA_DETALLE_COLUMN_LABELS.get(key, key)
                cell = ws.cell(row=header_row, column=col_idx, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = gold_border
                cell.alignment = align_center
            ws.row_dimensions[header_row].height = 24

            # Data rows
            for row_idx, row_data in enumerate(resultados, start=data_start_row):
                is_zebra = (row_idx - data_start_row) % 2 == 1
                for col_idx, key in enumerate(visible_keys, start=1):
                    val = row_data.get(key)
                    if val is None:
                        val = ""
                    elif not isinstance(val, (int, float, bool)):
                        val = str(val)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = gold_border
                    cell.alignment = align_center if key in MOV_POS_MONO_COLUMNS else align_left
                    if is_zebra:
                        cell.fill = zebra_fill
                ws.row_dimensions[row_idx].height = 20

            # Auto-fit column widths (con piso extra para las columnas de
            # detalle de vacancia/insubsistencia, que en el export de Vacantes
            # también cargan el texto largo de la portada en las filas 1-8)
            for col_idx, key in enumerate(visible_keys, start=1):
                col_letter = get_column_letter(col_idx)
                header_len = len(MOV_POS_COLUMN_LABELS.get(key) or VACANCIA_DETALLE_COLUMN_LABELS.get(key, key))
                max_len = max(
                    (len(str(r.get(key, "") or "")) for r in resultados),
                    default=0,
                )
                width = min(max(max_len, header_len) + 4, 60)
                if is_vacantes_only:
                    width = max(width, VACANCIA_EXPORT_MIN_COL_WIDTH.get(key, 0))
                ws.column_dimensions[col_letter].width = width

            if is_vacantes_only:
                for col_letter in ("F", "G", "H"):
                    ws.column_dimensions[col_letter].width = max(
                        ws.column_dimensions[col_letter].width or 0,
                        VACANCIA_LEGEND_MIN_COL_WIDTH.get(col_letter, 0),
                    )

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="Movimientos_Posiciones.xlsx"'
            return response

        except Exception:
            logger.exception("Error generando Excel de movimientos de posiciones")
            return Response(
                {"error": "Error generando Excel"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CadenaMandoView(APIView):
    """
    Vista para buscar la cadena de mando jerárquica en EMPLEADOS_COMPLETOS_SIG.
    Busca por posición, nombre completo o número de empleado.

    `direction` (query param, default "arriba"):
      - "arriba": Bottom-Up clásico — sube de la posición buscada hasta la
        cúspide, siguiendo `DependenciaDirecta` (un solo camino, 1 jefe por
        nivel). Campo de profundidad: `Nivel_Hacia_Arriba`.
      - "abajo" (8.5 QA): Top-Down — todos los subordinados directos e
        indirectos de la posición buscada (árbol completo, N hijos por
        nivel). Campo de profundidad: `Nivel_Hacia_Abajo`. Incluye la propia
        posición base en `Nivel_Hacia_Abajo = 0` para que el front arme el
        árbol completo (raíz + descendientes) a partir de un solo arreglo
        plano, agrupando por `Jefe_Directo`.
    """

    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
    )

    # Tope defensivo de niveles: evita recursión sin fin si datos sucios
    # generan un ciclo (p. ej. A depende de B y B depende de A por error de
    # captura). Ningún organigrama real de ANAM tiene más de ~10 niveles.
    MAX_PROFUNDIDAD = 20

    def get(self, request):
        query = request.query_params.get("q", "").strip()
        direction = request.query_params.get("direction", "arriba").strip().lower()
        if not query:
            return Response(
                {"error": "Se requiere el parámetro 'q' para buscar."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if direction not in ("arriba", "abajo"):
            return Response(
                {"error": "direction debe ser 'arriba' o 'abajo'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1. Buscar la posición base
        base_employee = EmpleadosCompletosSig.objects.filter(
            Q(posicion=query) | Q(nombres__icontains=query) | Q(id_empleado=query)
        ).first()

        if not base_employee:
            return Response(
                {"error": f"No se encontró un empleado con el criterio '{query}'."},
                status=status.HTTP_404_NOT_FOUND,
            )

        base_posicion = base_employee.posicion

        # 2. Ejecutar CTE recursivo (hacia arriba o hacia abajo)
        # Columnas adicionales (Estado Nómina, UN/UA, departamento): permiten al
        # front pintar el estado real de cada posición (no inferirlo por nombre
        # vacío) y cruzar `Id_Departamento` con el catálogo de ORGANIGRAMA_ANAM
        # (isSIGInfo=1) client-side vía useOrganigramaCatalog — sin JOIN aquí,
        # que además truena por collations distintas entre ambas tablas.
        _CAMPOS_NODO = """
                        `Posición` AS Posicion,
                        `Nombres` AS Empleado,
                        `Nombre Puesto Funcional` AS Puesto_Funcional,
                        `Nivel` AS Nivel,
                        `Estado Nómina` AS Estado_Nomina,
                        `Cd UN` AS Cd_UN,
                        `Unidad de Negocio` AS Unidad_Negocio,
                        `Cd UA` AS Cd_UA,
                        `Unidad Administrativa` AS Unidad_Administrativa,
                        `Id Departamento` AS Id_Departamento,
                        `Departamento` AS Departamento,
                        `DependenciaDirecta` AS Jefe_Directo,
        """
        _CAMPOS_FINALES = (
            "Posicion, Empleado, Puesto_Funcional, Nivel, Estado_Nomina, "
            "Cd_UN, Unidad_Negocio, Cd_UA, Unidad_Administrativa, "
            "Id_Departamento, Departamento, Jefe_Directo"
        )

        if direction == "arriba":
            sql = """
                WITH RECURSIVE CadenaHaciaArriba AS (
                    SELECT
                        {campos}
                        1 AS Nivel_Hacia_Arriba
                    FROM EMPLEADOS_COMPLETOS_SIG
                    WHERE `Posición` = %s

                    UNION ALL

                    SELECT
                        jefe.`Posición`,
                        jefe.`Nombres`,
                        jefe.`Nombre Puesto Funcional`,
                        jefe.`Nivel`,
                        jefe.`Estado Nómina`,
                        jefe.`Cd UN`,
                        jefe.`Unidad de Negocio`,
                        jefe.`Cd UA`,
                        jefe.`Unidad Administrativa`,
                        jefe.`Id Departamento`,
                        jefe.`Departamento`,
                        jefe.`DependenciaDirecta`,
                        empleado.Nivel_Hacia_Arriba + 1
                    FROM EMPLEADOS_COMPLETOS_SIG jefe
                    INNER JOIN CadenaHaciaArriba empleado ON jefe.`Posición` = empleado.Jefe_Directo
                    WHERE empleado.Jefe_Directo IS NOT NULL
                      AND empleado.Jefe_Directo != ''
                      AND empleado.Jefe_Directo != '0'
                      AND jefe.`Posición` != empleado.Posicion
                      AND empleado.Nivel_Hacia_Arriba < %s
                )
                SELECT
                    {finales}, Nivel_Hacia_Arriba
                FROM CadenaHaciaArriba
                ORDER BY Nivel_Hacia_Arriba ASC;
            """.format(campos=_CAMPOS_NODO, finales=_CAMPOS_FINALES)
        else:
            sql = """
                WITH RECURSIVE CadenaHaciaAbajo AS (
                    SELECT
                        {campos}
                        0 AS Nivel_Hacia_Abajo
                    FROM EMPLEADOS_COMPLETOS_SIG
                    WHERE `Posición` = %s

                    UNION ALL

                    SELECT
                        subordinado.`Posición`,
                        subordinado.`Nombres`,
                        subordinado.`Nombre Puesto Funcional`,
                        subordinado.`Nivel`,
                        subordinado.`Estado Nómina`,
                        subordinado.`Cd UN`,
                        subordinado.`Unidad de Negocio`,
                        subordinado.`Cd UA`,
                        subordinado.`Unidad Administrativa`,
                        subordinado.`Id Departamento`,
                        subordinado.`Departamento`,
                        subordinado.`DependenciaDirecta`,
                        padre.Nivel_Hacia_Abajo + 1
                    FROM EMPLEADOS_COMPLETOS_SIG subordinado
                    INNER JOIN CadenaHaciaAbajo padre ON subordinado.`DependenciaDirecta` = padre.Posicion
                    WHERE subordinado.`Posición` != padre.Posicion
                      AND padre.Nivel_Hacia_Abajo < %s
                )
                SELECT
                    {finales}, Nivel_Hacia_Abajo
                FROM CadenaHaciaAbajo
                ORDER BY Nivel_Hacia_Abajo ASC;
            """.format(campos=_CAMPOS_NODO, finales=_CAMPOS_FINALES)

        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, [base_posicion, self.MAX_PROFUNDIDAD])
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

                # 3. Posición activa (independiente del estado del empleado que la
                # ocupa): una posición se considera activa cuando su movimiento MÁS
                # RECIENTE en MOV_POS tiene `Estado Psn` = 'A'. Esto puede diferir
                # de `Estado_Nomina` (EMPLEADOS_COMPLETOS_SIG) — p. ej. una posición
                # puede seguir marcada como ocupada ahí sin que el catálogo de
                # movimientos la reconozca como vigente. Restringido con IN a las
                # posiciones de esta cadena (no escanea las ~53k filas de MOV_POS).
                posiciones = [str(r["Posicion"]).strip() for r in results if r.get("Posicion")]
                posiciones_activas = set()
                if posiciones:
                    placeholders = ",".join(["%s"] * len(posiciones))
                    activas_sql = """
                        SELECT DISTINCT TRIM(m.`Nº Pos Actual`) AS Posicion
                        FROM MOV_POS m
                        INNER JOIN (
                            SELECT id FROM (
                                SELECT id, ROW_NUMBER() OVER (
                                    PARTITION BY `Nº Pos Actual`
                                    ORDER BY `F Efva` DESC, `Fecha Captura` DESC, `F/H Últ Actz` DESC, id DESC
                                ) as rn
                                FROM MOV_POS
                                WHERE `Nº Pos Actual` IN ({ph})
                            ) ranked WHERE rn = 1
                        ) latest ON m.id = latest.id
                        WHERE m.`Estado Psn` = 'A';
                    """.format(ph=placeholders)
                    with connection.cursor() as cursor2:
                        cursor2.execute(activas_sql, posiciones)
                        posiciones_activas = {row[0] for row in cursor2.fetchall()}

                for r in results:
                    r["Posicion_Activa"] = str(r.get("Posicion", "")).strip() in posiciones_activas

                return Response(
                    {
                        "empleado_base": {
                            "posicion": base_employee.posicion,
                            "nombres": base_employee.nombres,
                            "puesto_funcional": base_employee.nombre_puesto_funcional
                            if hasattr(base_employee, "nombre_puesto_funcional")
                            else "",
                            "nivel": base_employee.nivel
                            if hasattr(base_employee, "nivel")
                            else "",
                        },
                        "direction": direction,
                        "cadena": results,
                    },
                    status=status.HTTP_200_OK,
                )
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


from .models import ZafiroBitacora


class ZafiroBitacoraView(APIView):
    """
    Endpoint para obtener el historial de ejecuciones de ZAFIRO.
    """

    view_permission = "authentication.view_monitoreo_zafiro"

    def get(self, request):
        limit = request.query_params.get("limit")
        logs = ZafiroBitacora.objects.all()
        if limit is not None:
            logs = logs[: int(limit)]

        data = []
        for log in logs:
            data.append(
                {
                    "id": log.id,
                    "fecha_ejecucion": log.fecha_ejecucion.isoformat(),
                    "duracion_segundos": log.duracion_segundos,
                    "registros_posiciones": log.registros_posiciones,
                    "registros_completos": log.registros_completos,
                    "registros_bajas": log.registros_bajas,
                    "registros_historial": log.registros_historial,
                    "status": log.status,
                    "error_message": log.error_message,
                    "es_historico": log.es_historico,
                    "logs_en_vivo": log.logs_en_vivo,
                }
            )

        return Response(data, status=status.HTTP_200_OK)


class ZafiroDuracionPromedioPorHoraView(APIView):
    """
    Endpoint para obtener el promedio de duración (segundos) de las
    importaciones exitosas de ZAFIRO, agrupado por hora del día (hora
    local de Ciudad de México).
    """

    view_permission = "authentication.view_monitoreo_zafiro"

    # `fecha_ejecucion` se guarda en UTC (USE_TZ=True), pero el servidor
    # MySQL no tiene cargadas las tablas de zona horaria (CONVERT_TZ
    # regresa NULL), así que no se puede pedirle a la DB que convierta a
    # America/Mexico_City. Como esa zona ya no aplica horario de verano,
    # el desfase es un entero fijo y se resta en Python.
    OFFSET_HORAS_CDMX = -6

    def get(self, request):
        import datetime

        resultados = (
            ZafiroBitacora.objects.filter(status="EXITO")
            .annotate(
                hora=ExtractHour("fecha_ejecucion", tzinfo=datetime.timezone.utc)
            )
            .values("hora")
            .annotate(
                duracion_promedio=Avg("duracion_segundos"),
                total_casos=Count("id"),
            )
            .order_by("hora")
        )

        data = [
            {
                "hora": (r["hora"] + self.OFFSET_HORAS_CDMX) % 24,
                "duracion_promedio_segundos": round(r["duracion_promedio"], 2)
                if r["duracion_promedio"] is not None
                else None,
                "total_casos": r["total_casos"],
            }
            for r in resultados
        ]
        data.sort(key=lambda d: d["hora"])

        return Response(data, status=status.HTTP_200_OK)


def _calcular_ultima_actualizacion_zafiro():
    """
    Busca la última corrida de ZafiroBitacora (EXITO, o "OK" / cualquiera como
    fallback) y devuelve (fecha_fin_iso, status), o (None, None) si no hay
    ninguna bitácora. Compartido por UltimaActualizacionZafiroView (polling
    REST) y ZafiroSSEView (fallback por BD, ver comentario ahí).
    """
    last_success = (
        ZafiroBitacora.objects.filter(status="EXITO")
        .order_by("-fecha_ejecucion")
        .first()
    )
    if not last_success:
        last_success = (
            ZafiroBitacora.objects.filter(status="OK")
            .order_by("-fecha_ejecucion")
            .first()
        )
    if not last_success:
        last_success = ZafiroBitacora.objects.all().first()

    if not last_success:
        return None, None

    from datetime import timedelta
    fecha_fin = last_success.fecha_ejecucion
    if last_success.duracion_segundos:
        fecha_fin += timedelta(seconds=last_success.duracion_segundos)
    return fecha_fin.isoformat(), last_success.status


class UltimaActualizacionZafiroView(APIView):
    """
    Endpoint público para obtener la fecha y estatus de la última actualización exitosa de ZAFIRO.
    """

    permission_classes = [AllowAny]

    def get(self, request):
        fecha_iso, last_status = _calcular_ultima_actualizacion_zafiro()
        if fecha_iso:
            return Response(
                {"fecha": fecha_iso, "status": last_status},
                status=status.HTTP_200_OK,
            )
        return Response({"fecha": None, "status": None}, status=status.HTTP_200_OK)


class IniciarSincronizacionZafiroView(APIView):
    """
    Endpoint para arrancar manualmente la sincronización de ZAFIRO.
    """

    view_permission = "authentication.view_monitoreo_zafiro"

    def post(self, request):
        if ZafiroBitacora.objects.filter(status="RUNNING").exists():
            return Response(
                {"error": "Ya hay una sincronización en ejecución en este momento."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .tasks import importar_zafiro

        importar_zafiro.delay()
        return Response(
            {"message": "Sincronización manual iniciada correctamente."},
            status=status.HTTP_200_OK,
        )


class InvalidarCacheManualView(APIView):
    """
    Endpoint para uso humano desde el front (botón "Borrar caché del
    servidor" en Monitoreo ZAFIRO). Borra la caché ya, sin bitácora_id ni
    token de servicio — a diferencia de `InvalidarCacheZafiroView`, que es
    exclusivo del worker de Celery remoto.
    """

    view_permission = "authentication.view_monitoreo_zafiro"

    def post(self, request):
        from .cache_invalidation import invalidar_todo_el_cache_servidor

        borradas = invalidar_todo_el_cache_servidor()
        return Response(
            {"status": "ok", "cache_keys_borradas": borradas},
            status=status.HTTP_200_OK,
        )


class InvalidarCacheZafiroView(APIView):
    """
    Endpoint interno: lo llama el worker de Celery de `copia_back` (PC
    Windows) justo al terminar `importar_zafiro`, para que ESTE servidor
    invalide su propia caché (Redis local del servidor, distinto al de la
    PC Windows) y publique el evento SSE — ver docstring de
    `cache_invalidation.invalidar_todo_el_cache_servidor`.

    Autenticación: Token de DRF de una cuenta de servicio dedicada (no un
    usuario humano). No usa `HasModulePermission` (el default del proyecto)
    porque esa cuenta no tiene permisos de módulo asignados.
    """

    from rest_framework.authentication import TokenAuthentication

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from datetime import timedelta

        from .cache_invalidation import invalidar_todo_el_cache_servidor

        bitacora_id = request.data.get("bitacora_id")
        if not bitacora_id:
            return Response(
                {"error": "Falta 'bitacora_id'."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            bitacora = ZafiroBitacora.objects.get(pk=bitacora_id)
        except ZafiroBitacora.DoesNotExist:
            return Response(
                {"error": f"No existe ZafiroBitacora con id={bitacora_id}."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if bitacora.status != "EXITO":
            return Response(
                {"error": f"ZafiroBitacora {bitacora_id} tiene status '{bitacora.status}', se esperaba 'EXITO'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Recalcula FECHA_OCUPACION/ID_REGISTRO_DES_FECHA_OCUPACION en
        # MOV_POS (nacen NULL tras el swap Blue-Green de importar_zafiro,
        # bulk_create no las conoce). Antes se disparaba de forma lazy desde
        # MovPosDetalleView.get la primera vez que alguien abría el tab
        # "Movimientos de Posiciones" tras un import — sin lock, así que
        # requests concurrentes podían disparar el SP en paralelo y su
        # UPDATE (window functions sobre cp_tbl_mov_completo_29_05_26)
        # sostenía locks de fila en MOV_POS varios minutos, chocando con
        # `_reaplicar_prioridad_nivel_jerarquico` y los pasos de vacancia del
        # propio pipeline de ZAFIRO (error 1205 Lock wait timeout). Aquí se
        # dispara una sola vez, justo cuando Celery avisa que el import ya
        # terminó, sin competir con ningún HTTP request de usuario.
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_actualizar_fecha_ocupacion_mov_pos();")
        except Exception:
            logger.exception(
                "Error ejecutando sp_actualizar_fecha_ocupacion_mov_pos tras import ZAFIRO"
            )

        # Recalcula MOV_POS.DIAS_OCUPADA/DIAS_VACANTE para las plazas activas
        # (sp_dias_ocupacion_masivo, ~12-15s sobre las 11,451 plazas activas).
        # DEBE correr antes de invalidar_todo_el_cache_servidor() de abajo:
        # el SP tiene su propio GET_LOCK BLOQUEANTE (a diferencia del de
        # fecha_ocupacion, que se salta si el lock está ocupado) precisamente
        # para que, si dos requests a este endpoint se traslapan, la 2a
        # espere a que la 1a termine de escribir en MOV_POS antes de seguir
        # — así ningún request invalida el caché mientras las columnas
        # todavía no reflejan el cálculo más reciente.
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_dias_ocupacion_masivo();")
        except Exception:
            logger.exception(
                "Error ejecutando sp_dias_ocupacion_masivo tras import ZAFIRO"
            )

        # Geolocaliza empleados activos sin latitud/longitud (ej. personal
        # administrativo) — igual que el SP de arriba, corre aquí y no en
        # tasks.py porque este endpoint SIEMPRE se ejecuta en este servidor,
        # sin importar en qué máquina corrió importar_zafiro (ver docstring
        # de plantilla.geocodificacion). Antes de invalidar el caché para
        # que "empleados_distribucion_geografica" se recalcule ya con las
        # coordenadas nuevas.
        try:
            from .geocodificacion import geocodificar_empleados_sin_coordenadas

            geocodificar_empleados_sin_coordenadas(bitacora)
        except Exception:
            logger.exception("Error geolocalizando empleados tras import ZAFIRO")

        borradas = invalidar_todo_el_cache_servidor()

        # Notifica "avísame cuando esta posición quede vacante/se ocupe" —
        # aquí porque este es el punto donde ESTE servidor (el que sirve a
        # los usuarios) se entera de que hay datos frescos, cuando
        # importar_zafiro corrió en la PC Windows remota (ver docstring de
        # esta vista). Si importar_zafiro corre en este mismo servidor, se
        # dispara en tasks.py (_procesar_notificaciones_posicion) en vez de
        # aquí.
        try:
            from .notificaciones_posicion import procesar_suscripciones_posicion

            procesar_suscripciones_posicion()
        except Exception:
            logger.exception("Error procesando notificaciones de posición tras invalidación remota de caché")

        fecha_fin = bitacora.fecha_ejecucion
        if bitacora.duracion_segundos:
            fecha_fin += timedelta(seconds=bitacora.duracion_segundos)
        try:
            import redis as redis_lib

            r = redis_lib.Redis.from_url(settings.CELERY_BROKER_URL)
            r.publish("zafiro_updates", fecha_fin.isoformat())
        except Exception:
            logger.exception("Error al publicar evento zafiro_updates tras invalidación remota de caché")

        return Response(
            {"status": "ok", "cache_keys_borradas": borradas},
            status=status.HTTP_200_OK,
        )


from django.views import View

# Vida máxima de un stream SSE antes de forzar el cierre (el cliente
# reconecta solo vía EventSource.onerror con backoff, ver
# ZafiroUpdatesContext.jsx / useCeldaUpdatesRealtime.js). Sin este tope, un
# hilo de streaming vive mientras la pestaña siga abierta y Django nunca
# libera su conexión MySQL (close_old_connections solo corre en
# request_started/request_finished, que no disparan hasta que el generator
# termina) — con varias pestañas/horas esto agota max_connections del
# servidor compartido.
SSE_MAX_LIFETIME_SECONDS = 300


class ZafiroSSEView(View):
    """
    Endpoint de Server-Sent Events (SSE) para notificar actualizaciones en tiempo real a clientes.

    El Celery worker que corre `importar_zafiro` puede vivir en una máquina
    aparte del backend Linux (ver copia_back/.env.example: "esta PC Windows
    corre su propio Redis, separado del de Linux"). El `r.publish(
    "zafiro_updates", ...)` de esa tarea llega al Redis LOCAL de esa máquina,
    no al de este servidor, así que el pubsub de abajo nunca recibe nada en
    ese escenario — el navbar solo se enteraba de la actualización hasta el
    siguiente refresh completo de página. Como respaldo, en cada timeout del
    pubsub (~20s sin mensaje) se revisa directo en la BD compartida (misma
    para ambas máquinas) si hay una corrida EXITO más reciente que la última
    enviada, y si la hay, se emite igual por este canal.
    """

    def get(self, request):
        import time

        import redis
        from django.db import close_old_connections, connections
        from django.http import StreamingHttpResponse

        def event_stream():
            r = redis.Redis.from_url(settings.CELERY_BROKER_URL)
            pubsub = r.pubsub()
            pubsub.subscribe("zafiro_updates")

            # Enviamos evento de inicialización de conexión
            yield "data: init\n\n"

            close_old_connections()
            last_sent, _ = _calcular_ultima_actualizacion_zafiro()
            start = time.monotonic()

            try:
                while True:
                    if time.monotonic() - start > SSE_MAX_LIFETIME_SECONDS:
                        break
                    # Esperar mensajes en el canal de redis con timeout de 20s
                    message = pubsub.get_message(
                        ignore_subscribe_messages=True, timeout=20.0
                    )
                    if message:
                        date_str = message["data"].decode("utf-8")
                        last_sent = date_str
                        yield f"data: {date_str}\n\n"
                    else:
                        # Sin mensaje de Redis: fallback por BD (ver docstring
                        # de la clase) antes de caer al ping de "sigo vivo".
                        close_old_connections()
                        current, _ = _calcular_ultima_actualizacion_zafiro()
                        if current and current != last_sent:
                            last_sent = current
                            yield f"data: {current}\n\n"
                        else:
                            yield ": ping\n\n"
            finally:
                try:
                    pubsub.unsubscribe("zafiro_updates")
                    pubsub.close()
                except Exception:
                    pass
                # Cierre explícito: garantiza liberar la conexión MySQL de
                # este hilo al terminar el stream, sin esperar a que
                # CONN_MAX_AGE la marque como vieja en un próximo request.
                connections.close_all()

        response = StreamingHttpResponse(
            event_stream(), content_type="text/event-stream"
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        # GZipMiddleware bufferea internamente (zlib sin flush por chunk) y
        # retiene todos los mensajes hasta que la conexión se cierra,
        # rompiendo el streaming en tiempo real. Content-Encoding ya seteado
        # hace que GZipMiddleware.process_response la omita.
        response["Content-Encoding"] = "identity"
        return response


class CeldaUpdatesSSEView(View):
    """
    SSE dedicado a cambios de celdas de EMPLEADOS_COMPLETOS_SIG (tab Detalle
    de Plantilla), para reflejar ediciones de otros usuarios en tiempo real
    (ver plantilla.celda_override.notificar_cambio_celda). A diferencia de
    ZafiroSSEView (bitácora, sin datos sensibles), este stream lleva datos de
    personal, así que exige el permiso view_plantilla_detalle vía token antes
    de aceptar la conexión — EventSource no puede mandar headers, por eso el
    token viaja como query param.
    """

    def get(self, request):
        import time

        import redis
        from django.db import connections
        from django.http import HttpResponseForbidden, StreamingHttpResponse
        from rest_framework.authtoken.models import Token

        token_key = request.GET.get("token")
        token_obj = (
            Token.objects.filter(key=token_key).select_related("user").first()
            if token_key else None
        )
        user = token_obj.user if token_obj else None
        if not user or not user.is_active or not user.has_perm("authentication.view_plantilla_detalle"):
            return HttpResponseForbidden("No autorizado.")

        def event_stream():
            r = redis.Redis.from_url(settings.CELERY_BROKER_URL)
            pubsub = r.pubsub()
            pubsub.subscribe("plantilla_celda_updates")

            yield "data: init\n\n"
            start = time.monotonic()

            try:
                while True:
                    if time.monotonic() - start > SSE_MAX_LIFETIME_SECONDS:
                        break
                    message = pubsub.get_message(
                        ignore_subscribe_messages=True, timeout=20.0
                    )
                    if message:
                        data_str = message["data"].decode("utf-8")
                        yield f"data: {data_str}\n\n"
                    else:
                        yield ": ping\n\n"
            finally:
                try:
                    pubsub.unsubscribe("plantilla_celda_updates")
                    pubsub.close()
                except Exception:
                    pass
                # Ver comentario equivalente en ZafiroSSEView: libera la
                # conexión MySQL de este hilo (auth Token query) al terminar
                # el stream, en vez de dejarla viva hasta que expire.
                connections.close_all()

        response = StreamingHttpResponse(
            event_stream(), content_type="text/event-stream"
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        response["Content-Encoding"] = "identity"
        return response


class AnuenciaAnexoUpdatesSSEView(View):
    """
    SSE dedicado a cambios de AnuenciaAnexo (sub-tab "Anuencia" de Mov.
    Posiciones), para reflejar en vivo cuando alguien agrega/quita plazas de
    un Anexo 2 desde OTRA sesión — típicamente, agregarlas desde el menú
    contextual de la tabla de Mov. Posiciones (ver AgregarAAnexo2Modal.jsx)
    mientras ese mismo Anexo 2 está abierto en AnuenciaTab.jsx en otra
    pestaña/usuario (ver `_notificar_actualizacion_anuencia_anexo`).

    Mismo patrón exacto que CeldaUpdatesSSEView (token por query param, sin
    fallback por BD porque no hay una "última corrida" que consultar aquí).
    """

    def get(self, request):
        import time

        import redis
        from django.db import connections
        from django.http import HttpResponseForbidden, StreamingHttpResponse
        from rest_framework.authtoken.models import Token

        token_key = request.GET.get("token")
        token_obj = (
            Token.objects.filter(key=token_key).select_related("user").first()
            if token_key else None
        )
        user = token_obj.user if token_obj else None
        if not user or not user.is_active or not user.has_perm("authentication.view_plantilla_mov_posiciones"):
            return HttpResponseForbidden("No autorizado.")

        def event_stream():
            r = redis.Redis.from_url(settings.CELERY_BROKER_URL)
            pubsub = r.pubsub()
            pubsub.subscribe("anuencia_anexo_updates")

            yield "data: init\n\n"
            start = time.monotonic()

            try:
                while True:
                    if time.monotonic() - start > SSE_MAX_LIFETIME_SECONDS:
                        break
                    message = pubsub.get_message(
                        ignore_subscribe_messages=True, timeout=20.0
                    )
                    if message:
                        data_str = message["data"].decode("utf-8")
                        yield f"data: {data_str}\n\n"
                    else:
                        yield ": ping\n\n"
            finally:
                try:
                    pubsub.unsubscribe("anuencia_anexo_updates")
                    pubsub.close()
                except Exception:
                    pass
                connections.close_all()

        response = StreamingHttpResponse(
            event_stream(), content_type="text/event-stream"
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        response["Content-Encoding"] = "identity"
        return response


from .models import BajasSig


class BajasSigListView(APIView):
    """
    Endpoint para obtener todos los registros de bajas sin paginación.
    """

    view_permission = "authentication.view_plantilla_bajas"

    def get(self, request):
        oficio = request.query_params.get("oficio")
        nivel = request.query_params.get("nivel")

        if oficio or nivel:
            cache_key = f"bajas_sig_list_{oficio}_{nivel}"
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return _paginated_or_full_response(request, cached_data)

            try:
                # Obtener posiciones de Plantilla1800Plazas que cumplan los filtros
                posiciones_qs = Plantilla1800Plazas.objects.all()
                if oficio:
                    if oficio == "(vacío)":
                        posiciones_qs = posiciones_qs.filter(
                            Q(of_de_solicitud__isnull=True) | Q(of_de_solicitud="")
                        )
                    else:
                        posiciones_qs = posiciones_qs.filter(of_de_solicitud=oficio)
                if nivel:
                    posiciones_qs = posiciones_qs.filter(nivel=nivel)

                posiciones_list = list(posiciones_qs.values_list("posición", flat=True))

                bajas = list(
                    BajasSig.objects.filter(posicion__in=posiciones_list).values()
                )
                cache.set(cache_key, bajas, 300)
                return _paginated_or_full_response(request, bajas)
            except Exception:
                logger.exception("Error inesperado en {}".format(request.path))
                return Response(
                    {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        cache_key = "bajas_sig_list"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return _paginated_or_full_response(request, cached_data)

        bajas = list(BajasSig.objects.all().values())
        cache.set(cache_key, bajas, 1200)
        return _paginated_or_full_response(request, bajas)


class BajasMotivosPieView(APIView):
    """
    Devuelve el conteo de bajas agrupado por Motivo para la gráfica de pastel.
    Respuesta: [{"motivo": "...", "total": N}, ...] ordenado por total descendente.
    """

    view_permission = "authentication.view_plantilla_bajas"

    def get(self, request):
        cache_key = "bajas_motivos_pie"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        data = (
            BajasSig.objects.exclude(motivo_descr__isnull=True)
            .exclude(motivo_descr__exact="")
            .values("motivo_descr")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        result = [
            {"motivo": row["motivo_descr"], "total": row["total"]} for row in data
        ]
        cache.set(cache_key, result, 1200)
        return Response(result, status=status.HTTP_200_OK)


class BajasHistoricoView(APIView):
    """
    Devuelve la evolución histórica de bajas_sig obtenida de ZAFIRO_BITACORA.
    Agrupado por día (el registro más reciente de cada día donde registros_bajas > 0).
    """

    view_permission = "authentication.view_plantilla_bajas"

    def get(self, request):
        cache_key = "bajas_historico"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        from .models import ZafiroBitacora

        queryset = ZafiroBitacora.objects.filter(registros_bajas__gt=0).order_by(
            "fecha_ejecucion"
        )
        bajas_por_dia = {}
        for r in queryset:
            bajas_por_dia[str(r.fecha_ejecucion.date())] = r.registros_bajas

        resultado = [
            {"fecha": fecha, "registros_bajas": count}
            for fecha, count in sorted(bajas_por_dia.items())
        ]
        cache.set(cache_key, resultado, 1200)
        return Response(resultado, status=status.HTTP_200_OK)


class ExportarEstatusExcelView(APIView):
    """
    Genera y exporta un archivo Excel (.xlsx) estructurado e interactivo.
    Si ya existe en caché para esa consulta exacta, lo retorna instantáneamente.
    Si no, lo genera de forma síncrona en el hilo de la petición y lo retorna.
    """

    view_permission = "authentication.view_plantilla_estatus_nomina"

    def get(self, request):
        from django.core.cache import cache
        from django.utils import timezone

        from plantilla.tasks import generar_excel_estatus_task

        uas_param = request.query_params.get("uas", "")
        levels_param = request.query_params.get("levels", "")
        group_by = request.query_params.get("group_by", "ua")

        # Consultar si ya existe el archivo Excel final generado en caché para esta consulta exacta
        import hashlib

        raw_key = f"excel_estatus_file_{uas_param}_{levels_param}_{group_by}"
        cache_key_excel = (
            f"excel_estatus_file_{hashlib.md5(raw_key.encode('utf-8')).hexdigest()}"
        )
        cached_excel_data = cache.get(cache_key_excel)
        if cached_excel_data is not None:
            filename = (
                f"Reporte_Plantilla_Estatus_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
            )
            response = HttpResponse(
                cached_excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        try:
            # Ejecutar la generación de forma síncrona
            generar_excel_estatus_task.__wrapped__(uas_param, levels_param, group_by)
        except Exception:
            logger.exception("Error generando el reporte de Excel de estatus")
            return HttpResponse("Error generando el reporte de Excel", status=500)

        # Recuperar el archivo generado desde la caché
        file_data = cache.get(cache_key_excel)
        if not file_data:
            return HttpResponse(
                "Error: No se pudo recuperar el archivo generado de la caché.",
                status=500,
            )

        filename = (
            f"Reporte_Plantilla_Estatus_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        )
        response = HttpResponse(
            file_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class OrganigramaSearchView(APIView):
    """
    Retorna el catálogo completo de ORGANIGRAMA_ANAM (útil para caché en
    memoria) para que el frontend filtre en cliente por descripcion_larga,
    departamento u ocupante (ver búsqueda "Nombre o código" en
    organigrama/page.jsx). El parámetro `q`, si se envía, filtra también
    server-side por los mismos tres campos (no lo usa el front hoy, que
    siempre pide el catálogo entero).
    Retorna la unidad_negocio para que el frontend sepa qué JSON cargar.
    """

    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
        "authentication.view_organigrama_sig",
    )

    def get(self, request):
        query = request.GET.get("q", "").strip()

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT departamento, descripcion_larga, unidad_negocio, nivel_direccion, isSIGInfo, num_posicion_gerente
                FROM ORGANIGRAMA_ANAM
                """
            )
            rows = cursor.fetchall()

        # Nombre del ocupante de la plaza titular de cada depto: cruce de
        # posiciones activas (MOV_POS_LATEST, que ya es la materialización
        # de "última fila de MOV_POS por Nº Pos Actual", ver
        # _materializar_mov_pos_latest en tasks.py) contra EMPLEADOS_COMPLETOS_SIG
        # por Posición — mismo dataset/join que el reporte de referencia
        # (CRUZA LAS POSICIONES ACTIVAS CON EMPLEADOS COMPLETOS SIG), sin
        # filtrar por Estado Nómina.
        posiciones = sorted({
            r[5] for r in rows if r[5] and r[5] != "(en blanco)"
        })
        ocupante_por_posicion = {}
        if posiciones:
            placeholders = ", ".join(["%s"] * len(posiciones))
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    SELECT e.`Posición`, e.`Nombres`
                    FROM EMPLEADOS_COMPLETOS_SIG e
                    INNER JOIN MOV_POS_LATEST m ON e.`Posición` = m.`Nº Pos Actual`
                    WHERE m.`Estado Psn` = 'A' AND e.`Posición` IN ({placeholders})
                    """,
                    posiciones,
                )
                for posicion, nombre in cursor.fetchall():
                    if posicion not in ocupante_por_posicion:
                        ocupante_por_posicion[posicion] = nombre

        # isSIGInfo se incluye para que el front pueda filtrar el catálogo
        # global de búsqueda según la vista activa (Vista SIG solo debe
        # sugerir/saltar a nodos con isSIGInfo=1, ver organigrama/page.jsx).
        results = [
            {
                "departamento": r[0],
                "descripcion_larga": r[1],
                "unidad_negocio": r[2],
                "nivel_direccion": r[3],
                "isSIGInfo": bool(r[4]),
                "ocupante_nombre": ocupante_por_posicion.get(r[5]),
            }
            for r in rows
        ]

        if query:
            q = query.lower()
            results = [
                r for r in results
                if q in r["departamento"].lower()
                or q in (r["descripcion_larga"] or "").lower()
                or q in (r["ocupante_nombre"] or "").lower()
            ][:50]

        return Response(results)


class TorreCaballito3DView(APIView):
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
    )

    def get(self, request):
        from django.db import connection

        query = """
            SELECT 
                e.`Descripción ubicación`,
                e.`Unidad Administrativa`,
                COUNT(*) as Total
            FROM EMPLEADOS_COMPLETOS_SIG e
            INNER JOIN MOV_POS_LATEST activas
                ON e.`Posición` = activas.`Nº Pos Actual` AND activas.`Estado Psn` = 'A'
            WHERE e.`Descripción ubicación` IS NOT NULL
              AND (
                  e.`Descripción ubicación` LIKE '%Caballito Reforma 10 P%'
                  OR e.`Descripción ubicación` LIKE '%Torre Caballito Reforma 10 P%'
              )
            GROUP BY e.`Descripción ubicación`, e.`Unidad Administrativa`
            ORDER BY e.`Descripción ubicación`, Total DESC;
        """

        with connection.cursor() as cursor:
            cursor.execute(query)
            results = cursor.fetchall()

        # Aggregate by floor
        floors_dict = {}
        for row in results:
            piso = row[0]
            ua = row[1] if row[1] else "No Asignada"
            count = row[2]

            if piso not in floors_dict:
                floors_dict[piso] = {"piso": piso, "count": 0, "uas": []}

            floors_dict[piso]["count"] += count
            floors_dict[piso]["uas"].append({"nombre": ua, "count": count})

        return Response(list(floors_dict.values()))


class TorreCaballitoEmpleadosView(APIView):
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
    )

    def get(self, request):
        piso = request.query_params.get("piso", None)
        ua = request.query_params.get("ua", None)

        if not piso:
            return Response({"error": "Falta el parametro piso"}, status=400)

        from django.db import connection

        if ua and ua.strip():
            query = """
                SELECT 
                    e.`Posición`,
                    e.`Numempleado`,
                    e.`Nombres`,
                    e.`Unidad Administrativa`,
                    e.`Descripción ubicación`,
                    e.`Estado Nómina`
                FROM EMPLEADOS_COMPLETOS_SIG e
                INNER JOIN MOV_POS_LATEST activas
                    ON e.`Posición` = activas.`Nº Pos Actual` AND activas.`Estado Psn` = 'A'
                WHERE e.`Descripción ubicación` = %s 
                  AND e.`Unidad Administrativa` = %s
                ORDER BY e.`Nombres`;
            """
            params = [piso, ua]
        else:
            query = """
                SELECT 
                    e.`Posición`,
                    e.`Numempleado`,
                    e.`Nombres`,
                    e.`Unidad Administrativa`,
                    e.`Descripción ubicación`,
                    e.`Estado Nómina`
                FROM EMPLEADOS_COMPLETOS_SIG e
                INNER JOIN MOV_POS_LATEST activas
                    ON e.`Posición` = activas.`Nº Pos Actual` AND activas.`Estado Psn` = 'A'
                WHERE e.`Descripción ubicación` = %s 
                ORDER BY e.`Nombres`;
            """
            params = [piso]

        with connection.cursor() as cursor:
            cursor.execute(query, params)
            results = cursor.fetchall()

        data = []
        for row in results:
            raw_estatus = row[5]
            if not raw_estatus or str(raw_estatus).strip() == "":
                estatus = "Vacante"
            else:
                val = str(raw_estatus).strip().upper()
                if val == "A":
                    estatus = "Activo"
                elif val == "S":
                    estatus = "Suspendido"
                elif val == "L":
                    estatus = "Licencia"
                elif val == "P":
                    estatus = "Licencia Médica"
                else:
                    estatus = "Vacante"

            data.append(
                {
                    "posicion": row[0],
                    "num_empleado": row[1],
                    "nombre": row[2],
                    "ua": row[3],
                    "ubicacion": row[4],
                    "estado_nomina": estatus,
                }
            )

        return Response(data)


class TorreCaballitoSearchView(APIView):
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
    )

    def get(self, request):
        q = request.query_params.get("q", "").strip()
        if not q or len(q) < 3:
            return Response({"results": []})

        from django.db import connection

        query = """
            SELECT
                e.`Posición`,
                e.`Numempleado`,
                e.`Nombres`,
                e.`Unidad Administrativa`,
                e.`Descripción ubicación`
            FROM EMPLEADOS_COMPLETOS_SIG e
            INNER JOIN MOV_POS_LATEST activas
                ON e.`Posición` = activas.`Nº Pos Actual` AND activas.`Estado Psn` = 'A'
            WHERE e.`Descripción ubicación` IS NOT NULL
              AND (
                  e.`Descripción ubicación` LIKE '%%Caballito Reforma 10 P%%'
                  OR e.`Descripción ubicación` LIKE '%%Torre Caballito Reforma 10 P%%'
              )
              AND (e.`Nombres` LIKE %s OR e.`Numempleado` LIKE %s)
            LIMIT 20;
        """

        like_q = f"%{q}%"
        with connection.cursor() as cursor:
            cursor.execute(query, [like_q, like_q])
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Parse piso for frontend convenience
        import re

        for r in results:
            match = re.search(
                r"10\s*P(?:iso)?\s*(\d+)",
                r["Descripción ubicación"] or "",
                re.IGNORECASE,
            )
            if match:
                r["piso_num"] = match.group(1)
            else:
                r["piso_num"] = None

        return Response({"results": results})


class EmpleadosGeografiaSearchView(APIView):
    """
    Busca empleados activos por nombre o número, devolviendo su ubicación
    geográfica (lat/long) para poder centrar el mapa en el resultado.
    """

    view_permission = "authentication.view_plantilla_geografia"

    def get(self, request):
        q = request.query_params.get("q", "").strip()
        if not q or len(q) < 3:
            return Response({"results": []})

        from django.db import connection

        query = """
            SELECT
                e.`Posición`,
                e.`Numempleado`,
                e.`Nombres`,
                e.`Unidad Administrativa`,
                e.`Descripción ubicación`,
                e.`latitud`,
                e.`longitud`,
                e.`Aduana`,
                e.`tipo`
            FROM EMPLEADOS_COMPLETOS_SIG e
            INNER JOIN MOV_POS_LATEST activas
                ON e.`Posición` = activas.`Nº Pos Actual` AND activas.`Estado Psn` = 'A'
            WHERE e.`latitud` IS NOT NULL AND e.`latitud` != ''
              AND e.`longitud` IS NOT NULL AND e.`longitud` != ''
              AND (e.`Nombres` LIKE %s OR e.`Numempleado` LIKE %s)
            LIMIT 20;
        """

        like_q = f"%{q}%"
        with connection.cursor() as cursor:
            cursor.execute(query, [like_q, like_q])
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return Response({"results": results})


from rest_framework.pagination import PageNumberPagination


class MovimientosPersonalPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 10000


def _finalize_mov_completo_rows(rows):
    """
    Replica dos conversiones que CpTblMovCompleto290526Serializer hacía gratis
    vía DRF y que `.values()` deja crudas:

    - `fecha_ult_actz` es el único DateTimeField del modelo (USE_TZ=True) —
      el resto son DateField (sin hora ni tz, no necesitan esto). DRF lo
      convierte a hora local antes de formatear; sin pasar por un campo de
      DRF, el JSONEncoder lo deja en UTC ("Z") — correría las fechas ~6h
      contra lo que ya mostraba el front.
    - `sal_base` (DecimalField) — DRF por default serializa Decimal como
      STRING con 2 decimales fijos (`COERCE_DECIMAL_TO_STRING`, default
      True) y `None` como `''` (cadena vacía, no `null`); el JSONEncoder de
      `.values()` lo deja como número JSON crudo. Distinto tipo para el
      front si no se replica.
    """
    from decimal import Decimal

    from django.utils import timezone

    for r in rows:
        if r.get("fecha_ult_actz") is not None:
            r["fecha_ult_actz"] = timezone.localtime(r["fecha_ult_actz"])
        sal_base = r.get("sal_base")
        r["sal_base"] = (
            format(sal_base.quantize(Decimal("0.01")), "f") if sal_base is not None else ""
        )
    return rows


class MovimientosPersonalListView(APIView):
    view_permission = "authentication.view_plantilla_movimientos"
    pagination_class = MovimientosPersonalPagination

    def get(self, request):
        from .models import CpTblMovCompleto290526
        from .serializers import CP_TBL_MOV_COMPLETO_FIELDS

        queryset = CpTblMovCompleto290526.objects.all()

        # Check if requesting distinct values for a field
        distinct_field = request.query_params.get("distinct_field", "").strip()

        # Search query. Además de los campos sueltos, se anota `full_name`
        # (nombre + ap_pat + ap_mat unidos por espacio, igual que
        # `buildFullName` en MovimientosPersonalTab.jsx) para que buscar el
        # nombre completo ("Juan Pérez López") también matchee aunque esté
        # repartido en los 3 campos reales.
        queryset = _annotate_full_name(queryset, "full_name")
        queryset = apply_text_search(
            queryset,
            request.query_params.get("search", ""),
            [
                "posicion",
                "num_empleado",
                "nombre",
                "ap_pat",
                "ap_mat",
                "full_name",
                "accion_nombre",
                "motivo_nombre",
                "un_admin",
            ],
        )

        # Dynamic Column Filters
        valid_fields = [f.name for f in CpTblMovCompleto290526._meta.get_fields()]
        text_fields = [
            f.name
            for f in CpTblMovCompleto290526._meta.get_fields()
            if f.get_internal_type() in ["CharField", "TextField"]
        ]
        from django.db.models.functions import Trim

        # full_name_column="nombre": el front unificó nombre + ap_pat + ap_mat
        # en una sola columna "Nombre Completo" (columna key `nombre`, ver
        # `buildFullName` en MovimientosPersonalTab.jsx); la BD no tiene ese
        # campo combinado, así que filtros/sugerencias sobre `nombre` deben
        # buscar en los tres campos reales a la vez.
        queryset = apply_dynamic_column_filters(
            queryset, request, CpTblMovCompleto290526, full_name_column="nombre"
        )

        # Advanced filters (built from the "Filtros Avanzados" modal). Se aplica
        # ANTES de la rama distinct_field: si no, "Vista actual" en el dropdown
        # de columna ignoraría las condiciones del modal (bug reportado: el
        # dropdown de Motivo mostraba valores fuera del rango filtrado).
        queryset = apply_advanced_filters(
            queryset, request, CpTblMovCompleto290526, full_name_column="nombre"
        )

        # If distinct_field requested, return distinct values directly
        if distinct_field in valid_fields:
            is_text = distinct_field in text_fields
            # "nombre" es la columna unificada "Nombre Completo" en el front
            # (nombre + ap_pat + ap_mat, sin campo combinado en BD): las
            # sugerencias de autocompletado deben listar el nombre completo,
            # no sólo el campo `nombre` (primer nombre).
            if distinct_field == "nombre":
                target_distinct_field = "full_name"
                queryset = _annotate_full_name(queryset, target_distinct_field)
            else:
                target_distinct_field = (
                    f"trimmed_{distinct_field}" if is_text else distinct_field
                )
                if is_text and target_distinct_field not in queryset.query.annotations:
                    queryset = queryset.annotate(
                        **{target_distinct_field: Trim(distinct_field)}
                    )

            # Apply search filter on the distinct field if present
            distinct_search = request.query_params.get("distinct_search", "").strip()
            if distinct_search:
                if is_text:
                    queryset = queryset.filter(
                        **{f"{target_distinct_field}__icontains": distinct_search}
                    )
                else:
                    # DateTimeField guarda hora exacta: un match exacto contra
                    # solo "YYYY-MM-DD" (lo unico que manda el front) nunca
                    # coincide salvo que la hora sea 00:00:00. `__date` no
                    # sirve: convierte a TIME_ZONE local (America/Mexico_City,
                    # UTC-6) antes de truncar, y estos timestamps casi siempre
                    # se guardaron como medianoche UTC -> se corren un dia
                    # hacia atras (mismo bug de -1 dia documentado en todo el
                    # front para columnas de fecha). Se trunca en UTC, sin
                    # conversion de zona horaria, para comparar la fecha tal
                    # cual esta guardada.
                    field_internal_type = CpTblMovCompleto290526._meta.get_field(
                        distinct_field
                    ).get_internal_type()
                    if field_internal_type == "DateTimeField":
                        utc_date_field = f"{target_distinct_field}__utc_date"
                        if utc_date_field not in queryset.query.annotations:
                            queryset = queryset.annotate(
                                **{utc_date_field: TruncDate(
                                    distinct_field, tzinfo=datetime.timezone.utc
                                )}
                            )
                        search_lookup = utc_date_field
                    else:
                        search_lookup = target_distinct_field
                    try:
                        queryset = queryset.filter(
                            **{search_lookup: distinct_search}
                        )
                    except (ValueError, exceptions.ValidationError):
                        queryset = queryset.none()

            distinct_qs = (
                queryset.values(target_distinct_field)
                .annotate(count=Count("*"))
                .order_by(target_distinct_field)
            )

            results = []
            for item in distinct_qs:
                val = item[target_distinct_field]
                results.append(
                    {"value": val if val is not None else "", "count": item["count"]}
                )
            return Response(results)

        # Sorting
        sort_by_param = request.query_params.get("sort_by", "").strip()
        sort_order = request.query_params.get("sort_order", "asc").strip().lower()
        if sort_by_param:
            sort_fields = [f.strip() for f in sort_by_param.split(",")]
            order_by_args = []
            for field in sort_fields:
                if field in valid_fields:
                    is_text = field in text_fields
                    target_sort_field = f"trimmed_{field}" if is_text else field
                    if is_text and target_sort_field not in queryset.query.annotations:
                        queryset = queryset.annotate(**{target_sort_field: Trim(field)})
                    if sort_order == "desc":
                        order_by_args.append(f"-{target_sort_field}")
                    else:
                        order_by_args.append(target_sort_field)
            if order_by_args:
                queryset = queryset.order_by(*order_by_args)
            else:
                queryset = queryset.order_by("-fecha_efectiva", "-sec")
        else:
            # Default ordering
            queryset = queryset.order_by("-fecha_efectiva", "-sec")

        # A partir de aquí la respuesta ya no necesita instancias de modelo —
        # .values() evita fetch de objetos completos + to_representation por
        # fila de CpTblMovCompleto290526Serializer (ModelSerializer plano,
        # sin SerializerMethodField, pero igual paga overhead de DRF por cada
        # una de hasta 155k filas). ~2.7x más rápido medido con datos reales.
        queryset = queryset.values(*CP_TBL_MOV_COMPLETO_FIELDS)

        # Excel download or full list without pagination
        no_pagination = (
            request.query_params.get("no_pagination", "false").strip().lower() == "true"
        )
        if no_pagination:
            return Response(_finalize_mov_completo_rows(list(queryset)))

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(queryset, request, view=self)
        if page is not None:
            return paginator.get_paginated_response(_finalize_mov_completo_rows(list(page)))

        return Response(_finalize_mov_completo_rows(list(queryset)))


class OrganigramaDeptoView(APIView):
    """
    Catálogo departamento→descripcion_larga/nivel_direccion de ORGANIGRAMA_ANAM.
    Respuesta: [{"departamento": "00100000000", "descripcion_larga": "...", "nivel_direccion": "..."}, ...]
    """
    view_permission = "authentication.view_plantilla_catalogos"

    def get(self, request):
        from django.db import connection
        sql = """
            SELECT departamento, descripcion_larga, nivel_direccion
            FROM ORGANIGRAMA_ANAM
            WHERE isSIGInfo = 1
            ORDER BY departamento
        """
        with connection.cursor() as cursor:
            cursor.execute(sql)
            rows = cursor.fetchall()
        return Response([
            {"departamento": r[0], "descripcion_larga": r[1], "nivel_direccion": r[2]}
            for r in rows
        ])


class OrganigramaTreeView(APIView):
    """
    Árbol jerárquico de ORGANIGRAMA_ANAM para una unidad_negocio, con la misma
    forma que los antiguos JSON estáticos (nodo raíz con `subordinados` anidados).
    El parentesco por default NO se recalcula en cada request: se lee directo
    de la columna `subordinados` (poblada por el comando de management
    poblar_subordinados_organigrama, que sí corre la lógica de segmentación
    del determinante) — "Vista Institucional" (manual/curada, editable).

    Query param opcional `vista`: `"institucional"` (default) lee de
    ORGANIGRAMA_ANAM con isSIGInfo=0 (manual/curada, editable). `"alineacion"`
    y `"sig"` son de solo lectura y fuerzan el recálculo en vivo desde el
    determinante real (`forzar_recalculo=True`, ignoran `subordinados`) — la
    diferencia entre ambas es el SUBCONJUNTO de filas: Alineación recalcula
    sobre isSIGInfo=0 (todo lo editado a mano incluido); SIG recalcula sobre
    isSIGInfo=1, un snapshot congelado por el management command
    `congelar_organigrama_sig` (antes en la tabla aparte ORGANIGRAMA_ANAM_SIG,
    ahora fusionado en ORGANIGRAMA_ANAM) — así SIG es inmune a CUALQUIER
    edición manual (estructura, orden y contenido), no solo a nodos creados a
    mano. Ver organigrama_tree.build_tree y plantilla.models.OrganigramaAnam.
    """
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
        "authentication.view_organigrama_sig",
    )

    VISTA_PERMISSIONS = {
        "alineacion": "authentication.view_organigrama_alineacion",
        "sig": "authentication.view_organigrama_sig",
        "institucional": "authentication.view_organigrama_institucional",
    }

    def get(self, request):
        from .organigrama_tree import build_tree

        unidad_negocio = request.GET.get("unidad_negocio", "").strip()
        if not unidad_negocio:
            return Response({"error": "Falta el parámetro unidad_negocio"}, status=400)

        vista = request.GET.get("vista", "institucional").strip().lower()
        # El class-level view_permission solo exige AL MENOS UNO de los tres
        # (para no bloquear el entry point); aquí se exige el específico de
        # la vista pedida, ya que institucional/alineación/sig son permisos
        # independientes (ver ModulePermission).
        vista_permission = self.VISTA_PERMISSIONS.get(vista, self.VISTA_PERMISSIONS["institucional"])
        if not request.user.has_perm(vista_permission):
            return Response({"error": "No tienes permiso para ver esta vista del organigrama."}, status=403)

        if vista == "sig":
            # Snapshot SIG de solo lectura, ahora fusionado dentro de
            # ORGANIGRAMA_ANAM (isSIGInfo=1) — antes era la tabla aparte
            # ORGANIGRAMA_ANAM_SIG.
            sql = """
                SELECT departamento, descripcion_larga, nivel_direccion, unidad_negocio,
                       unidad_administrativa, doaf, num_posicion_gerente, posicion_director
                FROM ORGANIGRAMA_ANAM
                WHERE unidad_negocio = %s AND isSIGInfo = 1
            """
        else:
            # institucional/alineación: siempre el conjunto editable
            # (isSIGInfo=0), nunca las filas del snapshot SIG.
            sql = """
                SELECT departamento, descripcion_larga, nivel_direccion, unidad_negocio,
                       unidad_administrativa, doaf, num_posicion_gerente, posicion_director,
                       subordinados
                FROM ORGANIGRAMA_ANAM
                WHERE unidad_negocio = %s AND isSIGInfo = 0
            """
        with connection.cursor() as cursor:
            cursor.execute(sql, [unidad_negocio])
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

        if not rows:
            return Response({"error": f"Sin datos para unidad_negocio={unidad_negocio}"}, status=404)

        occupant_map = self._build_occupant_map(rows)
        tree = build_tree(rows, occupant_map, forzar_recalculo=(vista in ("alineacion", "sig")))
        return Response(tree)

    def _build_occupant_map(self, rows):
        """
        Batchea el ocupante (nombre/nivel/SMB) de la plaza titular de cada
        depto (num_posicion_gerente) en 2 queries en vez de una por nodo.
        Misma lógica que OrganigramaPosicionInfoView pero para N plazas.
        """
        posiciones = sorted({
            r["num_posicion_gerente"] for r in rows
            if r.get("num_posicion_gerente") and r["num_posicion_gerente"] != "(en blanco)"
        })
        if not posiciones:
            return {}

        placeholders = ", ".join(["%s"] * len(posiciones))
        with connection.cursor() as cursor:
            cursor.execute(
                f"SELECT `Nº Pos Actual` FROM MOV_POS_LATEST "
                f"WHERE `Nº Pos Actual` IN ({placeholders}) AND `Estado Psn` = 'A'",
                posiciones,
            )
            activas = {r[0] for r in cursor.fetchall()}

        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT `Posición`, `Nombres`, `Nivel`, `SMB`, `Estado Nómina`, `Numempleado`
                FROM EMPLEADOS_COMPLETOS_SIG
                WHERE `Posición` IN ({placeholders})
                """,
                posiciones,
            )
            ocupantes = {}
            for posicion, nombre, nivel, smb, estado_nomina, numempleado in cursor.fetchall():
                if posicion not in ocupantes:
                    ocupantes[posicion] = (nombre, nivel, smb, estado_nomina, numempleado)

        occupant_map = {}
        for posicion in posiciones:
            if posicion not in activas:
                occupant_map[posicion] = {"activa": False, "vacante": None}
                continue
            row = ocupantes.get(posicion)
            if not row or not str(row[3] or "").strip():
                occupant_map[posicion] = {"activa": True, "vacante": True}
                continue
            nombre, nivel, smb, estado_nomina, numempleado = row
            occupant_map[posicion] = {
                "activa": True,
                "vacante": False,
                "nombre": nombre,
                "nivel": nivel,
                "smb": smb,
                "numempleado": numempleado,
            }
        return occupant_map


class OrganigramaPosicionInfoView(APIView):
    """
    Dado un número de plaza (posición), indica si está activa al día de hoy
    (según MOV_POS_LATEST, que ya trae la última fila por posición) y, de
    estarlo, el ocupante actual en EMPLEADOS_COMPLETOS_SIG.
    """
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
        "authentication.view_organigrama_sig",
    )

    def get(self, request):
        from django.db import connection

        posicion = request.GET.get("posicion", "").strip()
        if not posicion or posicion in ("(en blanco)",):
            return Response({"error": "Falta el parámetro posicion"}, status=400)

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM MOV_POS_LATEST WHERE `Nº Pos Actual` = %s AND `Estado Psn` = 'A'",
                [posicion],
            )
            activa = cursor.fetchone() is not None

        if not activa:
            return Response({"posicion": posicion, "activa": False})

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT `Nombres`, `Numempleado`, `Estado Nómina`
                FROM EMPLEADOS_COMPLETOS_SIG
                WHERE `Posición` = %s
                LIMIT 1
                """,
                [posicion],
            )
            row = cursor.fetchone()

        if not row or not str(row[2] or "").strip():
            return Response({"posicion": posicion, "activa": True, "vacante": True})

        nombre, num_empleado, estado_nomina = row
        return Response({
            "posicion": posicion,
            "activa": True,
            "vacante": False,
            "nombre": nombre,
            "num_empleado": num_empleado,
            "estado_nomina": estado_nomina,
        })


class OrganigramaUnidadesView(APIView):
    """
    Catálogo dinámico de unidades de negocio para el selector del front
    (reemplaza el arreglo estático UNIDADES que antes vivía en
    organigrama/page.jsx). Una "unidad de negocio" = un lienzo del
    organigrama; su raíz se resuelve con la misma heurística que
    organigrama_tree.build_tree (find_root), por si existen varias filas
    General/Titular sueltas bajo el mismo unidad_negocio (dato legado).
    """
    view_permission = (
        "authentication.view_organigrama_institucional",
        "authentication.view_organigrama_alineacion",
        "authentication.view_organigrama_sig",
    )

    def get(self, request):
        from .organigrama_tree import find_root

        rows = list(
            OrganigramaAnam.objects.filter(isSIGInfo=False).values(
                "departamento", "descripcion_larga", "nivel_direccion", "unidad_negocio"
            )
        )
        by_unidad = {}
        for r in rows:
            by_unidad.setdefault(r["unidad_negocio"], []).append(r)

        unidades = [
            {"id": unidad_negocio, "label": find_root(group)["descripcion_larga"]}
            for unidad_negocio, group in by_unidad.items()
        ]
        unidades.sort(key=lambda u: u["id"])
        return Response(unidades)


class OrganigramaCrearNodoView(APIView):
    """
    Alta de un nuevo nodo en ORGANIGRAMA_ANAM aplicando la regla de negocio
    del determinante (ver organigrama_tree.py / Webwright_runs/generar_organigramas.py):

      Nivel        Segmento
      General      G
      Central      C
      Director     A (Área)
      Subdir.      S
      Jefe Depto   D

    "Enlace" no tiene tramo propio (el código solo tiene 5, y Jefe Depto ya
    ocupa el último) — se codifica aparte, ver `_crear_enlace`.

    Tres modos:
      - tipo="General": alta de una nueva Dirección General (raíz, sin padre,
        abre lienzo nuevo). unidad_negocio y departamento se capturan a mano
        porque son códigos asignados externamente (SAT/SIG), no hay fórmula.
      - tipo="Enlace": ver `_crear_enlace` — nivel puramente manual, solo
        visible en Vista Institucional (excluido de Vista Alineación).
      - cualquier otro tipo: requiere parent_departamento. El código se
        autogenera heredando el esquema de longitud del padre (10 u 11 chars,
        ver organigrama_tree.parse_code), reutilizando sus segmentos hasta el
        nivel del padre, poniendo a cero los niveles intermedios saltados y
        asignando el siguiente número de 2 dígitos libre en el segmento
        objetivo. Actualiza también `subordinados` del padre para que el
        árbol se refleje sin esperar a poblar_subordinados_organigrama.
    """
    # Solo POST (alta de nodo): exige el permiso de edición, no basta con ver.
    edit_permission = "authentication.edit_organigrama"

    LEVEL_SEGPOS = {"General": 0, "Central": 1, "Director": 2, "Subdir.": 3, "Jefe Depto": 4, "Enlace": 5}
    WIDTHS_11 = [3, 2, 2, 2, 2]
    WIDTHS_10 = [2, 2, 2, 2, 2]

    def post(self, request):
        data = request.data
        tipo = (data.get("tipo") or "").strip()
        if tipo not in self.LEVEL_SEGPOS:
            return Response(
                {"detail": f"tipo inválido. Debe ser uno de: {', '.join(self.LEVEL_SEGPOS)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        descripcion_larga = (data.get("descripcion_larga") or "").strip()
        if not descripcion_larga:
            return Response({"detail": "Falta descripcion_larga."}, status=status.HTTP_400_BAD_REQUEST)

        if tipo == "General":
            return self._crear_general(request, data, descripcion_larga)
        return self._crear_hijo(request, data, tipo, descripcion_larga)

    def _crear_general(self, request, data, descripcion_larga):
        unidad_negocio = (data.get("unidad_negocio") or "").strip()
        departamento = (data.get("departamento") or "").strip()
        if not unidad_negocio or not departamento:
            return Response(
                {"detail": "unidad_negocio y departamento son obligatorios para crear una Dirección General."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if OrganigramaAnam.objects.filter(isSIGInfo=False, departamento=departamento).exists():
            return Response({"detail": f"Ya existe el departamento {departamento}."}, status=status.HTTP_409_CONFLICT)
        if OrganigramaAnam.objects.filter(isSIGInfo=False, unidad_negocio=unidad_negocio).exists():
            return Response({"detail": f"Ya existe la unidad_negocio {unidad_negocio}."}, status=status.HTTP_409_CONFLICT)

        nuevo = OrganigramaAnam.objects.create(
            departamento=departamento,
            unidad_negocio=unidad_negocio,
            estado_fecha_efectiva="Activo",
            descripcion_larga=descripcion_larga,
            nivel_direccion="General",
            unidad_administrativa=(data.get("unidad_administrativa") or "").strip(),
            doaf=(data.get("doaf") or "").strip(),
            num_posicion_gerente=(data.get("num_posicion_gerente") or "(en blanco)").strip(),
            posicion_director="(en blanco)",
            modificado_por=request.user.username,
        )
        return Response(OrganigramaAnamSerializer(nuevo).data, status=status.HTTP_201_CREATED)

    def _crear_hijo(self, request, data, tipo, descripcion_larga):
        from django.shortcuts import get_object_or_404

        parent_code = (data.get("parent_departamento") or "").strip()
        if not parent_code:
            return Response({"detail": "Falta parent_departamento."}, status=status.HTTP_400_BAD_REQUEST)
        parent = get_object_or_404(OrganigramaAnam, isSIGInfo=False, departamento=parent_code)

        if tipo == "Enlace":
            return self._crear_enlace(request, data, parent, descripcion_larga)

        parent_nivel = parent.nivel_direccion
        if parent_nivel == "Titular":
            parent_pos = 0
        elif parent_nivel in self.LEVEL_SEGPOS:
            parent_pos = self.LEVEL_SEGPOS[parent_nivel]
        else:
            return Response(
                {"detail": f"El padre ({parent_code}) tiene nivel_direccion '{parent_nivel}', no reconocido para derivar un hijo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        target_pos = self.LEVEL_SEGPOS[tipo]
        if target_pos <= parent_pos:
            return Response(
                {"detail": f"'{tipo}' debe ser un nivel más profundo que el padre ({parent_nivel})."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        length = len(parent_code)
        if length == 11:
            widths = self.WIDTHS_11
        elif length == 10:
            widths = self.WIDTHS_10
        else:
            return Response(
                {"detail": f"El código del padre ({parent_code}) no tiene una longitud reconocida (10 u 11)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        bounds = []
        idx = 0
        for w in widths:
            bounds.append((idx, idx + w))
            idx += w

        segs = [parent_code[a:b] for a, b in bounds]

        # Niveles intermedios saltados (p.ej. Depto directo bajo Director,
        # sin Subdirección): quedan en cero, igual que el resto de códigos
        # legados que hoy resuelve organigrama_tree.candidate_parents.
        for i in range(parent_pos + 1, target_pos):
            segs[i] = "0" * widths[i]

        prefix = "".join(segs[:target_pos])
        w_target = widths[target_pos]
        siblings = OrganigramaAnam.objects.filter(
            isSIGInfo=False,
            unidad_negocio=parent.unidad_negocio,
            departamento__startswith=prefix,
        ).exclude(departamento=parent_code).values_list("departamento", flat=True)

        max_num = 0
        seg_start, seg_end = bounds[target_pos]
        for dep in siblings:
            if len(dep) != length:
                continue
            seg_val = dep[seg_start:seg_end]
            if seg_val.isdigit():
                max_num = max(max_num, int(seg_val))
        next_num = max_num + 1
        if next_num >= 10 ** w_target:
            return Response(
                {"detail": f"Se agotó la numeración disponible ({w_target} dígitos) bajo {parent_code} para el nivel '{tipo}'."},
                status=status.HTTP_409_CONFLICT,
            )
        segs[target_pos] = str(next_num).zfill(w_target)

        for i in range(target_pos + 1, len(widths)):
            segs[i] = "0" * widths[i]

        new_code = "".join(segs)
        if OrganigramaAnam.objects.filter(isSIGInfo=False, departamento=new_code).exists():
            return Response(
                {"detail": f"Colisión al generar el código {new_code}, intenta de nuevo."},
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            nuevo = OrganigramaAnam.objects.create(
                departamento=new_code,
                unidad_negocio=parent.unidad_negocio,
                estado_fecha_efectiva="Activo",
                descripcion_larga=descripcion_larga,
                nivel_direccion=tipo,
                unidad_administrativa=(data.get("unidad_administrativa") or parent.unidad_administrativa or "").strip(),
                doaf=(data.get("doaf") or parent.doaf or "").strip(),
                num_posicion_gerente=(data.get("num_posicion_gerente") or "(en blanco)").strip(),
                posicion_director=parent.num_posicion_gerente or "(en blanco)",
                modificado_por=request.user.username,
            )
            existentes = [c for c in (parent.subordinados or "").split(",") if c]
            existentes.append(new_code)
            parent.subordinados = ",".join(existentes)
            parent.save(update_fields=["subordinados"])

        return Response(OrganigramaAnamSerializer(nuevo).data, status=status.HTTP_201_CREATED)

    def _crear_enlace(self, request, data, parent, descripcion_larga):
        """
        'Enlace' no ocupa un tramo del determinante (solo hay 5 tramos, y
        'Jefe Depto' ya ocupa el último) — se codifica apendizando 2 dígitos
        al código completo del padre, sea cual sea su nivel_direccion (no es
        regla estricta que cuelgue de un Jefe Depto). El código resultante
        (12-13 chars) nunca colisiona con un código real de ZAFIRO (siempre
        10 u 11 chars exactos). Por eso Vista Alineación (organigrama_tree.
        build_tree con forzar_recalculo=True) excluye por completo los nodos
        'Enlace': no tienen tramo que parsear en el determinante oficial.
        """
        parent_code = parent.departamento
        suffix_len = len(parent_code) + 2
        siblings = OrganigramaAnam.objects.filter(
            isSIGInfo=False,
            departamento__startswith=parent_code,
        ).values_list("departamento", flat=True)

        max_num = 0
        for dep in siblings:
            if len(dep) != suffix_len:
                continue
            seg_val = dep[len(parent_code):]
            if seg_val.isdigit():
                max_num = max(max_num, int(seg_val))
        next_num = max_num + 1
        if next_num >= 100:
            return Response(
                {"detail": f"Se agotó la numeración disponible (2 dígitos) bajo {parent_code} para 'Enlace'."},
                status=status.HTTP_409_CONFLICT,
            )
        new_code = f"{parent_code}{str(next_num).zfill(2)}"
        if OrganigramaAnam.objects.filter(isSIGInfo=False, departamento=new_code).exists():
            return Response(
                {"detail": f"Colisión al generar el código {new_code}, intenta de nuevo."},
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            nuevo = OrganigramaAnam.objects.create(
                departamento=new_code,
                unidad_negocio=parent.unidad_negocio,
                estado_fecha_efectiva="Activo",
                descripcion_larga=descripcion_larga,
                nivel_direccion="Enlace",
                unidad_administrativa=(data.get("unidad_administrativa") or parent.unidad_administrativa or "").strip(),
                doaf=(data.get("doaf") or parent.doaf or "").strip(),
                num_posicion_gerente=(data.get("num_posicion_gerente") or "(en blanco)").strip(),
                posicion_director=parent.num_posicion_gerente or "(en blanco)",
                modificado_por=request.user.username,
            )
            existentes = [c for c in (parent.subordinados or "").split(",") if c]
            existentes.append(new_code)
            parent.subordinados = ",".join(existentes)
            parent.save(update_fields=["subordinados"])

        return Response(OrganigramaAnamSerializer(nuevo).data, status=status.HTTP_201_CREATED)


class EmpleadosBusquedaView(APIView):
    """
    Búsqueda de empleados en EMPLEADOS_COMPLETOS_SIG por nombre, posición o
    número de empleado, usada para reasignar plazas (titular/superior) en el
    organigrama. A diferencia de TorreCaballitoSearchView, no filtra por
    ubicación física.
    """
    view_permission = "authentication.view_organigrama_institucional"

    def get(self, request):
        query = request.GET.get("q", "").strip()
        if len(query) < 3:
            return Response({"results": []})

        empleados = (
            EmpleadosCompletosSig.objects.filter(
                Q(nombres__icontains=query)
                | Q(posicion__icontains=query)
                | Q(numempleado__icontains=query),
                estado_nomina="A",
            )
            .values("posicion", "numempleado", "nombres", "nivel", "smb", "unidad_administrativa", "departamento")[:20]
        )

        results = [
            {
                "posicion": e["posicion"],
                "num_empleado": e["numempleado"],
                "nombre": e["nombres"],
                "nivel": e["nivel"],
                "smb": e["smb"],
                "unidad_administrativa": e["unidad_administrativa"],
                "departamento": e["departamento"],
            }
            for e in empleados
        ]
        return Response({"results": results})


class AuditedViewSetMixin:
    """
    Llena `modificado_por` con el usuario autenticado en cada create/update.
    `fecha_modificacion` se autollena vía `auto_now=True` en el modelo.
    """

    def perform_create(self, serializer):
        serializer.save(modificado_por=self.request.user.username)

    def perform_update(self, serializer):
        serializer.save(modificado_por=self.request.user.username)


class CatAccionesViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD del catálogo cat_acciones (action→action_description/descripcion).
    El listado (GET) mantiene la misma forma que el APIView previo; los
    consumidores existentes (`useAccionesMotivosCatalog.js`) leen los campos
    por nombre y no se ven afectados por los campos extra (effective_status,
    modificado_por, fecha_modificacion).
    """
    queryset = CatAcciones.objects.all()
    serializer_class = CatAccionesSerializer
    view_permission = "authentication.view_plantilla_catalogos"


class CatAccionesMotivosViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD del catálogo cat_acciones_motivos (accion+cd_motivo→descripcion).
    """
    queryset = CatAccionesMotivos.objects.all()
    serializer_class = CatAccionesMotivosSerializer
    view_permission = "authentication.view_plantilla_catalogos"


class CatPtoFuncViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD del catálogo CAT_PTO_FUNC (Cd Pto Funcional → Nombre Puesto
    Funcional), usado por el SP `sp_llenar_nombre_puesto` tras cada
    importación de ZAFIRO.
    """
    queryset = CatPtoFunc.objects.all()
    serializer_class = CatPtoFuncSerializer
    view_permission = "authentication.view_plantilla_catalogos"

    # `_get_mapa_puesto_funcional` cachea este catálogo completo 30 min (ver
    # su docstring) — sin invalidar aquí, un código recién agregado/editado
    # tardaría hasta 30 min en verse reflejado en la columna "Puesto" de
    # Mov. Posiciones / Anexo 2, en vez de al momento.
    def perform_create(self, serializer):
        super().perform_create(serializer)
        cache.delete("mapa_puesto_funcional_completo")

    def perform_update(self, serializer):
        super().perform_update(serializer)
        cache.delete("mapa_puesto_funcional_completo")

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        cache.delete("mapa_puesto_funcional_completo")


class RcCatCodPresupuestalViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD del catálogo rc_cat_cod_presupuestal (SMB/SMN/nivel jerárquico por
    código presupuestal + escala), usado por los SPs
    `sp_corregir_smb_smn_empleados` y `sp_llenar_niveles_vacios_pos_activas`.

    Pk compuesta (codigo_presupuestal, escala): las rutas de detalle llevan
    ambos valores como segmentos de URL en vez de un solo `pk`.
    """
    queryset = RcCatCodPresupuestal.objects.all()
    serializer_class = RcCatCodPresupuestalSerializer
    view_permission = "authentication.view_plantilla_catalogos"

    # Mismo motivo que en CatPtoFuncViewSet — ver `_get_mapa_cod_presupuestal`.
    def perform_create(self, serializer):
        super().perform_create(serializer)
        cache.delete("mapa_cod_presupuestal_completo")

    def perform_update(self, serializer):
        super().perform_update(serializer)
        cache.delete("mapa_cod_presupuestal_completo")

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        cache.delete("mapa_cod_presupuestal_completo")

    def get_object(self):
        from django.shortcuts import get_object_or_404
        obj = get_object_or_404(
            self.get_queryset(),
            codigo_presupuestal=self.kwargs["codigo_presupuestal"],
            escala=self.kwargs["escala"],
        )
        self.check_object_permissions(self.request, obj)
        return obj


def _pivotar_correccion_posicion(qs):
    """Pivota filas (posicion, columna, valor) de CatCorreccionPosicion a un
    renglón por posición con una key por columna — para el tab "Catálogos"
    del frontend (edición fila-única, no clave-valor cruda)."""
    pivot = {}
    for r in qs.values("posicion", "columna", "valor", "modificado_por", "fecha_modificacion"):
        entry = pivot.setdefault(
            r["posicion"],
            {"posicion": r["posicion"], **{c: "" for c in COLUMNAS_CORRECCION_VALIDAS},
             "modificado_por": "", "fecha_modificacion": None},
        )
        entry[r["columna"]] = r["valor"] or ""
        if entry["fecha_modificacion"] is None or (
            r["fecha_modificacion"] and r["fecha_modificacion"] > entry["fecha_modificacion"]
        ):
            entry["fecha_modificacion"] = r["fecha_modificacion"]
            entry["modificado_por"] = r["modificado_por"] or ""
    return sorted(pivot.values(), key=lambda x: x["posicion"])


def _invalidar_cache_correccion_posicion():
    cache.delete_many([f"cat_correccion_posicion_mapa_{c}" for c in COLUMNAS_CORRECCION_VALIDAS])
    _invalidar_cache_detalle_plantilla()


class CatCorreccionPosicionListView(APIView):
    """
    CRUD (vista "ancha": un renglón por posición) del catálogo
    cat_correccion_posicion — Código, Tipo de Aduana, DG de Aduana
    compactada. Pivota la tabla clave-valor real (ver
    `_pivotar_correccion_posicion`) para que editarlo desde el tab
    "Catálogos" sea igual de simple que los demás catálogos, sin exponer el
    diseño clave-valor interno. Ej. de uso real: una posición se realinea a
    otra unidad administrativa y hay que corregir su Tipo de
    Aduana/DG de Aduana compactada a mano.

    Solo lectura para el resto del sistema (Plantilla Detalle la consulta
    como fallback vía `_get_mapa_correccion`, nunca la escribe) — este es el
    único punto de edición.
    """

    view_permission = "authentication.view_plantilla_catalogos"

    def get(self, request):
        qs = CatCorreccionPosicion.objects.filter(columna__in=COLUMNAS_CORRECCION_VALIDAS)
        return Response(_pivotar_correccion_posicion(qs))

    def post(self, request):
        posicion = str(request.data.get("posicion", "")).strip()
        if not posicion:
            return Response({"detail": "posicion es requerida."}, status=status.HTTP_400_BAD_REQUEST)
        if CatCorreccionPosicion.objects.filter(posicion=posicion, columna__in=COLUMNAS_CORRECCION_VALIDAS).exists():
            return Response(
                {"detail": f"La posición '{posicion}' ya existe en el catálogo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for columna in COLUMNAS_CORRECCION_VALIDAS:
            valor = str(request.data.get(columna) or "").strip()
            if valor:
                CatCorreccionPosicion.objects.create(
                    posicion=posicion, columna=columna, valor=valor, modificado_por=request.user.username,
                )

        _invalidar_cache_correccion_posicion()
        fila = _pivotar_correccion_posicion(
            CatCorreccionPosicion.objects.filter(posicion=posicion, columna__in=COLUMNAS_CORRECCION_VALIDAS)
        )
        return Response(
            fila[0] if fila else {"posicion": posicion, **{c: "" for c in COLUMNAS_CORRECCION_VALIDAS}},
            status=status.HTTP_201_CREATED,
        )


class CatCorreccionPosicionDetailView(APIView):
    """PUT/DELETE de una posición del catálogo — ver CatCorreccionPosicionListView."""

    view_permission = "authentication.view_plantilla_catalogos"

    def put(self, request, posicion):
        existe = CatCorreccionPosicion.objects.filter(
            posicion=posicion, columna__in=COLUMNAS_CORRECCION_VALIDAS
        ).exists()
        if not existe:
            return Response(
                {"detail": f"La posición '{posicion}' no existe en el catálogo."},
                status=status.HTTP_404_NOT_FOUND,
            )

        for columna in COLUMNAS_CORRECCION_VALIDAS:
            if columna not in request.data:
                continue
            valor = str(request.data.get(columna) or "").strip()
            if valor:
                CatCorreccionPosicion.objects.update_or_create(
                    posicion=posicion, columna=columna,
                    defaults={"valor": valor, "modificado_por": request.user.username},
                )
            else:
                CatCorreccionPosicion.objects.filter(posicion=posicion, columna=columna).delete()

        _invalidar_cache_correccion_posicion()
        fila = _pivotar_correccion_posicion(
            CatCorreccionPosicion.objects.filter(posicion=posicion, columna__in=COLUMNAS_CORRECCION_VALIDAS)
        )
        return Response(fila[0] if fila else {"posicion": posicion, **{c: "" for c in COLUMNAS_CORRECCION_VALIDAS}})

    def delete(self, request, posicion):
        CatCorreccionPosicion.objects.filter(posicion=posicion, columna__in=COLUMNAS_CORRECCION_VALIDAS).delete()
        _invalidar_cache_correccion_posicion()
        return Response(status=status.HTTP_204_NO_CONTENT)


class OrganigramaAnamViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD del catálogo ORGANIGRAMA_ANAM (unidad_negocio/departamento →
    descripción, nivel de dirección, DOAF, posiciones de gerente/director),
    usado por organigrama_tree.build_tree y las vistas de búsqueda de
    organigrama.

    Solo opera sobre el conjunto institucional (isSIGInfo=False): el
    snapshot SIG (isSIGInfo=True) es de solo lectura, congelado por
    `congelar_organigrama_sig`, y no debe poder editarse/borrarse desde acá.
    `departamento` ya no es la PK de la tabla (ahora es `id`, porque un mismo
    `departamento` puede repetirse entre institucional y SIG), pero sigue
    siendo único DENTRO del conjunto institucional, así que `lookup_field`
    se mantiene en `departamento` para no romper el contrato con el frontend
    (`cat-organigrama-anam/<departamento>/`).
    """
    queryset = OrganigramaAnam.objects.filter(isSIGInfo=False)
    lookup_field = "departamento"
    # URL conf usa <str:pk>/ (no <str:departamento>/); sin esto DRF busca
    # self.kwargs["departamento"] y truena con AssertionError en get_object.
    lookup_url_kwarg = "pk"
    serializer_class = OrganigramaAnamSerializer
    # Se usa desde 2 superficies: tab Catálogos de Plantilla (tabla cruda) y
    # el botón "Editar departamento"/alta-baja de nodo en el módulo
    # Organigrama ANAM (organigrama/page.jsx). Lectura: cualquiera de los 2
    # permisos basta. Escritura: igual, para no romper a quien ya editaba
    # desde Catálogos — edit_organigrama solo suma una vía adicional para
    # roles enfocados en el organigrama que no tengan Catálogos.
    view_permission = (
        "authentication.view_plantilla_catalogos",
        "authentication.view_organigrama_institucional",
    )
    edit_permission = (
        "authentication.view_plantilla_catalogos",
        "authentication.edit_organigrama",
    )

    # Campos bloqueados en update: `departamento` es la PK y codifica el
    # determinante (ver organigrama_tree.parse_code); `nivel_direccion` y
    # `unidad_negocio` están atados a esa codificación y a la resolución
    # padre-hijo (candidate_parents/resolve_by_position). Cambiarlos sin
    # regenerar el código (y el de todos los descendientes) deja el árbol
    # inconsistente. Para "mover" o "re-nivelar" un nodo: crear uno nuevo
    # bajo el padre correcto y eliminar el viejo.
    LOCKED_UPDATE_FIELDS = ("departamento", "nivel_direccion", "unidad_negocio")

    def perform_update(self, serializer):
        for field in self.LOCKED_UPDATE_FIELDS:
            serializer.validated_data.pop(field, None)

        # `subordinados` sí es editable (lo usa el drag-and-drop de
        # reordenamiento de hermanos), pero solo para REORDENAR — debe seguir
        # siendo el mismo conjunto de códigos. Agregar/quitar hijos por esta
        # vía dejaría el árbol inconsistente sin pasar por las validaciones
        # de crear/eliminar nodo (OrganigramaCrearNodoView / destroy).
        if "subordinados" in serializer.validated_data:
            nuevo = {c for c in (serializer.validated_data["subordinados"] or "").split(",") if c}
            actual = {c for c in (serializer.instance.subordinados or "").split(",") if c}
            if nuevo != actual:
                raise ValidationError({
                    "subordinados": "Solo se puede reordenar a los subordinados existentes — usa crear/eliminar nodo para agregar o quitar.",
                })

        serializer.save(modificado_por=self.request.user.username)

    def destroy(self, request, *args, **kwargs):
        """
        No permite borrar un nodo con subordinados (quedarían huérfanos: ya
        no aparecerían en ningún `subordinados` CSV y se volverían
        invisibles en el árbol, ver build_children_map_from_column). Al
        borrar, limpia también la referencia en el `subordinados` del padre.
        """
        instance = self.get_object()
        departamento = instance.departamento
        if (instance.subordinados or "").strip():
            n_hijos = len([c for c in instance.subordinados.split(",") if c])
            return Response(
                {"detail": f"No se puede eliminar '{departamento}': tiene {n_hijos} subordinado(s). Elimínalos o reasígnalos primero."},
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            for padre in OrganigramaAnam.objects.filter(isSIGInfo=False, subordinados__contains=departamento):
                codigos = [c for c in (padre.subordinados or "").split(",") if c and c != departamento]
                nuevo_valor = ",".join(codigos)
                if nuevo_valor != (padre.subordinados or ""):
                    padre.subordinados = nuevo_valor
                    padre.save(update_fields=["subordinados"])
            instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# Campos del cuadro del Anexo 2 sobre los que se detectan ediciones manuales
# para el historial de cambios (ver _diff_hojas_anuencia) — mismo conjunto
# que ANEXO2_COLUMNAS en el front (anexo2Schema.js), sin las columnas
# puramente informativas (_unidadDeNegocioResuelta, _movPosId, etc.) que no
# captura el usuario a mano.
_CAMPOS_FILA_ANUENCIA_ANEXO2 = [
    ("ramo", "Ramo"),
    ("unidad_responsable", "Unidad Responsable"),
    ("codigo", "Código"),
    ("denominacion_puesto", "Denominación del puesto"),
    ("nivel_salarial", "Nivel Salarial"),
    ("rango_salarial", "Rango Salarial"),
    ("numero_plazas", "Número de plazas"),
    ("numero_horas", "Número de Horas"),
    ("tipo_contratacion", "Tipo de contratación"),
    ("fecha_inicio_vacancia", "Fecha de inicio de la vacancia"),
    ("fecha_alta_solicitada", "Fecha de alta solicitada"),
    ("oficio_autorizacion", "Oficio de autorización presupuestaria"),
]

# Tope defensivo: un guardado con cientos de ediciones a la vez (p. ej. el
# backfill silencioso de Unidad de Negocio/Rango que corre AnuenciaTab.jsx al
# abrir un anexo viejo, ver AnuenciaTab.jsx/backfillUnidadNegocioYRango) no
# debe producir una entrada de historial de cientos de líneas ilegibles.
_MAX_LINEAS_CAMBIO_ANUENCIA = 40


def _resumen_fila_anuencia(fila):
    codigo = str(fila.get("codigo") or "").strip()
    denominacion = str(fila.get("denominacion_puesto") or "").strip()
    if codigo and denominacion:
        return f"{codigo} — {denominacion}"
    return codigo or denominacion or "(plaza sin código)"


def _diff_hojas_anuencia(hojas_antes, hojas_despues):
    """
    Compara el `hojas` de un AnuenciaAnexo antes/después de un guardado y
    arma una lista de líneas legibles para el historial de cambios (ver
    AnuenciaAnexoCambio) — hojas agregadas/eliminadas/renombradas, plazas
    agregadas/eliminadas, y campos editados a mano en una plaza que ya
    existía. Nunca lanza: ante cualquier forma inesperada de los datos, esa
    comparación puntual simplemente se salta (mejor un historial incompleto
    que un guardado que falla por esto) — de ahí el try/except al llamarla.
    """
    lineas = []

    def por_id(lista):
        return {h.get("_id"): h for h in (lista or []) if h.get("_id")}

    mapa_antes = por_id(hojas_antes)
    mapa_despues = por_id(hojas_despues)

    for hoja_id, hoja in mapa_despues.items():
        if hoja_id in mapa_antes:
            continue
        n = len(hoja.get("filas") or [])
        nombre = hoja.get("nombre") or "(sin nombre)"
        extra = f" con {n} {'plaza' if n == 1 else 'plazas'}" if n else ""
        lineas.append(f'Agregó la hoja "{nombre}"{extra}')

    for hoja_id, hoja in mapa_antes.items():
        if hoja_id in mapa_despues:
            continue
        n = len(hoja.get("filas") or [])
        nombre = hoja.get("nombre") or "(sin nombre)"
        extra = f" (tenía {n} {'plaza' if n == 1 else 'plazas'})" if n else ""
        lineas.append(f'Eliminó la hoja "{nombre}"{extra}')

    for hoja_id, hoja_d in mapa_despues.items():
        hoja_a = mapa_antes.get(hoja_id)
        if not hoja_a:
            continue
        nombre_d = hoja_d.get("nombre") or "(sin nombre)"
        nombre_a = hoja_a.get("nombre") or "(sin nombre)"
        if nombre_a != nombre_d:
            lineas.append(f'Renombró la hoja "{nombre_a}" a "{nombre_d}"')
        if (hoja_a.get("unidad_administrativa") or "") != (hoja_d.get("unidad_administrativa") or ""):
            lineas.append(f'Cambió la Unidad Administrativa de "{nombre_d}"')
        if (hoja_a.get("justificacion") or "") != (hoja_d.get("justificacion") or ""):
            lineas.append(f'Editó la justificación de "{nombre_d}"')

        filas_a = por_id(hoja_a.get("filas"))
        filas_d = por_id(hoja_d.get("filas"))

        agregadas = [f for fid, f in filas_d.items() if fid not in filas_a]
        eliminadas = [f for fid, f in filas_a.items() if fid not in filas_d]

        if len(agregadas) == 1:
            lineas.append(f'Agregó la plaza {_resumen_fila_anuencia(agregadas[0])} a "{nombre_d}"')
        elif agregadas:
            lineas.append(f'Agregó {len(agregadas)} plazas a "{nombre_d}"')

        if len(eliminadas) == 1:
            lineas.append(f'Eliminó la plaza {_resumen_fila_anuencia(eliminadas[0])} de "{nombre_d}"')
        elif eliminadas:
            lineas.append(f'Eliminó {len(eliminadas)} plazas de "{nombre_d}"')

        for fid, fila_d in filas_d.items():
            fila_a = filas_a.get(fid)
            if not fila_a:
                continue
            for campo, etiqueta in _CAMPOS_FILA_ANUENCIA_ANEXO2:
                antes = fila_a.get(campo)
                despues = fila_d.get(campo)
                if (antes or "") == (despues or ""):
                    continue
                if len(lineas) >= _MAX_LINEAS_CAMBIO_ANUENCIA:
                    continue
                lineas.append(
                    f'Editó {etiqueta} de {_resumen_fila_anuencia(fila_d)} en "{nombre_d}": '
                    f'"{antes or "—"}" → "{despues or "—"}"'
                )

    if len(lineas) > _MAX_LINEAS_CAMBIO_ANUENCIA:
        lineas = lineas[:_MAX_LINEAS_CAMBIO_ANUENCIA] + ["… y más cambios (se omiten por espacio)"]

    return lineas


def _registrar_cambio_anuencia_anexo(anexo, usuario, cambios):
    """Guarda una entrada de historial si `cambios` trae algo — nunca lanza
    (ver docstring de _diff_hojas_anuencia): un fallo aquí no debe tumbar un
    guardado que por lo demás salió bien."""
    if not cambios:
        return
    try:
        AnuenciaAnexoCambio.objects.create(anexo=anexo, usuario=usuario, cambios=cambios)
    except Exception:
        logger.exception("Error al registrar historial de cambios del Anexo 2 %s", anexo.id)


def _notificar_actualizacion_anuencia_anexo(anexo, usuario):
    """
    Publica en el canal Redis "anuencia_anexo_updates" para que
    AnuenciaAnexoUpdatesSSEView lo reenvíe a cualquier pestaña con este mismo
    Anexo 2 abierto en AnuenciaTab.jsx — mismo patrón que
    `celda_override.notificar_cambio_celda`/"plantilla_celda_updates".

    El Anexo 2 es colaborativo: dos personas pueden tenerlo abierto a la vez
    (una capturándolo directo en Anuencia, otra agregándole plazas desde el
    menú contextual de Mov. Posiciones — ver AgregarAAnexo2Modal.jsx, que
    hace PATCH sobre este mismo ViewSet). Sin este aviso, la pestaña que lo
    tiene abierto nunca se entera de las plazas nuevas hasta recargar la
    página a mano.

    Va sólo el id (no `hojas`, que puede pesar cientos de KB): el cliente
    decide si conviene traer el detalle completo o no (ver
    useAnuenciaAnexoUpdatesRealtime.js — si hay cambios sin guardar en curso,
    no se pisa el borrador local con un refetch silencioso).
    """
    import redis as redis_lib

    r = redis_lib.Redis.from_url(settings.CELERY_BROKER_URL)
    r.publish(
        "anuencia_anexo_updates",
        json.dumps({
            "type": "anexo_update",
            "anexo_id": anexo.id,
            "usuario": usuario.username,
            "usuario_nombre": usuario.get_full_name() or usuario.username,
        }),
    )


class AnuenciaAnexoViewSet(viewsets.ModelViewSet):
    """
    Historial del ANEXO 2 (sub-tab "Anuencia" de Mov. Posiciones, ver
    AnuenciaTab.jsx). No usa `AuditedViewSetMixin` (ese sólo lleva un
    "modificado_por" genérico): aquí se necesitan 3 eventos de auditoría
    distintos — creó / modificó / generó — ver docstring de `AnuenciaAnexo`.

    Sin `destroy` real (ver `http_method_names`, sin "delete"/"put" — sólo
    `partial_update`, nunca reemplazo completo): "eliminar" un Anexo 2 es un
    soft delete vía la acción `eliminar` (más abajo), nunca un DELETE ni un
    borrado real de la fila — sirve para "cerrar" una solicitud ya resuelta
    (sus plazas dejan de contar como "en anuencia") sin perder el registro
    para auditoría. `get_queryset` excluye siempre los eliminados: una vez
    eliminado, un Anexo 2 deja de ser alcanzable por esta API (list/retrieve/
    patch/generar), aunque su fila siga en la base de datos para siempre.

    `GET` (list/retrieve) usa el mismo permiso que el resto de Anuencia
    (`AnuenciaLookupView`/`AnuenciaSugerenciasView`); `POST`/`PATCH`/`generar`/
    `eliminar` caen al mismo permiso por no declarar `edit_permission` (ver
    `HasModulePermission`) — no hay un permiso "editar anuencia" separado hoy.

    `eliminados`/`reactivar` exigen el permiso aparte `view_anuencia_eliminados`
    (ver `action_permissions`): ver y poder recuperar algo que alguien más
    marcó como eliminado es más sensible que el uso normal de Anuencia, así
    que no basta con el permiso general — un rol puede ver/editar Anexo 2
    corrientes sin poder husmear ni reactivar los eliminados.
    """

    queryset = AnuenciaAnexo.objects.select_related(
        "creado_por", "actualizado_por", "generado_por", "eliminado_por"
    ).filter(eliminado=False)
    view_permission = "authentication.view_plantilla_mov_posiciones"
    http_method_names = ["get", "post", "patch", "head", "options"]
    action_permissions = {
        "eliminados": "authentication.view_anuencia_eliminados",
        "eliminado_detalle": "authentication.view_anuencia_eliminados",
        "reactivar": "authentication.view_anuencia_eliminados",
    }

    def get_queryset(self):
        qs = self.queryset
        if self.action == "list":
            # El listado ordena por -actualizado_en (Meta.ordering) y
            # AnuenciaAnexoListSerializer sólo necesita agregados sobre
            # `hojas` (ver total_hojas/total_filas/unidades_administrativas),
            # no su contenido. Sin `defer`, MySQL tiene que hacer el filesort
            # del ORDER BY cargando esa columna JSON completa por fila —con
            # anexos de 800+ plazas eso pesa cientos de KB por renglón y
            # agota sort_buffer_size (error 1038, "Out of sort memory").
            # Diferirla evita el filesort pesado; Django la trae aparte,
            # fila por fila y por PK (sin sort), sólo para las que se sirvan.
            qs = qs.defer("hojas")
        return qs

    def get_serializer_class(self):
        return AnuenciaAnexoListSerializer if self.action == "list" else AnuenciaAnexoDetailSerializer

    def perform_create(self, serializer):
        # Crear ya NO implica generar: el botón "Guardar" de AnuenciaTab.jsx
        # crea/actualiza el registro sin descargar nada, así que un anexo
        # puede existir con `veces_generado=0` (nunca se ha bajado su .xlsx).
        # El flujo de "Descargar Anexo 2" sigue estampando generado_por/
        # generado_en aparte, con una llamada explícita a `generar/` justo
        # después de crear (ver handleExportar en el front).
        serializer.save(creado_por=self.request.user, actualizado_por=self.request.user)
        # Un Anexo 2 nuevo puede traer plazas que Mov. Posiciones debe marcar
        # de inmediato como "ya en anuencia" (ver `_get_mapa_codigos_en_anuencia`).
        _invalidar_cache_anuencia()
        _notificar_actualizacion_anuencia_anexo(serializer.instance, self.request.user)
        total_hojas = len(serializer.instance.hojas or [])
        total_plazas = sum(len(h.get("filas") or []) for h in (serializer.instance.hojas or []))
        _registrar_cambio_anuencia_anexo(
            serializer.instance,
            self.request.user,
            [
                f"Creó el Anexo 2 con {total_plazas} {'plaza' if total_plazas == 1 else 'plazas'} "
                f"en {total_hojas} {'hoja' if total_hojas == 1 else 'hojas'}."
            ],
        )

    def perform_update(self, serializer):
        # `serializer.instance` todavía trae el `hojas` de ANTES en este
        # punto — `.save()` lo reemplaza (no lo muta), así que capturarlo
        # aquí antes de guardar es suficiente para diffear sin una query
        # extra (ver _diff_hojas_anuencia).
        hojas_antes = serializer.instance.hojas
        serializer.save(actualizado_por=self.request.user)
        _invalidar_cache_anuencia()
        _notificar_actualizacion_anuencia_anexo(serializer.instance, self.request.user)
        try:
            cambios = _diff_hojas_anuencia(hojas_antes, serializer.instance.hojas)
        except Exception:
            logger.exception("Error al comparar hojas de Anexo 2 %s para el historial", serializer.instance.id)
            cambios = []
        _registrar_cambio_anuencia_anexo(serializer.instance, self.request.user, cambios)

    @action(detail=True, methods=["post"])
    def generar(self, request, pk=None):
        """Descarga de un anexo YA guardado (re-generar desde el historial o
        volver a bajar el .xlsx de uno abierto) — estampa quién y cuándo,
        e incrementa el contador, sin tocar el contenido del anexo."""
        from django.utils import timezone

        anexo = self.get_object()
        anexo.generado_por = request.user
        anexo.generado_en = timezone.now()
        anexo.veces_generado = F("veces_generado") + 1
        anexo.save(update_fields=["generado_por", "generado_en", "veces_generado"])
        anexo.refresh_from_db()
        return Response(AnuenciaAnexoDetailSerializer(anexo).data)

    @action(detail=True, methods=["post"])
    def eliminar(self, request, pk=None):
        """Soft delete: nunca se borra la fila (ver docstring de la clase y
        de `AnuenciaAnexo.eliminado`) — sólo se marca, para que deje de
        aparecer en el historial y sus plazas dejen de contar como "en
        anuencia" en Mov. Posiciones, liberándolas para poder volver a
        solicitarlas en un Anexo 2 nuevo."""
        from django.utils import timezone

        anexo = self.get_object()
        anexo.eliminado = True
        anexo.eliminado_por = request.user
        anexo.eliminado_en = timezone.now()
        anexo.save(update_fields=["eliminado", "eliminado_por", "eliminado_en"])
        _invalidar_cache_anuencia()
        _notificar_actualizacion_anuencia_anexo(anexo, request.user)
        _registrar_cambio_anuencia_anexo(anexo, request.user, ["Eliminó el Anexo 2."])
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"])
    def eliminados(self, request):
        """Anexo 2 eliminados (soft delete) — botón "Anexos eliminados" del
        sub-tab Anuencia. Fuera de `get_queryset` a propósito: ese método
        siempre filtra `eliminado=False` para el resto de acciones, así que
        aquí se arma el queryset de cero en vez de reutilizarlo."""
        qs = (
            AnuenciaAnexo.objects.filter(eliminado=True)
            .select_related("creado_por", "actualizado_por", "generado_por", "eliminado_por")
            .defer("hojas")
        )
        return Response(AnuenciaAnexoListSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def eliminado_detalle(self, request, pk=None):
        """Detalle completo (con `hojas`) de un Anexo 2 eliminado — para el
        botón "Ver" del modal de eliminados, que sólo consulta sin reactivar
        (a diferencia de `reactivar`, que sí cambia el estado)."""
        from django.shortcuts import get_object_or_404

        anexo = get_object_or_404(
            AnuenciaAnexo.objects.select_related("creado_por", "actualizado_por", "generado_por", "eliminado_por"),
            pk=pk, eliminado=True,
        )
        return Response(AnuenciaAnexoDetailSerializer(anexo).data)

    @action(detail=True, methods=["post"])
    def reactivar(self, request, pk=None):
        """Deshace el soft delete: el Anexo 2 vuelve a aparecer en el
        historial normal y sus plazas vuelven a contar como "en anuencia" en
        Mov. Posiciones."""
        from django.shortcuts import get_object_or_404

        anexo = get_object_or_404(AnuenciaAnexo, pk=pk, eliminado=True)
        anexo.eliminado = False
        anexo.eliminado_por = None
        anexo.eliminado_en = None
        anexo.save(update_fields=["eliminado", "eliminado_por", "eliminado_en"])
        _invalidar_cache_anuencia()
        _notificar_actualizacion_anuencia_anexo(anexo, request.user)
        _registrar_cambio_anuencia_anexo(anexo, request.user, ["Reactivó el Anexo 2."])
        return Response(AnuenciaAnexoDetailSerializer(anexo).data)

    @action(detail=True, methods=["get"])
    def historial(self, request, pk=None):
        """Historial de cambios del Anexo 2 (ver AnuenciaAnexoCambio) — botón
        "Historial de cambios" de AnuenciaTab.jsx. A diferencia de
        actualizado_por/actualizado_en (sólo el último editor), aquí se ve
        CADA guardado con un resumen legible de qué cambió."""
        anexo = self.get_object()
        cambios = anexo.cambios.select_related("usuario").all()
        return Response(AnuenciaAnexoCambioSerializer(cambios, many=True).data)


class AnuenciaAnexo3VersionViewSet(viewsets.ModelViewSet):
    """
    Historial de versiones guardadas del Anexo 3 (FUMP) de un Anexo 2 — el
    Anexo 3 se puede corregir a mano arrastrando plazas entre hojas del
    resultado (ver Anexo3Editor.jsx), y como esa corrección es editable, un
    mismo Anexo 2 puede tener varias versiones guardadas. Mismo patrón que
    AnuenciaAnexoViewSet (sin `destroy`/`put`, sólo 2 eventos de auditoría).

    `list` es SIEMPRE por Anexo 2 (`?anexo=<id>`) — no tiene sentido listar
    versiones de Anexo 3 de todos los anexos juntos.
    """

    queryset = AnuenciaAnexo3Version.objects.select_related("creado_por", "actualizado_por").all()
    view_permission = "authentication.view_plantilla_mov_posiciones"
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_queryset(self):
        qs = self.queryset
        if self.action == "list":
            anexo_id = self.request.query_params.get("anexo")
            if not anexo_id:
                return qs.none()
            qs = qs.filter(anexo_id=anexo_id)
            # Mismo motivo que en AnuenciaAnexoViewSet: no cargar el JSON
            # grande (`grupos`) en el SELECT que hace el ORDER BY del listado.
            qs = qs.defer("grupos")
        return qs

    def list(self, request, *args, **kwargs):
        if not request.query_params.get("anexo"):
            return Response(
                {"error": "Falta el parámetro 'anexo' para listar versiones de Anexo 3."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().list(request, *args, **kwargs)

    def get_serializer_class(self):
        return AnuenciaAnexo3VersionListSerializer if self.action == "list" else AnuenciaAnexo3VersionDetailSerializer

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user, actualizado_por=self.request.user)

    def perform_update(self, serializer):
        serializer.save(actualizado_por=self.request.user)


class AnuenciaJustificacionCatalogoViewSet(viewsets.ModelViewSet):
    """
    Catálogo de justificaciones reutilizables del Anexo 2 (ver
    AnuenciaJustificacionCatalogo) — alta y baja desde el modal de catálogo
    en AnuenciaTab.jsx. Sin `put`/`patch`: no hay UI de edición, sólo agregar
    una entrada nueva o eliminar una existente.
    """

    queryset = AnuenciaJustificacionCatalogo.objects.select_related("creado_por").all()
    serializer_class = AnuenciaJustificacionCatalogoSerializer
    view_permission = "authentication.view_plantilla_mov_posiciones"
    http_method_names = ["get", "post", "delete", "head", "options"]

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user)


class CatNivelJerarquicoPlazaViewSet(AuditedViewSetMixin, viewsets.ModelViewSet):
    """
    CRUD + acciones en bloque del catálogo cat_nivel_jerarquico_plaza.

    La siembra/actualización de plazas desde MOV_POS (`nvl_direc_origen`) ya
    no es una acción manual: corre automáticamente en cada import ZAFIRO (ver
    `plantilla.tasks._sincronizar_plazas_nivel_jerarquico`), justo antes de
    reaplicar la prioridad. Se quitó el trigger manual porque dispararlo a
    mitad de ciclo (con MOV_POS ya sobreescrito por una prioridad
    `nivel_jerarquico` recién aplicada) contaminaba `nvl_direc_origen` con el
    propio override, en vez de con el dato original de ZAFIRO.

    - `bulk-assign`: asigna una misma descripción de nivel jerárquico (enum)
      a varias plazas seleccionadas desde el frontend, y de inmediato
      propaga ese nivel a EMPLEADOS_COMPLETOS_SIG y MOV_POS (fila vigente)
      para esas plazas, sin importar la `fuente` de prioridad configurada
      (ver `nivel_jerarquico_sync.aplicar_nivel_a_plazas`).
    """
    queryset = CatNivelJerarquicoPlaza.objects.all()
    serializer_class = CatNivelJerarquicoPlazaSerializer
    view_permission = "authentication.view_plantilla_catalogos"
    lookup_value_regex = "[^/]+"

    def get_object(self):
        from django.shortcuts import get_object_or_404
        obj = get_object_or_404(self.get_queryset(), plaza=self.kwargs["pk"])
        self.check_object_permissions(self.request, obj)
        return obj

    @action(detail=False, methods=["get"])
    def niveles(self, request):
        """Catálogo estático (NJ, descripción) para poblar el select del frontend."""
        return Response([
            {"descripcion_nivel_jerarquico": value, "label": label}
            for value, label in DESCRIPCION_NJ_CHOICES
        ])

    @action(detail=False, methods=["post"], url_path="bulk-assign")
    def bulk_assign(self, request):
        plazas = request.data.get("plazas") or []
        descripcion = request.data.get("descripcion_nivel_jerarquico")
        if not plazas or not isinstance(plazas, list):
            return Response({"detail": "Se requiere una lista no vacía de plazas."}, status=status.HTTP_400_BAD_REQUEST)
        if descripcion not in dict(DESCRIPCION_NJ_CHOICES):
            return Response({"detail": "descripcion_nivel_jerarquico inválido."}, status=status.HTTP_400_BAD_REQUEST)

        usuario = request.user.username
        actualizadas = 0
        plaza_a_nivel = {}
        with transaction.atomic():
            for plaza in plazas:
                obj, _ = CatNivelJerarquicoPlaza.objects.get_or_create(plaza=plaza)
                obj.descripcion_nivel_jerarquico = descripcion
                obj.modificado_por = usuario
                obj.save()
                actualizadas += 1
                if obj.nivel_jerarquico is not None:
                    plaza_a_nivel[obj.plaza] = obj.nivel_jerarquico

            # Fija "nivel_jerarquico" como fuente de prioridad (si no lo era
            # ya) para que el próximo import ZAFIRO la reaplique — si no,
            # el truncado/recarga de MOV_POS/EMPLEADOS_COMPLETOS_SIG en cada
            # import (ver plantilla.tasks) borraría este cambio en <30 min.
            config, _ = NivelJerarquicoPrioridadConfig.objects.get_or_create(pk=1)
            if config.fuente != "nivel_jerarquico":
                config.fuente = "nivel_jerarquico"
                config.modificado_por = usuario
                config.save()

        from .nivel_jerarquico_sync import aplicar_nivel_a_plazas

        stats = aplicar_nivel_a_plazas(plaza_a_nivel)
        return Response({"actualizadas": actualizadas, **stats})

    @action(detail=False, methods=["get"], url_path="prioridad")
    def prioridad(self, request):
        """Fuente de prioridad configurada actualmente (o null si no se ha fijado)."""
        config = NivelJerarquicoPrioridadConfig.objects.first()
        return Response({"fuente": config.fuente if config else None})

    @action(detail=False, methods=["post"], url_path="aplicar-prioridad")
    def aplicar_prioridad(self, request):
        """
        Fija qué columna manda como fuente de verdad del nivel jerárquico
        ("nivel_jerarquico" o "nvl_direc_origen") y de inmediato cruza
        cat_nivel_jerarquico_plaza contra MOV_POS y EMPLEADOS_COMPLETOS_SIG,
        sobreescribiendo sus columnas de nivel jerárquico donde la posición
        coincida. La misma fuente se reaplica automáticamente en cada import
        de ZAFIRO (esas dos tablas se truncan y recargan completas cada 30
        min, ver plantilla.tasks._reaplicar_prioridad_nivel_jerarquico).
        """
        fuente = request.data.get("fuente")
        if fuente not in ("nivel_jerarquico", "nvl_direc_origen"):
            return Response({"detail": "fuente inválida."}, status=status.HTTP_400_BAD_REQUEST)

        usuario = request.user.username
        with transaction.atomic():
            config, _ = NivelJerarquicoPrioridadConfig.objects.get_or_create(pk=1)
            config.fuente = fuente
            config.modificado_por = usuario
            config.save()
            stats = aplicar_prioridad_nivel_jerarquico(fuente)

        return Response({"fuente": fuente, **stats})


class MovimientosPersonalHistorialView(APIView):
    """
    Devuelve el historial completo (sin filtro de año) de los empleados indicados,
    ordenado por num_empleado, fecha_efectiva, sec ASC.
    Consulta directa a cp_tbl_mov_completo_29_05_26 via raw SQL.

    Usar POST con body {"num_empleado": [123, 456, 789]} para listas grandes
    (evita el límite de longitud de URL de un GET). Se mantiene ?num_empleado__in=
    por compatibilidad con clientes existentes.

    Además de su uso original (módulo Movimientos), alimenta el tab "Historial
    de movimientos" del expediente de empleado (EmployeesModal.jsx) — se
    amplía el permiso a la misma tupla que DatosPersonalesEmpleadoView para
    que cualquiera que pueda abrir el expediente vea este tab, sin importar
    desde qué módulo de Plantilla lo abrió.
    """
    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
        "authentication.view_plantilla_bajas",
        "authentication.view_plantilla_geografia",
    )

    def get(self, request):
        raw_param = request.query_params.get("num_empleado__in", "").strip()
        emp_ids = [e.strip() for e in raw_param.split(",") if e.strip()]
        return self._historial(emp_ids)

    def post(self, request):
        emp_ids = request.data.get("num_empleado", [])
        if isinstance(emp_ids, str):
            emp_ids = [e.strip() for e in emp_ids.split(",") if e.strip()]
        else:
            emp_ids = [str(e).strip() for e in emp_ids if str(e).strip()]
        return self._historial(emp_ids)

    def _historial(self, emp_ids):
        from django.db import connection

        if not emp_ids:
            return Response([])

        placeholders = ", ".join(["%s"] * len(emp_ids))
        sql = f"""
            SELECT
                id, posicion, num_empleado,
                nombre, ap_pat, ap_mat,
                accion, accion_nombre,
                motivo, motivo_nombre,
                fecha_efectiva, sec, fecha_captura,
                est_hr, estado_pago, partida_presup,
                un, un_admin,
                id_depto, depen_direc,
                plan_sal, grado, escala,
                puesto_ptal, nivel_tabular,
                gp_pago, prog_benef, sal_base,
                cd_puesto, ubicacion, id_estbl,
                salida_prevista, fecha_ult_actz, por,
                ult_inicio, fecha_inicial, gp_trabajo,
                grupo_cd_sal, antiguo_empr,
                rfc, curp, id_persona,
                desc_larga_p, nv_jerarquico, desc_larga_un,
                sexo, fecha_entrada, fecha_posicion
            FROM cp_tbl_mov_completo_29_05_26
            WHERE num_empleado IN ({placeholders})
            ORDER BY num_empleado ASC, fecha_efectiva ASC, sec ASC
        """

        with connection.cursor() as cursor:
            cursor.execute(sql, emp_ids)
            cols = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        results = []
        for row in rows:
            record = {}
            for i, col in enumerate(cols):
                val = row[i]
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                record[col] = val
            results.append(record)

        return Response(results)


class MovimientosPosicionHistorialView(APIView):
    """
    Devuelve el historial completo (sin filtro de año) de las posiciones
    indicadas, ordenado por posición ASC y fecha efectiva DESCENDENTE (el
    movimiento más reciente primero, el más viejo al final) — a diferencia
    de MovimientosPersonalHistorialView (ASC), a pedido explícito del
    usuario para el tab "Historial de posición" del expediente de plaza
    (EmployeesModal.jsx). Consulta directa a MOV_POS via raw SQL, mismo
    patrón que MovimientosPersonalHistorialView.

    Usar POST con body {"posicion": ["10300009", ...]} para listas grandes
    (evita el límite de longitud de URL de un GET); también acepta
    ?posicion__in= por GET, igual que el endpoint de empleados.

    Cada carril de este tab en pantalla es una unidad administrativa
    (`Unidad Adva#` de MOV_POS, no `Unidad de Negocio` ni `Cd UN` — ver
    comprobación 2026-09-02: el código de 3 dígitos de `Unidad Adva#`
    coincide con la parte numérica de `Cd UN`, y `Unidad de Negocio` ya es
    el nombre legible de ese mismo código), así que se exponen ambos como
    `unidad_adva` (código) y `unidad_administrativa` (nombre).
    """
    view_permission = (
        "authentication.view_plantilla_detalle",
        "authentication.view_plantilla_estatus_nomina",
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
        "authentication.view_plantilla_bajas",
        "authentication.view_plantilla_geografia",
    )

    def get(self, request):
        raw_param = request.query_params.get("posicion__in", "").strip()
        posiciones = [p.strip() for p in raw_param.split(",") if p.strip()]
        return self._historial(posiciones)

    def post(self, request):
        posiciones = request.data.get("posicion", [])
        if isinstance(posiciones, str):
            posiciones = [p.strip() for p in posiciones.split(",") if p.strip()]
        else:
            posiciones = [str(p).strip() for p in posiciones if str(p).strip()]
        return self._historial(posiciones)

    def _historial(self, posiciones):
        from django.db import connection

        if not posiciones:
            return Response([])

        placeholders = ", ".join(["%s"] * len(posiciones))
        sql = f"""
            SELECT
                id,
                `Nº Pos Actual` AS posicion,
                `F Efva` AS fecha_efectiva,
                `Estado Psn` AS estado_posicion,
                `Fecha Captura` AS fecha_captura,
                `FECHA VACANCIA` AS fecha_vacancia,
                `Cd Motivo` AS cd_motivo,
                `Motivo` AS motivo,
                `Cd UN` AS cd_un,
                `Unidad de Negocio` AS unidad_administrativa,
                `Unidad Adva#` AS unidad_adva,
                `Cd Departamento` AS cd_departamento,
                `Cd Puesto` AS cd_puesto,
                `Estado Ptal` AS estado_ptal,
                `Fecha Est` AS fecha_establecimiento,
                `Máximo` AS maximo,
                `Depnd Drt` AS dependencia_directa,
                `Depnd Indrt` AS dependencia_indirecta,
                `Ubicación` AS ubicacion,
                `Nvl Direc` AS nivel_direccion,
                `Plan Sal` AS plan_salarial,
                `Grado` AS grado,
                `Esc` AS escala,
                `Puesto Ptal` AS puesto_presupuestal,
                `Partida Ptal` AS partida_presupuestal,
                `Gp Pago` AS grupo_pago,
                `Prog Beneficios` AS programa_beneficios,
                `F/H Últ Actz` AS fecha_ultima_actualizacion,
                `Por` AS por,
                `Hr Estd/Semn` AS horas_estandar_semana,
                `Descr` AS descripcion,
                `Gp Trabajo` AS grupo_trabajo,
                `Org Code` AS codigo_organizacional,
                `Grupo Cd Sal` AS grupo_codigo_salarial,
                `FormalDesc` AS descripcion_formal,
                `Pto Compt` AS puesto_compartido,
                `Posn Clv` AS posicion_clave,
                `Presupuesto` AS presupuesto,
                `Nombre Puesto` AS nombre_puesto_funcional,
                `CATEGORIA_VACANCIA` AS categoria_vacancia,
                `idRegistroDesicivo` AS id_registro_decisivo,
                `TUVO_INSUBSISTENCIA` AS tuvo_insubsistencia,
                `idInsubsistenciaDetectada` AS id_insubsistencia_detectada,
                `FECHA_OCUPACION` AS fecha_ocupacion,
                `ID_REGISTRO_DES_FECHA_OCUPACION` AS id_registro_des_fecha_ocupacion
            FROM MOV_POS
            WHERE `Nº Pos Actual` IN ({placeholders})
            ORDER BY `Nº Pos Actual` ASC, `F Efva` DESC, `Fecha Captura` DESC, id DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(sql, posiciones)
            cols = [d[0] for d in cursor.description]
            rows = cursor.fetchall()

        results = []
        for row in rows:
            record = {}
            for i, col in enumerate(cols):
                val = row[i]
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                record[col] = val
            results.append(record)

        return Response(results)


class MovimientosPersonalStatsView(APIView):
    """
    Devuelve la estadística de movimientos de personal agrupado por accion_nombre y año.
    Respuesta: {
        "by_year": {
            "2026": [{"accion_nombre": "REINGRESO", "total": 10}, ...],
            ...
        },
        "all": [{"accion_nombre": "REINGRESO", "total": 150}, ...]
    }
    """

    view_permission = "authentication.view_plantilla_movimientos"

    def get(self, request):
        accion_nombre = request.query_params.get("accion_nombre")
        fecha_captura__in = request.query_params.get("fecha_captura__in")

        from django.db.models import Count, F, Value
        from django.db.models.functions import Coalesce, ExtractYear, NullIf

        from .models import CpTblMovCompleto290526

        queryset = CpTblMovCompleto290526.objects

        if fecha_captura__in:
            val_list = [v.strip() for v in fecha_captura__in.split(",") if v.strip()]
            from django.db.models import Q

            q_objects = Q()
            for val in val_list:
                q_objects |= Q(fecha_captura__startswith=val)
            queryset = queryset.filter(q_objects)

        if accion_nombre:
            queryset = queryset.filter(accion_nombre=accion_nombre)
            group_field = "motivo_nombre"
            fallback_label = "Sin motivo"
        else:
            group_field = "accion_nombre"
            fallback_label = "Sin acción"

        # BUG-08 QA: antes se excluían del agrupamiento los registros con
        # `group_field` nulo/vacío (`.exclude(isnull=True).exclude(exact="")`),
        # así que el total de esta gráfica quedaba 1 registro por debajo del
        # total real de la tabla (que sí los cuenta). Se agrupan bajo un bucket
        # explícito en vez de descartarlos, para que ambos totales coincidan.
        queryset = queryset.annotate(
            group_val=Coalesce(NullIf(F(group_field), Value("")), Value(fallback_label))
        )

        # Fetch stats grouped by year and group_field
        stats_by_year = (
            queryset.annotate(year=ExtractYear("fecha_efectiva"))
            .values("year", "group_val")
            .annotate(total=Count("*"))
            .order_by("-year", "-total")
        )

        # Fetch stats for ALL years combined
        stats_all = (
            queryset.values("group_val")
            .annotate(total=Count("*"))
            .order_by("-total")
        )

        by_year_dict = {}
        for row in stats_by_year:
            year_val = row["year"]
            year_str = str(year_val) if year_val is not None else "Sin Año"
            if year_str not in by_year_dict:
                by_year_dict[year_str] = []
            by_year_dict[year_str].append(
                {group_field: row["group_val"], "total": row["total"]}
            )

        all_list = [
            {group_field: row["group_val"], "total": row["total"]} for row in stats_all
        ]

        result = {"by_year": by_year_dict, "all": all_list}

        return Response(result, status=status.HTTP_200_OK)


class CuadroVacanciaView(APIView):
    # Sub-tab "Cuadros" dentro de Mov. Posiciones (CuadrosVacanciaTab en el front).
    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        try:
            resultados = CuadroVacancia.objects.all().order_by("-fecha").values()
            return Response(list(resultados), status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# Ramo presupuestal de la ANAM — fijo para todo el sistema (una sola
# dependencia, no un catálogo). No hay ningún campo "ramo" en la BD (ni en
# EMPLEADOS_COMPLETOS_SIG, ni en el catálogo de corrección, ni en MOV_POS):
# se confirmó contra los ~11 mil códigos reales de cat_correccion_posicion
# que el primer segmento SIEMPRE es "06", así que hardcodearlo es más
# confiable que parsearlo del código (un código mal capturado sin ese
# prefijo rompería el parseo en silencio; el ramo nunca cambia).
RAMO_ANAM = "06"


def _parse_unidad_responsable(codigo):
    """``"06-H00-001794"`` → ``"H00"``; ``"EV-2026-06-H00-034236"`` → ``"H00"``.

    A diferencia de Ramo (ver ``RAMO_ANAM``), la Unidad Responsable sí varía
    por código, así que se sigue leyendo del segundo segmento. Las plazas
    eventuales anteponen ``EV-<año>-``. Devuelve ``""`` si el formato no es
    reconocible — el usuario siempre puede escribir esa celda a mano.
    """
    partes = [p for p in str(codigo or "").strip().split("-") if p]
    if partes and partes[0].upper() == "EV":
        partes = partes[2:]  # descarta "EV" y el año
    return partes[1] if len(partes) >= 2 else ""


# El Anexo 2 sólo admite estas 3 opciones en "Tipo de contratación"; los
# subtipos internos (PASEM, "Eventual N.C.") se colapsan al tipo que les
# corresponde en el formato oficial.
_TIPO_CONTRATACION_ANEXO2 = {
    "11301": "Permanente",
    "11401": "Permanente",
    "12201": "Eventual",
}


class AnuenciaSugerenciasView(APIView):
    """Autocompletado del Código Federal de Puesto mientras el usuario escribe
    (sub-tab "Anuencia", ver AnuenciaTab.jsx) — propone los códigos que
    coinciden para que no tenga que copiar el código exacto de otro lado.

    Sólo propone posiciones VACANTES: el Anexo 2 es para solicitar ocupación,
    así que sugerir una ya ocupada sería un error casi seguro. Si el usuario
    de todas formas escribe (o pega) el código exacto de una ocupada, eso
    sigue resolviéndose normal en ``AnuenciaLookupView``, que sí avisa
    "Plaza OCUPADA" — aquí sólo se filtra la LISTA de sugerencias.

    Filtra sobre ``_get_mapa_codigos()`` (cacheado 30 min, ~11 mil códigos),
    no contra la BD, así que es seguro llamarlo en cada tecla con debounce.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"
    LIMITE_DEFAULT = 10
    LIMITE_MAXIMO = 20

    def get(self, request, *args, **kwargs):
        termino = (request.query_params.get("q") or "").strip().upper()
        if len(termino) < 2:
            return Response([], status=status.HTTP_200_OK)

        try:
            limite = min(int(request.query_params.get("limit", self.LIMITE_DEFAULT)), self.LIMITE_MAXIMO)
        except (TypeError, ValueError):
            limite = self.LIMITE_DEFAULT

        try:
            mapa = _get_mapa_codigos()
            posiciones_ocupadas = _get_posiciones_ocupadas_set()

            # "Empieza con" antes que "contiene en medio": lo más probable es
            # que el usuario esté escribiendo el código desde el principio,
            # pero no se descarta el resto (las eventuales anteponen
            # "EV-<año>-", así que buscar "H00" no debe fallar por no estar
            # al inicio).
            empieza_con, contiene = [], []
            for pos, cod in mapa.items():
                if pos in posiciones_ocupadas:
                    continue
                cod_str = str(cod or "")
                if not cod_str:
                    continue
                cod_upper = cod_str.upper()
                if cod_upper.startswith(termino):
                    empieza_con.append((pos, cod_str))
                elif termino in cod_upper:
                    contiene.append((pos, cod_str))

            coincidencias = (sorted(empieza_con, key=lambda t: t[1]) + sorted(contiene, key=lambda t: t[1]))[:limite]

            posiciones = [pos for pos, _ in coincidencias]
            mapa_denominacion = dict(
                EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values_list(
                    "posicion", "nombre_puesto_funcional"
                )
            )

            resultados = [
                {"codigo": cod, "posicion": pos, "denominacion_puesto": mapa_denominacion.get(pos) or ""}
                for pos, cod in coincidencias
            ]
            return Response(resultados, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def _resolver_datos_anuencia(codigos):
    """
    Resuelve el autollenado de Anuencia para VARIOS Códigos Federales de
    Puesto de golpe — el mismo cruce que hace ``AnuenciaLookupView`` para uno
    solo, factorizado aquí para poder reusarlo en bulk (ver
    ``AnuenciaLookupBulkView``, que rellena Unidad de Negocio/Rango Salarial
    al ABRIR un Anexo 2 ya guardado, sin pegarle a la BD código por código).

    :returns: ``{codigo: {...los mismos campos que devuelve el GET de uno
        solo, salvo "codigo"...} }`` — un código que no resuelve simplemente
        no aparece en el dict (no es un error por sí solo en el caso bulk).
    """
    codigos_por_upper = {str(c).strip().upper(): str(c).strip() for c in codigos if str(c or "").strip()}
    if not codigos_por_upper:
        return {}

    mapa = _get_mapa_codigos()
    posicion_por_codigo = {}
    for pos, cod in mapa.items():
        cod_upper = str(cod or "").strip().upper()
        if cod_upper in codigos_por_upper:
            posicion_por_codigo[codigos_por_upper[cod_upper]] = pos

    posiciones = list(posicion_por_codigo.values())
    if not posiciones:
        return {}

    datos_mov_pos = _get_datos_mov_pos_bulk_map(posiciones)
    mapa_puesto_funcional = _get_mapa_puesto_funcional()
    mapa_presupuestal = _get_mapa_cod_presupuestal()
    mapa_fecha_vacancia = _get_fecha_vacancia_bulk_map(posiciones)
    mapa_mov_pos_id = _get_mov_pos_id_bulk_map(posiciones)
    posiciones_ocupadas = _get_posiciones_ocupadas_set()
    mapa_nombres = {
        f["posicion"]: f["nombres"]
        for f in EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values("posicion", "nombres")
    }

    resultado = {}
    for codigo, posicion in posicion_por_codigo.items():
        datos_mp = datos_mov_pos.get(posicion)
        if not datos_mp:
            continue
        cd_puesto = str(datos_mp.get("cd_puesto") or "").strip()
        puesto_ptal = str(datos_mp.get("puesto_ptal") or "").strip()
        try:
            escala = int(str(datos_mp.get("esc") or "").strip())
        except ValueError:
            escala = None
        datos_presupuestal = mapa_presupuestal.get((puesto_ptal, escala), {})
        ocupada = posicion in posiciones_ocupadas
        resultado[codigo] = {
            "codigo": codigo,
            "posicion": posicion,
            "ramo": RAMO_ANAM,
            "unidad_responsable": _parse_unidad_responsable(codigo),
            "denominacion_puesto": mapa_puesto_funcional.get(cd_puesto, ""),
            "nivel_salarial": datos_presupuestal.get("nivel") or "",
            "rango_salarial": (
                str(datos_presupuestal["smn"]) if datos_presupuestal.get("smn") is not None else "0"
            ),
            "numero_plazas": 1,
            "numero_horas": "",
            "tipo_contratacion": _TIPO_CONTRATACION_ANEXO2.get(str(datos_mp.get("partida_ptal") or "").strip(), ""),
            "fecha_inicio_vacancia": mapa_fecha_vacancia.get(posicion, ""),
            "mov_pos_id": mapa_mov_pos_id.get(posicion),
            "unidad_de_negocio": str(datos_mp.get("unidad_de_negocio") or "").strip(),
            "ocupada": ocupada,
            "ocupante": (mapa_nombres.get(posicion) or "") if ocupada else "",
        }
    return resultado


class AnuenciaLookupView(APIView):
    """Autollenado de una fila del Anexo 2 (Solicitud de Ocupación de Plazas) a
    partir del Código Federal de Puesto — sub-tab "Anuencia" dentro de
    Mov. Posiciones (ver AnuenciaTab.jsx).

    El código es la llave que el usuario captura; de ahí se resuelve la
    posición (``cat_correccion_posicion``, misma fuente que la columna "Código"
    de Plantilla Detalle) y con ella el resto de las columnas, que se sacan de
    MOV_POS (no de Plantilla Detalle): Unidad de Negocio viene directo de esa
    tabla; el puesto se cruza por ``cd_puesto`` contra CAT_PTO_FUNC; el nivel
    salarial y el Salario Mensual Neto se cruzan por ``puesto_ptal`` + ``esc``
    contra ``rc_cat_cod_presupuestal`` (catálogo "Códigos Presupuestales"); el
    tipo de contratación sale de ``partida_ptal`` (11301 Permanente, 12201
    Eventual). Todo lo devuelto es editable en el front: esto es una
    comodidad de captura, no una restricción.

    La resolución en sí vive en ``_resolver_datos_anuencia`` (compartida con
    ``AnuenciaLookupBulkView``, que resuelve muchos códigos de golpe para
    rellenar Unidad de Negocio/Rango Salarial al abrir un Anexo 2 guardado
    ANTES de que existieran esas dos columnas).
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        codigo = (request.query_params.get("codigo") or "").strip()
        if not codigo:
            return Response(
                {"error": "El parámetro 'codigo' es requerido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            resultado = _resolver_datos_anuencia([codigo]).get(codigo)
            if not resultado:
                return Response(
                    {"error": f"No se encontró ninguna posición con el código '{codigo}'."},
                    status=status.HTTP_404_NOT_FOUND,
                )
            return Response(resultado, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AnuenciaLookupBulkView(APIView):
    """
    Como ``AnuenciaLookupView`` pero para VARIOS códigos a la vez — usada al
    ABRIR un Anexo 2 ya guardado (ver ``AnuenciaTab.jsx``,
    ``backfillUnidadNegocioYRango``) para rellenar Unidad de Negocio y Rango
    Salarial en filas capturadas ANTES de que esas dos columnas existieran
    (antes, Unidad de Negocio no se guardaba y Rango Salarial se dejaba en
    "0" fijo). Sólo rellena lo que venga vacío/"0" en el front — aquí siempre
    se manda la resolución completa y actual, sin distinguir "ya se había
    resuelto antes"; esa decisión de qué pisar es del front, no de aquí.

    POST body: ``{"codigos": ["06-H00-001589", ...]}`` (hasta 500 por
    llamada). Devuelve ``{"resultados": {"<codigo>": {...}}}`` — un código
    que no resuelve simplemente no aparece en el dict.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def post(self, request, *args, **kwargs):
        codigos = request.data.get("codigos") or []
        if not isinstance(codigos, list) or not codigos:
            return Response(
                {"error": "'codigos' debe ser una lista no vacía."}, status=status.HTTP_400_BAD_REQUEST
            )
        codigos = codigos[:500]

        try:
            return Response({"resultados": _resolver_datos_anuencia(codigos)}, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def _mapa_columnas_plantilla_historica():
    """db_column crudo (tal como lo devuelve `sp_plantilla_historica`) -> field
    name snake_case, derivado de `EmpleadosCompletosSig` (mismo esquema que el
    SP reconstruye) + las 2 columnas que el SP agrega y que no existen en ese
    modelo (`Estado Plaza`, calculado por posición; `Salario base (mov)`,
    tomado del movimiento del empleado). Derivarlo del modelo (en vez de
    mantener un diccionario manual, como mapVacanteRow.js en el front) evita
    que ambos se desincronicen si el modelo gana/pierde una columna."""
    mapa = {field.column: field.name for field in EmpleadosCompletosSig._meta.fields}
    mapa["Estado Plaza"] = "estado_plaza"
    mapa["Salario base (mov)"] = "salario_base_mov"
    return mapa


def _mapear_fila_plantilla_historica(row, cols, mapa_columnas):
    from decimal import Decimal

    fila = {}
    for key, value in zip(cols, row):
        if isinstance(value, Decimal):
            value = float(value)
        fila[mapa_columnas.get(key, key)] = value
    # Etiqueta lista para pintar (Activa/Inactiva) — evita que cada consumidor
    # del front tenga que traducir el código A/I de `estado_plaza`.
    fila["estado_plaza_label"] = "Activa" if fila.get("estado_plaza") == "A" else "Inactiva"
    return fila


class PlantillaHistoricaView(APIView):
    """Reconstruye la plantilla completa (una fila por plaza, con su estado
    Activa/Inactiva y su ocupante si aplica) a una fecha pasada arbitraria —
    botón "Consultar plantillas pasadas" en el tab Plantilla Detalle.

    Combina 2 stored procedures con el mismo criterio de corte:
      - `sp_conteo_plazas_historico`: desglose plazas totales/activas/
        inactivas/ocupadas/vacantes (barato, sin reconstruir cada fila) — se
        usa como resumen "oficial" en las tarjetas del front.
      - `sp_plantilla_historica`: la reconstrucción fila por fila (pesado:
        ~90s, joins completos sobre MOV_POS/cp_tbl_mov_completo_29_05_26) +
        su propio resumen (se toma sólo para las 2 columnas de anomalías,
        que `sp_conteo_plazas_historico` no calcula).

    GET ?fecha=YYYY-MM-DD (obligatoria, entre 2022-01-01 y hoy — mismo límite
    inferior que exigen ambos SPs: inicio de MOV_POS/ANAM).
    """

    view_permission = "authentication.view_plantilla_historico"

    _FECHA_MINIMA = datetime.date(2022, 1, 1)

    def get(self, request, *args, **kwargs):
        from django.db import connection

        fecha_str = (request.query_params.get("fecha") or "").strip()
        try:
            fecha = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Parámetro 'fecha' inválido u omitido, se espera formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        hoy = datetime.date.today()
        if fecha < self._FECHA_MINIMA or fecha > hoy:
            return Response(
                {
                    "error": (
                        f"'fecha' debe estar entre {self._FECHA_MINIMA.isoformat()} y "
                        f"{hoy.isoformat()} (inicio de MOV_POS/ANAM)."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        cache_key = f"plantilla_historica_{fecha.isoformat()}"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_conteo_plazas_historico(%s)", [fecha])
                conteo_cols = [col[0] for col in cursor.description]
                conteo_row = cursor.fetchone()
                conteo = dict(zip(conteo_cols, conteo_row)) if conteo_row else {}
                cursor.nextset()

                cursor.execute("CALL sp_plantilla_historica(%s)", [fecha])
                resumen_cols = [col[0] for col in cursor.description]
                resumen_row = cursor.fetchone()
                resumen_sp = dict(zip(resumen_cols, resumen_row)) if resumen_row else {}
                cursor.nextset()

                filas_cols = [col[0] for col in cursor.description]
                filas_rows = cursor.fetchall()
                cursor.nextset()

            mapa_columnas = _mapa_columnas_plantilla_historica()
            filas = [
                _mapear_fila_plantilla_historica(row, filas_cols, mapa_columnas)
                for row in filas_rows
            ]

            def _entero(valor):
                return int(valor) if valor is not None else None

            resumen = {
                "plazas_totales": _entero(conteo.get("Plazas totales", resumen_sp.get("Plazas totales"))),
                "plazas_activas": _entero(conteo.get("Plazas activas", resumen_sp.get("Plazas activas"))),
                "plazas_inactivas": _entero(conteo.get("Plazas inactivas", resumen_sp.get("Plazas inactivas"))),
                "ocupadas": _entero(conteo.get("Ocupadas", resumen_sp.get("Ocupadas"))),
                "vacantes": _entero(conteo.get("Vacantes", resumen_sp.get("Vacantes"))),
                "anomalia_ocupante_en_plaza_inactiva": _entero(resumen_sp.get("Anomalia ocupante en plaza inactiva")),
                "anomalia_ocupante_sin_plaza": _entero(resumen_sp.get("Anomalia ocupante sin plaza")),
            }

            payload = _serializar_fechas({
                "fecha": fecha.isoformat(),
                "resumen": resumen,
                "filas": filas,
            })
            # 24h: una fecha pasada cerrada nunca cambia de resultado (MOV_POS/
            # cp_tbl_mov_completo_29_05_26 son históricos append-only); si
            # `fecha` es hoy, el corte sí puede moverse durante el día, pero
            # 24h de margen es aceptable frente al costo (~90s) de repetir el
            # cálculo en cada consulta.
            cache.set(cache_key, payload, 60 * 60 * 24)
            return Response(payload, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ConteoPlazasHistoricoSerieView(APIView):
    # Serie mensual (corte a fin de cada mes desde 2022-01) de plazas
    # totales/activas/inactivas/ocupadas/vacantes, para la gráfica histórica
    # de CuadrosVacanciaTab. El SP recorre mes a mes con joins pesados sobre
    # MOV_POS/cp_tbl_mov_completo_29_05_26 (~25s), de ahí el cache largo: el
    # histórico de meses cerrados no cambia, solo el corte del mes en curso.
    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from django.db import connection

        cache_key = "conteo_plazas_historico_serie"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_conteo_plazas_historico_serie();")
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

            cache.set(cache_key, results, 3600)
            return Response(results, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PlazasMovimientoMesView(APIView):
    # Detalle de qué posiciones concretas se crearon (activas) o desactivaron
    # (inactivas) entre el corte de fin de mes anterior y el corte de fin de
    # mes actual — para el click en las franjas verde/guinda de la gráfica
    # "Plazas Totales vs Activas vs Inactivas" en CuadrosVacanciaTab. Misma
    # técnica de snapshot por posición (ROW_NUMBER sobre MOV_POS con cutoff)
    # que sp_conteo_plazas_historico_serie, pero sin el join a
    # cp_tbl_mov_completo_29_05_26 (no hace falta: aquí solo importa el
    # tránsito Activa/Inactiva, no Ocupada/Vacante) y acotado a las dos
    # posiciones cuyo estado cambió, no al conteo agregado.
    view_permission = "authentication.view_plantilla_mov_posiciones"

    # A pedido del usuario (2026-08-12): la fila devuelta es el registro
    # MOV_POS COMPLETO de la posición (todas sus columnas), sin tocar
    # EMPLEADOS_COMPLETOS_SIG ni ua_unidadadministrativa para nada — esa
    # tabla no es la fuente de verdad de estos movimientos y tenía huecos
    # (p.ej. `Nivel` en null para posiciones desactivadas). El registro
    # elegido por posición es el de `s_curr` (el snapshot vigente al corte
    # fecha_actual): para creación es la fila que puso el estado en 'A', para
    # desactivación es la fila que puso el estado en 'I' — en ambos casos el
    # movimiento que efectivamente causó la transición.
    # `_RAW_COLUMN_MAP` traduce cada columna cruda de MOV_POS a una key
    # snake_case estable para el front (ver ALL_AVAILABLE_COLUMNS en
    # EmployeesModal.jsx, categoría "Movimiento de Posición"); las keys
    # `posicion`, `nombre_puesto_funcional`, `unidad_administrativa`,
    # `fecha_efectiva_mov_pos`, `fecha_de_captura` y `capturado_por` se
    # reutilizan tal cual porque ya existían en ALL_AVAILABLE_COLUMNS con el
    # mismo significado. `grado_escala` es un campo calculado (Grado + Esc)
    # que reemplaza a `nivel` — MOV_POS no tiene nivel, tiene grado/escala.
    _RAW_COLUMN_MAP = {
        "Nº Pos Actual": "posicion",
        "F Efva": "fecha_efectiva_mov_pos",
        "Estado Psn": "estado_posicion",
        "Fecha Captura": "fecha_de_captura",
        "Cd Motivo": "cd_motivo",
        "Motivo": "motivo",
        "Cd UN": "cd_un",
        "Unidad de Negocio": "unidad_administrativa",
        "Unidad Adva#": "unidad_adva",
        "Cd Departamento": "cd_departamento",
        "Cd Puesto": "cd_puesto",
        "Estado Ptal": "estado_ptal",
        "Fecha Est": "fecha_establecimiento",
        "Máximo": "maximo",
        "Depnd Drt": "dependencia_directa",
        "Depnd Indrt": "dependencia_indirecta",
        "Ubicación": "ubicacion",
        "Nvl Direc": "nivel_direccion",
        "Plan Sal": "plan_salarial",
        "Grado": "grado",
        "Esc": "escala",
        "Puesto Ptal": "puesto_presupuestal",
        "Partida Ptal": "partida_presupuestal",
        "Gp Pago": "grupo_pago",
        "Prog Beneficios": "programa_beneficios",
        "F/H Últ Actz": "fecha_ultima_actualizacion",
        "Por": "capturado_por",
        "Hr Estd/Semn": "horas_estandar_semana",
        "Descr": "descripcion",
        "Gp Trabajo": "grupo_trabajo",
        "Org Code": "codigo_organizacional",
        "Grupo Cd Sal": "grupo_codigo_salarial",
        "FormalDesc": "descripcion_formal",
        "Pto Compt": "puesto_compartido",
        "Posn Clv": "posicion_clave",
        "Presupuesto": "presupuesto",
        "Nombre Puesto": "nombre_puesto_funcional",
        "FECHA VACANCIA": "fecha_vacancia_mov_pos",
        "CATEGORIA_VACANCIA": "categoria_vacancia",
        "idRegistroDesicivo": "id_registro_decisivo",
        "TUVO_INSUBSISTENCIA": "tuvo_insubsistencia",
        "idInsubsistenciaDetectada": "id_insubsistencia_detectada",
        "FECHA_OCUPACION": "fecha_ocupacion",
        "ID_REGISTRO_DES_FECHA_OCUPACION": "id_registro_des_fecha_ocupacion",
    }

    # 2026-08-12: MOV_POS perdió 5 de sus 6 índices en algún punto de esta
    # sesión (quedó solo PRIMARY + índice sobre `Estado Psn`; se sospecha un
    # swap de tabla de un ETL externo — no se tocó el schema a pedido del
    # usuario). Sin índice sobre `Nº Pos Actual` ni sobre `F Efva`, el
    # optimizer de MySQL 8 además cae en un plan patológico: intenta fusionar
    # (`derived_merge`) las CTEs s_curr/s_prev —de forma idéntica salvo el
    # cutoff— y arma un hash join SIN condición (cross join de ~90 millones
    # de filas) en vez de usar `s_curr.posicion = s_prev.posicion` como llave.
    # `s_curr` ya trae el registro COMPLETO (`m.*`) para no necesitar un join
    # de vuelta a MOV_POS. El toggle de `derived_merge` en el get() de abajo
    # fuerza a MySQL a materializar cada CTE por separado — evita el cross
    # join sin tocar el schema (queda ~1-3s en vez de timeout/minutos).
    _SNAPSHOT_CTE = """
        WITH s_curr AS (
            SELECT m.*
            FROM MOV_POS m
            JOIN (
                SELECT id FROM (
                    SELECT id, ROW_NUMBER() OVER (
                        PARTITION BY `Nº Pos Actual`
                        ORDER BY `F Efva` DESC, `Fecha Captura` DESC, `F/H Últ Actz` DESC, id DESC
                    ) rn
                    FROM MOV_POS WHERE `F Efva` <= %s
                ) r WHERE rn = 1
            ) l ON l.id = m.id
        ),
        s_prev AS (
            SELECT `Nº Pos Actual` AS posicion, TRIM(`Estado Psn`) AS estado
            FROM MOV_POS m
            JOIN (
                SELECT id FROM (
                    SELECT id, ROW_NUMBER() OVER (
                        PARTITION BY `Nº Pos Actual`
                        ORDER BY `F Efva` DESC, `Fecha Captura` DESC, `F/H Últ Actz` DESC, id DESC
                    ) rn
                    FROM MOV_POS WHERE `F Efva` <= %s
                ) r WHERE rn = 1
            ) l ON l.id = m.id
        )
    """

    _QUERY_CREACION = _SNAPSHOT_CTE + """
        SELECT s_curr.*
        FROM s_curr
        LEFT JOIN s_prev ON s_prev.posicion = s_curr.`Nº Pos Actual`
        WHERE TRIM(s_curr.`Estado Psn`) = 'A' AND (s_prev.estado IS NULL OR s_prev.estado <> 'A')
        ORDER BY s_curr.`F Efva` DESC, s_curr.`Nº Pos Actual`;
    """

    _QUERY_DESACTIVACION = _SNAPSHOT_CTE + """
        SELECT s_curr.*
        FROM s_prev
        JOIN s_curr ON s_curr.`Nº Pos Actual` = s_prev.posicion
        WHERE s_prev.estado = 'A' AND TRIM(s_curr.`Estado Psn`) = 'I'
        ORDER BY s_curr.`F Efva` DESC, s_curr.`Nº Pos Actual`;
    """

    def get(self, request, *args, **kwargs):
        import time
        from django.db import connection

        tipo = request.query_params.get("tipo")
        fecha_actual = request.query_params.get("fecha_actual")
        fecha_anterior = request.query_params.get("fecha_anterior")

        if tipo not in ("creacion", "desactivacion"):
            return Response(
                {"error": "tipo debe ser 'creacion' o 'desactivacion'"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            datetime.date.fromisoformat(fecha_actual)
            datetime.date.fromisoformat(fecha_anterior)
        except (TypeError, ValueError):
            return Response(
                {"error": "fecha_actual y fecha_anterior deben tener formato YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cache_key = f"plazas_movimiento_mes_{tipo}_{fecha_anterior}_{fecha_actual}"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        query = self._QUERY_CREACION if tipo == "creacion" else self._QUERY_DESACTIVACION

        t0 = time.monotonic()
        try:
            with connection.cursor() as cursor:
                # Ámbito acotado a esta query nada más — CONN_MAX_AGE=60
                # reutiliza la conexión entre requests, así que se revierte el
                # toggle en el finally para no afectar otras queries.
                cursor.execute("SET SESSION optimizer_switch='derived_merge=off'")
                try:
                    cursor.execute(query, [fecha_actual, fecha_anterior])
                    raw_columns = [col[0] for col in cursor.description]
                    rows = cursor.fetchall()
                finally:
                    cursor.execute("SET SESSION optimizer_switch='derived_merge=on'")

                results = []
                for row in rows:
                    raw = dict(zip(raw_columns, row))
                    mapped = {
                        self._RAW_COLUMN_MAP[col]: val
                        for col, val in raw.items()
                        if col in self._RAW_COLUMN_MAP
                    }
                    grado = (raw.get("Grado") or "").strip()
                    escala = (raw.get("Esc") or "").strip()
                    mapped["grado_escala"] = f"{grado}-{escala}" if grado or escala else None
                    results.append(mapped)

            elapsed_ms = round((time.monotonic() - t0) * 1000)
            logger.info(
                "plazas_movimiento_mes tipo=%s %s->%s: %dms, %d filas",
                tipo, fecha_anterior, fecha_actual, elapsed_ms, len(results),
            )
            cache.set(cache_key, results, 3600)
            return Response(results, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DesgloseJerarquicoView(APIView):
    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from django.db import connection

        cache_key = "desglose_jerarquico"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        # SELECT amplía todas las columnas de EMPLEADOS_COMPLETOS_SIG expuestas
        # en ALL_AVAILABLE_COLUMNS del front (EmployeesModal.jsx) — antes solo
        # traía 19, y el botón "Columnas" del modal ofrecía ~60, generando un
        # mismatch (columnas seleccionables que siempre salían vacías porque
        # el backend nunca las mandaba). Si el front agrega una columna nueva
        # a ALL_AVAILABLE_COLUMNS, hay que sumarla aquí también.
        query = """
        SELECT
            e.NJ,
            e.`Nombre Puesto Funcional`,
            e.`Nivel`,
            e.`Posición`,
            m.`FECHA VACANCIA` AS `Fecha Vacancia`,
            m.id AS mov_pos_id,
            e.`Unidad de Negocio`,
            e.`Cd UA`,
            COALESCE(u.nombre, e.`Cd UA`) AS `nombre_ua`,
            e.`Cd UN`,
            e.`Código Presupuestal`,
            e.`Escala`,
            e.`Partida`,
            e.`TIPO DE CONTRATACIÓN`,
            e.`Sindicato`,
            e.`Entidad Federativa`,
            e.`nombreNJ`,
            e.`Id Departamento`,
            e.`Departamento`,
            e.`SMB`,
            e.`SMN`,
            e.`Id Empleado`,
            e.`Nombres`,
            e.`RFC`,
            e.`CURP`,
            e.`Fecha de ingreso`,
            e.`Estado Nómina`,
            e.`Aduana`,
            e.`Tipo de Aduana`,
            e.`municipio`,
            e.`Ubicación`,
            e.`Descripción ubicación`,
            e.`tipo`,
            e.`Val_estat`,
            e.`Estado en nomina`,
            e.`UA Validación`,
            e.`Validando de posición por documento`,
            e.`Status Jefe Inm Posición`,
            e.`numeral`,
            e.`OBSERVACIONES`,
            e.`Personal Militar o Civil`,
            e.`Rango`,
            e.`Posición _Civil / SEDENA / SEMAR`,
            e.`Tipo de personal SEDENA / SEMAR`,
            e.`DG o Aduana compactada`,
            e.`Proyecto 2024 Reducción de plazas Eventuales`,
            e.`Fecha efectiva (Personal)`,
            e.`Fecha de captura`,
            e.`Qna`,
            e.`Fecha prevista de salida`,
            e.`DependenciaDirecta`,
            e.`Numempleado`,
            e.`Id_campo`,
            e.`cent`,
            e.`dir`,
            e.`subd`,
            e.`jd`,
            e.`depto`,
            e.`id tipo`,
            e.`ua2`,
            e.`latitud`,
            e.`longitud`,
            e.`Cd Pto Funcional`
        FROM EMPLEADOS_COMPLETOS_SIG e
        INNER JOIN MOV_POS m
            ON e.`Posición` = m.`Nº Pos Actual`
        INNER JOIN MOV_POS_LATEST latest ON m.id = latest.id
        LEFT JOIN ua_unidadadministrativa u
            ON TRIM(e.`Cd UA`) = TRIM(u.codigo)
        WHERE m.`Estado Psn` = 'A'
          AND e.`Estado Nómina` = ' '
          AND m.`Nº Pos Actual` NOT LIKE '103L%%'
          AND m.`Nº Pos Actual` NOT LIKE '1039%%'
          AND m.`Partida Ptal` <> '11401';
        """

        try:
            with connection.cursor() as cursor:
                cursor.execute(query)
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

            cache.set(cache_key, results, 1200)
            return Response(results, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DesgloseJerarquicoOcupadosView(APIView):
    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        from django.db import connection

        cache_key = "desglose_jerarquico_ocupados"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        # SELECT amplía todas las columnas de EMPLEADOS_COMPLETOS_SIG expuestas
        # en ALL_AVAILABLE_COLUMNS del front (EmployeesModal.jsx) — ver mismo
        # comentario en DesgloseJerarquicoView arriba.
        query = """
        SELECT
            e.NJ,
            e.`Nombre Puesto Funcional`,
            e.`Nivel`,
            e.`Posición`,
            e.`Unidad de Negocio`,
            e.`Cd UA`,
            COALESCE(u.nombre, e.`Cd UA`) AS `nombre_ua`,
            e.`Cd UN`,
            e.`Código Presupuestal`,
            e.`Escala`,
            e.`Partida`,
            e.`TIPO DE CONTRATACIÓN`,
            e.`Sindicato`,
            e.`Entidad Federativa`,
            e.`nombreNJ`,
            e.`Id Departamento`,
            e.`Departamento`,
            e.`SMB`,
            e.`SMN`,
            e.`Id Empleado`,
            e.`Nombres`,
            e.`RFC`,
            e.`CURP`,
            e.`Fecha de ingreso`,
            e.`Estado Nómina`,
            e.`Aduana`,
            e.`Tipo de Aduana`,
            e.`municipio`,
            e.`Ubicación`,
            e.`Descripción ubicación`,
            e.`tipo`,
            e.`Val_estat`,
            e.`Estado en nomina`,
            e.`UA Validación`,
            e.`Validando de posición por documento`,
            e.`Status Jefe Inm Posición`,
            e.`numeral`,
            e.`OBSERVACIONES`,
            e.`Personal Militar o Civil`,
            e.`Rango`,
            e.`Posición _Civil / SEDENA / SEMAR`,
            e.`Tipo de personal SEDENA / SEMAR`,
            e.`DG o Aduana compactada`,
            e.`Proyecto 2024 Reducción de plazas Eventuales`,
            e.`Fecha efectiva (Personal)`,
            e.`Fecha de captura`,
            e.`Qna`,
            e.`Fecha prevista de salida`,
            e.`DependenciaDirecta`,
            e.`Numempleado`,
            e.`Id_campo`,
            e.`cent`,
            e.`dir`,
            e.`subd`,
            e.`jd`,
            e.`depto`,
            e.`id tipo`,
            e.`ua2`,
            e.`latitud`,
            e.`longitud`,
            e.`Cd Pto Funcional`
        FROM EMPLEADOS_COMPLETOS_SIG e
        INNER JOIN MOV_POS m
            ON e.`Posición` = m.`Nº Pos Actual`
        INNER JOIN MOV_POS_LATEST latest ON m.id = latest.id
        LEFT JOIN ua_unidadadministrativa u
            ON TRIM(e.`Cd UA`) = TRIM(u.codigo)
        WHERE m.`Estado Psn` = 'A'
          AND e.`Estado Nómina` <> ' '
          AND m.`Nº Pos Actual` NOT LIKE '103L%%'
          AND m.`Nº Pos Actual` NOT LIKE '1039%%'
          AND m.`Partida Ptal` <> '11401';
        """

        try:
            with connection.cursor() as cursor:
                cursor.execute(query)
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

            cache.set(cache_key, results, 1200)
            return Response(results, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AduanasOcupacionVacanciaView(APIView):
    """
    Subtab "Aduanas Ocupación vs Vacantes" (Movimientos > Mov. Posiciones).
    Agrega conteos de personal por Aduana x Nivel Jerárquico x Nivel, separando
    ocupación y vacancia, con el mismo criterio de "plaza activa" que
    DesgloseJerarquicoView/DesgloseJerarquicoOcupadosView (Cuadros Vacancia),
    para que los totales cuadren entre subtabs.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def get(self, request, *args, **kwargs):
        cache_key = "aduanas_ocupacion_vacancia"
        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data, status=status.HTTP_200_OK)

        def agg_query(estado_nomina_op):
            return f"""
            SELECT
                TRIM(e.`Aduana`) AS aduana,
                TRIM(e.`tipo`) AS tipo,
                TRIM(e.`latitud`) AS latitud,
                TRIM(e.`longitud`) AS longitud,
                TRIM(e.`NJ`) AS nj,
                TRIM(e.`Nivel`) AS nivel,
                COUNT(*) AS cantidad
            FROM EMPLEADOS_COMPLETOS_SIG e
            INNER JOIN MOV_POS m ON e.`Posición` = m.`Nº Pos Actual`
            INNER JOIN MOV_POS_LATEST latest ON m.id = latest.id
            WHERE m.`Estado Psn` = 'A'
              AND e.`Estado Nómina` {estado_nomina_op} ' '
              AND m.`Nº Pos Actual` NOT LIKE '103L%%'
              AND m.`Nº Pos Actual` NOT LIKE '1039%%'
              AND m.`Partida Ptal` <> '11401'
              AND e.`Unidad Administrativa` LIKE 'Aduana %%'
              AND e.`Aduana` IS NOT NULL AND e.`Aduana` <> ''
            GROUP BY TRIM(e.`Aduana`), TRIM(e.`tipo`), TRIM(e.`latitud`), TRIM(e.`longitud`), TRIM(e.`NJ`), TRIM(e.`Nivel`);
            """

        try:
            with connection.cursor() as cursor:
                cursor.execute(agg_query("="))
                columns = [col[0] for col in cursor.description]
                vac_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

                cursor.execute(agg_query("<>"))
                columns = [col[0] for col in cursor.description]
                ocu_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

            # dict() sobre tuplas con clave duplicada se queda con la última:
            # NIVELES_JERARQUICOS define (3, "Director") antes que (3, "Titular de
            # Aduana"), así que nj_labels[3] resuelve a "Titular de Aduana", que es
            # lo correcto aquí porque esta vista siempre filtra Unidad
            # Administrativa LIKE 'Aduana %'.
            nj_labels = dict(NIVELES_JERARQUICOS)

            combos = {}
            for r in vac_rows + ocu_rows:
                nj = r["nj"] or ""
                nivel = r["nivel"] or ""
                combos.setdefault(nj, set()).add(nivel)

            def nj_sort_key(nj):
                try:
                    return (0, int(nj))
                except (TypeError, ValueError):
                    return (1, nj)

            grupos_nj = []
            for nj in sorted(combos.keys(), key=nj_sort_key):
                try:
                    nombre_nj = nj_labels.get(int(nj), nj or "Sin Nivel Jerárquico")
                except (TypeError, ValueError):
                    nombre_nj = nj or "Sin Nivel Jerárquico"
                label = f"{nj}. {nombre_nj}" if nj else nombre_nj
                niveles = sorted(n for n in combos[nj] if n) or [""]
                grupos_nj.append({"nj": nj, "label": label, "niveles": niveles})

            all_keys = [f"{g['nj']}|{n}" for g in grupos_nj for n in g["niveles"]]

            def build_pivot(rows):
                por_aduana = {}
                for r in rows:
                    nombre = r["aduana"]
                    if not nombre:
                        continue
                    entry = por_aduana.get(nombre)
                    if entry is None:
                        entry = {
                            "aduana": nombre,
                            "tipo": r["tipo"] or "",
                            "latitud": None,
                            "longitud": None,
                            "valores": {},
                        }
                        por_aduana[nombre] = entry
                    if entry["latitud"] is None and r["latitud"]:
                        try:
                            entry["latitud"] = float(r["latitud"])
                            entry["longitud"] = float(r["longitud"])
                        except (TypeError, ValueError):
                            pass
                    key = f"{r['nj'] or ''}|{r['nivel'] or ''}"
                    entry["valores"][key] = entry["valores"].get(key, 0) + r["cantidad"]

                for entry in por_aduana.values():
                    for k in all_keys:
                        entry["valores"].setdefault(k, 0)

                return sorted(por_aduana.values(), key=lambda a: a["aduana"] or "")

            resultado = {
                "grupos_nj": grupos_nj,
                "ocupacion": build_pivot(ocu_rows),
                "vacancia": build_pivot(vac_rows),
            }

            cache.set(cache_key, resultado, 1200)
            return Response(resultado, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# --- Anexo 3 (FUMP) a partir del Anexo 2 -------------------------------------

# Códigos presupuestales que NUNCA se valúan (confirmado con el área): las
# plazas de laudo no tienen tabulador, y las PASEM/SEM son personal
# SEDENA/SEMAR con un tabulador militar que este sistema no administra.
# No es un error del cruce — se excluyen a propósito y se reportan como aviso.
_CODIGOS_NO_VALUABLES = {
    "LAUDO": "Plaza de laudo — no se valúa presupuestalmente.",
}
_PREFIJOS_NO_VALUABLES = (
    ("SEM", "Plaza PASEM (SEDENA/SEMAR) — tabulador militar, fuera de este sistema."),
)


def _motivo_no_valuable(codigo_presupuestal):
    """Devuelve el motivo por el que un código está excluido de valuación, o
    ``None`` si sí debería valuarse."""
    cod = str(codigo_presupuestal or "").strip().upper()
    if cod in _CODIGOS_NO_VALUABLES:
        return _CODIGOS_NO_VALUABLES[cod]
    for prefijo, motivo in _PREFIJOS_NO_VALUABLES:
        if cod.startswith(prefijo):
            return motivo
    return None


_PALABRAS_IGNORADAS_SIGLAS = {
    "de", "del", "la", "las", "el", "los", "y", "e", "en", "con", "sede", "a", "al",
    "oficina",
}


def _siglas_unidad(nombre):
    """Siglas de una Unidad Administrativa a partir de sus iniciales.

    "Dirección General de Modernización Equipamiento e Infraestructura
    Aduanera" → "DGMEIA". Sólo se usa como respaldo cuando la posición no
    trae "DG o Aduana compactada" (que es el nombre corto oficial y siempre
    gana). Es una propuesta: el usuario puede corregirla en la pantalla de
    previa antes de generar.
    """
    palabras = [p for p in re.split(r"[\s,\.]+", str(nombre or "")) if p]
    iniciales = [
        p[0].upper() for p in palabras if p.lower() not in _PALABRAS_IGNORADAS_SIGLAS
    ]
    return "".join(iniciales)[:20]


class AnuenciaAnexo3View(APIView):
    """Agrupa y valúa las plazas de un Anexo 2 para generar el Anexo 3 (FUMP).

    El Anexo 3 se calcula POR HOJA DEL ANEXO 2 Y POR PERÍODO: cada hoja del
    libro de salida cubre plazas de UNA sola hoja del Anexo 2 con UNA sola
    fecha de alta, porque la valuación sólo puede correr contra un período a
    la vez. Se toma como cierto que cada hoja del Anexo 2 ya es una sola
    Unidad Administrativa (así la organiza el capturista, p.ej. una hoja
    "UAF"); si esa hoja mezcla varias fechas de alta, se reparte aquí en
    varios grupos —uno por fecha—, y cada grupo será una hoja del Anexo 3,
    nombrada "<nombre de la hoja del Anexo 2> <fecha de alta>" (p.ej.
    "UAF 01-07-2026", "UAF 16-07-2026").

    A diferencia de una versión anterior, la Unidad Administrativa YA NO se
    usa para decidir el agrupamiento (antes se resolvía desde el Código
    Federal de Puesto precisamente para no confiar en el campo de texto
    libre de la hoja); ahora la identidad de la hoja del Anexo 2 manda. La UA
    resuelta por código (Unidad de Negocio de MOV_POS, ver
    `_get_datos_mov_pos_bulk_map`) se sigue usando únicamente para los datos
    que sí dependen de la posición (encabezado del documento,
    unidad_responsable, etc.), no para decidir qué plazas comparten hoja.

    POST body::

        {
          "hojas": [ {"filas": [{"codigo": ..., "fecha_alta_solicitada": "YYYY-MM-DD",
                                 "numero_plazas": 1}, ...]}, ... ],
          "overrides": {"<clave_grupo>": {"fecha_fin": "YYYY-MM-DD", "nombre_hoja": "...", "orden": 1.5}}
        }

    Devuelve los grupos ya valuados más el detalle de todo lo que quedó fuera
    (``avisos``), para que el front pueda decir exactamente qué no se pudo
    incluir y por qué en vez de generar un Anexo 3 incompleto en silencio.
    """

    view_permission = "authentication.view_plantilla_mov_posiciones"

    def post(self, request, *args, **kwargs):
        from presupuesto.views import _build_catalogo_index, _resolver_catalogo
        from presupuesto.valuacion import calcular_meses_periodo, calcular_valuacion

        hojas = request.data.get("hojas") or []
        overrides = request.data.get("overrides") or {}
        # {"<codigo>": "<clave_de_grupo_destino>"} — reacomodo manual hecho en
        # Anexo3Editor.jsx (arrastrar una plaza a otra hoja del resultado).
        # Sólo se respeta si el grupo destino es del MISMO período que la
        # plaza (ver más abajo); si no, se ignora en silencio. Si la clave de
        # destino no corresponde a ninguna hoja del Anexo 2 (se arrastró la
        # plaza a la zona "+ Agregar una hoja"), se crea una hoja nueva sin
        # heredar nombre ni Unidad Administrativa de nadie (ver
        # `identidades_nuevas` más abajo).
        reasignaciones = request.data.get("reasignaciones") or {}

        try:
            # 1) Aplanar las filas capturadas y quedarse sólo con las que
            #    traen código (una fila en blanco del cuadro no es una plaza).
            #    El Anexo 3 SÓLO valúa plazas Eventuales: las de carácter
            #    Permanente se cubren con el presupuesto regularizable ya
            #    autorizado y nunca pasan por esta calculadora, aunque sí
            #    puedan capturarse en el Anexo 2 (que cubre ambos tipos).
            filas = []
            avisos = []
            for indice_hoja, hoja in enumerate(hojas):
                nombre_hoja_anexo2 = str(hoja.get("nombre") or "").strip() or f"Hoja {indice_hoja + 1}"
                for fila in (hoja.get("filas") or []):
                    codigo = str(fila.get("codigo") or "").strip()
                    if not codigo:
                        continue
                    tipo_contratacion = str(fila.get("tipo_contratacion") or "").strip()
                    if tipo_contratacion and tipo_contratacion.lower() != "eventual":
                        avisos.append({
                            "codigo": codigo, "hoja": nombre_hoja_anexo2,
                            "motivo": f"Plaza de tipo {tipo_contratacion!r} — el Anexo 3 sólo valúa plazas Eventuales.",
                        })
                        continue
                    filas.append((indice_hoja, nombre_hoja_anexo2, fila, codigo))

            if not filas:
                return Response(
                    {"error": "El Anexo 2 no tiene ninguna plaza capturada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # 2) Código Federal de Puesto -> posición (mismo mapa que usa el
            #    autollenado del Anexo 2).
            mapa = _get_mapa_codigos()
            codigo_a_posicion = {
                str(cod or "").strip().upper(): pos for pos, cod in mapa.items()
            }

            resueltas = []
            for indice_hoja, nombre_hoja, fila, codigo in filas:
                posicion = codigo_a_posicion.get(codigo.upper())
                if not posicion:
                    avisos.append({
                        "codigo": codigo, "hoja": nombre_hoja,
                        "motivo": "No se encontró ninguna posición con ese Código Federal de Puesto.",
                    })
                    continue
                fecha_alta = str(fila.get("fecha_alta_solicitada") or "").strip()
                if not fecha_alta:
                    avisos.append({
                        "codigo": codigo, "hoja": nombre_hoja,
                        "motivo": "Sin 'Fecha de alta solicitada' — es la que define el período a valuar.",
                    })
                    continue
                try:
                    cantidad = int(fila.get("numero_plazas") or 1)
                except (TypeError, ValueError):
                    cantidad = 1
                resueltas.append({
                    "codigo": codigo, "posicion": posicion, "hoja": nombre_hoja, "hoja_indice": indice_hoja,
                    "fecha_alta": fecha_alta, "cantidad": max(1, cantidad),
                })

            # 3) Datos de plantilla detalle para todas las posiciones de golpe
            #    (el tabulador/valuación sigue viniendo de aquí, sin cambios).
            posiciones = [r["posicion"] for r in resueltas]
            detalle = {
                d["posicion"]: d
                for d in EmpleadosCompletosSig.objects.filter(posicion__in=posiciones).values(
                    "posicion", "codigo_presupuestal", "nivel", "escala", "partida",
                    "dg_o_aduana_compactada",
                )
            }
            # "DG o Aduana compactada" viene vacía en ~2/3 de las filas crudas
            # de ZAFIRO; el catálogo de corrección la completa (mismo mecanismo
            # que Plantilla Detalle, ver `_completar_aduana_row`).
            mapa_dg = _get_mapa_correccion("dg_o_aduana_compactada")
            # La Unidad Administrativa que se imprime en cada hoja del Anexo 3
            # ya no sale de Plantilla Detalle — sale de MOV_POS (Unidad de
            # Negocio), igual que el autollenado del Anexo 2 (ver
            # AnuenciaLookupView / `_get_datos_mov_pos_bulk_map`).
            datos_mov_pos = _get_datos_mov_pos_bulk_map(posiciones)

            # 4) Agrupar por (hoja del Anexo 2, fecha de alta) y resolver el
            #    tabulador de cada plaza — la UA resuelta por código ya NO
            #    decide el agrupamiento (ver docstring de la clase).
            #
            #    Identidad "natural" de cada clave ANTES de aplicar
            #    `reasignaciones` — se calcula aparte porque una plaza
            #    reasignada puede procesarse antes que la plaza que
            #    naturalmente "es dueña" de esa hoja; sin este mapa, el grupo
            #    nuevo heredaría por accidente el nombre de quien llegó
            #    primero en vez del de su hoja de origen real.
            identidad_por_clave = {}
            for r in resueltas:
                clave_natural = f"{r['hoja_indice']}||{r['fecha_alta']}"
                identidad_por_clave.setdefault(
                    clave_natural, {"hoja_indice": r["hoja_indice"], "nombre_hoja_anexo2": r["hoja"]}
                )

            # Hojas creadas al soltar una plaza en un hueco entre hojas (o en
            # el menú contextual "Generar nueva hoja a partir de esta plaza",
            # ver Anexo3Editor.jsx): no existen en `identidad_por_clave`
            # porque no las "posee" ninguna plaza de forma natural, así que
            # se les da nombre/UA en blanco para que el usuario los ponga a
            # mano. El front incrusta en la propia clave la posición donde se
            # soltó ("nueva:<posición>:<uuid>||<fecha>") para que ese lugar
            # en la lista se mantenga en los siguientes recálculos de esta
            # misma sesión (la posición viaja con `reasignaciones`); si por
            # algún motivo no se puede leer, cae hasta el final.
            NUEVA_HOJA_BASE_INDICE = 10 ** 9
            identidades_nuevas = {}

            def _posicion_incrustada(clave):
                coincide = re.match(r"^nueva:(-?[\d.]+):", clave)
                if not coincide:
                    return None
                try:
                    return float(coincide.group(1))
                except ValueError:
                    return None

            index = _build_catalogo_index()
            grupos = {}
            for r in resueltas:
                d = detalle.get(r["posicion"])
                if not d:
                    avisos.append({
                        "codigo": r["codigo"], "hoja": r["hoja"],
                        "motivo": "La posición no existe en la plantilla.",
                    })
                    continue

                cod_presupuestal = str(d.get("codigo_presupuestal") or "").strip()
                motivo = _motivo_no_valuable(cod_presupuestal)
                if motivo:
                    avisos.append({
                        "codigo": r["codigo"], "hoja": r["hoja"],
                        "posicion": r["posicion"], "codigo_presupuestal": cod_presupuestal,
                        "motivo": motivo,
                    })
                    continue

                plaza = _resolver_catalogo(index, cod_presupuestal, d.get("escala"), d.get("nivel"))
                if not plaza:
                    avisos.append({
                        "codigo": r["codigo"], "hoja": r["hoja"],
                        "posicion": r["posicion"], "codigo_presupuestal": cod_presupuestal,
                        "motivo": (
                            f"El código presupuestal {cod_presupuestal} (nivel {d.get('nivel') or '—'}, "
                            f"escala {d.get('escala') or '—'}) no está en el catálogo de tabuladores."
                        ),
                    })
                    continue

                ua = str((datos_mov_pos.get(r["posicion"]) or {}).get("unidad_de_negocio") or "").strip()
                clave_natural = f"{r['hoja_indice']}||{r['fecha_alta']}"
                clave_destino = str(reasignaciones.get(r["codigo"]) or "").strip()
                # Un reacomodo manual sólo se respeta si el grupo destino es
                # del MISMO período que la plaza — nunca se mezclan períodos,
                # ni aunque el front mande un estado desactualizado (p.ej. el
                # usuario cambió la fecha de alta después de haber arrastrado
                # la plaza a otra hoja).
                clave = clave_destino if clave_destino and clave_destino.endswith(f"||{r['fecha_alta']}") else clave_natural

                es_hoja_nueva = False
                if clave == clave_natural:
                    identidad = identidad_por_clave.get(
                        clave, {"hoja_indice": r["hoja_indice"], "nombre_hoja_anexo2": r["hoja"]}
                    )
                elif clave in identidad_por_clave:
                    identidad = identidad_por_clave[clave]
                else:
                    # Clave que no pertenece a ninguna hoja del Anexo 2: se
                    # arrastró una plaza a la zona "+ Agregar una hoja". No
                    # hereda nombre ni UA de nadie — el usuario los define.
                    es_hoja_nueva = True
                    posicion_incrustada = _posicion_incrustada(clave)
                    identidad = identidades_nuevas.setdefault(
                        clave,
                        {
                            "hoja_indice": posicion_incrustada if posicion_incrustada is not None
                            else NUEVA_HOJA_BASE_INDICE + len(identidades_nuevas),
                            "nombre_hoja_anexo2": "Hoja nueva",
                        },
                    )

                grupo = grupos.get(clave)
                if grupo is None:
                    corta = "" if es_hoja_nueva else (
                        str(d.get("dg_o_aduana_compactada") or "").strip()
                        or mapa_dg.get(r["posicion"], "").strip()
                        or _siglas_unidad(ua)
                    )
                    grupo = grupos[clave] = {
                        "clave": clave,
                        "hoja_indice": identidad["hoja_indice"],
                        "nombre_hoja_anexo2": identidad["nombre_hoja_anexo2"],
                        "unidad_administrativa": "" if es_hoja_nueva else ua,
                        "unidad_corta": corta,
                        "fecha_inicio": r["fecha_alta"],
                        "plazas": [],
                        "_cantidades": {},
                        "_cuadro": {},
                    }
                unidad_responsable = _parse_unidad_responsable(r["codigo"])
                partida = str(d.get("partida") or "").strip()

                # Detalle fila por fila del Anexo 2 — es lo que arma la lista
                # arrastrable de Anexo3Editor.jsx (`detalle_plazas` en la
                # respuesta), NO el cuadro impreso agregado (`plazas`, más
                # abajo). `unidad_administrativa` es sólo informativa en esa
                # lista (ver columna junto a "Código presupuestal").
                grupo["plazas"].append({
                    "codigo": r["codigo"], "posicion": r["posicion"],
                    "codigo_presupuestal": cod_presupuestal,
                    "catalogo_id": plaza.id, "catalogo_codigo": plaza.codigo,
                    "nivel": plaza.nivel, "zona": plaza.zona,
                    "denominacion": plaza.denominacion,
                    "partida": partida,
                    "unidad_responsable": unidad_responsable,
                    "unidad_administrativa": ua,
                    "sueldo": float(plaza.sueldo),
                    "cantidad": r["cantidad"],
                })

                # Renglón del cuadro impreso: el Anexo 3 no lista una línea por
                # posición, sino una por tipo de plaza (mismo código
                # presupuestal + nivel + zona) con el total de plazas sumado.
                clave_fila = (unidad_responsable, cod_presupuestal, plaza.id)
                cuadro = grupo["_cuadro"].get(clave_fila)
                if cuadro is None:
                    cuadro = grupo["_cuadro"][clave_fila] = {
                        "unidad_responsable": unidad_responsable,
                        # Sólo informativa en la previa (ver Anexo3PreviewModal,
                        # columna junto a "Código presupuestal"): ya NO decide
                        # el agrupamiento (ver docstring de la clase), así que
                        # una misma hoja puede legítimamente traer renglones
                        # con UA distinta si el capturista mezcló varias en la
                        # misma hoja del Anexo 2.
                        "unidad_administrativa": ua,
                        "codigo_presupuestal": cod_presupuestal,
                        "nivel": plaza.nivel,
                        "zona": plaza.zona,
                        "denominacion": plaza.denominacion,
                        "partida": partida,
                        "cantidad": 0,
                        "sueldo": float(plaza.sueldo),
                        "compensacion": float(plaza.compensacion_garantizada),
                    }
                cuadro["cantidad"] += r["cantidad"]
                grupo["_cantidades"][plaza.id] = grupo["_cantidades"].get(plaza.id, 0) + r["cantidad"]

            # 5) Valuar cada grupo. Por default el período corre hasta el 31 de
            #    diciembre del año de la fecha de alta (igual que el Anexo 3
            #    oficial de referencia); el front puede mandar otro `fecha_fin`.
            #
            #    Orden de las hojas: por defecto es `hoja_indice` (el orden en
            #    que aparecen en el Anexo 2), pero el usuario puede arrastrar
            #    una hoja completa en Anexo3Editor.jsx para reacomodarla —
            #    ese reacomodo se guarda como `overrides[clave].orden` (un
            #    número) y, si existe, manda sobre el orden natural. No toca
            #    identidad/nombre/UA de la hoja, sólo en qué lugar se dibuja.
            def _orden_de_hoja(clave, hoja_indice):
                orden = (overrides.get(clave) or {}).get("orden")
                return orden if isinstance(orden, (int, float)) else hoja_indice

            salida = []
            for clave, g in sorted(
                grupos.items(),
                key=lambda kv: (_orden_de_hoja(kv[0], kv[1]["hoja_indice"]), kv[1]["fecha_inicio"]),
            ):
                ov = overrides.get(clave) or {}
                fecha_inicio = datetime.datetime.strptime(g["fecha_inicio"], "%Y-%m-%d").date()
                fecha_fin_str = str(ov.get("fecha_fin") or "").strip() or f"{fecha_inicio.year}-12-31"
                try:
                    fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
                except ValueError:
                    fecha_fin = datetime.date(fecha_inicio.year, 12, 31)
                    fecha_fin_str = fecha_fin.isoformat()

                meses = calcular_meses_periodo(fecha_inicio, fecha_fin)
                plazas_input = [
                    {"catalogo_id": cid, "plazas": qty} for cid, qty in g["_cantidades"].items()
                ]
                valuacion = calcular_valuacion(meses, plazas_input) if meses > 0 else None

                nombre_hoja = str(ov.get("nombre_hoja") or "").strip() or (
                    f"{g['nombre_hoja_anexo2']} {fecha_inicio.strftime('%d-%m-%Y')}"
                )

                # Los importes "por período" del cuadro dependen de `meses`, que
                # sólo se conoce aquí (ya con el override aplicado).
                filas_cuadro = []
                for fila_cuadro in g["_cuadro"].values():
                    filas_cuadro.append({
                        **fila_cuadro,
                        "sueldo_periodo": fila_cuadro["sueldo"] * fila_cuadro["cantidad"] * meses,
                        "compensacion_periodo": fila_cuadro["compensacion"] * fila_cuadro["cantidad"] * meses,
                    })
                filas_cuadro.sort(key=lambda f: (f["nivel"] or "", f["codigo_presupuestal"] or ""))

                salida.append({
                    "clave": clave,
                    "nombre_hoja": nombre_hoja,
                    "unidad_administrativa": g["unidad_administrativa"],
                    "unidad_corta": g["unidad_corta"],
                    "fecha_inicio": g["fecha_inicio"],
                    "fecha_fin": fecha_fin_str,
                    "meses": meses,
                    "total_plazas": sum(g["_cantidades"].values()),
                    "plazas": filas_cuadro,
                    "detalle_plazas": g["plazas"],
                    "valuacion": valuacion,
                    "periodo_invalido": meses <= 0,
                })

            # Ya viene en orden (hoja del Anexo 2, fecha de alta): se construyó
            # iterando `grupos` ordenado así arriba — no hace falta reordenar
            # por UA resuelta, que ya no es lo que define el agrupamiento.

            return Response(
                {"grupos": salida, "avisos": avisos, "total_grupos": len(salida)},
                status=status.HTTP_200_OK,
            )
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RotacionTitularesAduanasView(APIView):
    """
    Línea de tiempo de titularidad de las 50 aduanas del país.

    Devuelve, por aduana, las gestiones de cada titular (entrada, salida, motivo,
    movimientos ocurridos durante su gestión) y los periodos en que la aduana
    quedó acéfala. Alimenta el subtab "Rotación de titulares de Aduanas" de
    MovimientosPersonalTab.

    El algoritmo vive en plantilla/rotacion_aduanas.py — ahí está documentado por
    qué la aduana se resuelve por `un_admin` contra el catálogo de su época y no
    por `desc_larga_un`. Esta vista solo hace las tres consultas y arma la
    respuesta.

    La derivación necesita la trayectoria COMPLETA de cada titular (no solo sus
    movimientos en los dos puestos de titularidad), porque la fecha de salida de
    una aduana es la del siguiente movimiento del empleado, que ya está fuera.

    ?refrescar=1 salta el caché.
    """

    view_permission = (
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
    )

    CACHE_KEY = "rotacion_titulares_aduanas_v1"
    # Sin expiración propia (`timeout=None` en cache.set más abajo) — se
    # borra solo cuando algo la invalida explícitamente: `?refrescar=1` de
    # esta misma vista, o /invalidar-cache/ (InvalidarCacheManualView,
    # invalidar_todo_el_cache_servidor barre TODAS las keys de caché del
    # servidor, esta incluida). Antes expiraba sola a los 30 min, lo que
    # servía datos viejos hasta esa ventana aun después de invalidar a mano.

    def get(self, request):
        from django.core.cache import cache

        from .rotacion_aduanas import PUESTOS_TITULARIDAD, construir_rotacion

        if request.query_params.get("refrescar") not in ("1", "true", "True"):
            cacheado = cache.get(self.CACHE_KEY)
            if cacheado is not None:
                return Response(cacheado, status=status.HTTP_200_OK)

        try:
            catalogo = self._catalogo_unidades()
            catalogo_puestos = {
                (cd or "").strip(): nombre
                for cd, nombre in CatPtoFunc.objects.exclude(
                    cd_pto_funcional__isnull=True
                ).values_list("cd_pto_funcional", "nombre_puesto_funcional")
            }
            movimientos = self._trayectorias_de_titulares()
            adscripcion = self._adscripcion_por_plaza(
                sorted({
                    m["posicion"] for m in movimientos
                    if m.get("posicion") and m.get("cd_puesto") in PUESTOS_TITULARIDAD
                })
            )

            resultado = construir_rotacion(
                movimientos,
                catalogo_vigente=catalogo,
                hoy=datetime.date.today(),
                adscripcion_plaza=adscripcion,
                catalogo_puestos=catalogo_puestos,
            )
            resultado = _serializar_fechas(resultado)

            cache.set(self.CACHE_KEY, resultado, timeout=None)
            return Response(resultado, status=status.HTTP_200_OK)
        except Exception:
            logger.exception("Error inesperado en {}".format(request.path))
            return Response(
                {"error": "Error interno del servidor"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _catalogo_unidades(self):
        """{cd_ua: unidad_administrativa} desde EMPLEADOS_COMPLETOS_SIG.

        Es el catálogo autoritativo: 65 códigos, de los cuales 101-150 son
        exactamente las 50 aduanas.
        """
        from django.db import connection

        from .rotacion_aduanas import normalizar_unidad

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT DISTINCT `Cd UA`, `Unidad Administrativa` "
                "FROM EMPLEADOS_COMPLETOS_SIG "
                "WHERE `Cd UA` IS NOT NULL AND `Unidad Administrativa` IS NOT NULL"
            )
            return {
                (cd or "").strip(): normalizar_unidad(unidad)
                for cd, unidad in cursor.fetchall()
                if (cd or "").strip() and normalizar_unidad(unidad)
            }

    def _trayectorias_de_titulares(self):
        """Trayectoria completa de todo empleado que alguna vez fue titular.

        Dos pasos deliberados: primero los num_empleado distintos de los dos
        puestos, después TODOS sus movimientos. Traer solo las filas de los dos
        puestos dejaría sin fecha de salida a quien se fue a otro puesto.
        """
        from django.db import connection

        from .rotacion_aduanas import PUESTOS_TITULARIDAD

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT DISTINCT num_empleado FROM cp_tbl_mov_completo_29_05_26 "
                "WHERE cd_puesto IN (%s, %s) AND num_empleado IS NOT NULL",
                list(PUESTOS_TITULARIDAD),
            )
            empleados = [row[0] for row in cursor.fetchall()]

            if not empleados:
                return []

            marcadores = ", ".join(["%s"] * len(empleados))
            cursor.execute(
                f"""
                SELECT
                    id, posicion, num_empleado,
                    nombre, ap_pat, ap_mat,
                    accion, accion_nombre,
                    motivo, motivo_nombre,
                    fecha_efectiva, sec, fecha_captura,
                    est_hr, estado_pago, partida_presup,
                    un, un_admin,
                    id_depto, depen_direc,
                    plan_sal, grado, escala,
                    puesto_ptal, nivel_tabular,
                    gp_pago, prog_benef, sal_base,
                    cd_puesto, ubicacion, id_estbl,
                    salida_prevista, fecha_ult_actz, por,
                    ult_inicio, fecha_inicial, gp_trabajo,
                    grupo_cd_sal, antiguo_empr,
                    rfc, curp, id_persona,
                    desc_larga_p, nv_jerarquico, desc_larga_un,
                    sexo, fecha_entrada, fecha_posicion
                FROM cp_tbl_mov_completo_29_05_26
                WHERE num_empleado IN ({marcadores})
                  AND fecha_efectiva IS NOT NULL
                ORDER BY num_empleado ASC, fecha_efectiva ASC, sec ASC, id ASC
                """,
                empleados,
            )
            columnas = [d[0] for d in cursor.description]
            return [dict(zip(columnas, fila)) for fila in cursor.fetchall()]

    def _adscripcion_por_plaza(self, plazas):
        """Devuelve un callable (posicion, fecha) -> (cd_ua, motivo) según MOV_POS.

        Sirve para detectar las filas de personal que conservaron el código de
        unidad viejo después de que su plaza ya había sido renumerada.
        """
        from django.db import connection

        if not plazas:
            return lambda posicion, fecha: None

        marcadores = ", ".join(["%s"] * len(plazas))
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT `Nº Pos Actual`, `F Efva`, `Unidad Adva#`, `Motivo`
                FROM MOV_POS
                WHERE `Nº Pos Actual` IN ({marcadores})
                """,
                list(plazas),
            )
            filas = cursor.fetchall()

        historia = {}
        for posicion, f_efva, cd_ua, motivo in filas:
            fecha = _parsear_fecha_mov_pos(f_efva)
            if fecha is None:
                continue
            historia.setdefault(posicion, []).append((fecha, (cd_ua or "").strip(), motivo or ""))
        for registros in historia.values():
            registros.sort(key=lambda r: r[0])

        def adscripcion(posicion, fecha):
            vigente = None
            for f_registro, cd_ua, motivo in historia.get(posicion, ()):
                if f_registro <= fecha:
                    vigente = (cd_ua, motivo)
                else:
                    break
            return vigente

        return adscripcion


def _parsear_fecha_mov_pos(valor):
    """MOV_POS guarda las fechas como varchar; conviven dos formatos."""
    if isinstance(valor, datetime.date):
        return valor
    if not valor:
        return None
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(valor.strip(), formato).date()
        except (ValueError, AttributeError):
            continue
    return None


def _serializar_fechas(valor):
    """date/datetime -> ISO, recursivo. DRF no lo hace dentro de dicts anidados."""
    if isinstance(valor, dict):
        return {k: _serializar_fechas(v) for k, v in valor.items()}
    if isinstance(valor, list):
        return [_serializar_fechas(v) for v in valor]
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    return valor


# =============================================================================
# Árbol de movimientos — subtab de MovimientosPersonalTab
# =============================================================================
# Dos vistas deliberadamente separadas porque el árbol se carga POR NODO, bajo
# demanda: cargarlo completo crecería de forma explosiva (cada plaza abre N
# empleados y cada empleado N plazas). El front pide el tronco al escribir la
# plaza y sólo pide la rama de un empleado cuando el usuario la expande.
#
# Todo el algoritmo vive en los SPs (ver PLAN_SP_HISTORIA_PLAZA_2026-09-03.md);
# estas vistas sólo llaman e hidratan la respuesta.


def _filas_de_sp(cursor):
    """Filas del cursor como lista de dicts con fechas ya en ISO."""
    columnas = [col[0] for col in cursor.description]
    return [_serializar_fechas(dict(zip(columnas, fila))) for fila in cursor.fetchall()]


class PlazaSugerenciasView(APIView):
    """Autocompletado del número de plaza mientras se escribe (input de
    "Árbol de movimientos", ver ArbolMovimientosSubTab.jsx).

    Filtra sobre EMPLEADOS_COMPLETOS_SIG (posición indexada, ~40 mil filas)
    en vez de cp_tbl_mov_completo_29_05_26 (histórico, sin índice sobre
    posición) — cubre las plazas vigentes en el dataset de ZAFIRO, que es
    el caso de uso real de este buscador.

    GET /plantilla/plazas/sugerencias/?q=<prefijo>

    Si la plaza está ocupada, incluye "ocupante" (nombre completo) para que
    se pueda distinguir sin tener que abrir el árbol.
    """

    view_permission = (
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
    )
    LIMITE_DEFAULT = 8
    LIMITE_MAXIMO = 20

    def get(self, request, *args, **kwargs):
        termino = (request.query_params.get("q") or "").strip()
        if len(termino) < 2:
            return Response([], status=status.HTTP_200_OK)

        try:
            limite = min(int(request.query_params.get("limit", self.LIMITE_DEFAULT)), self.LIMITE_MAXIMO)
        except (TypeError, ValueError):
            limite = self.LIMITE_DEFAULT

        filas = (
            EmpleadosCompletosSig.objects.filter(posicion__istartswith=termino)
            .exclude(posicion__isnull=True)
            .exclude(posicion="")
            .values("posicion", "nombre_puesto_funcional", "nombres")
            .order_by("posicion")
            .distinct()[:limite]
        )
        resultados = [
            {
                "posicion": f["posicion"],
                "puesto": f["nombre_puesto_funcional"],
                "ocupada": bool((f["nombres"] or "").strip()),
                "ocupante": (f["nombres"] or "").strip() or None,
            }
            for f in filas
        ]
        return Response(resultados, status=status.HTTP_200_OK)


class HistoriaPlazaView(APIView):
    """Pila cronológica continua de una plaza: creación, ocupaciones, vacancias,
    insubsistencias y movimientos de tránsito.

    Es el TRONCO del árbol de movimientos. Devuelve `sp_historia_plaza` tal cual,
    más un encabezado con el conteo de gestiones (sólo los periodos de tipo
    `ocupacion` cuentan: insubsistencias y tránsitos no).

    GET /plantilla/historia-plaza/<posicion>/
    """

    view_permission = (
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
    )

    def get(self, request, posicion, *args, **kwargs):
        posicion = (posicion or "").strip()
        if not posicion:
            return Response(
                {"detail": "La posición es obligatoria."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cache_key = f"historia_plaza_{posicion}"
        if request.query_params.get("refrescar") != "1":
            en_cache = cache.get(cache_key)
            if en_cache is not None:
                return Response(en_cache, status=status.HTTP_200_OK)

        with connection.cursor() as cursor:
            cursor.execute("CALL sp_historia_plaza(%s)", [posicion])
            periodos = _filas_de_sp(cursor)
            while cursor.nextset():
                pass

        if not periodos:
            return Response(
                {"detail": f"La posición {posicion} no existe o no tiene historia."},
                status=status.HTTP_404_NOT_FOUND,
            )

        actual = next((p for p in periodos if p.get("es_ocupante_actual")), None)
        datos = {
            "posicion": posicion,
            "ocupada": actual is not None,
            "ocupante_actual": actual.get("num_empleado") if actual else None,
            "num_gestiones": sum(1 for p in periodos if p.get("tipo_periodo") == "ocupacion"),
            "num_insubsistencias": sum(
                1 for p in periodos if p.get("tipo_periodo") == "insubsistencia"
            ),
            # El front lo usa para avisar que el dato de origen se contradice.
            "tiene_inconsistencias": any(p.get("inconsistente") for p in periodos),
            "periodos": periodos,
        }

        # 5 min: el SP tarda ~30ms, pero el usuario expande y colapsa ramas
        # repetidamente y no tiene sentido repetir la consulta cada vez.
        cache.set(cache_key, datos, timeout=300)
        return Response(datos, status=status.HTTP_200_OK)


class HistoriaEmpleadoView(APIView):
    """Tramos de un empleado en todas las plazas por las que pasó.

    Es una RAMA del árbol: se pide sólo cuando el usuario expande un empleado.
    No incluye periodos de vacancia — la vacancia es propiedad de la plaza, no
    del empleado, y la aporta el tronco (HistoriaPlazaView).

    GET /plantilla/historia-empleado/<num_empleado>/
    """

    view_permission = (
        "authentication.view_plantilla_mov_posiciones",
        "authentication.view_plantilla_movimientos",
    )

    def get(self, request, num_empleado, *args, **kwargs):
        num_empleado = (num_empleado or "").strip()
        if not num_empleado:
            return Response(
                {"detail": "El número de empleado es obligatorio."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        cache_key = f"historia_empleado_{num_empleado}"
        if request.query_params.get("refrescar") != "1":
            en_cache = cache.get(cache_key)
            if en_cache is not None:
                return Response(en_cache, status=status.HTTP_200_OK)

        with connection.cursor() as cursor:
            cursor.execute("CALL sp_historia_empleado(%s)", [num_empleado])
            tramos = _filas_de_sp(cursor)
            while cursor.nextset():
                pass

        if not tramos:
            return Response(
                {"detail": f"El empleado {num_empleado} no tiene movimientos registrados."},
                status=status.HTTP_404_NOT_FOUND,
            )

        nombre = ""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT MAX(TRIM(CONCAT(COALESCE(nombre,''), ' ',
                                       COALESCE(ap_pat,''), ' ',
                                       COALESCE(ap_mat,''))))
                FROM cp_tbl_mov_completo_29_05_26
                WHERE num_empleado = %s
                """,
                [num_empleado],
            )
            fila = cursor.fetchone()
            if fila and fila[0]:
                nombre = fila[0]

        datos = {
            "num_empleado": num_empleado,
            "nombre_completo": nombre,
            "num_plazas": len({t["posicion"] for t in tramos}),
            "num_tramos": len(tramos),
            "tramos": tramos,
        }
        cache.set(cache_key, datos, timeout=300)
        return Response(datos, status=status.HTTP_200_OK)
