"""Membretado institucional compartido por TODOS los generadores de Excel del
backend (ExportExcelView, MovPosExportExcelView, ExportarEstatusExcelView,
generar_excel_estatus_task). Contraparte de excelLetterhead.js en el front.

Diseño apilado (logo -> título -> leyenda) para no depender de cuántas
columnas tenga cada hoja.
"""
import os
from datetime import datetime

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

LOGO_PATH = os.path.join(os.path.dirname(__file__), "static", "branding", "hacienda_aduanas_logos.png")

TITLE_LINES = [
    "AGENCIA NACIONAL DE ADUANAS DE MÉXICO",
    "UNIDAD DE ADMINISTRACIÓN Y FINANZAS",
    "DIRECCIÓN DE RECURSOS HUMANOS",
]

LOGO_DISPLAY_WIDTH = 260

# Filas que ocupa el membretado: 1=logo, 2=título (3 líneas envueltas),
# 3=leyenda de generación, 4=separador en blanco. El contenido real de cada
# hoja debe arrancar en la fila LETTERHEAD_ROWS + 1.
LETTERHEAD_ROWS = 4


def _fecha_hora_generacion():
    ahora = datetime.now()
    dias = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
             "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    fecha = f"{ahora.day} de {meses[ahora.month - 1]} de {ahora.year}"
    hora = ahora.strftime("%H:%M:%S")
    return f"{hora} horas del {fecha}"


def add_excel_letterhead(ws, num_cols):
    """Inserta el membretado (logo Hacienda/ANAM + título + leyenda de
    generación) en las primeras filas de `ws`. Debe llamarse ANTES de escribir
    cualquier otro contenido — el contenido real debe empezar en la fila
    LETTERHEAD_ROWS + 1 (usar el valor devuelto, nunca un número mágico).

    Devuelve el número de filas ocupadas por el membretado.
    """
    num_cols = max(num_cols, 1)
    last_col = get_column_letter(num_cols)

    img = XLImage(LOGO_PATH)
    orig_w, orig_h = img.width, img.height
    display_h = round(LOGO_DISPLAY_WIDTH * orig_h / orig_w)
    img.width = LOGO_DISPLAY_WIDTH
    img.height = display_h
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = max(round(display_h * 0.85), 34)

    ws.merge_cells(f"A2:{last_col}2")
    title_cell = ws["A2"]
    title_cell.value = "\n".join(TITLE_LINES)
    title_cell.font = Font(name="Calibri", bold=True, size=11, color="FF621F32")
    title_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 54

    ws.merge_cells(f"A3:{last_col}3")
    legend_cell = ws["A3"]
    legend_cell.value = f"Reporte generado por el sistema de control de plazas a las {_fecha_hora_generacion()}."
    legend_cell.font = Font(name="Calibri", italic=True, size=9, color="FF64748B")
    legend_cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 8

    return LETTERHEAD_ROWS
