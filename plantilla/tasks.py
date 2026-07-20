"""
Tarea Celery para descargar e importar datos de ZAFIRO a la BD.

Flujo (tarea `importar_zafiro`), por cada uno de los 4 reportes
(Posiciones, Empleados Completos, Empleados Bajas, Historial Posición):
  1. Ejecuta el script Node.js index.js con el argIndex correspondiente
     para descargar el CSV.
  2. Corrige el CSV con el corrector heurístico (binario externo), si existe.
  3. Trunca la tabla *_STAGING correspondiente e inserta los registros con
     bulk_create (quitando antes los índices secundarios para acelerar la
     carga masiva, y recreándolos al terminar).

Tras importar los 4 reportes:
  4. Hace un swap atómico (RENAME TABLE) entre las tablas activas y las de
     staging (Blue-Green), dejando los datos nuevos en producción al
     instante, y trunca las tablas staging (que quedan con los datos viejos).
  5. Ejecuta stored procedures de post-proceso: llenar Nombre Puesto en
     MOV_POS, corregir SMB/SMN en EMPLEADOS_COMPLETOS_SIG, y calcular/
     actualizar fechas de vacancia.
  6. Regenera el Cuadro de Vacancia Diario (management command).
  7. Publica evento en Redis (canal "zafiro_updates") para notificar a
     clientes vía SSE, e invalida las cachés de dashboard/plantilla.

Si `es_historico` (ejecución después de las 23:00), además guarda copia de
cada lote importado en su tabla *_HISTORICO correspondiente.

Toda la ejecución queda registrada en vivo en ZafiroBitacora (campo
logs_en_vivo), incluyendo errores. Usa un lock distribuido en Redis para que
solo una instancia corra a la vez; reintenta hasta 2 veces si falla.

Configuración relevante en settings.py:
  - ZAFIRO_DOWNLOAD_DIR : directorio donde Node.js guarda los CSVs
  - ZAFIRO_SCRIPT_PATH  : ruta absoluta a index.js
  - CELERY_BEAT_SCHEDULE: horarios de ejecución (Cada 30 minutos)
"""

import csv
import os
import re
import subprocess
import sys
import time
from contextlib import contextmanager
from pathlib import Path

csv.field_size_limit(sys.maxsize)

from celery import shared_task
from celery.utils.log import get_task_logger
from django.conf import settings
from django.db import connection, transaction

from .models import (
    BajasSigHistorico,
    BajasSigStaging,
    CpTblMovCompleto290526Historico,
    CpTblMovCompleto290526Staging,
    EmpleadosCompletosSig,
    EmpleadosCompletosSigHistorico,
    EmpleadosCompletosSigStaging,
    MovPosHistorico,
    MovPosStaging,
    ZafiroBitacora,
)

logger = get_task_logger(__name__)

# ---------------------------------------------------------------------------
# Nombres de archivo que genera index.js según el argIndex
# ---------------------------------------------------------------------------
ZAFIRO_FILES = {
    6: "zafiro_info_Empleados_Completo.csv",  # argIndex=6 → Empleados Completos
    3: "zafiro_info_Empleados_Bajas.csv",  # argIndex=3 → Empleados Bajas
    1: "zafiro_info_Posiciones.csv",  # argIndex=1 → Posiciones
    0: "zafiro_info_Historial_Posición.csv",  # argIndex=0 → Historial Posición
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

from django.utils import timezone


def _append_log(bitacora, mensaje, is_error=False):
    if is_error:
        logger.error(mensaje)
    else:
        logger.info(mensaje)

    if not bitacora:
        return

    timestamp = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    prefix = "[ERROR]" if is_error else "[INFO]"
    linea = f"{timestamp} {prefix} {mensaje}\n"

    if bitacora.logs_en_vivo is None:
        bitacora.logs_en_vivo = linea
    else:
        bitacora.logs_en_vivo += linea
    bitacora.save(update_fields=["logs_en_vivo"])


def _ejecutar_script_node(
    arg_index: int, download_dir: str, script_path: str, bitacora=None
) -> str:
    """
    Lanza index.js con el argIndex dado y espera a que termine.
    Retorna la ruta completa del CSV descargado.

    Args:
        arg_index:    Número de checkbox (6=Completos, 3=Bajas)
        download_dir: Directorio de descarga
        script_path:  Ruta absoluta a index.js

    Returns:
        Ruta absoluta al CSV descargado.

    Raises:
        RuntimeError: Si el script falla o el CSV no aparece.
    """
    import shutil
    import tempfile

    csv_name = ZAFIRO_FILES[arg_index]
    final_csv_path = Path(download_dir) / csv_name

    if final_csv_path.exists():
        final_csv_path.unlink()
        _append_log(bitacora, f"CSV previo eliminado: {final_csv_path}")

    _append_log(bitacora, f"Ejecutando script Node.js: argIndex={arg_index}")

    with tempfile.TemporaryDirectory(dir=download_dir) as temp_dir:
        result = subprocess.run(
            ["node", script_path, str(arg_index), temp_dir],
            capture_output=True,
            text=True,
            timeout=1200,  # 20 minutos máximo
            cwd=str(Path(script_path).parent),
        )

        if result.returncode != 0:
            _append_log(
                bitacora,
                f"Script Node.js falló (argIndex={arg_index}):\n{result.stderr}",
                is_error=True,
            )
            raise RuntimeError(
                f"Script Node.js falló con código {result.returncode}: {result.stderr}"
            )

        _append_log(bitacora, f"Script Node.js completado (argIndex={arg_index})")

        # The node script saved it in temp_dir / csv_name
        temp_csv_path = Path(temp_dir) / csv_name
        if not temp_csv_path.exists():
            # If the node script saved it with a different name somehow, let's just grab the first CSV in temp_dir
            csvs = list(Path(temp_dir).glob("*.csv"))
            if csvs:
                temp_csv_path = csvs[0]
            else:
                raise RuntimeError(
                    f"El script terminó exitosamente pero no se encontró ningún CSV en {temp_dir}"
                )

        shutil.move(str(temp_csv_path), str(final_csv_path))

    return str(final_csv_path)


def _corregir_csv(csv_path: str, script_path: str, bitacora=None) -> str:
    """
    Ejecuta el corrector heurístico sobre el CSV dado.
    Si el corrector genera un archivo de salida corregido, retorna su ruta.
    De lo contrario, retorna la ruta del CSV original.
    """

    ext = ".exe" if os.name == "nt" else ""
    corrector_exe = Path(script_path).parent / f"corregir_heuristico{ext}"
    if not corrector_exe.exists():
        _append_log(
            bitacora,
            f"Corrector heurístico no encontrado en {corrector_exe}. Se usará el archivo original.",
        )
        return csv_path

    csv_path_obj = Path(csv_path)
    output_path = csv_path_obj.parent / f"{csv_path_obj.stem}_Corregido.csv"

    if output_path.exists():
        try:
            output_path.unlink()
        except Exception as e:
            _append_log(
                bitacora,
                f"No se pudo eliminar el archivo de salida previo {output_path}: {e}",
            )

    _append_log(bitacora, f"Ejecutando corrector heurístico para: {csv_path}")
    try:
        result = subprocess.run(
            [
                str(corrector_exe),
                csv_path,
                "--corregir",
                "--salida",
                str(output_path),
            ],
            capture_output=True,
            text=True,
            timeout=1200,  # 20 minutos máximo
            cwd=str(Path(script_path).parent),
        )
        if result.returncode == 0:
            if output_path.exists():
                _append_log(
                    bitacora,
                    f"Corrector completado. Se usará el archivo corregido: {output_path}",
                )
                return str(output_path)
            else:
                _append_log(
                    bitacora,
                    "Corrector completado (no se encontraron errores, se usará el original).",
                )
                return csv_path
        else:
            _append_log(
                bitacora,
                f"Corrector heurístico falló (código {result.returncode}):\n{result.stderr}",
                is_error=True,
            )
            return csv_path
    except Exception as e:
        _append_log(
            bitacora, f"Error al ejecutar corrector heurístico: {e}", is_error=True
        )
        return csv_path


def _capturar_indices_secundarios(table):
    """Captura la DDL de los índices secundarios (no PRIMARY) de la tabla desde
    SHOW CREATE TABLE. Drift-proof: recrea exactamente lo que existe, incluidos
    los índices funcionales (TRIM)."""
    with connection.cursor() as cursor:
        cursor.execute(f"SHOW CREATE TABLE `{table}`")
        ddl = cursor.fetchone()[1]
    indices = []
    for line in ddl.split("\n"):
        line = line.strip().rstrip(",")
        if line.startswith("PRIMARY KEY"):
            continue
        m = re.match(r"^(UNIQUE )?KEY `([^`]+)` (.+)$", line)
        if m:
            unique = "UNIQUE " if m.group(1) else ""
            name = m.group(2)
            cols = m.group(3)
            indices.append((name, f"CREATE {unique}INDEX `{name}` ON `{table}` {cols}"))
    return indices


@contextmanager
def _staging_sin_indices(table, bitacora=None):
    """Quita los índices secundarios de una tabla antes de una carga masiva y los
    recrea al terminar (en ``finally``). Insertar sin índices y construir el
    índice de una pasada es más rápido que mantenerlo fila por fila. El DDL va
    fuera de cualquier transacción (MySQL auto-commitea el DDL)."""
    indices = _capturar_indices_secundarios(table)
    if indices:
        with connection.cursor() as cursor:
            for name, _ in indices:
                cursor.execute(f"ALTER TABLE `{table}` DROP INDEX `{name}`")
        _append_log(
            bitacora, f"{table}: {len(indices)} índices quitados para la carga."
        )
    try:
        yield
    finally:
        if indices:
            with connection.cursor() as cursor:
                for _, ddl in indices:
                    cursor.execute(ddl)
            _append_log(bitacora, f"{table}: {len(indices)} índices recreados.")


def _truncar_tabla(model_class, bitacora=None):
    """
    Trunca la tabla del modelo dado usando SQL directo (más rápido que .all().delete()).
    Esto es seguro porque Celery recarga la tabla completa en cada ejecución.
    """
    table = model_class._meta.db_table
    with connection.cursor() as cursor:
        cursor.execute(f"TRUNCATE TABLE `{table}`")
    _append_log(bitacora, f"Tabla truncada: {table}")


def _preparar_columnas_csv(fieldnames) -> list:
    """
    Limpia espacios en blanco de los encabezados del CSV y renombra
    columnas duplicadas agregando un sufijo numérico (ej. NIVEL -> NIVEL1).
    """
    seen = {}
    new_fieldnames = []
    for name in fieldnames:
        name_stripped = name.strip()
        if not name_stripped:
            new_fieldnames.append("")
            continue
        if name_stripped in seen:
            seen[name_stripped] += 1
            new_fieldnames.append(f"{name_stripped}{seen[name_stripped]}")
        else:
            seen[name_stripped] = 0
            new_fieldnames.append(name_stripped)
    return new_fieldnames


def _importar_csv_empleados_completos(
    csv_path: str, guardar_historico: bool = False, bitacora=None
) -> int:
    """
    Lee el CSV de Empleados Completos y hace bulk_create en EMPLEADOS_COMPLETOS_SIG.
    Retorna el número de registros insertados.
    """
    registros = []
    registros = []
    registros_historico = []
    with open(csv_path, encoding="cp1252", newline="") as f:
        reader = csv.DictReader(f, delimiter="|")
        reader.fieldnames = _preparar_columnas_csv(reader.fieldnames)
        for row in reader:
            registros.append(
                EmpleadosCompletosSigStaging(
                    id_field=row.get("Id") or None,
                    numeral=row.get("numeral") or None,
                    ua=row.get("ua") or None,
                    cent=row.get("cent") or None,
                    dir=row.get("dir") or None,
                    subd=row.get("subd") or None,
                    jd=row.get("jd") or None,
                    depto=row.get("depto") or None,
                    aduana=row.get("Aduana") or None,
                    id_tipo=row.get("id tipo") or None,
                    tipo=row.get("tipo") or None,
                    estado=row.get("estado") or None,
                    municipio=row.get("municipio") or None,
                    latitud=row.get("latitud") or None,
                    longitud=row.get("longitud") or None,
                    ua2=row.get("ua2") or None,
                    posicion=row.get("Posición") or None,
                    estado_nomina=row.get("Estado Nómina") or None,
                    id_empleado=row.get("Id Empleado") or None,
                    rfc=row.get("RFC") or None,
                    curp=row.get("CURP") or None,
                    nombres=row.get("Nombres") or None,
                    motivo=row.get("Motivo") or None,
                    fecha_efectiva_personal=row.get("Fecha efectiva (Personal)")
                    or None,
                    fecha_de_captura=row.get("Fecha de captura") or None,
                    qna=row.get("Qna") or None,
                    fecha_prevista_de_salida=row.get("Fecha prevista de salida")
                    or None,
                    nj=row.get("NJ") or None,
                    codigo_presupuestal=row.get("Código Presupuestal") or None,
                    nivel=row.get("Nivel") or None,
                    escala=row.get("Escala") or None,
                    smb=row.get("SMB") or None,
                    smn=row.get("SMN") or None,
                    partida=row.get("Partida") or None,
                    tipo_de_contratacion=row.get("TIPO DE CONTRATACIÓN") or None,
                    cd_un=row.get("Cd UN") or None,
                    unidad_de_negocio=row.get("Unidad de Negocio") or None,
                    cd_ua=row.get("Cd UA") or None,
                    unidad_administrativa=row.get("Unidad Administrativa") or None,
                    cd_pto_funcional=row.get("Cd Pto Funcional") or None,
                    nombre_puesto_funcional=row.get("Nombre Puesto Funcional") or None,
                    id_departamento=row.get("Id Departamento") or None,
                    departamento=row.get("Departamento") or None,
                    dependencia_directa=row.get("DependenciaDirecta") or None,
                    observaciones=row.get("OBSERVACIONES") or None,
                    ubicacion=row.get("Ubicación") or None,
                    descripcion_ubicacion=row.get("Descripción ubicación") or None,
                    posicion_civil_sedena_semar=row.get(
                        "Posición _Civil / SEDENA / SEMAR"
                    )
                    or None,
                    personal_militar_o_civil=row.get("Personal Militar o Civil")
                    or None,
                    tipo_de_personal_sedena_semar=row.get(
                        "Tipo de personal SEDENA / SEMAR"
                    )
                    or None,
                    rango=row.get("Rango") or None,
                    fecha_de_ingreso=row.get("Fecha de ingreso") or None,
                    val_estat=row.get("Val_estat") or None,
                    status_jefe_inm_posicion=row.get("Status Jefe Inm Posición")
                    or None,
                    numempleado=row.get("Numempleado") or None,
                    sindicato=row.get("Sindicato") or None,
                    entidad_federativa=row.get("Entidad Federativa") or None,
                    tipo_de_aduana=row.get("Tipo de Aduana") or None,
                    dg_o_aduana_compactada=row.get("DG o Aduana compactada") or None,
                    proyecto_2024_reduccion_plazas_eventuales=row.get(
                        "Proyecto 2024 Reducción de plazas Eventuales"
                    )
                    or None,
                    estado_en_nomina=row.get("Estado en nomina") or None,
                    ua_validacion=row.get("UA Validación") or None,
                    validando_posicion_por_documento=row.get(
                        "Validando de posición por documento"
                    )
                    or None,
                    val_estatx=row.get("Val_estatx") or None,
                    nj_comp=row.get("NJ COMP") or None,
                    nj_ok=row.get("NJ OK") or None,
                    columna=row.get("Columna") or None,
                    nombre_nj=row.get("nombreNJ") or None,
                    nj_operativo_comb=row.get("NJOperativoComb") or None,
                )
            )
            if guardar_historico:
                registros_historico.append(
                    EmpleadosCompletosSigHistorico(
                        id_field=row.get("Id") or None,
                        numeral=row.get("numeral") or None,
                        ua=row.get("ua") or None,
                        cent=row.get("cent") or None,
                        dir=row.get("dir") or None,
                        subd=row.get("subd") or None,
                        jd=row.get("jd") or None,
                        depto=row.get("depto") or None,
                        aduana=row.get("Aduana") or None,
                        id_tipo=row.get("id tipo") or None,
                        tipo=row.get("tipo") or None,
                        estado=row.get("estado") or None,
                        municipio=row.get("municipio") or None,
                        latitud=row.get("latitud") or None,
                        longitud=row.get("longitud") or None,
                        ua2=row.get("ua2") or None,
                        posicion=row.get("Posición") or None,
                        estado_nomina=row.get("Estado Nómina") or None,
                        id_empleado=row.get("Id Empleado") or None,
                        rfc=row.get("RFC") or None,
                        curp=row.get("CURP") or None,
                        nombres=row.get("Nombres") or None,
                        motivo=row.get("Motivo") or None,
                        fecha_efectiva_personal=row.get("Fecha efectiva (Personal)")
                        or None,
                        fecha_de_captura=row.get("Fecha de captura") or None,
                        qna=row.get("Qna") or None,
                        fecha_prevista_de_salida=row.get("Fecha prevista de salida")
                        or None,
                        nj=row.get("NJ") or None,
                        codigo_presupuestal=row.get("Código Presupuestal") or None,
                        nivel=row.get("Nivel") or None,
                        escala=row.get("Escala") or None,
                        smb=row.get("SMB") or None,
                        smn=row.get("SMN") or None,
                        partida=row.get("Partida") or None,
                        tipo_de_contratacion=row.get("TIPO DE CONTRATACIÓN") or None,
                        cd_un=row.get("Cd UN") or None,
                        unidad_de_negocio=row.get("Unidad de Negocio") or None,
                        cd_ua=row.get("Cd UA") or None,
                        unidad_administrativa=row.get("Unidad Administrativa") or None,
                        cd_pto_funcional=row.get("Cd Pto Funcional") or None,
                        nombre_puesto_funcional=row.get("Nombre Puesto Funcional")
                        or None,
                        id_departamento=row.get("Id Departamento") or None,
                        departamento=row.get("Departamento") or None,
                        dependencia_directa=row.get("DependenciaDirecta") or None,
                        observaciones=row.get("OBSERVACIONES") or None,
                        ubicacion=row.get("Ubicación") or None,
                        descripcion_ubicacion=row.get("Descripción ubicación") or None,
                        posicion_civil_sedena_semar=row.get(
                            "Posición _Civil / SEDENA / SEMAR"
                        )
                        or None,
                        personal_militar_o_civil=row.get("Personal Militar o Civil")
                        or None,
                        tipo_de_personal_sedena_semar=row.get(
                            "Tipo de personal SEDENA / SEMAR"
                        )
                        or None,
                        rango=row.get("Rango") or None,
                        fecha_de_ingreso=row.get("Fecha de ingreso") or None,
                        val_estat=row.get("Val_estat") or None,
                        status_jefe_inm_posicion=row.get("Status Jefe Inm Posición")
                        or None,
                        numempleado=row.get("Numempleado") or None,
                        sindicato=row.get("Sindicato") or None,
                        entidad_federativa=row.get("Entidad Federativa") or None,
                        tipo_de_aduana=row.get("Tipo de Aduana") or None,
                        dg_o_aduana_compactada=row.get("DG o Aduana compactada")
                        or None,
                        proyecto_2024_reduccion_plazas_eventuales=row.get(
                            "Proyecto 2024 Reducción de plazas Eventuales"
                        )
                        or None,
                        estado_en_nomina=row.get("Estado en nomina") or None,
                        ua_validacion=row.get("UA Validación") or None,
                        validando_posicion_por_documento=row.get(
                            "Validando de posición por documento"
                        )
                        or None,
                        val_estatx=row.get("Val_estatx") or None,
                        nj_comp=row.get("NJ COMP") or None,
                        nj_ok=row.get("NJ OK") or None,
                        columna=row.get("Columna") or None,
                        nombre_nj=row.get("nombreNJ") or None,
                        nj_operativo_comb=row.get("NJOperativoComb") or None,
                    )
                )

    with (
        _staging_sin_indices(EmpleadosCompletosSigStaging._meta.db_table, bitacora),
        transaction.atomic(),
    ):
        _truncar_tabla(EmpleadosCompletosSigStaging, bitacora)
        EmpleadosCompletosSigStaging.objects.bulk_create(registros, batch_size=1000)
        if guardar_historico:
            EmpleadosCompletosSigHistorico.objects.bulk_create(
                registros_historico, batch_size=1000
            )

    _append_log(
        bitacora,
        f"EmpleadosCompletosSigStaging: {len(registros)} registros insertados en staging.",
    )
    return len(registros)


def _importar_csv_bajas(
    csv_path: str, guardar_historico: bool = False, bitacora=None
) -> int:
    """
    Lee el CSV de Empleados Bajas y hace bulk_create en BAJAS_SIG.
    Retorna el número de registros insertados.
    """
    registros = []
    registros = []
    registros_historico = []
    with open(csv_path, encoding="cp1252", newline="") as f:
        reader = csv.DictReader(f, delimiter="|")
        reader.fieldnames = _preparar_columnas_csv(reader.fieldnames)
        for row in reader:
            registros.append(
                BajasSigStaging(
                    posicion=row.get("POSICION") or None,
                    no_empleado=row.get("NO_EMPLEADO") or None,
                    nombre_completo=row.get("NOMBRE_COMPLETO") or None,
                    primer_apellido=row.get("PRIMER_APELLIDO") or None,
                    segundo_apellido=row.get("SEGUNDO_APELLIDO") or None,
                    accion=row.get("ACCION") or None,
                    accion_descr=row.get("ACCION_DESCR") or None,
                    motivo=row.get("MOTIVO") or None,
                    motivo_descr=row.get("MOTIVO_DESCR") or None,
                    fecha_efectiva=row.get("FECHA_EFECTIVA") or None,
                    sequencia_efectiva=row.get("SEQUENCIA_EFECTIVA") or None,
                    fecha_aplicacion=row.get("FECHA_APLICACION") or None,
                    humanos_status=row.get("HUMANOS_STATUS") or None,
                    nomina_status=row.get("NOMINA_STATUS") or None,
                    partida=row.get("PARTIDA") or None,
                    unidad_general=row.get("UNIDAD_GENERAL") or None,
                    unidad_admon=row.get("UNIDAD_ADMON") or None,
                    departamento=row.get("DEPARTAMENTO") or None,
                    dependencia_directa=row.get("DEPENDENCIA_DIRECTA") or None,
                    plan_salarial=row.get("PLAN_SALARIAL") or None,
                    grado=row.get("GRADO") or None,
                    escala=row.get("ESCALA") or None,
                    puesto_presupuestal=row.get("PUESTO_PRESUPUESTAL") or None,
                    nivel_tabular=row.get("NIVEL_TABULAR") or None,
                    grupo_de_pago=row.get("GRUPO_DE_PAGO") or None,
                    beneficios=row.get("BENEFICIOS") or None,
                    smb=row.get("SMB") or None,
                    puesto=row.get("PUESTO") or None,
                    ubicacion=row.get("UBICACION") or None,
                    inmueble=row.get("INMUEBLE") or None,
                    fecha_prevista=row.get("FECHA_PREVISTA") or None,
                    ultima_actualizacion=row.get("ULTIMA_ACTUALIZACION") or None,
                    ultimo_operador=row.get("ULTIMO_OPERADOR") or None,
                    ultima_fecha_ingreso=row.get("ULTIMA_FECHA_INGRESO") or None,
                    fecha_ingreso=row.get("FECHA_INGRESO") or None,
                    grupo_trabajo=row.get("GRUPO_TRABAJO") or None,
                    codigo_grupo=row.get("CODIGO_GRUPO") or None,
                    fecha_asignacion=row.get("FECHA_ASIGNACION") or None,
                    rfc=row.get("RFC") or None,
                    curp=row.get("CURP") or None,
                    id_persona=row.get("ID_PERSONA") or None,
                    nivel=row.get("NIVEL") or None,
                    nivel1=row.get("NIVEL1") or None,
                    unidad_administrativa=row.get("UNIDAD_ADMINISTRATIVA") or None,
                    genero=row.get("GENERO") or None,
                    fecha_entrada_posicion=row.get("FECHA_ENTRADA_POSICION") or None,
                    fecha_posicion=row.get("FECHA_POSICION") or None,
                )
            )

    with (
        _staging_sin_indices(BajasSigStaging._meta.db_table, bitacora),
        transaction.atomic(),
    ):
        _truncar_tabla(BajasSigStaging, bitacora)
        BajasSigStaging.objects.bulk_create(registros, batch_size=1000)
        if guardar_historico:
            BajasSigHistorico.objects.bulk_create(registros_historico, batch_size=1000)

    _append_log(
        bitacora, f"BajasSigStaging: {len(registros)} registros insertados en staging."
    )
    return len(registros)


def _importar_csv_posiciones(
    csv_path: str, guardar_historico: bool = False, bitacora=None
) -> int:
    """
    Lee el CSV de Posiciones y hace bulk_create en MovPos.
    Retorna el número de registros insertados.
    """
    registros = []
    registros = []
    registros_historico = []
    with open(csv_path, encoding="cp1252", newline="") as f:
        reader = csv.DictReader(f, delimiter="|")
        reader.fieldnames = _preparar_columnas_csv(reader.fieldnames)
        for row in reader:
            registros.append(
                MovPosStaging(
                    no_pos_actual=row.get("Nº Pos Actual") or None,
                    f_efva=row.get("F Efva") or None,
                    estado_psn=row.get("Estado Psn") or None,
                    fecha_captura=row.get("Fecha Captura") or None,
                    cd_motivo=row.get("Cd Motivo") or None,
                    motivo=row.get("Motivo") or None,
                    cd_un=row.get("Cd UN") or None,
                    unidad_de_negocio=row.get("Unidad de Negocio") or None,
                    unidad_adva=row.get("Unidad Adva.") or None,
                    cd_departamento=row.get("Cd Departamento") or None,
                    cd_puesto=row.get("Cd Puesto") or None,
                    estado_ptal=row.get("Estado Ptal") or None,
                    fecha_est=row.get("Fecha Est") or None,
                    maximo=row.get("Máximo") or None,
                    depnd_drt=row.get("Depnd Drt") or None,
                    depnd_indrt=row.get("Depnd Indrt") or None,
                    ubicacion=row.get("Ubicación") or None,
                    nvl_direc=row.get("Nvl Direc") or None,
                    plan_sal=row.get("Plan Sal") or None,
                    grado=row.get("Grado") or None,
                    esc=row.get("Esc") or None,
                    puesto_ptal=row.get("Puesto Ptal") or None,
                    partida_ptal=row.get("Partida Ptal") or None,
                    gp_pago=row.get("Gp Pago") or None,
                    prog_beneficios=row.get("Prog Beneficios") or None,
                    fh_ult_actz=row.get("F/H Últ Actz") or None,
                    por=row.get("Por") or None,
                    hr_estd_semn=row.get("Hr Estd/Semn") or None,
                    descr=row.get("Descr") or None,
                    gp_trabajo=row.get("Gp Trabajo") or None,
                    org_code=row.get("Org Code") or None,
                    grupo_cd_sal=row.get("Grupo Cd Sal") or None,
                    formal_desc=row.get("FormalDesc") or None,
                    pto_compt=row.get("Pto Compt") or None,
                    posn_clv=row.get("Posn Clv") or None,
                    presupuesto=row.get("Presupuesto") or None,
                    nombre_puesto=row.get("Nombre Puesto") or None,
                )
            )
            if guardar_historico:
                registros_historico.append(
                    MovPosHistorico(
                        no_pos_actual=row.get("Nº Pos Actual") or None,
                        f_efva=row.get("F Efva") or None,
                        estado_psn=row.get("Estado Psn") or None,
                        fecha_captura=row.get("Fecha Captura") or None,
                        cd_motivo=row.get("Cd Motivo") or None,
                        motivo=row.get("Motivo") or None,
                        cd_un=row.get("Cd UN") or None,
                        unidad_de_negocio=row.get("Unidad de Negocio") or None,
                        unidad_adva=row.get("Unidad Adva.") or None,
                        cd_departamento=row.get("Cd Departamento") or None,
                        cd_puesto=row.get("Cd Puesto") or None,
                        estado_ptal=row.get("Estado Ptal") or None,
                        fecha_est=row.get("Fecha Est") or None,
                        maximo=row.get("Máximo") or None,
                        depnd_drt=row.get("Depnd Drt") or None,
                        depnd_indrt=row.get("Depnd Indrt") or None,
                        ubicacion=row.get("Ubicación") or None,
                        nvl_direc=row.get("Nvl Direc") or None,
                        plan_sal=row.get("Plan Sal") or None,
                        grado=row.get("Grado") or None,
                        esc=row.get("Esc") or None,
                        puesto_ptal=row.get("Puesto Ptal") or None,
                        partida_ptal=row.get("Partida Ptal") or None,
                        gp_pago=row.get("Gp Pago") or None,
                        prog_beneficios=row.get("Prog Beneficios") or None,
                        fh_ult_actz=row.get("F/H Últ Actz") or None,
                        por=row.get("Por") or None,
                        hr_estd_semn=row.get("Hr Estd/Semn") or None,
                        descr=row.get("Descr") or None,
                        gp_trabajo=row.get("Gp Trabajo") or None,
                        org_code=row.get("Org Code") or None,
                        grupo_cd_sal=row.get("Grupo Cd Sal") or None,
                        formal_desc=row.get("FormalDesc") or None,
                        pto_compt=row.get("Pto Compt") or None,
                        posn_clv=row.get("Posn Clv") or None,
                        presupuesto=row.get("Presupuesto") or None,
                        nombre_puesto=row.get("Nombre Puesto") or None,
                    )
                )

    with (
        _staging_sin_indices(MovPosStaging._meta.db_table, bitacora),
        transaction.atomic(),
    ):
        _truncar_tabla(MovPosStaging, bitacora)
        MovPosStaging.objects.bulk_create(registros, batch_size=1000)
        if guardar_historico:
            MovPosHistorico.objects.bulk_create(registros_historico, batch_size=1000)

    _append_log(
        bitacora, f"MovPosStaging: {len(registros)} registros insertados en staging."
    )
    return len(registros)


def _importar_csv_historial_posicion(
    csv_path: str, guardar_historico: bool = False, bitacora=None
) -> int:
    """
    Lee el CSV de Historial Posición.
    Usa el sistema Staging: Copia los datos < 2026 de producción, y luego inserta el nuevo CSV (que trae 2026).
    """
    registros = []
    registros_historico = []

    def clean_val(v, max_len=None):
        if v is None:
            return None
        v = str(v).strip()
        if v == "" or v == "(vacío)":
            return None
        if max_len:
            return v[:max_len]
        return v

    with open(csv_path, encoding="cp1252", newline="") as f:
        reader = csv.reader(f, delimiter="|")
        headers = next(reader)

        for row in reader:
            if not row or not row[0]:
                continue

            def get_val(idx, max_len=None):
                if idx < len(row):
                    return clean_val(row[idx], max_len)
                return None

            def get_date(idx):
                val = get_val(idx)
                if not val:
                    return None
                try:
                    from datetime import datetime

                    if "/" in val:
                        return datetime.strptime(val[:10], "%d/%m/%Y").strftime(
                            "%Y-%m-%d"
                        )
                    return val[:10]
                except:
                    return None

            def get_datetime(idx):
                val = get_val(idx)
                if not val:
                    return None
                # Ejemplo: '2025-12-04-12.38.35.000000'
                if len(val) >= 19 and val[10] == "-":
                    date_part = val[:10]
                    time_part = val[11:19].replace(".", ":")
                    try:
                        from datetime import datetime

                        from django.utils.timezone import make_aware

                        dt_naive = datetime.strptime(
                            f"{date_part} {time_part}", "%Y-%m-%d %H:%M:%S"
                        )
                        return make_aware(dt_naive)
                    except:
                        return f"{date_part} {time_part}"
                return get_date(idx)

            kwargs = dict(
                posicion=get_val(0, 50),
                num_empleado=get_val(1, 20),
                columna_c=get_val(2, 100),
                columna_d=get_val(3, 100),
                nombre=get_val(4, 100),
                ap_pat=get_val(5, 100),
                ap_mat=get_val(6, 100),
                accion=get_val(7, 50),
                accion_nombre=get_val(8, 100),
                motivo=get_val(9, 50),
                motivo_nombre=get_val(10, 100),
                fecha_efectiva=get_date(11),
                sec=get_val(12),
                fecha_captura=get_date(13),
                est_hr=get_val(14, 50),
                estado_pago=get_val(15, 50),
                partida_presup=get_val(16, 50),
                un=get_val(17, 50),
                un_admin=get_val(18, 100),
                id_depto=get_val(19, 50),
                depen_direc=get_val(20, 100),
                plan_sal=get_val(21, 50),
                grado=get_val(22, 50),
                escala=get_val(23, 50),
                puesto_ptal=get_val(24, 100),
                nivel_tabular=get_val(25, 50),
                gp_pago=get_val(26, 50),
                prog_benef=get_val(27, 100),
                sal_base=get_val(28),
                cd_puesto=get_val(29, 50),
                ubicacion=get_val(30, 100),
                id_estbl=get_val(31, 50),
                salida_prevista=get_date(32),
                fecha_ult_actz=get_datetime(33),
                por=get_val(34, 50),
                ult_inicio=get_date(35),
                fecha_inicial=get_date(36),
                gp_trabajo=get_val(37, 100),
                grupo_cd_sal=get_val(38, 50),
                antiguo_empr=get_val(39),
                rfc=get_val(40, 13),
                curp=get_val(41, 18),
                id_persona=get_val(42, 50),
                desc_larga_p=get_val(43),
                nv_jerarquico=get_val(44, 50),
                desc_larga_un=get_val(45),
                sexo=get_val(46, 20),
                fecha_entrada=get_date(47),
                fecha_posicion=get_date(48),
            )

            if kwargs["sal_base"]:
                try:
                    kwargs["sal_base"] = float(kwargs["sal_base"].replace(",", ""))
                except:
                    kwargs["sal_base"] = None

            if kwargs["sec"]:
                try:
                    kwargs["sec"] = int(float(kwargs["sec"]))
                except:
                    kwargs["sec"] = None

            if kwargs["antiguo_empr"]:
                try:
                    kwargs["antiguo_empr"] = int(float(kwargs["antiguo_empr"]))
                except:
                    kwargs["antiguo_empr"] = None

            registros.append(CpTblMovCompleto290526Staging(**kwargs))
            if guardar_historico:
                registros_historico.append(CpTblMovCompleto290526Historico(**kwargs))

    with (
        _staging_sin_indices(CpTblMovCompleto290526Staging._meta.db_table, bitacora),
        transaction.atomic(),
    ):
        _truncar_tabla(CpTblMovCompleto290526Staging, bitacora)

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO cp_tbl_mov_completo_29_05_26_staging 
                (id, posicion, num_empleado, columna_C, columna_D, nombre, ap_pat, ap_mat, 
                accion, accion_nombre, motivo, motivo_nombre, fecha_efectiva, sec, 
                fecha_captura, est_hr, estado_pago, partida_presup, un, un_admin, 
                id_depto, depen_direc, plan_sal, grado, escala, puesto_ptal, nivel_tabular, 
                gp_pago, prog_benef, sal_base, cd_puesto, ubicacion, id_estbl, 
                salida_prevista, fecha_ult_actz, por, ult_inicio, fecha_inicial, 
                gp_trabajo, grupo_cd_sal, antiguo_empr, rfc, curp, id_persona, 
                desc_larga_p, nv_jerarquico, desc_larga_un, sexo, fecha_entrada, fecha_posicion)
                
                SELECT id, posicion, num_empleado, columna_C, columna_D, nombre, ap_pat, ap_mat, 
                accion, accion_nombre, motivo, motivo_nombre, fecha_efectiva, sec, 
                fecha_captura, est_hr, estado_pago, partida_presup, un, un_admin, 
                id_depto, depen_direc, plan_sal, grado, escala, puesto_ptal, nivel_tabular, 
                gp_pago, prog_benef, sal_base, cd_puesto, ubicacion, id_estbl, 
                salida_prevista, fecha_ult_actz, por, ult_inicio, fecha_inicial, 
                gp_trabajo, grupo_cd_sal, antiguo_empr, rfc, curp, id_persona, 
                desc_larga_p, nv_jerarquico, desc_larga_un, sexo, fecha_entrada, fecha_posicion
                FROM cp_tbl_mov_completo_29_05_26
                WHERE YEAR(fecha_efectiva) != 2026 OR fecha_efectiva IS NULL
            """)

        CpTblMovCompleto290526Staging.objects.bulk_create(registros, batch_size=1000)

        if guardar_historico:
            CpTblMovCompleto290526Historico.objects.bulk_create(
                registros_historico, batch_size=1000
            )

    _append_log(
        bitacora,
        f"CpTblMovCompleto290526Staging: {len(registros)} registros (del año 2026) insertados en staging.",
    )
    return len(registros)


def _swap_blue_green_tables(bitacora=None):
    """
    Realiza un swap atómico de las tablas activas y las de staging en MySQL.
    Luego trunca las tablas de staging (que ahora contienen los datos viejos).
    """
    _append_log(bitacora, "Iniciando swap atómico de tablas (Blue-Green)...")
    sql = """
        RENAME TABLE 
            MOV_POS TO MOV_POS_TEMP, 
            MOV_POS_STAGING TO MOV_POS, 
            MOV_POS_TEMP TO MOV_POS_STAGING,
            EMPLEADOS_COMPLETOS_SIG TO EMPLEADOS_COMPLETOS_SIG_TEMP,
            EMPLEADOS_COMPLETOS_SIG_STAGING TO EMPLEADOS_COMPLETOS_SIG,
            EMPLEADOS_COMPLETOS_SIG_TEMP TO EMPLEADOS_COMPLETOS_SIG_STAGING,
            BAJAS_SIG TO BAJAS_SIG_TEMP,
            BAJAS_SIG_STAGING TO BAJAS_SIG,
            BAJAS_SIG_TEMP TO BAJAS_SIG_STAGING,
            cp_tbl_mov_completo_29_05_26 TO cp_tbl_mov_completo_29_05_26_temp,
            cp_tbl_mov_completo_29_05_26_staging TO cp_tbl_mov_completo_29_05_26,
            cp_tbl_mov_completo_29_05_26_temp TO cp_tbl_mov_completo_29_05_26_staging;
    """
    with connection.cursor() as cursor:
        cursor.execute(sql)
    _append_log(
        bitacora,
        "Swap atómico completado con éxito. Tablas en producción actualizadas.",
    )

    # Limpieza: Truncamos las tablas staging (que ahora contienen los datos viejos)
    _truncar_tabla(MovPosStaging, bitacora)
    _truncar_tabla(EmpleadosCompletosSigStaging, bitacora)
    _truncar_tabla(BajasSigStaging, bitacora)
    _truncar_tabla(CpTblMovCompleto290526Staging, bitacora)


def _materializar_mov_pos_latest(bitacora=None):
    """
    Reconstruye MOV_POS_LATEST: una fila por `Nº Pos Actual` con el registro
    más reciente de MOV_POS. Antes esta window function (ROW_NUMBER() OVER)
    se recalculaba en cada request en 6+ endpoints (~300ms c/u, ver
    AUDITORIA_BUGS_BACK.md BE2); ahora se paga una sola vez aquí, por import.
    """
    _append_log(bitacora, "Materializando MOV_POS_LATEST...")
    with connection.cursor() as cursor:
        cursor.execute("TRUNCATE TABLE MOV_POS_LATEST")
        cursor.execute("""
            INSERT INTO MOV_POS_LATEST (`Nº Pos Actual`, id, `Estado Psn`)
            SELECT `Nº Pos Actual`, id, `Estado Psn`
            FROM (
                SELECT `Nº Pos Actual`, id, `Estado Psn`, ROW_NUMBER() OVER (
                    PARTITION BY `Nº Pos Actual`
                    ORDER BY `F Efva` DESC, `Fecha Captura` DESC, `F/H Últ Actz` DESC, id DESC
                ) as rn
                FROM MOV_POS
            ) ranked
            WHERE rn = 1
        """)
    _append_log(bitacora, "MOV_POS_LATEST reconstruido.")


def _llenar_nombre_puesto(bitacora):
    """
    Ejecuta el Stored Procedure sp_llenar_nombre_puesto, el cual llena
    la columna `Nombre Puesto` de MOV_POS haciendo match contra el
    catálogo CAT_PTO_FUNC (por `Cd Puesto` = `Cd Pto Funcional`).
    """
    _append_log(
        bitacora, "Llenando Nombre Puesto en MOV_POS desde catálogo CAT_PTO_FUNC..."
    )
    try:
        with connection.cursor() as cursor:
            cursor.execute("CALL sp_llenar_nombre_puesto();")
            _append_log(bitacora, "Nombre Puesto actualizado exitosamente.")
    except Exception as e:
        _append_log(bitacora, f"Error llenando Nombre Puesto: {str(e)}")
        logger.error(f"Error en _llenar_nombre_puesto: {str(e)}", exc_info=True)


def _corregir_smb_smn_empleados(bitacora):
    """
    Ejecuta el Stored Procedure sp_corregir_smb_smn_empleados, el cual
    corrige SMB y SMN de EMPLEADOS_COMPLETOS_SIG con base en el catálogo
    rc_cat_cod_presupuestal (match por Código Presupuestal, Nivel y Escala).
    """
    _append_log(
        bitacora,
        "Corrigiendo SMB y SMN en EMPLEADOS_COMPLETOS_SIG desde catálogo rc_cat_cod_presupuestal...",
    )
    try:
        with connection.cursor() as cursor:
            cursor.execute("CALL sp_corregir_smb_smn_empleados();")
            _append_log(bitacora, "SMB y SMN corregidos exitosamente.")
    except Exception as e:
        _append_log(bitacora, f"Error corrigiendo SMB y SMN: {str(e)}")
        logger.error(f"Error en _corregir_smb_smn_empleados: {str(e)}", exc_info=True)


def _llenar_niveles_vacios_pos_activas(bitacora):
    """
    Ejecuta el Stored Procedure sp_llenar_niveles_vacios_pos_activas, el cual
    llena la columna `Nivel` de EMPLEADOS_COMPLETOS_SIG para empleados con
    posición activa y Nivel NULL, haciendo match contra rc_cat_cod_presupuestal
    por Código Presupuestal y Escala.
    """
    _append_log(
        bitacora,
        "Llenando Niveles vacíos en EMPLEADOS_COMPLETOS_SIG desde catálogo rc_cat_cod_presupuestal...",
    )
    try:
        with connection.cursor() as cursor:
            cursor.execute("CALL sp_llenar_niveles_vacios_pos_activas();")
            _append_log(bitacora, "Niveles vacíos actualizados exitosamente.")
    except Exception as e:
        _append_log(bitacora, f"Error llenando Niveles vacíos: {str(e)}", is_error=True)
        logger.error(f"Error en _llenar_niveles_vacios_pos_activas: {str(e)}", exc_info=True)


def _sincronizar_plazas_nivel_jerarquico(bitacora):
    """
    Siembra/actualiza cat_nivel_jerarquico_plaza desde las posiciones activas
    de MOV_POS (recién recargado por ZAFIRO): crea plazas nuevas y refresca
    `nvl_direc_origen`, sin tocar `nivel_jerarquico` (asignación manual) ni
    borrar plazas que hoy no estén activas.

    Debe correr ANTES de `_reaplicar_prioridad_nivel_jerarquico`: así, si una
    posición se dio de baja y se reactivó, el catálogo ya la trae de vuelta
    (con su `nivel_jerarquico` previo intacto, porque la fila nunca se borró)
    antes de que se cruce contra MOV_POS/EMPLEADOS_COMPLETOS_SIG.
    """
    from .nivel_jerarquico_sync import sincronizar_plazas_desde_mov_pos

    _append_log(bitacora, "Sincronizando cat_nivel_jerarquico_plaza desde MOV_POS...")
    try:
        stats = sincronizar_plazas_desde_mov_pos()
        _append_log(
            bitacora,
            f"Plazas sincronizadas: {stats['creadas']} creada(s), {stats['actualizadas']} actualizada(s) "
            f"de {stats['total_activas']} posición(es) activa(s).",
        )
    except Exception as e:
        _append_log(bitacora, f"Error sincronizando plazas de nivel jerárquico: {str(e)}", is_error=True)
        logger.error("Error en _sincronizar_plazas_nivel_jerarquico: %s", e, exc_info=True)


def _reaplicar_prioridad_nivel_jerarquico(bitacora):
    """
    Si hay una prioridad de nivel jerárquico configurada (checkbox en el
    frontend, ver CatNivelJerarquicoPlazaViewSet.aplicar_prioridad), la
    reaplica sobre las tablas recién recargadas por ZAFIRO: MOV_POS y
    EMPLEADOS_COMPLETOS_SIG se truncan y recargan completas en cada import,
    así que cualquier escritura manual anterior se pierde si no se reaplica
    aquí.
    """
    from .models import NivelJerarquicoPrioridadConfig
    from .nivel_jerarquico_sync import aplicar_prioridad_nivel_jerarquico

    config = NivelJerarquicoPrioridadConfig.objects.first()
    if not config or not config.fuente:
        return
    _append_log(bitacora, f"Reaplicando prioridad de nivel jerárquico ({config.fuente})...")
    try:
        stats = aplicar_prioridad_nivel_jerarquico(config.fuente)
        _append_log(
            bitacora,
            f"Prioridad reaplicada: {stats['empleados_actualizados']} empleado(s) en "
            f"EMPLEADOS_COMPLETOS_SIG, {stats['posiciones_actualizadas']} posición(es) en MOV_POS.",
        )
    except Exception as e:
        _append_log(bitacora, f"Error reaplicando prioridad de nivel jerárquico: {str(e)}", is_error=True)
        logger.error("Error en _reaplicar_prioridad_nivel_jerarquico: %s", e, exc_info=True)


def _reaplicar_celda_overrides_empleados(bitacora):
    """
    Reaplica las ediciones manuales de celda (CeldaOverride, ver
    plantilla/celda_override.py) sobre EMPLEADOS_COMPLETOS_SIG, que se acaba
    de truncar y recargar completa por el swap Blue-Green. Mismo motivo que
    `_reaplicar_prioridad_nivel_jerarquico`: sin esto, cualquier edición
    manual hecha desde el tab Detalle se perdería en la siguiente importación.
    """
    from .celda_override import aplicar_overrides_empleados_completos

    try:
        stats = aplicar_overrides_empleados_completos(bitacora)
        _append_log(
            bitacora,
            f"CeldaOverride reaplicados sobre EMPLEADOS_COMPLETOS_SIG: "
            f"{stats['aplicados']} aplicado(s), {stats['huerfanos']} huérfano(s).",
        )
    except Exception as e:
        _append_log(bitacora, f"Error reaplicando CeldaOverride: {str(e)}", is_error=True)
        logger.error("Error en _reaplicar_celda_overrides_empleados: %s", e, exc_info=True)


def _calcular_y_actualizar_vacancias(bitacora):
    """
    Ejecuta el Stored Procedure sp_obtener_todas_vacancias, el cual
    calcula la fecha de vacancia de cada posición activa e internamente
    hace el UPDATE del campo FECHA VACANCIA en MOV_POS.
    """
    _append_log(
        bitacora, "Calculando y asignando fechas de vacancia directamente en BD..."
    )
    try:
        with connection.cursor() as cursor:
            cursor.execute("CALL sp_obtener_todas_vacancias();")
            _append_log(
                bitacora, "Actualización de fechas de vacancia completada exitosamente."
            )
    except Exception as e:
        _append_log(bitacora, f"Error calculando fechas de vacancia: {str(e)}")
        logger.error(
            f"Error en _calcular_y_actualizar_vacancias: {str(e)}", exc_info=True
        )


def _actualizar_historico_alineacion_general(bitacora=None):
    """Recalcula el % de Alineación General (mismo cruce MOV_POS x
    EMPLEADOS_COMPLETOS_SIG que MovPosAlineacionView, ver plantilla/views.py)
    y hace upsert de la fila del día en ALINEACION_ORGANIZACIONAL_HISTORICO:
    1 fila por día (no un log por cada corrida de Celery cada 30 min), solo
    se actualiza si el % cambió respecto al ya guardado hoy."""
    from .models import AlineacionOrganizacionalHistorico
    from .views import _construir_dataset_alineacion, get_mov_pos_alineacion_stats

    try:
        dataset = _construir_dataset_alineacion()
        stats = get_mov_pos_alineacion_stats(dataset)
        porcentaje = stats["porcentaje_alineacion_general"]
        hoy = timezone.localdate()

        registro, created = AlineacionOrganizacionalHistorico.objects.get_or_create(
            fecha=hoy,
            defaults={
                "porcentaje_alineacion_general": porcentaje,
                "total_activas": stats["total_activas"],
                "total_alineadas": stats["total_alineadas"],
                "total_con_diferencias": stats["total_con_diferencias"],
            },
        )
        if not created and float(registro.porcentaje_alineacion_general) != porcentaje:
            registro.porcentaje_alineacion_general = porcentaje
            registro.total_activas = stats["total_activas"]
            registro.total_alineadas = stats["total_alineadas"]
            registro.total_con_diferencias = stats["total_con_diferencias"]
            registro.save()

        _append_log(
            bitacora,
            f"Histórico de alineación general actualizado: {porcentaje}% ({hoy.isoformat()})",
        )
    except Exception as e:
        logger.error("Error actualizando histórico de alineación general: %s", e, exc_info=True)
        _append_log(bitacora, f"Error actualizando histórico de alineación general: {e}", is_error=True)


# ---------------------------------------------------------------------------
# Tarea principal
# ---------------------------------------------------------------------------


@shared_task(
    bind=True,
    name="plantilla.tasks.importar_zafiro",
    max_retries=2,
    default_retry_delay=300,  # 5 minutos entre reintentos
)
def importar_zafiro(self):
    """
    Tarea Celery que orquesta la sincronización completa de ZAFIRO:

      1. Descarga e importa CSV de Posiciones (argIndex=1) → MOV_POS_STAGING
      2. Descarga e importa CSV de Empleados Completos (argIndex=6) →
         EMPLEADOS_COMPLETOS_SIG_STAGING
      3. Descarga e importa CSV de Empleados Bajas (argIndex=3) →
         BAJAS_SIG_STAGING
      4. Descarga e importa CSV de Historial Posición (argIndex=0) →
         cp_tbl_mov_completo_29_05_26_staging (conservando además los
         registros previos a 2026 que ya estaban en producción)

         Cada CSV es procesado por el corrector heurístico antes de la
         importación. Si `es_historico` (run después de las 23:00), cada
         lote también se guarda en su tabla *_HISTORICO.

      5. Swap atómico (Blue-Green) entre tablas activas y staging, y
         truncado de staging (que queda con los datos viejos).
      6. Post-proceso vía stored procedures: Nombre Puesto en MOV_POS,
         corrección de SMB/SMN en EMPLEADOS_COMPLETOS_SIG, y cálculo de
         fechas de vacancia.
      7. Sincroniza cat_nivel_jerarquico_plaza desde las posiciones activas
         de MOV_POS (siembra/actualiza `nvl_direc_origen`, conserva el
         `nivel_jerarquico` ya asignado) y reaplica la fuente de prioridad
         de nivel jerárquico fijada (si hay una).
      8. Regenera el Cuadro de Vacancia Diario.
      9. Publica evento en Redis para SSE e invalida cachés de
         dashboard/plantilla/reportes Excel.

    Todo el progreso y errores se registran en vivo en un ZafiroBitacora.

    Programada para correr cada 30 minutos (48 veces al día).
    En caso de error, reintenta hasta 2 veces con 5 minutos de espera.

    Usa un lock distribuido en Redis para garantizar que solo una instancia
    corra a la vez. Si otra instancia ya está en ejecución, esta tarea
    se descarta sin reintentar.
    """
    import redis as redis_lib

    # ── Lock distribuido: solo 1 instancia a la vez ────────────────────────
    LOCK_KEY = "lock:importar_zafiro"
    LOCK_TTL = 1800  # 30 minutos (safety net por si el worker muere)

    r = redis_lib.Redis.from_url(settings.CELERY_BROKER_URL)
    lock_acquired = r.set(LOCK_KEY, self.request.id, nx=True, ex=LOCK_TTL)

    if not lock_acquired:
        logger.warning(
            "importar_zafiro ya está en ejecución (lock=%s). "
            "Descartando tarea duplicada %s.",
            r.get(LOCK_KEY),
            self.request.id,
        )
        return {"status": "skipped", "reason": "otra instancia ya está en ejecución"}

    download_dir = settings.ZAFIRO_DOWNLOAD_DIR
    script_path = settings.ZAFIRO_SCRIPT_PATH
    resultados = {}

    logger.info("=== Iniciando tarea importar_zafiro ===")
    inicio = time.time()

    # Asegurar que el directorio de descarga existe
    Path(download_dir).mkdir(parents=True, exist_ok=True)

    from datetime import datetime

    ahora = datetime.now()
    # Si son las 23:00 o más tarde, consideramos que es el último run del día para el histórico
    es_historico = ahora.hour >= 23

    bitacora = ZafiroBitacora.objects.create(
        status="RUNNING",
        es_historico=es_historico,
        logs_en_vivo="",
        registros_posiciones=0,
        registros_completos=0,
        registros_bajas=0,
        registros_historial=0,
    )
    _append_log(bitacora, "=== Iniciando tarea importar_zafiro ===")

    try:
        # ── 1. Posiciones (argIndex=1) ─────────────────────────────────────
        _append_log(bitacora, "Descargando Posiciones (argIndex=1)...")

        csv_posiciones = _ejecutar_script_node(1, download_dir, script_path, bitacora)
        csv_posiciones_corregido = _corregir_csv(csv_posiciones, script_path, bitacora)
        total_posiciones = _importar_csv_posiciones(
            csv_posiciones_corregido, es_historico, bitacora
        )
        resultados["posiciones"] = total_posiciones

        # ── 2. Empleados Completos (argIndex=6) ────────────────────────────
        _append_log(bitacora, "Descargando Empleados Completos (argIndex=6)...")
        csv_completos = _ejecutar_script_node(6, download_dir, script_path, bitacora)
        csv_completos_corregido = _corregir_csv(csv_completos, script_path, bitacora)
        total_completos = _importar_csv_empleados_completos(
            csv_completos_corregido, es_historico, bitacora
        )
        resultados["empleados_completos"] = total_completos

        # ── 3. Empleados Bajas (argIndex=3) ────────────────────────────────
        _append_log(bitacora, "Descargando Empleados Bajas (argIndex=3)...")
        csv_bajas = _ejecutar_script_node(3, download_dir, script_path, bitacora)
        csv_bajas_corregido = _corregir_csv(csv_bajas, script_path, bitacora)
        total_bajas = _importar_csv_bajas(csv_bajas_corregido, es_historico, bitacora)
        resultados["empleados_bajas"] = total_bajas

        # ── 4. Historial Posición (argIndex=0) ─────────────────────────────
        _append_log(bitacora, "Descargando Historial Posición (argIndex=0)...")
        csv_historial = _ejecutar_script_node(0, download_dir, script_path, bitacora)
        csv_historial_corregido = _corregir_csv(csv_historial, script_path, bitacora)
        total_historial = _importar_csv_historial_posicion(
            csv_historial_corregido, es_historico, bitacora
        )
        resultados["historial_posicion"] = total_historial

        # Realizar el intercambio atómico (Blue-Green Swap)
        _swap_blue_green_tables(bitacora)

        # Reconstruir tabla resumen MOV_POS_LATEST (evita recalcular la
        # window function de "posiciones activas más recientes" por request)
        _materializar_mov_pos_latest(bitacora)

        # ── 5. Llenar Nombre Puesto en MOV_POS desde CAT_PTO_FUNC ──────────
        _llenar_nombre_puesto(bitacora)

        # ── 6. Corregir SMB y SMN en EMPLEADOS_COMPLETOS_SIG desde catálogo ──
        _corregir_smb_smn_empleados(bitacora)

        # ── 7. Llenar Niveles vacíos en EMPLEADOS_COMPLETOS_SIG ────────────
        _llenar_niveles_vacios_pos_activas(bitacora)

        # ── 8. Calcular y Actualizar Fechas de Vacancia ─────────────────────
        _calcular_y_actualizar_vacancias(bitacora)

        # ── 9. Sincronizar cat_nivel_jerarquico_plaza desde MOV_POS ────────
        _sincronizar_plazas_nivel_jerarquico(bitacora)

        # ── 10. Reaplicar prioridad de nivel jerárquico (si hay una fijada) ─
        _reaplicar_prioridad_nivel_jerarquico(bitacora)

        # ── 10.5. Reaplicar ediciones manuales (CeldaOverride) sobre
        # EMPLEADOS_COMPLETOS_SIG, que se acaba de truncar y recargar ───────
        _reaplicar_celda_overrides_empleados(bitacora)

        # ── 11. Generar/Actualizar Cuadro de Vacancia ──────────────────────
        _append_log(bitacora, "Generando/Actualizando Cuadro de Vacancia Diario...")
        try:
            from django.core.management import call_command

            call_command("generar_cuadro_vacancia")
        except Exception as e:
            _append_log(
                bitacora, f"Error generando cuadro vacancia: {e}", is_error=True
            )
            logger.error("Error generando cuadro vacancia: %s", e)

        duracion = round(time.time() - inicio, 1)

        # Guardar éxito en ZafiroBitacora
        bitacora.duracion_segundos = duracion
        bitacora.registros_posiciones = total_posiciones
        bitacora.registros_completos = total_completos
        bitacora.registros_bajas = total_bajas
        bitacora.registros_historial = total_historial
        bitacora.status = "EXITO"
        bitacora.save()

        # Invalida cache de posiciones activas y dashboard ANTES de publicar el
        # evento SSE: los clientes reaccionan al publish con un router.refresh()
        # casi inmediato, así que si el cache siguiera vivo en ese instante
        # podrían recargar y seguir viendo datos viejos.
        try:
            from django.core.cache import cache

            cache_keys = [
                "plantilla_1800_list_json",
                "latest_movpos_sub_ids",
                "active_position_codes",
                "active_employees_filtered",
                "plantilla_vacantes_por_nivel",
                "plantilla_vacantes_por_nivel_resumen",
                "empleados_completos_estatus_resumen",
                "empleados_completos_activos_detalle",
                "empleados_estatus_por_nivel_ua",
                "empleados_distribucion_geografica",
                "mov_pos_detalle",
                "mov_pos_card_stats",
                "mov_pos_ocupadas_set",
                "desglose_jerarquico",
                "desglose_jerarquico_ocupados",
                "bajas_sig_list",
                "bajas_motivos_pie",
                "bajas_historico",
                "mov_pos_alineacion_dataset",
                "mov_pos_alineacion_stats",
            ]
            cache.delete_many(cache_keys)

            # Invalidar por patrón las claves hasheadas por parámetros de request
            for pattern in ("*mov_stats_*", "*bajas_sig_list_*", "*empleados_completos_activos_detalle_*", "*excel_estatus_file_*"):
                for key in r.scan_iter(pattern):
                    r.delete(key)
        except Exception:
            logger.exception("Error al invalidar caché tras importación de ZAFIRO")

        # Publicar evento de actualización exitosa a Redis para tiempo real (SSE)
        try:
            from datetime import timedelta

            fecha_fin = bitacora.fecha_ejecucion
            if bitacora.duracion_segundos:
                fecha_fin += timedelta(seconds=bitacora.duracion_segundos)
            r.publish("zafiro_updates", fecha_fin.isoformat())
        except Exception as e:
            logger.error("Error al publicar evento de actualización en Redis: %s", e)

        # ── 12. Actualizar histórico diario de % Alineación General ────────
        _actualizar_historico_alineacion_general(bitacora)

        _append_log(
            bitacora,
            f"=== Tarea completada en {duracion}s | Posiciones: {total_posiciones} | Completos: {total_completos} | Bajas: {total_bajas} | Historial: {total_historial} ===",
        )

        return {
            "status": "ok",
            "duracion_segundos": duracion,
            **resultados,
        }

    except Exception as exc:
        duracion = round(time.time() - inicio, 1)
        bitacora.duracion_segundos = duracion
        bitacora.status = "ERROR"
        bitacora.error_message = str(exc)
        bitacora.save()
        _append_log(bitacora, f"Error en importar_zafiro: {exc}", is_error=True)
        raise self.retry(exc=exc)

    finally:
        # Siempre liberar el lock al terminar (éxito, error, o retry)
        r.delete(LOCK_KEY)


@shared_task(bind=True, name="plantilla.tasks.generar_excel_estatus_task")
def generar_excel_estatus_task(self, uas_param, levels_param, group_by):
    is_sync = not getattr(self, "request", None) or not getattr(
        self.request, "id", None
    )
    if is_sync:
        self.update_state = lambda *args, **kwargs: None

    import io

    from django.core.cache import cache
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from .views import obtener_posiciones_activas

    self.update_state(
        state="PROGRESS",
        meta={"progress": 5, "message": "Obteniendo datos de plantilla..."},
    )

    # Parsers
    selected_uas = [ua.strip() for ua in uas_param.split(",") if ua.strip()]
    selected_levels = [lvl.strip() for lvl in levels_param.split(",") if lvl.strip()]

    # 1. Obtener empleados activos (usando caché para optimizar velocidad)
    cache_key = "active_employees_filtered"
    employees_qs = cache.get(cache_key)

    if employees_qs is None:
        active_position_codes = set(obtener_posiciones_activas())
        all_employees = list(EmpleadosCompletosSig.objects.all().values())
        employees_qs = [
            emp for emp in all_employees if emp["posicion"] in active_position_codes
        ]
        cache.set(cache_key, employees_qs, 1200)

    self.update_state(
        state="PROGRESS",
        meta={"progress": 20, "message": "Preparando libro de Excel..."},
    )

    # Precalcular estatus amigable, nivel y UA si vienen de caché anterior sin procesar
    def get_friendly_status_fallback(val):
        if not val or val.strip() == "":
            return "Vacante"
        val_upper = val.strip().upper()
        if val_upper == "A":
            return "Activo"
        elif val_upper == "S":
            return "Suspendido"
        elif val_upper == "L":
            return "Licencia"
        elif val_upper == "P":
            return "Licencia Médica"
        else:
            return "Vacante"

    for emp in employees_qs:
        if "_ua_lower" not in emp:
            emp["_friendly_status"] = get_friendly_status_fallback(
                emp.get("estado_nomina")
            )
            emp["_level_lower"] = (emp.get("nivel") or "").strip().lower()
            emp["_ua_lower"] = (emp.get("unidad_administrativa") or "").strip().lower()

    # Configuración del Libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = True

    # Estilos
    ua_title_fill = PatternFill(
        start_color="FF1B365D", end_color="FF1B365D", fill_type="solid"
    )
    ua_title_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(
        start_color="FF2B4C7E", end_color="FF2B4C7E", fill_type="solid"
    )
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=9)
    link_font = Font(name="Segoe UI", size=9, color="FF0056B3", underline="single")
    bold_link_font = Font(
        name="Segoe UI", size=10, bold=True, color="FF0056B3", underline="single"
    )

    gold_side = Side(style="thin", color="FFBC955C")
    thin_border = Border(
        left=gold_side, right=gold_side, top=gold_side, bottom=gold_side
    )
    double_bottom_gold = Side(style="double", color="FFBC955C")
    total_border = Border(
        left=gold_side, right=gold_side, top=gold_side, bottom=double_bottom_gold
    )
    zebra_fill = PatternFill(
        start_color="FFF4F7FA", end_color="FFF4F7FA", fill_type="solid"
    )
    row_total_fill = PatternFill(
        start_color="FFEAEFF8", end_color="FFEAEFF8", fill_type="solid"
    )
    total_row_fill = PatternFill(
        start_color="FFE6ECF5", end_color="FFE6ECF5", fill_type="solid"
    )
    grand_total_fill = PatternFill(
        start_color="FFD5E1F2", end_color="FFD5E1F2", fill_type="solid"
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    status_list = ["Activo", "Vacante", "Suspendido", "Licencia", "Licencia Médica"]

    model_fields = EmpleadosCompletosSig._meta.fields
    field_names = [f.name for f in model_fields]
    detail_headers = [
        (f.db_column or f.name) if f.name != "id" else "ID" for f in model_fields
    ]
    last_col_letter = get_column_letter(len(model_fields))

    col_alignments = []
    for f in model_fields:
        if f.name in [
            "nombres",
            "ua",
            "ua2",
            "unidad_administrativa",
            "nombre_puesto_funcional",
            "departamento",
            "unidad_de_negocio",
        ]:
            col_alignments.append(align_left)
        else:
            col_alignments.append(align_center)

    def create_detail_sheet(title_text, employees_subset, parent_cell, parent_font):
        sheet_index = len(wb.sheetnames) + 1
        sheet_name = f"Det_{sheet_index}"
        det_ws = wb.create_sheet(title=sheet_name)
        det_ws.sheet_view.showGridLines = True

        back_cell = det_ws.cell(row=1, column=1, value="← Volver al Resumen")
        back_cell.font = link_font
        back_cell.hyperlink = f"#'Resumen'!{parent_cell.coordinate}"

        det_ws.merge_cells(f"A3:{last_col_letter}3")
        det_title = det_ws.cell(row=3, column=1, value=title_text)
        det_title.fill = ua_title_fill
        det_title.font = ua_title_font
        det_title.alignment = align_left

        col_widths = {}
        for d_col, dh_text in enumerate(detail_headers, 1):
            d_cell = det_ws.cell(row=5, column=d_col, value=dh_text)
            d_cell.fill = header_fill
            d_cell.font = header_font
            d_cell.alignment = align_center
            d_cell.border = thin_border
            col_widths[d_col] = len(str(dh_text))

        for r_idx, emp in enumerate(employees_subset, 6):
            emp_data = [
                str(emp.get(fname)) if emp.get(fname) is not None else ""
                for fname in field_names
            ]
            det_ws.append(emp_data)

            row_cells = det_ws[r_idx]
            for c_idx, cell in enumerate(row_cells, 1):
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = col_alignments[c_idx - 1]
                if r_idx % 2 == 0:
                    cell.fill = zebra_fill

                val_len = len(emp_data[c_idx - 1])
                if val_len > col_widths[c_idx]:
                    col_widths[c_idx] = val_len

        for c_idx, max_len in col_widths.items():
            col_letter = get_column_letter(c_idx)
            det_ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        parent_cell.hyperlink = f"#'{sheet_name}'!A1"
        parent_cell.font = parent_font

    current_row = 1

    if group_by == "level":
        ws.column_dimensions["A"].width = 30
        for col_idx in range(2, 8):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        total_levels = len(selected_levels)
        for idx, level in enumerate(selected_levels):
            lvl_lower = level.lower()

            # Envío de progreso
            progress_pct = 20 + int(70 * (idx / total_levels))
            self.update_state(
                state="PROGRESS",
                meta={
                    "progress": progress_pct,
                    "message": f"Generando hojas para Nivel: {level}...",
                },
            )

            if lvl_lower == "sin nivel":
                level_employees = [
                    emp for emp in employees_qs if not emp["_level_lower"]
                ]
            else:
                level_employees = [
                    emp for emp in employees_qs if emp["_level_lower"] == lvl_lower
                ]

            if selected_uas:
                uas_in_level = selected_uas
            else:
                uas_set = set()
                for emp in level_employees:
                    ua_name = emp.get("unidad_administrativa")
                    ua_clean = ua_name.strip() if ua_name else ""
                    if ua_clean:
                        uas_set.add(ua_clean)
                uas_in_level = sorted(list(uas_set))

            ws.merge_cells(
                start_row=current_row, start_column=1, end_row=current_row, end_column=8
            )
            title_cell = ws.cell(
                row=current_row, column=1, value=f"NIVEL: {level.upper()}"
            )
            title_cell.fill = ua_title_fill
            title_cell.font = ua_title_font
            title_cell.alignment = align_center
            current_row += 1

            headers = ["Unidad Administrativa"] + status_list + ["Total"]
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
            current_row += 1

            start_table_row = current_row
            level_employees_shown = []

            for ua_name in uas_in_level:
                ua_name_lower = ua_name.lower()
                cell_ua = ws.cell(row=current_row, column=1, value=ua_name)
                cell_ua.font = bold_font
                cell_ua.alignment = align_left
                cell_ua.border = thin_border

                ua_level_employees = [
                    emp for emp in level_employees if emp["_ua_lower"] == ua_name_lower
                ]
                level_employees_shown.extend(ua_level_employees)

                for status_idx, status_name in enumerate(status_list, 2):
                    subset = [
                        emp
                        for emp in ua_level_employees
                        if emp["_friendly_status"] == status_name
                    ]
                    count = len(subset)

                    cell_count = ws.cell(
                        row=current_row, column=status_idx, value=count
                    )
                    cell_count.font = regular_font
                    cell_count.alignment = align_center
                    cell_count.border = thin_border

                    if count > 0:
                        create_detail_sheet(
                            f"Nivel: {level} | UA: {ua_name} | Estatus: {status_name}",
                            subset,
                            cell_count,
                            link_font,
                        )

                cell_total = ws.cell(row=current_row, column=7)
                col_end_letter = get_column_letter(len(status_list) + 1)
                cell_total.value = f"=SUM(B{current_row}:{col_end_letter}{current_row})"
                cell_total.font = bold_font
                cell_total.alignment = align_center
                cell_total.border = thin_border
                cell_total.fill = row_total_fill

                total_count = len(ua_level_employees)
                if total_count > 0:
                    create_detail_sheet(
                        f"Nivel: {level} | UA: {ua_name} | Estatus: TODOS",
                        ua_level_employees,
                        cell_total,
                        bold_link_font,
                    )

                current_row += 1

            end_table_row = current_row - 1
            total_row_cell = ws.cell(row=current_row, column=1, value="Total")
            total_row_cell.font = bold_font
            total_row_cell.alignment = align_center
            total_row_cell.border = total_border
            total_row_cell.fill = total_row_fill

            for status_idx, status_name in enumerate(status_list, 2):
                col_let = get_column_letter(status_idx)
                cell_sum = ws.cell(row=current_row, column=status_idx)
                cell_sum.value = (
                    f"=SUM({col_let}{start_table_row}:{col_let}{end_table_row})"
                )
                cell_sum.font = bold_font
                cell_sum.alignment = align_center
                cell_sum.border = total_border
                cell_sum.fill = total_row_fill

                status_employees = [
                    emp
                    for emp in level_employees_shown
                    if emp["_friendly_status"] == status_name
                ]
                status_count = len(status_employees)

                if status_count > 0:
                    create_detail_sheet(
                        f"Nivel: {level} | UA: TODOS | Estatus: {status_name}",
                        status_employees,
                        cell_sum,
                        bold_link_font,
                    )

            cell_grand_total = ws.cell(row=current_row, column=7)
            col_end_letter = get_column_letter(len(status_list) + 1)
            cell_grand_total.value = (
                f"=SUM(B{current_row}:{col_end_letter}{current_row})"
            )
            cell_grand_total.font = bold_font
            cell_grand_total.alignment = align_center
            cell_grand_total.border = total_border
            cell_grand_total.fill = grand_total_fill

            grand_total_count = len(level_employees_shown)
            if grand_total_count > 0:
                create_detail_sheet(
                    f"Nivel: {level} | UA: TODOS | Estatus: TODOS",
                    level_employees_shown,
                    cell_grand_total,
                    bold_link_font,
                )

            current_row += 3

    else:
        total_uas = len(selected_uas)
        for idx, ua in enumerate(selected_uas):
            ua_lower = ua.lower()

            # Envío de progreso
            progress_pct = 20 + int(70 * (idx / total_uas))
            self.update_state(
                state="PROGRESS",
                meta={
                    "progress": progress_pct,
                    "message": f"Generando hojas para UA: {ua[:30]}...",
                },
            )

            ua_employees = [emp for emp in employees_qs if emp["_ua_lower"] == ua_lower]

            if selected_levels:
                levels_in_ua = selected_levels
            else:
                levels_set = set()
                for emp in ua_employees:
                    lvl = emp.get("nivel")
                    lvl_clean = lvl.strip() if lvl else ""
                    levels_set.add(lvl_clean if lvl_clean else "SIN NIVEL")
                levels_in_ua = sorted(list(levels_set))

            ws.merge_cells(
                start_row=current_row, start_column=1, end_row=current_row, end_column=8
            )
            title_cell = ws.cell(row=current_row, column=1, value=ua.upper())
            title_cell.fill = ua_title_fill
            title_cell.font = ua_title_font
            title_cell.alignment = align_center
            current_row += 1

            headers = ["Nivel"] + status_list + ["Total"]
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
            current_row += 1

            start_table_row = current_row
            ua_employees_shown = []

            for level_name in levels_in_ua:
                lvl_lower = level_name.lower()
                cell_lvl = ws.cell(row=current_row, column=1, value=level_name)
                cell_lvl.font = bold_font
                cell_lvl.alignment = align_center
                cell_lvl.border = thin_border

                if lvl_lower == "sin nivel":
                    level_employees = [
                        emp for emp in ua_employees if not emp["_level_lower"]
                    ]
                else:
                    level_employees = [
                        emp for emp in ua_employees if emp["_level_lower"] == lvl_lower
                    ]
                ua_employees_shown.extend(level_employees)

                for status_idx, status_name in enumerate(status_list, 2):
                    subset = [
                        emp
                        for emp in level_employees
                        if emp["_friendly_status"] == status_name
                    ]
                    count = len(subset)

                    cell_count = ws.cell(
                        row=current_row, column=status_idx, value=count
                    )
                    cell_count.font = regular_font
                    cell_count.alignment = align_center
                    cell_count.border = thin_border

                    if count > 0:
                        create_detail_sheet(
                            f"UA: {ua} | Nivel: {level_name} | Estatus: {status_name}",
                            subset,
                            cell_count,
                            link_font,
                        )

                cell_total = ws.cell(row=current_row, column=7)
                col_end_letter = get_column_letter(len(status_list) + 1)
                cell_total.value = f"=SUM(B{current_row}:{col_end_letter}{current_row})"
                cell_total.font = bold_font
                cell_total.alignment = align_center
                cell_total.border = thin_border
                cell_total.fill = row_total_fill

                total_count = len(level_employees)
                if total_count > 0:
                    create_detail_sheet(
                        f"UA: {ua} | Nivel: {level_name} | Estatus: TODOS",
                        level_employees,
                        cell_total,
                        bold_link_font,
                    )

                current_row += 1

            end_table_row = current_row - 1
            total_row_cell = ws.cell(row=current_row, column=1, value="Total")
            total_row_cell.font = bold_font
            total_row_cell.alignment = align_center
            total_row_cell.border = total_border
            total_row_cell.fill = total_row_fill

            for status_idx, status_name in enumerate(status_list, 2):
                col_let = get_column_letter(status_idx)
                cell_sum = ws.cell(row=current_row, column=status_idx)
                cell_sum.value = (
                    f"=SUM({col_let}{start_table_row}:{col_let}{end_table_row})"
                )
                cell_sum.font = bold_font
                cell_sum.alignment = align_center
                cell_sum.border = total_border
                cell_sum.fill = total_row_fill

                status_employees = [
                    emp
                    for emp in ua_employees_shown
                    if emp["_friendly_status"] == status_name
                ]
                status_count = len(status_employees)

                if status_count > 0:
                    create_detail_sheet(
                        f"UA: {ua} | Nivel: TODOS | Estatus: {status_name}",
                        status_employees,
                        cell_sum,
                        bold_link_font,
                    )

            cell_grand_total = ws.cell(row=current_row, column=7)
            col_end_letter = get_column_letter(len(status_list) + 1)
            cell_grand_total.value = (
                f"=SUM(B{current_row}:{col_end_letter}{current_row})"
            )
            cell_grand_total.font = bold_font
            cell_grand_total.alignment = align_center
            cell_grand_total.border = total_border
            cell_grand_total.fill = grand_total_fill

            grand_total_count = len(ua_employees_shown)
            if grand_total_count > 0:
                create_detail_sheet(
                    f"UA: {ua} | Nivel: TODOS | Estatus: TODOS",
                    ua_employees_shown,
                    cell_grand_total,
                    bold_link_font,
                )

            current_row += 3

        ws.column_dimensions["A"].width = 18
        for col_idx in range(2, 8):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    self.update_state(
        state="PROGRESS",
        meta={"progress": 95, "message": "Comprimiendo archivo Excel..."},
    )

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    file_data = output.read()

    # Guardar en caché permanente del reporte
    import hashlib

    raw_key = f"excel_estatus_file_{uas_param}_{levels_param}_{group_by}"
    cache_key_excel = (
        f"excel_estatus_file_{hashlib.md5(raw_key.encode('utf-8')).hexdigest()}"
    )
    cache.set(cache_key_excel, file_data, 1200)

    # Guardar temporalmente en caché el resultado de esta tarea de descarga
    if not is_sync:
        cache_key_result = f"excel_task_result_{self.request.id}"
        cache.set(cache_key_result, file_data, 300)  # Mantener 5 minutos

    self.update_state(
        state="SUCCESS", meta={"progress": 100, "message": "¡Completado!"}
    )
    return True
